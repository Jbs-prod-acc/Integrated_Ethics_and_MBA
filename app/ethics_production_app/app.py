
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()
from app.models import ArchivedEthicsForm, db_session, User, Rec, UserRole, UserInfo, FormA, FormB, FormC, FormD, FormUploads, Documents, FormARequirements, Watched
from app.models import db_session, User, Rec, UserRole, UserInfo, FormA, FormB, FormC, FormD, FormUploads, Documents, FormARequirements, Watched, UserActivityLog, LoginLog
from flask import jsonify
from flask import Flask, abort, flash, g, get_flashed_messages, make_response, render_template, request, redirect, url_for, session, jsonify, send_from_directory, send_file
from utils.helpers import generate_reset_token, send_email, validate_password
from utils.activity_logger import log_user_activity
from utils.document_files import (
    UploadValidationError,
    decode_legacy_binary,
    read_validated_upload,
    response_document_metadata,
)
import json
import base64
import math
import mimetypes
from db_queries import getFormAData, getSupervisorsList
import os
import zipfile
import pandas as pd
from werkzeug.utils import secure_filename
import os, traceback
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import extract
import io
import pdfkit
from werkzeug.utils import secure_filename
import secrets
import uuid
from datetime import date, datetime, timedelta, timezone
from datetime import time as dt_time
import pytz
from flask_cors import CORS
# Import CSRFProtect if available; provide a no-op fallback for environments
# where flask-wtf or its dependencies are incompatible (allows test client to run).
try:
    from flask_wtf.csrf import CSRFProtect
except Exception:
    class CSRFProtect:
        def __init__(self, app=None):
            self._app = app
        def init_app(self, app):
            return None
        def exempt(self, view):
            return view
from datetime import date
from sqlalchemy import desc,asc,cast ,Date,func,union_all,and_, not_, or_, extract, String
from sqlalchemy.orm import joinedload, defer
from collections import defaultdict
from mailtrap import configure_mail, send_email, mail
from flask_mail import Mail, Message
from flask import current_app
from sqlalchemy import create_engine
import time
from functools import wraps
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from sqlalchemy.exc import OperationalError
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
import traceback
from sqlalchemy import text

import sqlalchemy

app = Flask(__name__,static_folder='static')


@app.route('/favicon.ico')
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'favicon.ico',
        mimetype='image/vnd.microsoft.icon'
    )

# Apply Config (includes session security settings)

# Add robust DB connection option for Flask-SQLAlchemy
app.config.from_object('config.Config')
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}

# Configure CORS with restrictions from config
CORS(app, origins=app.config.get('CORS_ORIGINS', ['https://jbs-ethics.onrender.com']))

csrf = CSRFProtect(app)
configure_mail(app)


@app.errorhandler(413)
def upload_too_large(_error):
    max_mb = app.config.get('MAX_CONTENT_LENGTH', 536870912) // (1024 * 1024)
    message = (
        f"The selected documents exceed the {max_mb} MB total upload limit. "
        "Reduce the file sizes and try again."
    )
    if request.accept_mimetypes.best == 'application/json':
        return jsonify({'error': message}), 413
    category = 'admin-danger' if request.path.startswith('/admin/upload_student_docs') else 'danger'
    flash(message, category)
    return redirect(request.referrer or url_for('student_dashboard'))


@app.errorhandler(UploadValidationError)
def invalid_upload(error):
    if request.accept_mimetypes.best == 'application/json':
        return jsonify({'error': str(error)}), 400
    category = 'admin-danger' if request.path.startswith('/admin/upload_student_docs') else 'danger'
    flash(str(error), category)
    return redirect(request.referrer or url_for('student_dashboard'))

# --- Robust SQLAlchemy DB engine with pool_pre_ping ---
if hasattr(sqlalchemy, 'create_engine'):
    # If you use a custom engine, set pool_pre_ping=True
    import os
    db_url = os.getenv('DATABASE_URL') or app.config.get('SQLALCHEMY_DATABASE_URI')
    if db_url:
        try:
            engine = create_engine(db_url, pool_pre_ping=True)
        except Exception:
            pass

# Expose a `csrf_token()` helper to Jinja templates.
# If `flask_wtf` is installed we use `generate_csrf()` to produce a token,
# otherwise we provide a safe no-op that returns an empty string so templates
# calling `{{ csrf_token() }}` don't raise `UndefinedError` in test/dev envs.
try:
    from flask_wtf.csrf import generate_csrf
    app.jinja_env.globals['csrf_token'] = lambda: generate_csrf()
except Exception:
    app.jinja_env.globals['csrf_token'] = lambda: ''

# Add a Jinja filter to load JSON strings safely in templates (used for certificate conditions)
def _from_json_filter(s):
    try:
        return json.loads(s) if s else []
    except Exception:
        return []

app.jinja_env.filters['from_json'] = _from_json_filter

# Timezone helper - South African Standard Time (SAST) is UTC+2
def get_local_time():
    """Get current time in South African timezone (SAST - UTC+2)"""
    sa_tz = pytz.timezone('Africa/Johannesburg')
    return datetime.now(sa_tz)


def parse_admin_log_date(date_text):
    date_text = (date_text or '').strip()
    if not date_text:
        return None

    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_text, fmt)
        except ValueError:
            continue
    raise ValueError("Invalid date format")


def parse_html_date(date_text):
    if isinstance(date_text, datetime):
        return date_text
    if isinstance(date_text, date):
        return datetime.combine(date_text, dt_time.min)
    if date_text is None:
        return None

    date_text = str(date_text).strip()
    if not date_text:
        return None
    for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_text, fmt)
        except ValueError:
            continue
    raise ValueError("Invalid date format")


def create_user_activity_entry(user_id, action, page=None, target_user_id=None, details=None, duration_seconds=None, timestamp=None):
    if not user_id:
        return

    serialized_details = details
    if isinstance(details, (dict, list, tuple)):
        serialized_details = json.dumps(details, default=str)
    elif details is not None:
        serialized_details = str(details)

    entry = UserActivityLog(
        user_id=user_id,
        action=action,
        page=page,
        target_user_id=target_user_id,
        timestamp=timestamp or datetime.utcnow(),
        user_agent=request.user_agent.string if request else None,
        details=serialized_details,
        duration_seconds=duration_seconds
    )
    db_session.add(entry)
    db_session.commit()

def build_reviewer_name_lookup(forms):
    reviewer_ids = {
        reviewer_id
        for form in forms
        for reviewer_id in [
            getattr(form, 'reviewer_name1', None),
            getattr(form, 'reviewer_name2', None),
            getattr(form, 'form_reviewed_by', None),
            getattr(form, 'form_reviewed_by1', None),
        ]
        if reviewer_id
    }
    if not reviewer_ids:
        return {}

    reviewers = (
        db_session.query(User)
        .filter(User.user_id.in_(reviewer_ids))
        .all()
    )
    return {reviewer.user_id: reviewer.full_name for reviewer in reviewers}

# Generate strong secret key with: python -c "import secrets; print(secrets.token_hex(32))"
app.secret_key = os.getenv('SECRET_KEY')
if not app.secret_key:
    raise ValueError("SECRET_KEY environment variable is required!")

mail = Mail(app)

# =====================================================================================================
# DATABASE SESSION TEARDOWN - Clean up scoped_session after each request
# =====================================================================================================
@app.teardown_appcontext
def shutdown_session(exception=None):
    """Remove scoped_session at the end of each request to prevent connection issues"""
    from app.models import db_session
    db_session.remove()


AUTH_SESSION_KEYS = (
    'loggedin',
    'id',
    'name',
    'last_active',
    'role',
    'supervisor_role',
    'admin_role',
    'rec_role',
    'reviewer_role',
    'super_role',
    'active_forma_id',
)
ETHICS_SSO_SALT = 'mba-to-ethics-sso'


def clear_auth_session(*, clear_flashes=False):
    """Remove authentication state and optionally discard queued flash messages."""
    for key in AUTH_SESSION_KEYS:
        session.pop(key, None)
    if clear_flashes:
        session.pop('_flashes', None)


def get_current_user():
    """Return the authenticated user for the current session, if any."""
    user_id = session.get('id')
    if not user_id:
        return None
    return db_session.query(User).filter_by(user_id=user_id).first()


def role_value(user):
    role = getattr(user, 'role', None)
    return getattr(role, 'value', role)


def is_super_admin(user):
    return str(role_value(user) or '').upper() == 'SUPER_ADMIN'


def is_admin(user):
    return str(role_value(user) or '').upper() == 'ADMIN'


def can_manage_ethics_user(actor, target):
    if not actor or not target:
        return False
    actor_role = str(role_value(actor) or '').upper()
    target_role = str(role_value(target) or '').upper()
    if actor_role == 'SUPER_ADMIN':
        return True
    if actor_role == 'ADMIN':
        return target_role != 'SUPER_ADMIN'
    return False


def allowed_assignable_roles(actor):
    actor_role = str(role_value(actor) or '').upper()
    if actor_role == 'SUPER_ADMIN':
        return {'SUPER_ADMIN', 'ADMIN', 'SUPERVISOR', 'REVIEWER', 'REC'}
    if actor_role == 'ADMIN':
        return {'SUPERVISOR', 'REVIEWER', 'REC'}
    return set()


def redirect_to_shared_login(message=None):
    if message:
        flash(message, "danger")
    return redirect('/login?system=ethics')


def _complete_ethics_login(user, *, audit_action='login', audit_page='login', destination=''):
    clear_auth_session()
    session['loggedin'] = True
    session['id'] = user.user_id
    session['name'] = user.full_name
    session['last_active'] = datetime.utcnow().isoformat()
    session['role'] = user.role.value or 'student'

    db_session.add(UserActivityLog(
        user_id=user.user_id,
        action=audit_action,
        page=audit_page,
        timestamp=datetime.utcnow(),
        user_agent=request.user_agent.string
    ))
    db_session.commit()

    role = user.role.value or 'student'

    if destination == 'admin_upload_docs' and role in {'ADMIN', 'SUPER_ADMIN'}:
        if role == 'ADMIN':
            session['admin_role'] = 'ADMIN'
        else:
            session['super_role'] = 'SUPER_ADMIN'
        return redirect(url_for('admin_upload_student_docs'))

    if role == 'STUDENT':
        user_id = session.get('id')
        watched_video = db_session.query(Watched).filter_by(user_id=user_id).first()

        if watched_video:
            if user.supervisor_id:
                student_id = user.user_id
                for model in [FormA, FormB, FormC]:
                    student_details = db_session.query(model).filter_by(user_id=student_id).first()
                    if student_details:
                        return redirect(url_for('student_dashboard'))
                return render_template('ethics_pack.html', name=session['name'])
            if not user.supervisor_id and user.authenticate_student and str(user.authenticate_student).lower() not in ['false', '0', 'none']:
                return redirect(url_for('student_choose_supervisor'))
            flash("You are not yet Authenticated", "danger")
            return redirect(url_for('login_page'))

        if not user.supervisor_id and user.authenticate_student and str(user.authenticate_student).lower() not in ['false', '0', 'none']:
            return redirect(url_for('student_choose_supervisor'))
        return render_template('video.html')

    if role == 'SUPERVISOR':
        session['supervisor_role'] = 'SUPERVISOR'
        return redirect(url_for('supervisor_dashboard'))
    if role == 'ADMIN':
        session['admin_role'] = 'ADMIN'
        return redirect(url_for('chair_landing'))
    if role == 'REC':
        session['rec_role'] = 'REC'
        return redirect(url_for('rec_dashboard'))
    if role == 'REVIEWER':
        session['reviewer_role'] = 'REVIEWER'
        session['supervisor_role'] = 'REVIEWER'
        return redirect(url_for('review_dashboard'))
    if role == 'SUPER_ADMIN':
        session['super_role'] = 'SUPER_ADMIN'
        return redirect(url_for('chair_landing'))

    return render_template('video.html')


def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get('id'):
            flash("Your session has expired. Please log in again.", "warning")
            return redirect(url_for('login_page'))
        return view(*args, **kwargs)
    return wrapper


def role_required(*allowed_roles):
    allowed = {role.upper() for role in allowed_roles}

    def decorator(view):
        @wraps(view)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                flash("Your session has expired. Please log in again.", "warning")
                return redirect(url_for('login_page'))
            if str(role_value(user) or '').upper() not in allowed:
                abort(403)
            return view(*args, **kwargs)
        return wrapper
    return decorator


def has_supervisor_submitted_feedback(form):
    """Return True once the supervisor completed the current review cycle."""
    return any(
        marker not in (None, '', False)
        for marker in (
            getattr(form, 'supervisor_date', None),
            getattr(form, 'signature_date', None),
            getattr(form, 'ethics_signature_date', None),
            getattr(form, 'submitted_to_admin', None),
            getattr(form, 'rejected_or_accepted', None),
        )
    )


def has_current_reviewer_submitted_feedback(form, user_id):
    """Return True when the assigned reviewer already used a review slot."""
    if user_id is None:
        return False
    reviewer_id = str(user_id)
    return any(
        str(completed_by) == reviewer_id
        for completed_by in (
            getattr(form, 'form_reviewed_by', None),
            getattr(form, 'form_reviewed_by1', None),
        )
        if completed_by is not None
    )


def _find_latest_form_for_user(model, user_id):
    query = db_session.query(model).filter_by(user_id=user_id)
    if hasattr(model, 'submitted_at'):
        query = query.order_by(model.submitted_at.desc().nullslast(), model.created_at.desc().nullslast())
    elif hasattr(model, 'submission_date'):
        query = query.order_by(model.submission_date.desc().nullslast(), model.created_at.desc().nullslast())
    return query.first()


def _find_all_forms_for_user(model, user_id, *, options=None):
    query = db_session.query(model)
    if options:
        query = query.options(*options)
    query = query.filter_by(user_id=user_id)
    if hasattr(model, 'submitted_at'):
        query = query.order_by(
            model.submitted_at.is_(None).desc(),
            model.submitted_at.desc().nullslast(),
            model.created_at.desc().nullslast(),
        )
    elif hasattr(model, 'submission_date'):
        query = query.order_by(
            model.submission_date.is_(None).desc(),
            model.submission_date.desc().nullslast(),
            model.created_at.desc().nullslast(),
        )
    else:
        query = query.order_by(model.created_at.desc().nullslast())
    return query.all()


def _find_latest_editable_form_for_user(model, user_id, submission_field_name, *, options=None):
    base_query = db_session.query(model)
    if options:
        base_query = base_query.options(*options)
    base_query = base_query.filter(model.user_id == user_id)

    submission_field = getattr(model, submission_field_name)
    draft = (
        base_query
        .filter(submission_field.is_(None))
        .order_by(model.created_at.desc().nullslast())
        .first()
    )
    if draft:
        return draft

    return (
        base_query
        .order_by(submission_field.desc().nullslast(), model.created_at.desc().nullslast())
        .first()
    )


def has_reviewer_feedback(form):
    if not form:
        return False

    feedback_fields = (
        'review_date',
        'review_date1',
        'review_recommendation',
        'review_recommendation1',
        'review_form_comments',
        'review_form_comments1',
        'review_additional_comments',
        'review_additional_comments1',
        'form_review_comment',
        'form_review_comment1',
        'form_reviewed_by',
        'form_reviewed_by1',
    )
    return any(getattr(form, field_name, None) for field_name in feedback_fields)


def can_reuse_forma_draft(form):
    if not form:
        return False
    if getattr(form, 'submitted_at', None) is not None:
        return False
    return (not has_reviewer_feedback(form)) or is_student_correction_state(form)


def _copy_missing_review_slot_fields(source_form, target_form, source_suffix, target_suffix):
    slot_field_bases = (
        'review_date',
        'review_org_permission_status',
        'review_org_permission_comments',
        'review_waiver_status',
        'review_waiver_comments',
        'form_status',
        'review_form_status',
        'review_form_comments',
        'review_questions_status',
        'review_questions_comments',
        'review_consent_status',
        'review_consent_comments',
        'review_proposal_status',
        'review_proposal_comments',
        'review_additional_comments',
        'review_recommendation',
        'review_supervisor_signature',
        'review_signature_date',
        'form_review_comment',
        'form_reviewed_by',
        'review_status',
    )

    copied_any = False
    for base_name in slot_field_bases:
        source_field = f"{base_name}{source_suffix}"
        target_field = f"{base_name}{target_suffix}"
        if not hasattr(source_form, source_field) or not hasattr(target_form, target_field):
            continue

        source_value = getattr(source_form, source_field, None)
        target_value = getattr(target_form, target_field, None)
        if not source_value:
            continue

        if isinstance(source_value, bool):
            should_copy = bool(source_value) and not bool(target_value)
        else:
            should_copy = target_value in (None, '')

        if should_copy:
            setattr(target_form, target_field, source_value)
            copied_any = True

    return copied_any


def merge_reviewer_feedback_from_related_draft(form):
    if not form or not getattr(form, 'user_id', None):
        return form

    if is_submitted_form_record(form) and not is_student_correction_state(form):
        return form

    timestamp_field_name = None
    if hasattr(type(form), 'submitted_at'):
        timestamp_field_name = 'submitted_at'
    elif hasattr(type(form), 'submission_date'):
        timestamp_field_name = 'submission_date'

    if not timestamp_field_name or getattr(form, timestamp_field_name, None) is None:
        return form

    model = type(form)
    timestamp_column = getattr(model, timestamp_field_name)
    draft_candidates = (
        db_session.query(model)
        .filter(
            model.user_id == form.user_id,
            model.form_id != form.form_id,
            timestamp_column.is_(None),
        )
        .order_by(model.created_at.desc().nullslast())
        .all()
    )

    source_draft = next((draft for draft in draft_candidates if has_reviewer_feedback(draft)), None)
    if not source_draft:
        return form

    copied_any = False
    slot_configs = (
        ('', 'reviewer_name1', 'form_reviewed_by'),
        ('1', 'reviewer_name2', 'form_reviewed_by1'),
    )

    for source_suffix, source_reviewer_attr, source_reviewed_by_attr in slot_configs:
        reviewer_id = (
            getattr(source_draft, source_reviewed_by_attr, None)
            or getattr(source_draft, source_reviewer_attr, None)
        )
        if not reviewer_id:
            continue

        target_suffix = None
        if reviewer_id == getattr(form, 'reviewer_name1', None):
            target_suffix = ''
        elif reviewer_id == getattr(form, 'reviewer_name2', None):
            target_suffix = '1'
        elif reviewer_id == getattr(source_draft, source_reviewer_attr, None):
            target_suffix = source_suffix

        if target_suffix is None:
            continue

        if _copy_missing_review_slot_fields(source_draft, form, source_suffix, target_suffix):
            copied_any = True

    if copied_any and getattr(form, 'rejected_or_accepted', None) is False:
        form.rejected_or_accepted = True

    return form


def get_form_submission_field(model):
    if hasattr(model, 'submitted_at'):
        return getattr(model, 'submitted_at')
    if hasattr(model, 'submission_date'):
        return getattr(model, 'submission_date')
    return None


def is_submitted_form_record(form):
    if not form:
        return False
    return bool(
        getattr(form, 'submitted_at', None)
        or getattr(form, 'submission_date', None)
    )


def get_latest_student_form_record(model, user_id, *, options=None):
    query = db_session.query(model)
    if options:
        query = query.options(*options)
    query = query.filter(model.user_id == user_id)

    submission_field = get_form_submission_field(model)
    if submission_field is not None:
        query = query.order_by(
            submission_field.desc().nullslast(),
            model.created_at.desc().nullslast()
        )
    else:
        query = query.order_by(model.created_at.desc().nullslast())

    return query.first()


def student_form_has_meaningful_data(form):
    if not form:
        return False
    if is_submitted_form_record(form):
        return True

    key_fields = (
        'applicant_name',
        'student_number',
        'study_title',
        'project_title',
        'declaration_name',
        'full_name',
        'email',
        'email_address',
    )
    return any((getattr(form, field_name, None) or '').strip() for field_name in key_fields)


def has_blocking_student_form(model, user_id, *, options=None):
    form = get_latest_student_form_record(model, user_id, options=options)
    return student_form_has_meaningful_data(form)


def normalize_forma_sample_sizes(sample_sizes):
    """Accept positive integers or numeric intervals for Form A sample sizes."""
    cleaned_sizes = []
    for index, raw_value in enumerate(sample_sizes, start=1):
        value = str(raw_value or '').strip()
        if not value:
            return None, (
                f"Sample Size row {index} is required. Enter a number like 100 "
                "or an interval like 100-150."
            )
        compact_value = value.replace(' ', '')
        if compact_value.isdigit():
            if int(compact_value) < 1:
                return None, f"Sample Size row {index} must be greater than zero."
            cleaned_sizes.append(compact_value)
            continue
        parts = compact_value.split('-')
        if len(parts) == 2 and all(part.isdigit() for part in parts):
            start_value, end_value = map(int, parts)
            if start_value < 1 or end_value < 1:
                return None, f"Sample Size row {index} must use positive numbers."
            if end_value < start_value:
                return None, (
                    f"Sample Size row {index} must use an interval like 100-150, "
                    "where the second number is not smaller than the first."
                )
            if end_value > start_value * 2:
                return None, (
                    f"Sample Size row {index} is too wide. For example, an interval "
                    "starting at 100 may not end above 200."
                )
            cleaned_sizes.append(f"{start_value}-{end_value}")
            continue
        return None, (
            f"Sample Size row {index} may contain numbers and one hyphen only. "
            "Enter a number like 100 or an interval like 100-150."
        )
    return cleaned_sizes, None


def get_student_supervisor_or_flash(user, *, redirect_endpoint='student_dashboard'):
    if not user or not getattr(user, 'supervisor_id', None):
        flash("No supervisor is assigned to your account. Please contact admin.", "danger")
        return None

    supervisor = db_session.query(User).filter(User.user_id == user.supervisor_id).first()
    if not supervisor:
        flash("Supervisor not found. Please contact admin.", "danger")
        return None

    return supervisor


STUDENT_CORRECTION_STATUSES = {
    'Corrections Required',
    'Revisions required',
    'Resubmission Required',
    'Submitted to Student for Corrections',
}


def is_student_correction_state(form):
    if not form:
        return False

    if getattr(form, 'status', None) in STUDENT_CORRECTION_STATUSES:
        return True

    if bool(getattr(form, 'visible_to_student', False)):
        return True

    normalized_review_comments = {
        (getattr(form, 'form_review_comment', None) or '').strip().lower(),
        (getattr(form, 'form_review_comment1', None) or '').strip().lower(),
        (getattr(form, 'review_recommendation', None) or '').strip().lower(),
        (getattr(form, 'review_recommendation1', None) or '').strip().lower(),
    }
    if any(
        status == 'resubmission required' or 'approved with minor changes' in status
        for status in normalized_review_comments
    ):
        return True

    normalized_return_statuses = {
        (getattr(form, 'recommendation', None) or '').strip().lower(),
        (getattr(form, 'ethics_status', None) or '').strip().lower(),
        (getattr(form, 'form_supervisor_status', None) or '').strip().lower(),
    }
    return any(status in {'revisions required', 'revision required'} for status in normalized_return_statuses)


def is_student_form_locked(form):
    if not form:
        return False

    submission_timestamp = getattr(form, 'submitted_at', None) or getattr(form, 'submission_date', None)
    if not submission_timestamp:
        return False

    return not is_student_correction_state(form)


def get_student_submission_timestamp(form):
    if not form:
        return None
    return getattr(form, 'submitted_at', None) or getattr(form, 'submission_date', None)


def has_student_submission_timestamp(form):
    return get_student_submission_timestamp(form) is not None


def redirect_if_missing_student_submission(form, destination_label, redirect_endpoint, **route_values):
    if has_student_submission_timestamp(form):
        return None

    flash(
        f"This form cannot be sent to {destination_label} because the student's submission date is missing.",
        'danger'
    )
    return redirect(url_for(redirect_endpoint, **route_values))


def get_student_dashboard_status(form):
    if not form:
        return ''

    if getattr(form, 'status', None) == 'Submitted to Student for Corrections':
        return 'Form Was Returned. Please Review The Feedback and Resubmit'

    if getattr(form, 'status', None) == 'Corrections Required':
        return 'Corrections Required'

    if _has_issued_certificate(form):
        return 'Certificate Issued'

    if getattr(form, 'rec_status', None) == 'Approved':
        return 'Approved'

    if is_student_correction_state(form):
        return 'Form Was Returned. Please Review The Feedback and Resubmit'

    if getattr(form, 'submitted_to_rec', False):
        return 'Form Submitted To REC'

    if getattr(form, 'reviewer_name1', None) or getattr(form, 'reviewer_name2', None):
        return 'Form Submitted To Ethics Admin'

    if getattr(form, 'rejected_or_accepted', False) and getattr(form, 'supervisor_date', None):
        return 'Form Submitted To Ethics Admin'

    if getattr(form, 'supervisor_date', None):
        return 'Form Submitted To Ethics Admin'

    if get_student_submission_timestamp(form):
        return 'Form Submitted To Supervisor'

    return 'Draft saved - not yet submitted'


def _status_value(form_or_record, field_name, default=None):
    if isinstance(form_or_record, dict):
        return form_or_record.get(field_name, default)
    return getattr(form_or_record, field_name, default)


def _status_text(value):
    return value.strip() if isinstance(value, str) else value


def format_yes_no(value, empty='Not Applicable'):
    """Render legacy boolean/string answers consistently as Yes or No."""
    if value is None:
        return empty
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    normalized = str(value).strip().lower()
    if normalized in {'yes', 'true', '1', 'on', 'checked'}:
        return 'Yes'
    if normalized in {'no', 'false', '0', 'off'}:
        return 'No'
    if not normalized:
        return empty
    return str(value)


def _has_issued_certificate(form_or_record):
    certificate_issued = _status_value(form_or_record, 'certificate_issued')
    if isinstance(certificate_issued, str):
        return certificate_issued.strip().lower() not in ('', 'not issued', 'none', 'false', '0')
    return bool(certificate_issued)


def _has_received_certificate(form_or_record):
    return bool(_status_value(form_or_record, 'certificate_received'))


def has_reviewer_approval(form_or_record):
    """Return whether the single assigned reviewer has approved the form."""
    approved_statuses = {'Approved', 'Approved with Minor Changes'}
    assigned_reviewer = (
        _status_value(form_or_record, 'reviewer_name1')
        or _status_value(form_or_record, 'first_reviewer')
        or _status_value(form_or_record, 'reviewer_name2')
        or _status_value(form_or_record, 'second_reviewer')
    )
    if not assigned_reviewer:
        return False

    completed_reviews = [
        (
            _status_value(form_or_record, 'form_reviewed_by'),
            _status_text(
                _status_value(form_or_record, 'review_recommendation', '')
                or _status_value(form_or_record, 'first_reviewer_recommendation', '')
                or ''
            ),
        ),
        (
            _status_value(form_or_record, 'form_reviewed_by1'),
            _status_text(
                _status_value(form_or_record, 'review_recommendation1', '')
                or _status_value(form_or_record, 'second_reviewer_recommendation', '')
                or ''
            ),
        ),
    ]
    return any(
        reviewed_by == assigned_reviewer and recommendation in approved_statuses
        for reviewed_by, recommendation in completed_reviews
    )


def has_dual_reviewer_approval(form_or_record):
    reviewer_recommendations = [
        _status_text(_status_value(form_or_record, 'review_recommendation', '') or ''),
        _status_text(_status_value(form_or_record, 'review_recommendation1', '') or ''),
    ]
    reviewer_ids = [
        _status_value(form_or_record, 'form_reviewed_by'),
        _status_value(form_or_record, 'form_reviewed_by1'),
    ]
    return (
        all(reviewer_ids)
        and all(reviewer_recommendations)
        and all(value == 'Approved' for value in reviewer_recommendations)
    )


def _assigned_reviewer_count_any(form_or_record):
    if isinstance(form_or_record, dict):
        reviewer_values = [
            _status_value(form_or_record, 'reviewer_name1') or _status_value(form_or_record, 'first_reviewer'),
            _status_value(form_or_record, 'reviewer_name2') or _status_value(form_or_record, 'second_reviewer'),
        ]
        return len([value for value in reviewer_values if value])
    return assigned_reviewer_count(form_or_record)


def _completed_reviewer_count_any(form_or_record):
    if isinstance(form_or_record, dict):
        recommendations = [
            _status_value(form_or_record, 'first_reviewer_recommendation'),
            _status_value(form_or_record, 'second_reviewer_recommendation'),
        ]
        return len([value for value in recommendations if value])
    return completed_reviewer_count(form_or_record)


def get_workflow_stage(form_or_record):
    if not form_or_record:
        return 'draft'

    recommendation = _status_text(_status_value(form_or_record, 'recommendation', '') or '')
    status = _status_text(_status_value(form_or_record, 'status', '') or '')
    
    supervisor_status = _status_text(
         _status_value(form_or_record, 'recommendation')
        or _status_value(form_or_record, 'form_supervisor_status')
        or ''
    )
    rec_status = _status_text(_status_value(form_or_record, 'rec_status', '') or '')
    submission_timestamp = (
        _status_value(form_or_record, 'submitted_at')
        or _status_value(form_or_record, 'submission_date')
    )
    supervisor_date = (
        _status_value(form_or_record, 'supervisor_date')
        or _status_value(form_or_record, 'signature_date')
    )

    ethics_status = (
       _status_value(form_or_record, 'ethics_status')
    )

    normalized_review_comments = {
        (_status_text(_status_value(form_or_record, 'form_review_comment', '') or '') or '').lower(),
        (_status_text(_status_value(form_or_record, 'form_review_comment1', '') or '') or '').lower(),
    }
    normalized_return_statuses = {
        (_status_text(recommendation) or '').lower(),
        (_status_text(supervisor_status) or '').lower(),
        (_status_text(status) or '').lower(),
    }
    is_correction_state = (
        status in STUDENT_CORRECTION_STATUSES
        or bool(_status_value(form_or_record, 'visible_to_student', False))
        or 'resubmission required' in normalized_review_comments
        or any(item in {'revisions required', 'revision required'} for item in normalized_return_statuses)
    )

    if _has_issued_certificate(form_or_record) or rec_status in {
        'Approved'
        'Approved with suggestions without re-submission',
    }:
        return 'process-complete'

    if (
        is_correction_state
        or supervisor_status == 'Revisions required'
        or recommendation == 'Revisions required'
        or status == 'Revisions required'
        or ethics_status == 'Revisions required'
    ):
        return 'with-student-revisions'

    if _status_value(form_or_record, 'submitted_to_rec', False):
        return 'with-rec'

    if submission_timestamp and not supervisor_date:
        return 'with-supervisor'

    assigned_count = _assigned_reviewer_count_any(form_or_record)
    completed_count = _completed_reviewer_count_any(form_or_record)
    reviewer_recommendations = [
        _status_text(_status_value(form_or_record, 'first_reviewer_recommendation', '') or ''),
        _status_text(_status_value(form_or_record, 'second_reviewer_recommendation', '') or ''),
    ]

    if supervisor_date and assigned_count == 0:
        return 'with-ethics-admin'


    if assigned_count > 0 and completed_count >= assigned_count:
        if any(rec in {'Approved with Minor Changes', 'Resubmission Required', 'Reject'} for rec in reviewer_recommendations):
            return 'pending-revisions'
        return 'with-ethics-admin'

    if assigned_count > 0 and completed_count < assigned_count:
        return 'with-reviewers'

    if supervisor_date and (
        _status_value(form_or_record, 'submitted_to_admin', False)
        or _status_value(form_or_record, 'rejected_or_accepted', False)
    ):
        return 'with-ethics-admin'

    if submission_timestamp:
        return 'with-supervisor'
    
    return 'draft'


def get_workflow_location(form_or_record):
    stage = get_workflow_stage(form_or_record)
    stage_labels = {
        'process-complete': 'Process Complete',
        'with-rec': 'With REC Chair',
        'with-student-revisions': 'With Student (Revisions)',
        'with-supervisor': 'With Supervisor',
        'pending-reviewers': 'Pending Reviewers',
        'with-reviewers': 'With Reviewer',
        'pending-revisions': 'Pending Revisions',
        'with-ethics-admin': 'With Ethics Admin',
        'draft': 'Draft saved - not yet submitted',
    }
    return stage_labels.get(stage, 'Draft saved - not yet submitted')


def get_supervisor_dashboard_status(form):
    if not form:
        return ''

    reviewer_recommendations = {
        (_status_text(_status_value(form, 'review_recommendation', '') or '') or '').lower(),
        (_status_text(_status_value(form, 'review_recommendation1', '') or '') or '').lower(),
    }
    if is_student_correction_state(form):
        if any('approved with minor changes' in item for item in reviewer_recommendations):
            return 'Approved with minor changes'
        return 'Resubmission required'

    supervisor_date = (
        _status_value(form, 'supervisor_date')
        or _status_value(form, 'signature_date')
    )
    if supervisor_date:
        return 'With Ethics'

    if get_student_submission_timestamp(form):
        return 'Awaiting review'
    return 'Draft saved - not yet submitted'


def is_with_ethics(form):
    """Return whether supervisor review is complete and Ethics owns the form."""
    return get_supervisor_dashboard_status(form) == 'With Ethics'


def get_supervisor_row_status(form):
    stage = get_workflow_stage(form)
    if stage == 'process-complete':
        return 'certificate-issued'
    if stage in {'with-rec', 'with-ethics-admin', 'with-reviewers', 'pending-reviewers'}:
        return 'submitted-admin'
    if stage in {'with-student-revisions', 'pending-revisions'}:
        return 'rejected'
    return 'awaiting-review'


def get_reviewer_dashboard_status(form, reviewer_id=None):
    if reviewer_id and reviewer_id in {
        _status_value(form, 'form_reviewed_by'),
        _status_value(form, 'form_reviewed_by1'),
    }:
        return 'Reviewed'

    stage = get_workflow_stage(form)
    if stage == 'process-complete':
        return 'Certificate Issued'
    if stage == 'with-rec':
        return 'With REC Chair'
    if stage == 'with-ethics-admin':
        return 'With Ethics Admin'
    if stage == 'with-student-revisions':
        return 'With Student (Revisions)'
    if stage == 'pending-revisions':
        return 'Pending Revisions'
    if stage == 'pending-reviewers':
        return 'Pending Reviewers'
    if stage == 'with-reviewers':
        return 'Awaiting review'
    if stage == 'with-supervisor':
        return 'With Supervisor'
    return 'Awaiting for review'


def is_reviewed_by(form, reviewer_id):
    """Return whether this specific reviewer already submitted feedback."""
    return bool(
        form
        and reviewer_id
        and has_current_reviewer_submitted_feedback(form, reviewer_id)
    )


def get_ethics_dashboard_status(form):
    """Return the Ethics Admin queue status for a submitted form."""
    if not form:
        return ''

    if is_student_correction_state(form):
        return 'Form was sent back to student'

    assigned_count = _assigned_reviewer_count_any(form)
    completed_count = _completed_reviewer_count_any(form)
    if assigned_count == 0:
        return 'Awaiting review'
    if completed_count < assigned_count:
        return 'With reviewer'

    recommendations = [
        _status_text(_status_value(form, 'review_recommendation', '') or ''),
        _status_text(_status_value(form, 'review_recommendation1', '') or ''),
        _status_text(_status_value(form, 'first_reviewer_recommendation', '') or ''),
        _status_text(_status_value(form, 'second_reviewer_recommendation', '') or ''),
    ]
    normalized = {(item or '').lower() for item in recommendations if item}
    if any(item in {'resubmission required', 'reject'} for item in normalized):
        return 'Resubmission required'
    if any('approved with minor changes' in item for item in normalized):
        return 'Approved with minor changes'
    if 'approved' in normalized:
        return 'Approved'
    return 'Reviewed'


def get_reviewer_row_status(form, reviewer_id=None):
    stage = get_workflow_stage(form)
    if stage == 'process-complete':
        return 'certificate-issued'
    if stage in {'with-rec', 'with-ethics-admin'}:
        return 'submitted-admin'
    if stage in {'with-student-revisions', 'pending-revisions'}:
        return 'rejected'
    return 'awaiting-review'


def redirect_if_student_form_locked(form, form_label='This form'):
    if is_student_form_locked(form):
        flash(
            f"{form_label} has already been submitted to the supervisor and is locked until it is sent back to the student.",
            "warning",
        )
        return redirect(url_for("student_dashboard"))
    return None


def reset_form_review_feedback(form):
    if not form:
        return

    fields_to_clear = [
        'supervisor_feedback',
        'recommendation',
        'supervisor_signature',
        'supervisor_date',
        'ethics_signature_date',
        'signature_date',
        'review_form_status',
        'review_form_comments',
        'review_form_status1',
        'review_form_comments1',
        'review_org_permission_status',
        'review_org_permission_comments',
        'review_org_permission_status1',
        'review_org_permission_comments1',
        'review_waiver_status',
        'review_waiver_comments',
        'review_waiver_status1',
        'review_waiver_comments1',
        'review_questions_status',
        'review_questions_comments',
        'review_questions_status1',
        'review_questions_comments1',
        'review_consent_status',
        'review_consent_comments',
        'review_consent_status1',
        'review_consent_comments1',
        'review_proposal_status',
        'review_proposal_comments',
        'review_proposal_status1',
        'review_proposal_comments1',
        'review_additional_comments',
        'review_additional_comments1',
        'review_recommendation',
        'review_recommendation1',
        'review_supervisor_signature',
        'review_supervisor_signature1',
        'review_date',
        'review_date1',
        'review_signature_date',
        'review_signature_date1',
        'form_review_comment',
        'form_review_comment1',
        'form_reviewed_by',
        'form_reviewed_by1',
        'form_a_comment',
        'questions_comment',
        'consent_comment',
        'proposal_comment',
        'org_permission_comment',
        'waiver_comment',
    ]

    for field_name in fields_to_clear:
        if hasattr(form, field_name):
            setattr(form, field_name, None)

    for field_name in ('review_status', 'review_status1'):
        if hasattr(form, field_name):
            setattr(form, field_name, False)

    for field_name in (
        'reviewer_name1',
        'reviewer_name2',
        'submitted_to_admin',
        'submitted_to_reviewers',
        'submitted_to_rec',
        'rejected_or_accepted',
    ):
        if hasattr(form, field_name):
            setattr(form, field_name, False if field_name.startswith('submitted_to_') or field_name == 'rejected_or_accepted' else None)

    if hasattr(form, 'rec_status'):
        form.rec_status = None

    if hasattr(form, 'certificate_issued'):
        form.certificate_issued = None

    if hasattr(form, 'certificate_received'):
        form.certificate_received = False


def get_assigned_reviewer_ids(form):
    if not form:
        return []
    return [
        reviewer_id for reviewer_id in
        [getattr(form, 'reviewer_name1', None), getattr(form, 'reviewer_name2', None)]
        if reviewer_id
    ]


def get_completed_reviewer_ids(form):
    if not form:
        return []

    assigned_ids = set(get_assigned_reviewer_ids(form))
    completed_ids = []
    for reviewer_id in [getattr(form, 'form_reviewed_by', None), getattr(form, 'form_reviewed_by1', None)]:
        if reviewer_id and reviewer_id in assigned_ids and reviewer_id not in completed_ids:
            completed_ids.append(reviewer_id)
    return completed_ids


def assigned_reviewer_count(form):
    return len(get_assigned_reviewer_ids(form))


def completed_reviewer_count(form):
    return len(get_completed_reviewer_ids(form))


def has_all_required_reviews(form):
    assigned_count = assigned_reviewer_count(form)
    if assigned_count == 0:
        return False
    return completed_reviewer_count(form) >= assigned_count


def get_admin_reviewer_outcome(form):
    """Return the completed reviewer outcome that requires an admin decision."""
    if not form or not has_all_required_reviews(form):
        return ''

    completed_ids = set(get_completed_reviewer_ids(form))
    recommendations = []
    for reviewer_id, recommendation in (
        (getattr(form, 'form_reviewed_by', None), getattr(form, 'review_recommendation', None)),
        (getattr(form, 'form_reviewed_by1', None), getattr(form, 'review_recommendation1', None)),
    ):
        if reviewer_id in completed_ids and recommendation:
            recommendations.append(str(recommendation).strip().lower())

    if not recommendations:
        return ''
    if any('approved with minor changes' in value for value in recommendations):
        return 'approved_with_minor_changes'
    if all(value == 'approved' for value in recommendations):
        return 'approved'
    return 'other'


def apply_reviewer_recommendation_routing(form, recommendation):
    """Route a form immediately after its assigned reviewer completes a review."""
    if not has_all_required_reviews(form):
        return

    outcome = get_admin_reviewer_outcome(form)
    if outcome == 'approved':
        form.status = 'Approved'
        form.submitted_to_admin = True
        form.submitted_to_reviewers = False
        form.visible_to_student = False
        form.rejected_or_accepted = True
    elif outcome == 'approved_with_minor_changes':
        # A reviewer recommendation is advice to the ethics administrator.  The
        # administrator must still choose whether to return the form, escalate
        # it to REC, or issue a certificate.
        form.status = 'Reviewed - Pending Admin Decision'
        form.submitted_to_admin = True
        form.submitted_to_reviewers = False
        form.visible_to_student = False
        form.rejected_or_accepted = False


def apply_admin_correction_routing(form, recommendation, user_role):
    """Return an admin-reviewed form to the student without reviewer gating."""
    normalized_role = (user_role or '').strip().upper()
    normalized_recommendation = (recommendation or '').strip().lower()
    if normalized_role not in {'ADMIN', 'SUPER_ADMIN'}:
        return
    if normalized_recommendation not in {
        'approved with minor changes',
        'approved with minor changes, resubmission required',
        'resubmission required',
        'revisions required',
    }:
        return

    form.status = 'Submitted to Student for Corrections'
    form.submitted_to_admin = False
    form.submitted_to_reviewers = False
    form.visible_to_student = True
    form.rejected_or_accepted = False


def is_waiting_for_additional_reviewer(form):
    return assigned_reviewer_count(form) > 1 and completed_reviewer_count(form) < assigned_reviewer_count(form)


app.jinja_env.globals['assigned_reviewer_count'] = assigned_reviewer_count
app.jinja_env.globals['completed_reviewer_count'] = completed_reviewer_count
app.jinja_env.globals['has_all_required_reviews'] = has_all_required_reviews
app.jinja_env.globals['get_admin_reviewer_outcome'] = get_admin_reviewer_outcome
app.jinja_env.globals['is_waiting_for_additional_reviewer'] = is_waiting_for_additional_reviewer
app.jinja_env.globals['is_student_form_locked'] = is_student_form_locked
app.jinja_env.globals['is_student_correction_state'] = is_student_correction_state
app.jinja_env.globals['get_student_submission_timestamp'] = get_student_submission_timestamp
app.jinja_env.globals['get_student_dashboard_status'] = get_student_dashboard_status
app.jinja_env.globals['get_workflow_stage'] = get_workflow_stage
app.jinja_env.globals['get_workflow_location'] = get_workflow_location
app.jinja_env.globals['get_supervisor_dashboard_status'] = get_supervisor_dashboard_status
app.jinja_env.globals['is_with_ethics'] = is_with_ethics
app.jinja_env.globals['get_supervisor_row_status'] = get_supervisor_row_status
app.jinja_env.globals['get_reviewer_dashboard_status'] = get_reviewer_dashboard_status
app.jinja_env.globals['is_reviewed_by'] = is_reviewed_by
app.jinja_env.globals['get_reviewer_row_status'] = get_reviewer_row_status
app.jinja_env.globals['get_ethics_dashboard_status'] = get_ethics_dashboard_status
app.jinja_env.globals['has_reviewer_feedback'] = has_reviewer_feedback
app.jinja_env.globals['has_reviewer_approval'] = has_reviewer_approval
app.jinja_env.globals['has_dual_reviewer_approval'] = has_dual_reviewer_approval
app.jinja_env.globals['has_supervisor_submitted_feedback'] = has_supervisor_submitted_feedback
app.jinja_env.globals['has_current_reviewer_submitted_feedback'] = has_current_reviewer_submitted_feedback
app.jinja_env.filters['yes_no'] = format_yes_no


def can_access_form(user, form):
    """Object-level access for student forms and their uploaded documents."""
    if not user or not form:
        return False

    current_role = str(role_value(user) or '').upper()
    if current_role in {'ADMIN', 'SUPER_ADMIN', 'DEAN'}:
        return True

    if getattr(form, 'user_id', None) == user.user_id:
        return True

    form_owner = db_session.query(User).filter_by(user_id=getattr(form, 'user_id', None)).first()
    if form_owner and getattr(form_owner, 'supervisor_id', None) == user.user_id:
        return True

    if getattr(form, 'supervisor_email', None) and getattr(form, 'supervisor_email', None) == user.email:
        return True
    if getattr(form, 'supervisor_name', None) and current_role == 'SUPERVISOR' and getattr(form_owner, 'supervisor_id', None) == user.user_id:
        return True

    if getattr(form, 'reviewer_name1', None) == user.user_id:
        return True
    if getattr(form, 'reviewer_name2', None) == user.user_id:
        return True

    if current_role == 'REC' and getattr(form, 'submitted_to_rec', False):
        return True

    return False


def can_act_as_assigned_supervisor(user, form):
    """Allow supervisor decisions only for students assigned to this user."""
    if not user or not form:
        return False
    if str(role_value(user) or '').upper() not in {'SUPERVISOR', 'REVIEWER'}:
        return False
    form_owner = db_session.query(User).filter_by(
        user_id=getattr(form, 'user_id', None)
    ).first()
    return bool(
        form_owner and getattr(form_owner, 'supervisor_id', None) == user.user_id
    )


def can_access_as_assigned_reviewer(user, form):
    """Allow reviewer actions only for reviewers assigned to the form."""
    if not user or not form:
        return False
    if str(role_value(user) or '').upper() != 'REVIEWER':
        return False
    return (
        getattr(form, 'reviewer_name1', None) == user.user_id
        or getattr(form, 'reviewer_name2', None) == user.user_id
    )


def can_access_requirements(user, req):
    """Access control for FormARequirements documents shared by dashboards."""
    if not user or not req:
        return False

    current_role = str(role_value(user) or '').upper()
    if current_role in {'ADMIN', 'SUPER_ADMIN', 'DEAN'}:
        return True

    if getattr(req, 'user_id', None) == user.user_id:
        return True

    student = db_session.query(User).filter_by(user_id=getattr(req, 'user_id', None)).first()
    if student and getattr(student, 'supervisor_id', None) == user.user_id:
        return True

    for model in (FormA, FormB, FormC):
        form = _find_latest_form_for_user(model, getattr(req, 'user_id', None))
        if form and can_access_form(user, form):
            return True

    return False


FORM_FILE_FIELDS = {
    'A': {
        'proposal_path', 'proposal', 'permission_letter',
        'prior_clearance', 'ethics_evidence', 'pending_note'
    },
    'FORMA': {
        'proposal_path', 'proposal', 'permission_letter',
        'prior_clearance', 'ethics_evidence', 'pending_note'
    },
    'B': {
        'permission_letter', 'prior_clearance', 'ethics_evidence', 'proposal_path',
        'pending_note', 'private_permission_file', 'private_permission', 'pdf_file_path',
    },
    'FORMB': {
        'permission_letter', 'prior_clearance', 'ethics_evidence', 'proposal_path',
        'pending_note', 'private_permission_file', 'private_permission', 'pdf_file_path',
    },
    'C': {'files', 'proposal_path', 'pending_note'},
    'FORMC': {'files', 'proposal_path', 'pending_note'},
}


REQUIREMENT_FILE_FIELDS = {
    'proposal_path', 'proposal', 'permission_letter', 'ethics_evidence',
    'ethics_evidence_path', 'prior_clearance_path', 'prior_clearance',
    'prior_clearance1', 'need_jbs_clearance', 'need_jbs_clearance1',
    'research_tools_path', 'impact_assessment_path', 'participation_info_sheet',
    'pending_note', 'needs_permission_pending', 'files'
}


def _send_stored_document(data, filename=None, fallback_name="document"):
    """Serve document bytes using their detected format, not a stale extension."""
    data = decode_legacy_binary(data)
    mimetype, download_name, as_attachment = response_document_metadata(
        data, filename, fallback_name
    )
    return send_file(
        io.BytesIO(data),
        mimetype=mimetype,
        download_name=download_name,
        as_attachment=as_attachment,
    )


def _resolve_requirement_file_field(requirements, requested_field):
    """Return the stored field that contains an uploaded requirement document.

    Form C historically stored its proposal in ``files`` while the other upload
    screens stored the same document in ``proposal_path``.  Existing students
    can therefore have a valid proposal in either field.
    """
    if not requirements:
        return None

    candidates = (requested_field,)
    if requested_field == 'files':
        candidates = ('files', 'proposal_path')

    for candidate in candidates:
        filename = getattr(requirements, f"{candidate}_filename", None)
        data = getattr(requirements, candidate, None)
        if filename or data:
            return candidate
    return None


@app.template_global()
def requirement_file_available(requirements, requested_field):
    return _resolve_requirement_file_field(requirements, requested_field) is not None


@app.after_request
def add_no_cache_headers(response):
    """Prevent browsers from reusing cached auth pages after logout/history navigation."""
    if request.endpoint != 'static' and (
        request.endpoint in {'login_page', 'logout'} or session.get('id')
    ):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, private'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')
    response.headers.setdefault(
        'Content-Security-Policy-Report-Only',
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "img-src 'self' data: https:; "
        "font-src 'self' data: https://cdnjs.cloudflare.com; "
        "connect-src 'self'; "
        "frame-ancestors 'self';"
    )
    if request.is_secure:
        response.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    return response

# =====================================================================================================
# HELPER FUNCTION: Safe FormB Query (avoids binary column deserialization issues)
# =====================================================================================================
def safe_query_formb(query_builder):
    """
    Safely query FormB without triggering binary column deserialization errors.
    Converts query results to proxy objects with only necessary attributes.
    
    Args:
        query_builder: A callable that takes FormB and returns a query 
                      (e.g., lambda fb: db_session.query(fb).filter(...))
    Returns:
        List of FormBProxy objects with safe attributes
    """
    # Define only the columns we need (excluding problematic LargeBinary columns)
    safe_columns = [
        FormB.form_id,
        FormB.user_id,
        FormB.applicant_name,
        FormB.student_number,
        FormB.email,
        FormB.supervisor,
        FormB.supervisor_email,
        FormB.submitted_at,
        FormB.recommendation,
        FormB.supervisor_date,
        FormB.ethics_status,
        FormB.signature_date,
        FormB.review_supervisor_signature,
        FormB.review_date,
        FormB.review_supervisor_signature1,
        FormB.review_date1,
        FormB.created_at,
        FormB.declaration_date
    ]
    
    # Build the query with safe columns
    query = db_session.query(*safe_columns)
    
    # Apply the additional filters/ordering from the query_builder if provided
    if query_builder:
        # Get the original query to extract filters/ordering
        original_query = query_builder(FormB)
        # Copy statement attributes (whereami, order_by, etc.)
        if hasattr(original_query, 'statement'):
            stmt = original_query.statement
            if hasattr(stmt, '_where_criteria') and stmt._where_criteria:
                for criterion in stmt._where_criteria:
                    query = query.filter(criterion)
            if hasattr(stmt, '_order_by') and stmt._order_by:
                query = query.order_by(*stmt._order_by)
    
    results = query.all()
    
    # Convert tuples to proxy objects
    form_b_records = []
    for result in results:
        class FormBProxy:
            pass
        proxy = FormBProxy()
        proxy.form_id = result.form_id
        proxy.user_id = result.user_id
        proxy.applicant_name = result.applicant_name
        proxy.student_number = result.student_number
        proxy.email = result.email
        proxy.supervisor = result.supervisor
        proxy.supervisor_email = result.supervisor_email
        proxy.submitted_at = result.submitted_at
        proxy.recommendation = result.recommendation
        proxy.supervisor_date = result.supervisor_date
        proxy.ethics_status = result.ethics_status
        proxy.signature_date = result.signature_date
        proxy.review_supervisor_signature = result.review_supervisor_signature
        proxy.review_date = result.review_date
        proxy.review_supervisor_signature1 = result.review_supervisor_signature1
        proxy.review_date1 = result.review_date1
        proxy.created_at = result.created_at
        proxy.declaration_date = result.declaration_date
        form_b_records.append(proxy)
    
    return form_b_records

@app.route('/view_requirement_file')
@login_required
def view_requirement_file():
    # Retrieve all possible parameter names from both positional and query string
    # Since we removed positional from the route, url_for will pass them all as query params
    req_id = request.args.get('req_id')
    user_id = request.args.get('user_id')
    identifier = request.args.get('identifier')
    field = request.args.get('field')
    field_name = request.args.get('field_name')
    
    actual_id = req_id or user_id or identifier
    f_name = field or field_name
    
    if not actual_id or not f_name:
        return f"Missing parameters: req_id/user_id={actual_id}, field/field_name={f_name}", 400

    if f_name not in REQUIREMENT_FILE_FIELDS:
        abort(403)
        
    # Try finding by user_id first, then by primary key ID
    req = db_session.query(FormARequirements).filter(
        (FormARequirements.user_id == actual_id) | (FormARequirements.id == str(actual_id))
    ).first()
    
    if not req:
        return "Record not found", 404

    current_user = get_current_user()
    if not can_access_requirements(current_user, req):
        abort(403)
    
    # Use f_name from now on
    actual_field = _resolve_requirement_file_field(req, f_name) or f_name
    data = getattr(req, actual_field, None)
    
    # If data is a boolean (like ethics_evidence), it's not the file data. Try _path.
    if isinstance(data, bool) or data is None:
        if hasattr(req, f"{f_name}_path"):
            actual_field = f"{f_name}_path"
            data = getattr(req, actual_field)
            
    if not data:
        return "File content not found", 404

    if isinstance(data, memoryview):
        data = data.tobytes()

    # Determine filename
    filename = getattr(req, f"{actual_field}_filename", None) or \
               getattr(req, f"{f_name}_filename", None) or \
               getattr(req, f_name.replace('_path', '') + "_filename", None) or \
               getattr(req, f_name.replace('_sheet', '') + "_filename", None) or \
               getattr(req, f_name.replace('_file', '') + "_filename", None)

    # NEW: Support for file paths instead of BLOBs
    if isinstance(data, str) and not data.startswith('\\x'):
        # Check if it looks like a path (e.g., "uploads/form/...")
        potential_path = os.path.join(app.root_path, 'static', data.replace('\\', '/').strip())
        if os.path.exists(potential_path) and os.path.isfile(potential_path):
            # Use mimetypes to be sure
            import mimetypes
            mtype, _ = mimetypes.guess_type(potential_path)
            return send_file(potential_path, mimetype=mtype or 'application/pdf', as_attachment=False, download_name=filename or os.path.basename(potential_path))

    return _send_stored_document(data, filename, f_name)


# --- Send Back for Corrections endpoint for FormB ---
@app.route('/send_back_for_corrections_b/<id>', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def send_back_for_corrections_b(id):
    try:
        form = db_session.query(FormB).filter_by(form_id=id).first()
        if not form:
            flash('Form B not found.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_b'))
        if not has_all_required_reviews(form):
            flash('Form B can only be sent back after all assigned reviewers have submitted their reviews.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_b'))
        if get_admin_reviewer_outcome(form) != 'approved_with_minor_changes':
            flash('Only a form approved with minor changes can be returned for corrections.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_b'))
        feedback = (request.form.get('corrections_feedback') or '').strip()
        if not feedback:
            flash('A comment is required before sending Form B back for corrections.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_b'))
        form.status = 'Submitted to Student for Corrections'
        form.review_form_status = feedback
        form.visible_to_student = True
        db_session.commit()
        flash('Form B sent back for corrections.', 'warning')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash('Database error: {}'.format(str(e)), 'danger')
    return redirect(url_for('ethics_reviewer_committee_form_b'))

# --- Student resubmits corrected FormB ---
@app.route('/resubmit_formb/<id>', methods=['POST'])
def resubmit_formb(id):
    form_id = id
    user_id=session.get('id')
    formA = db_session.query(FormA).filter_by(user_id=user_id).first()
    if formA:
        print("[DEBUG] User has FormA, not permitted to submit FormB.")
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
    formC = db_session.query(FormC).filter_by(user_id=user_id).first()
    if formC:
        print("[DEBUG] User has FormC, not permitted to submit FormB.")
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
    if not user_id:
        print("[DEBUG] No user_id in session, unauthorized.")
        return jsonify({'error': 'Unauthorized'}), 401
    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor=db_session.query(User).filter(User.user_id == user.supervisor_id).first()
    print(f"[DEBUG] User: {user}")
    print(f"[DEBUG] Supervisor: {supervisor}")
    # Check if form exists in database - if yes UPDATE (mark as resubmitted), if no error
    form = db_session.query(FormB).filter(FormB.user_id==user_id, FormB.form_id==form_id).first()
    if not form:
        flash("Form not found for resubmission.", "danger")
        return redirect(url_for("student_dashboard"))

    if request.method == 'POST':
        if form.status not in ('Corrections Required', 'Submitted to Student for Corrections'):
            flash("Form B is not awaiting corrections.", "danger")
            return redirect(url_for("student_dashboard"))
        # Ensure reviewers are assigned before resubmission
        if assigned_reviewer_count(form) < 1:
            flash("You cannot resubmit Form B until at least one reviewer is assigned.", "danger")
            return redirect(url_for("student_dashboard"))
        # Mark as resubmitted, update timestamp, status, etc.
        form.ethics_status = None
        form.form_supervisor_status = "Resubmitted"
        form.submitted_at = get_local_time()
        form.status = "Resubmitted"
        form.visible_to_student = False
        reset_form_review_feedback(form)
        db_session.commit()
        flash("Form B resubmitted successfully.", "success")
        return redirect(url_for("student_dashboard"))


def preserve_single_reviewer_assignment(new_form, previous_form):
    """Keep the new version's assignment or inherit it from the previous version."""
    if not new_form:
        return False

    assigned_reviewer = (
        getattr(new_form, 'reviewer_name1', None)
        or getattr(new_form, 'reviewer_name2', None)
    )
    if not assigned_reviewer and previous_form:
        assigned_reviewer = (
            getattr(previous_form, 'reviewer_name1', None)
            or getattr(previous_form, 'reviewer_name2', None)
        )

    changed = (
        getattr(new_form, 'reviewer_name1', None) != assigned_reviewer
        or getattr(new_form, 'reviewer_name2', None) is not None
    )
    new_form.reviewer_name1 = assigned_reviewer
    new_form.reviewer_name2 = None
    return changed


def inherit_previous_reviewers(new_form, model, user_id, order_column):
    """
    When a new version of a form is created, carry forward the latest assigned
    reviewers from the student's previous version so version history is kept
    without losing reviewer assignments.
    """
    previous_forms = (
        db_session.query(model)
        .filter(model.user_id == user_id, order_column.isnot(None))
        .order_by(order_column.desc().nullslast(), model.created_at.desc().nullslast())
        .all()
    )

    for previous_form in previous_forms:
        reviewer_name1 = getattr(previous_form, 'reviewer_name1', None)
        reviewer_name2 = getattr(previous_form, 'reviewer_name2', None)
        if reviewer_name1 or reviewer_name2:
            preserve_single_reviewer_assignment(new_form, previous_form)
            break


def backfill_reviewer_from_previous_version(current_form, model, order_column):
    """Persist the newest prior reviewer assignment on an existing current version."""
    if not current_form:
        return False
    if getattr(current_form, 'reviewer_name1', None) or getattr(current_form, 'reviewer_name2', None):
        return False

    previous_form = (
        db_session.query(model)
        .filter(
            model.user_id == current_form.user_id,
            model.form_id != current_form.form_id,
            or_(
                model.reviewer_name1.isnot(None),
                model.reviewer_name2.isnot(None),
            ),
        )
        .order_by(
            order_column.desc().nullslast(),
            model.created_at.desc().nullslast(),
        )
        .first()
    )
    return preserve_single_reviewer_assignment(current_form, previous_form)


def get_reviewers_for_ethics_assignment(current_form, model, order_column):
    """
    On the ethics assignment screen, prefer reviewers already saved on the
    current version. If the current version is a fresh resubmission with no
    reviewers yet, fall back to the most recent previous version for the same
    student.
    """
    if not current_form:
        return []

    current_ids = [
        reviewer_id for reviewer_id in
        [getattr(current_form, 'reviewer_name1', None), getattr(current_form, 'reviewer_name2', None)]
        if reviewer_id
    ]
    if current_ids:
        return current_ids[:1]

    previous_forms = (
        db_session.query(model)
        .filter(
            model.user_id == current_form.user_id,
            model.form_id != current_form.form_id,
            order_column.isnot(None),
        )
        .order_by(order_column.desc().nullslast(), model.created_at.desc().nullslast())
        .all()
    )

    for previous_form in previous_forms:
        previous_ids = [
            reviewer_id for reviewer_id in
            [getattr(previous_form, 'reviewer_name1', None), getattr(previous_form, 'reviewer_name2', None)]
            if reviewer_id
        ]
        if previous_ids:
            return previous_ids[:1]

    return []

# --- Send Back for Corrections endpoint for FormC ---
@app.route('/send_back_for_corrections_c/<id>', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def send_back_for_corrections_c(id):
    try:
        form = db_session.query(FormC).filter_by(form_id=id).first()
        if not form:
            print('Form C not found.')
            flash('Form C not found.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_c'))
        if not has_all_required_reviews(form):
            flash('Form C can only be sent back after all assigned reviewers have submitted their reviews.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_c'))
        if get_admin_reviewer_outcome(form) != 'approved_with_minor_changes':
            flash('Only a form approved with minor changes can be returned for corrections.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_c'))
        print(f"Form C status before corrections: {form.status}")
        feedback = (request.form.get('corrections_feedback') or '').strip()
        if not feedback:
            flash('A comment is required before sending Form C back for corrections.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_c'))
        form.status = 'Submitted to Student for Corrections'
        form.review_form_status = feedback
        form.visible_to_student = True
        
        form.rejected_or_accepted = False
        db_session.commit()
        print('Form C sent back for corrections.')
        flash('Form C sent back for corrections.', 'warning')
    except SQLAlchemyError as e:
        db_session.rollback()
        print(f"❌ Database error in send_back_for_corrections_c: {str(e)}")
        flash('Database error: {}'.format(str(e)), 'danger')
    return redirect(url_for('ethics_reviewer_committee_form_c'))

# --- Student resubmits corrected FormC ---
@app.route('/resubmit_formc/<id>', methods=['POST'])
def resubmit_formc(id):
    try:
        form = db_session.query(FormC).filter_by(form_id=id).first()
        if not form:
            print('Form C not found.')
            flash('Form C not found.', 'danger')
            return redirect(url_for('student_dashboard'))
        print(f"Form C status before resubmit: {form.status}")
        if form.status not in ('Corrections Required', 'Submitted to Student for Corrections'):
            print('Form C is not awaiting corrections.')
            flash('Form C is not awaiting corrections.', 'danger')
            return redirect(url_for('student_dashboard'))
        # Ensure reviewers are assigned before resubmission
        if assigned_reviewer_count(form) < 1:
            flash("You cannot resubmit Form C until at least one reviewer is assigned.", "danger")
            return redirect(url_for("student_dashboard"))
        # Update form fields from student input as needed
        form.submission_date = get_local_time()
        form.ethics_status = None
        form.form_supervisor_status = "Resubmitted"
        form.status = 'Resubmitted'
        form.visible_to_student = False
        reset_form_review_feedback(form)
        db_session.commit()
        print('Form C resubmitted to admin and supervisor.')
        flash('Form C resubmitted to admin and supervisor.', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        print(f"❌ Database error in resubmit_formc: {str(e)}")
        flash('Database error: {}'.format(str(e)), 'danger')
    return redirect(url_for('student_dashboard'))  

    

@app.route('/student_autosave_forma', methods=['POST'])
@csrf.exempt
# Autosaves a logged-in student's Form A draft without submitting the form.
def student_autosave_forma():
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    _apply_forma_autosave_payload(form, request.form, section='all', include_declaration=False)

    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500



# --- AUTOSAVE ENDPOINT FOR FORM B ---
def _get_or_create_formb_draft(user_id, form_data):
    explicit_form_id = form_data.get('formb_id') or form_data.get('form_id') or ''
    requested_form_id = (
        explicit_form_id
        or session.get('active_formb_id')
        or ''
    )
    requested_form_id = str(requested_form_id).strip()
    if requested_form_id:
        requested_form = (
            db_session.query(FormB)
            .options(
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file),
            )
            .filter(FormB.form_id == requested_form_id, FormB.user_id == user_id)
            .first()
        )
        if requested_form:
            if requested_form.submitted_at is None:
                session['active_formb_id'] = requested_form.form_id
                return requested_form, None
            if is_student_correction_state(requested_form):
                return _get_or_create_formb_resubmission_draft(user_id, requested_form)
            if explicit_form_id:
                return None, (
                    jsonify({
                        'success': False,
                        'error': 'Autosave ignored because this Form B has already been submitted.',
                    }),
                    409,
                )

    form = (
        db_session.query(FormB)
        .filter(FormB.user_id == user_id, FormB.submitted_at.is_(None))
        .order_by(FormB.created_at.desc().nullslast())
        .first()
    )
    if form:
        session['active_formb_id'] = form.form_id
        return form, None

    latest_submission = (
        db_session.query(FormB)
        .options(*FORMB_DEFERRED_UPLOADS)
        .filter(FormB.user_id == user_id, FormB.submitted_at.isnot(None))
        .order_by(FormB.submitted_at.desc(), FormB.created_at.desc().nullslast())
        .first()
    )
    if latest_submission and is_student_correction_state(latest_submission):
        return _get_or_create_formb_resubmission_draft(user_id, latest_submission)

    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user:
        return None, (jsonify({'success': False, 'error': 'Unauthorized'}), 401)

    supervisor = db_session.query(User).filter(User.user_id == user.supervisor_id).first()

    form = FormB(
        user_id=user_id,
        applicant_name=form_data.get('applicant_name', ''),
        student_number=form_data.get('student_number', ''),
        institution=form_data.get('institution', 'University of Johannesburg'),
        department=form_data.get('department', ''),
        degree=form_data.get('degree', ''),
        study_title=form_data.get('study_title', ''),
        mobile=form_data.get('mobile', ''),
        email=(user.email or ''),
        supervisor=(supervisor.full_name if supervisor else ''),
        supervisor_email=(supervisor.email if supervisor else ''),
        submitted=False,
    )
    db_session.add(form)
    db_session.flush()
    session['active_formb_id'] = form.form_id
    return form, None

def _apply_formb_public_domain_answers(form, form_data):
    """Save only the Form B conditional answers that match their Yes/No choices."""
    answer = form_data.get('data_public')
    if answer is not None:
        is_public = str(answer).strip().lower() in ('yes', 'true', '1', 'on')
        form.data_public = is_public
        if is_public:
            form.public_evidence = None
            form.access_conditions = (form_data.get('access_conditions') or '').strip() or None
        else:
            form.public_evidence = (form_data.get('public_evidence') or '').strip() or None
            form.access_conditions = None

    personal_info_answer = form_data.get('personal_info')
    if personal_info_answer is not None:
        contains_personal_info = str(personal_info_answer).strip().lower() in (
            'yes', 'true', '1', 'on'
        )
        form.personal_info = contains_personal_info
        form.personal_info_comment = (
            (form_data.get('personal_info_comment') or '').strip() or None
            if contains_personal_info
            else 'Not Applicable'
        )

    permission_answer = form_data.get('private_permission')
    if permission_answer is not None:
        permission_required = str(permission_answer).strip().lower() in ('yes', 'true', '1', 'on')
        form.private_permission = permission_required
        if permission_required:
            form.permission_details = (form_data.get('permission_details') or '').strip() or None
        else:
            form.permission_details = None
            form.private_permission_file = None
            form.private_permission_filename = None

    clear_permission_file = str(
        form_data.get('clear_private_permission_file') or ''
    ).strip().lower() in ('yes', 'true', '1', 'on')
    if clear_permission_file:
        form.private_permission_file = None
        form.private_permission_filename = None


# Autosaves a logged-in student's Form B draft without submitting the form.
@app.route('/student_autosave_formb', methods=['POST'])
@csrf.exempt
def student_autosave_formb():
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    form, error_response = _get_or_create_formb_draft(user_id, request.form)
    if error_response:
        return error_response

    def str_to_bool(val):
        if isinstance(val, bool):
            return val
        if val is None:
            return None
        if isinstance(val, str):
            if val.lower() in ['yes', 'true', '1', 'on', 'checked']:
                return True
            if val.lower() in ['no', 'false', '0', 'off', '']:
                return False
        return val

    boolean_fields = [
        'data_public',
        'personal_info',
        'private_permission',
        'shortcomings_reported',
        'methodology_alignment'
    ]
    declaration_fields = {
        'declaration_name',
        'applicant_signature',
        'full_name',
        'declaration_date',
        'submission_date',
    }

    data = request.form.to_dict()
    for key, value in data.items():
        if key not in declaration_fields and hasattr(form, key):
            if key in boolean_fields:
                setattr(form, key, str_to_bool(value))
            elif key in ['supervisor_date', 'supervisor_signature']:
                # Set to None if empty string, to avoid invalid timestamp errors
                setattr(form, key, value if value else None)
            else:
                setattr(form, key, value)
    _apply_formb_public_domain_answers(form, request.form)
    permission_file = request.files.get('private_permission_file')
    if form.private_permission and permission_file and permission_file.filename:
        form.private_permission_file = permission_file.read()
        form.private_permission_filename = permission_file.filename
    try:
        db_session.commit()
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        db_session.rollback()
        print('\n' + '='*60)
        print('❌ Exception in student_autosave_formb:', str(e))
        print('--- TRACEBACK BELOW ---')
        traceback.print_exc()
        print('='*60 + '\n')
        return jsonify({'success': False, 'error': str(e)}), 500

# --- AUTOSAVE ENDPOINT FOR FORM C ---
def _get_or_create_formc_draft(user_id, form_data):
    explicit_form_id = form_data.get('formc_id') or form_data.get('form_id') or ''
    requested_form_id = (
        explicit_form_id
        or session.get('active_formc_id')
        or ''
    )
    requested_form_id = str(requested_form_id).strip()
    if requested_form_id:
        requested_form = (
            db_session.query(FormC)
            .filter(FormC.form_id == requested_form_id, FormC.user_id == user_id)
            .first()
        )
        if requested_form:
            if requested_form.submission_date is None:
                session['active_formc_id'] = requested_form.form_id
                return requested_form, None
            if is_student_correction_state(requested_form):
                return _get_or_create_formc_resubmission_draft(user_id, requested_form)
            if explicit_form_id:
                return None, (
                    jsonify({
                        'success': False,
                        'error': 'Autosave ignored because this Form C has already been submitted.',
                    }),
                    409,
                )

    form = (
        db_session.query(FormC)
        .filter(FormC.user_id == user_id, FormC.submission_date.is_(None))
        .order_by(FormC.created_at.desc().nullslast())
        .first()
    )
    if form:
        session['active_formc_id'] = form.form_id
        return form, None

    latest_submission = (
        db_session.query(FormC)
        .filter(FormC.user_id == user_id, FormC.submission_date.isnot(None))
        .order_by(FormC.submission_date.desc(), FormC.created_at.desc().nullslast())
        .first()
    )
    if latest_submission and is_student_correction_state(latest_submission):
        return _get_or_create_formc_resubmission_draft(user_id, latest_submission)

    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user:
        return None, (jsonify({'success': False, 'error': 'Unauthorized'}), 401)

    supervisor = db_session.query(User).filter(User.user_id == user.supervisor_id).first()

    form = FormC(
        user_id=user_id,
        applicant_name=form_data.get('applicant_name', ''),
        student_number=form_data.get('student_number', ''),
        institution=form_data.get('institution', 'University of Johannesburg'),
        department=form_data.get('department', ''),
        degree=form_data.get('degree', ''),
        project_title=form_data.get('project_title', ''),
        mobile_number=form_data.get('mobile_number', ''),
        email_address=(user.email or ''),
        supervisor_name=(supervisor.full_name if supervisor else ''),
        supervisor_email=(supervisor.email if supervisor else ''),
        submitted=False,
    )
    db_session.add(form)
    db_session.flush()
    session['active_formc_id'] = form.form_id
    return form, None


@app.route('/student_autosave_formc', methods=['POST'])
@csrf.exempt
# Autosaves a logged-in student's Form C draft without submitting the form.
def student_autosave_formc():
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    form, error_response = _get_or_create_formc_draft(user_id, request.form)
    if error_response:
        return error_response

    def str_to_bool(val):
        if isinstance(val, bool):
            return val
        if val is None:
            return None
        if isinstance(val, str):
            if val.lower() in ['yes', 'true', '1', 'on', 'checked']:
                return True
            if val.lower() in ['no', 'false', '0', 'off', '']:
                return False
        return val

    boolean_fields = [
        'vulnerable',
        'age_under_18_or_over_65',
        'uj_employees',
        'non_vulnerable_context',
        'non_english',
        'own_students',
        'poverty',
        'no_education',
        'consent_violation',
        'discomfiture',
        'deception',
        'sensitive_issues',
        'prejudicial_info',
        'intrusive',
        'illegal',
        'direct_social_info',
        'identifiable_records',
        'psychology_tests',
        'researcher_risk',
        'incentives',
        'participant_costs',
        'researcher_interest',
        'conflict_of_interest',
        'uj_premises',
        'uj_facilities',
        'uj_funding'
    ]
    declaration_fields = {
        'declaration_name',
        'applicant_signature',
        'full_name',
        'declaration_date',
        'submission_date',
    }

    data = request.form.to_dict()

    if 'uj_employee' in data and 'uj_employees' not in data:
        data['uj_employees'] = data['uj_employee']
    if 'own_student' in data and 'own_students' not in data:
        data['own_students'] = data['own_student']
    if 'non_education' in data and 'no_education' not in data:
        data['no_education'] = data['non_education']
    if 'prejuditial_info' in data and 'prejudicial_info' not in data:
        data['prejudicial_info'] = data['prejuditial_info']
    if 'reseacher_risk' in data and 'researcher_risk' not in data:
        data['researcher_risk'] = data['reseacher_risk']

    for key, value in data.items():
        if key not in declaration_fields and hasattr(form, key):
            if key in boolean_fields:
                setattr(form, key, str_to_bool(value))
            else:
                setattr(form, key, value)
    try:
        db_session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# --- AUTOSAVE ENDPOINTS FOR FORM A SECTIONS ---

def _create_forma_resubmission_draft(source_form):
    draft = FormA(user_id=source_form.user_id)
    skip_fields = {
        'form_id',
        'created_at',
        'updated_at',
    }

    for column in FormA.__table__.columns:
        field_name = column.name
        if field_name in skip_fields:
            continue
        setattr(draft, field_name, getattr(source_form, field_name, None))

    draft.submitted_at = None
    draft.declaration_name = None
    draft.applicant_signature = None
    draft.declaration_date = None
    draft.submitted = False
    draft.submitted_to_admin = False
    draft.submitted_to_reviewers = False
    draft.submitted_to_rec = False
    draft.rec_comments = None
    draft.rec_status = None
    draft.rec_date = None
    draft.certificate_code = None
    draft.certificate_issued = None
    draft.certificate_end_date = None
    draft.certificate_issuer = None
    draft.certificate_email = None
    draft.certificate_received = False
    draft.certificate_heading = None
    draft.certificate_modified = False
    draft.certificate_condition_1 = None
    draft.visible_to_student = True
    preserve_single_reviewer_assignment(draft, source_form)

    db_session.add(draft)
    db_session.flush()
    session['active_forma_id'] = draft.form_id
    return draft


def _get_or_create_forma_resubmission_draft(user_id, source_form):
    existing_draft = (
        db_session.query(FormA)
        .filter(FormA.user_id == user_id, FormA.submitted_at.is_(None))
        .order_by(FormA.created_at.desc().nullslast())
        .all()
    )
    for draft in existing_draft:
        if can_reuse_forma_draft(draft):
            if preserve_single_reviewer_assignment(draft, source_form):
                db_session.commit()
            session['active_forma_id'] = draft.form_id
            return draft, None

    if not source_form:
        return None, (jsonify({'success': False, 'error': 'No source form available for resubmission'}), 404)

    try:
        draft = _create_forma_resubmission_draft(source_form)
        db_session.commit()
        return draft, None
    except Exception as exc:
        db_session.rollback()
        app.logger.exception("Failed to create Form A resubmission draft")
        return None, (jsonify({'success': False, 'error': str(exc)}), 500)


def _create_resubmission_draft(source_form, model, submission_field_name, declaration_fields, session_key):
    """Copy a returned submission into a new editable version."""
    draft = model(user_id=source_form.user_id)
    skip_fields = {'form_id', 'created_at', 'updated_at'}

    for column in model.__table__.columns:
        field_name = column.name
        if field_name in skip_fields:
            continue
        setattr(draft, field_name, getattr(source_form, field_name, None))

    setattr(draft, submission_field_name, None)
    for field_name in declaration_fields:
        if hasattr(draft, field_name):
            setattr(draft, field_name, None)

    if hasattr(draft, 'submitted'):
        draft.submitted = False
    draft.submitted_to_admin = False
    draft.submitted_to_reviewers = False
    draft.submitted_to_rec = False
    draft.visible_to_student = True
    preserve_single_reviewer_assignment(draft, source_form)

    for field_name, empty_value in {
        'rec_comments': None,
        'rec_status': None,
        'rec_date': None,
        'certificate_code': None,
        'certificate_issued': None,
        'certificate_end_date': None,
        'certificate_issuer': None,
        'certificate_email': None,
        'certificate_received': False,
        'certificate_heading': None,
        'certificate_modified': False,
        'certificate_condition_1': None,
    }.items():
        if hasattr(draft, field_name):
            setattr(draft, field_name, empty_value)

    db_session.add(draft)
    db_session.flush()
    session[session_key] = draft.form_id
    return draft


def _get_or_create_resubmission_draft(
    user_id,
    source_form,
    model,
    submission_field_name,
    declaration_fields,
    session_key,
    *,
    options=None,
):
    query = db_session.query(model)
    if options:
        query = query.options(*options)
    submission_field = getattr(model, submission_field_name)
    existing_draft = (
        query
        .filter(model.user_id == user_id, submission_field.is_(None))
        .order_by(model.created_at.desc().nullslast())
        .first()
    )
    if existing_draft:
        if preserve_single_reviewer_assignment(existing_draft, source_form):
            db_session.commit()
        session[session_key] = existing_draft.form_id
        return existing_draft, None

    if not source_form:
        return None, (jsonify({'success': False, 'error': 'No source form available for resubmission'}), 404)

    try:
        draft = _create_resubmission_draft(
            source_form,
            model,
            submission_field_name,
            declaration_fields,
            session_key,
        )
        db_session.commit()
        return draft, None
    except Exception as exc:
        db_session.rollback()
        app.logger.exception("Failed to create %s resubmission draft", model.__name__)
        return None, (jsonify({'success': False, 'error': str(exc)}), 500)


FORMB_DEFERRED_UPLOADS = (
    defer(FormB.permission_letter),
    defer(FormB.prior_clearance),
    defer(FormB.ethics_evidence),
    defer(FormB.proposal_path),
    defer(FormB.pending_note),
    defer(FormB.private_permission_file),
)


def _get_or_create_formb_resubmission_draft(user_id, source_form):
    return _get_or_create_resubmission_draft(
        user_id,
        source_form,
        FormB,
        'submitted_at',
        {'declaration_name', 'full_name', 'declaration_date'},
        'active_formb_id',
        options=FORMB_DEFERRED_UPLOADS,
    )


def _get_or_create_formc_resubmission_draft(user_id, source_form):
    return _get_or_create_resubmission_draft(
        user_id,
        source_form,
        FormC,
        'submission_date',
        {'declaration_name', 'full_name'},
        'active_formc_id',
    )

def _get_or_create_forma_draft(user_id, form_data):
    form_data = form_data or {}
    explicit_form_id = form_data.get('forma_id') or form_data.get('form_id') or ''
    form_id = explicit_form_id or session.get('active_forma_id')

    # Prefer explicit draft id from the client when available.
    if form_id:
        form = db_session.query(FormA).filter_by(user_id=user_id, form_id=form_id).first()
        if form:
            if getattr(form, 'submitted_at', None) is not None and is_student_correction_state(form):
                return _get_or_create_forma_resubmission_draft(user_id, form)
            if can_reuse_forma_draft(form):
                session['active_forma_id'] = form.form_id
                return form, None
            if explicit_form_id and getattr(form, 'submitted_at', None) is not None:
                return None, (
                    jsonify({
                        'success': False,
                        'error': 'Autosave ignored because this Form A has already been submitted.',
                    }),
                    409,
                )

    # Otherwise, reuse the latest unsubmitted draft for this user.
    unsubmitted_drafts = (
        db_session.query(FormA)
        .filter(FormA.user_id == user_id, FormA.submitted_at.is_(None))
        .order_by(FormA.created_at.desc().nullslast())
        .all()
    )
    for form in unsubmitted_drafts:
        if can_reuse_forma_draft(form):
            session['active_forma_id'] = form.form_id
            return form, None

    form_requirements = db_session.query(FormARequirements).filter_by(user_id=user_id).first()

    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user:
        return None, (jsonify({'success': False, 'error': 'User not found'}), 404)

    supervisor = db_session.query(User).filter(User.user_id == user.supervisor_id).first()

    form = FormA(
        user_id=user_id,
        attachment_id=(form_requirements.id if form_requirements else (form_data.get('attachment_id') or 'AUTOSAVE_PENDING')),
        applicant_name=form_data.get('applicant_name', ''),
        student_number=form_data.get('student_number', ''),
        institution=form_data.get('institution', 'University of Johannesburg'),
        department=form_data.get('department', ''),
        degree=form_data.get('degree', ''),
        study_title=form_data.get('study_title', ''),
        mobile=form_data.get('mobile', ''),
        email=(user.email or ''),
        supervisor=(supervisor.full_name if supervisor else ''),
        supervisor_email=(supervisor.email if supervisor else ''),
    )
    db_session.add(form)
    db_session.flush()
    session['active_forma_id'] = form.form_id
    return form, None

def _get_latest_forma_for_user(user_id):
    if not user_id:
        return None
    active_form_id = session.get('active_forma_id')
    if active_form_id:
        active_form = db_session.query(FormA).filter_by(user_id=user_id, form_id=active_form_id).first()
        if active_form and (can_reuse_forma_draft(active_form) or getattr(active_form, 'submitted_at', None) is not None):
            session['active_forma_id'] = active_form.form_id
            return active_form

    latest_clean_draft = (
        db_session.query(FormA)
        .filter(FormA.user_id == user_id, FormA.submitted_at.is_(None))
        .order_by(FormA.created_at.desc().nullslast())
        .all()
    )
    for form in latest_clean_draft:
        if can_reuse_forma_draft(form):
            session['active_forma_id'] = form.form_id
            return form

    latest_form = (
        db_session.query(FormA)
        .filter(FormA.user_id == user_id)
        .order_by(FormA.submitted_at.desc().nullslast(), FormA.created_at.desc().nullslast())
        .first()
    )
    if latest_form:
        session['active_forma_id'] = latest_form.form_id
    return latest_form

def _autosave_str_to_bool(val):
    if isinstance(val, bool):
        return val
    if val is None:
        return None
    if isinstance(val, str):
        if val.lower() in ['yes', 'true', '1', 'on', 'checked']:
            return True
        if val.lower() in ['no', 'false', '0', 'off', '']:
            return False
    return val

def _build_autosave_data(form_payload):
    data = {}
    for key in form_payload.keys():
        if key in {'csrf_token', 'forma_id'}:
            continue
        values = form_payload.getlist(key)
        normalized_key = key[:-2] if key.endswith('[]') else key
        if key in ['quantitative[]', 'qualitative[]', 'mixed_methods[]']:
            data[normalized_key] = True if values else False
        elif len(values) > 1:
            data[normalized_key] = ','.join(values)
        else:
            data[normalized_key] = values[0] if values else ''
    return data


def _clear_forma_secondary_data_details(form):
    """Remove every answer that is conditional on Form A question 5.5.1."""
    for field in [
        'secondary_data_type', 'data_nature', 'data_origin', 'access_conditions',
        'personal_info', 'personal_info_comment', 'data_anonymized',
        'anonymization_comment', 'shortcomings_reported',
        'limitations_reporting', 'methodology_alignment', 'data_acknowledgment',
    ]:
        if hasattr(form, field):
            setattr(form, field, '')


def _apply_forma_autosave_payload(form, form_payload, section='all', include_declaration=False):
    data = _build_autosave_data(form_payload)

    if 'secondary_data' in data and 'uses_secondary_data' not in data:
        data['uses_secondary_data'] = data['secondary_data']
    if 'questionnaire_permission' in data:
        questionnaire_permission = (data.get('questionnaire_permission') or '').strip()
        if questionnaire_permission == 'Yes':
            data['permission_obtained'] = 'Yes'
            data['open_source'] = 'No'
        elif questionnaire_permission == 'Open Source':
            data['permission_obtained'] = 'No'
            data['open_source'] = 'Yes'
        else:
            data['permission_obtained'] = None
            data['open_source'] = None

    sec2_bool_fields = [
        'survey', 'focus_groups', 'observations', 'interviews', 'documents',
        'vulnerable_communities', 'age_range', 'uj_employees', 'vulnerable', 'non_english',
        'own_students', 'poverty', 'no_education', 'disclosure', 'discomfiture', 'deception',
        'sensitive', 'prejudice', 'intrusive_techniques', 'illegal_activities', 'personal',
        'available_records', 'inventories', 'risk_activities', 'incentives', 'financial_costs',
        'reward', 'conflict', 'uj_premises', 'uj_facilities', 'uj_funding'
    ]

    sec4_bool_fields = [
        'quantitative', 'qualitative', 'mixed_methods', 'conflict_interest',
        'intervention', 'translator', 'uses_secondary_data'
    ]

    sec5_bool_fields = [f'q6_9{key}' for key in ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s']]

    all_bool_fields = set(sec2_bool_fields + sec4_bool_fields + sec5_bool_fields + ['interviews_one', 'documents_one'])
    for key in list(data.keys()):
        if key in all_bool_fields:
            data[key] = _autosave_str_to_bool(data[key])

    if section in ['all', 'sec1']:
        for key in ['applicant_name', 'student_number', 'institution', 'department', 'degree', 'study_title', 'mobile', 'email', 'supervisor', 'supervisor_email']:
            if key in data and hasattr(form, key):
                setattr(form, key, data.get(key, ''))

    if section in ['all', 'sec2']:
        for key in sec2_bool_fields:
            if section == 'all':
                setattr(form, key, _autosave_str_to_bool(data[key]) if key in data else False)
            else:
                setattr(form, key, _autosave_str_to_bool(data[key]) if key in data else False)

        for key in ['assessment_other_specify', 'vulnerable_other_specify', 'activity_specify', 'vulnerable_comments_1', 'vulnerable_comments_2', 'vulnerable_comments_3', 'risk_rating', 'risk_justification', 'benefits_description', 'risk_mitigation', 'apply_comments', 'other_sec2']:
            if key in data and hasattr(form, key):
                setattr(form, key, data.get(key, ''))

        if hasattr(form, 'interviews_one'):
            form.interviews_one = _autosave_str_to_bool(data.get('interviews')) if 'interviews' in data else False
        if hasattr(form, 'documents_one'):
            form.documents_one = _autosave_str_to_bool(data.get('documents')) if 'documents' in data else False

    if section in ['all', 'sec3']:
        for key in ['title_provision', 'abstract', 'questions', 'purpose_objectives', 'grant_permission', 'researcher_affiliation', 'affiliation_details', 'collective_involvement', 'collective_details', 'is_funded', 'indemnity_arrangements', 'other_committee']:
            if key in data and hasattr(form, key):
                setattr(form, key, data.get(key, ''))

        for key in ['grant_permission', 'researcher_affiliation', 'collective_involvement', 'is_funded']:
            if key in data and hasattr(form, key):
                raw_value = data.get(key)
                if raw_value in [None, '']:
                    setattr(form, key, '')
                else:
                    setattr(form, key, 'Yes' if _autosave_str_to_bool(raw_value) is True else 'No')

        grant_permission_value = (getattr(form, 'grant_permission', None) or '').strip().lower()
        if grant_permission_value == 'yes':
            form.org_name = ','.join(form_payload.getlist('org_name[]'))
            form.org_contact = ','.join(form_payload.getlist('org_contact[]'))
            form.org_role = ','.join(form_payload.getlist('org_role[]'))
            form.org_permission = ','.join(form_payload.getlist('org_permission[]'))
        else:
            form.org_name = ''
            form.org_contact = ''
            form.org_role = ''
            form.org_permission = ''

        if _autosave_str_to_bool(getattr(form, 'researcher_affiliation', None)) is not True:
            form.affiliation_details = ''
        if _autosave_str_to_bool(getattr(form, 'collective_involvement', None)) is not True:
            form.collective_details = ''
        form.fund_org = ','.join(form_payload.getlist('fund_org[]'))
        form.fund_contact = ','.join(form_payload.getlist('fund_contact[]'))
        form.fund_role = ','.join(form_payload.getlist('fund_role[]'))
        form.fund_amount = ','.join(form_payload.getlist('fund_amount[]'))

    if section in ['all', 'sec4']:
        for key in sec4_bool_fields:
            if section == 'all':
                setattr(form, key, _autosave_str_to_bool(data[key]) if key in data else False)
            else:
                setattr(form, key, _autosave_str_to_bool(data[key]) if key in data else False)

        for paradigm_field in ['quantitative', 'qualitative', 'mixed_methods']:
            if section == 'all' and paradigm_field not in data:
                setattr(form, paradigm_field, False)

        for key in ['paradigm_explanation', 'design', 'participants_description', 'duration_timing', 'contact_details_method', 'conflict_explanation', 'questionnaire_type', 'permission_obtained', 'open_source', 'instrument_attachment_reason', 'interview_type', 'interview_recording', 'focus_recording', 'observation_details', 'documents_details', 'other_details', 'data_collection_procedure', 'data_collectors', 'intervention_details', 'sensitive_data', 'translator_procedure', 'data_nature', 'data_origin', 'access_conditions', 'personal_info', 'personal_info_comment', 'data_anonymized', 'anonymization_comment', 'shortcomings_reported', 'limitations_reporting', 'methodology_alignment', 'data_acknowledgment', 'secondary_data_type']:
            if key in data and hasattr(form, key):
                setattr(form, key, data.get(key, ''))

        form.population = ','.join(form_payload.getlist('population[]'))
        form.sampling_method = ','.join(form_payload.getlist('sampling_method[]'))
        form.sampling_size = ','.join(form_payload.getlist('sample_size[]'))
        form.inclusion_criteria = ','.join(form_payload.getlist('inclusion_criteria[]'))
        form.data_methods = ','.join(form_payload.getlist('data_methods[]'))

        selected_methods = [method.strip() for method in form.data_methods.split(',') if method.strip()]
        form.use_focus_groups = 'focus' in selected_methods

        if 'data_type' in data:
            form.secondary_data_type = data.get('data_type', '')

        if hasattr(form, 'uses_secondary_data') and not form.uses_secondary_data:
            _clear_forma_secondary_data_details(form)

    if section in ['all', 'sec5']:
        for key in ['informed_consent', 'study_benefits', 'participant_risks', 'adverse_steps', 'community_participation', 'community_effects', 'results_feedback', 'products_access', 'publication_plans', 'participant_comp', 'participant_costs', 'ethics_reporting']:
            if key in data and hasattr(form, key):
                setattr(form, key, data.get(key, ''))

        form.data_storage = ','.join(form_payload.getlist('data_storage[]'))
        form.privacy = ','.join(form_payload.getlist('privacy[]'))

        for key in sec5_bool_fields:
            if section == 'all':
                setattr(form, key, _autosave_str_to_bool(data[key]) if key in data else False)
            else:
                setattr(form, key, _autosave_str_to_bool(data[key]) if key in data else False)

    if section in ['all', 'sec6'] and include_declaration:
        if 'declaration_name' in data:
            form.declaration_name = data.get('declaration_name', '')
        if 'applicant_signature' in data:
            form.applicant_signature = data.get('applicant_signature', '')

        date_str = data.get('declaration_date', '')
        if date_str:
            try:
                form.declaration_date = parse_html_date(date_str)
            except (ValueError, TypeError):
                pass

@csrf.exempt
@app.route('/form_a_sec1_autosave', methods=['POST'])
def form_a_sec1_autosave():
    """Autosave for Form A Section 1 - Researcher's Details"""
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    _apply_forma_autosave_payload(form, request.form, section='sec1', include_declaration=False)
    
    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error in form_a_sec1_autosave: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@csrf.exempt
@app.route('/form_a_sec2_autosave', methods=['POST'])
def form_a_sec2_autosave():
    """Autosave for Form A Section 2 - Ethical Considerations"""
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    _apply_forma_autosave_payload(form, request.form, section='sec2', include_declaration=False)
    
    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error in form_a_sec2_autosave: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@csrf.exempt
@app.route('/form_a_sec3_autosave', methods=['POST'])
def form_a_sec3_autosave():
    """Autosave for Form A Section 3 - Project Information & Affiliations"""
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    _apply_forma_autosave_payload(form, request.form, section='sec3', include_declaration=False)
    
    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error in form_a_sec3_autosave: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@csrf.exempt
@app.route('/form_a_sec4_autosave', methods=['POST'])
def form_a_sec4_autosave():
    """Autosave for Form A Section 4 - Methodology"""
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    cleaned_sample_sizes, sample_size_error = normalize_forma_sample_sizes(
        request.form.getlist('sample_size[]')
    )
    if sample_size_error:
        return jsonify({'success': False, 'error': sample_size_error}), 422

    _apply_forma_autosave_payload(form, request.form, section='sec4', include_declaration=False)
    form.sampling_size = ','.join(cleaned_sample_sizes)
    
    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error in form_a_sec4_autosave: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@csrf.exempt
@app.route('/form_a_sec5_autosave', methods=['POST'])
def form_a_sec5_autosave():
    """Autosave for Form A Section 5 - Ethical Considerations"""
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    _apply_forma_autosave_payload(form, request.form, section='sec5', include_declaration=False)
    
    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error in form_a_sec5_autosave: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@csrf.exempt
@app.route('/form_a_sec6_autosave', methods=['POST'])
@app.route('/form_a_sec7_autosave', methods=['POST'])
def form_a_sec6_autosave():
    """Autosave for Form A Section 6/7 - Declaration"""
    user_id = session.get('id')
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    form, error_response = _get_or_create_forma_draft(user_id, request.form)
    if error_response:
        return error_response

    # Declaration details are submission-only and must never be autosaved.
    _apply_forma_autosave_payload(form, request.form, section='sec6', include_declaration=False)
    
    try:
        db_session.commit()
        return jsonify({'success': True, 'form_id': form.form_id})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error in form_a_sec6_autosave: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

    
@app.route('/view_form_file/<form_type>/<form_id>/<field_name>')
@app.route('/view_form_file/<form_type>/<form_id>')
@login_required
def view_form_file(form_type, form_id, field_name=None):
    f_name = field_name or request.args.get('field_name') or request.args.get('field')
    
    if not f_name:
        return "Missing field_name", 400

    if f_name not in FORM_FILE_FIELDS.get(str(form_type).upper(), set()):
        abort(403)

    model_map = {
        'A': FormA,
        'FormA': FormA,
        'B': FormB,
        'FormB': FormB,
        'C': FormC,
        'FormC': FormC
    }
    model = model_map.get(form_type)
    if not model:
        return "Invalid model", 400
        
    form = db_session.query(model).filter_by(form_id=form_id).first()
    if not form:
        return "Form not found", 404

    current_user = get_current_user()
    if not can_access_form(current_user, form):
        abort(403)
        
    data = getattr(form, f_name, None)
    if not data:
        return "File not found", 404

    if isinstance(data, memoryview):
        data = data.tobytes()

    # Better filename detection
    filename = getattr(form, f"{f_name}_filename", None) or \
               getattr(form, f_name.replace('_path', '') + "_filename", None) or \
               getattr(form, f_name.replace('_file', '') + "_filename", None)
               
    # Backward compatibility for path-based legacy uploads.
    if isinstance(data, str):
        if not data.startswith('\\x') and len(data) < 500:
            # It might be a legacy file path if it's short
            clean_path = data.replace('\\', '/')
            if clean_path.startswith('static/'):
                clean_path = clean_path.replace('static/', '', 1)
            file_path = os.path.join(app.root_path, 'static', clean_path)
            if os.path.exists(file_path):
                # Determine mimetype from filename or fallback
                mtype, _ = mimetypes.guess_type(file_path)
                return send_file(file_path, mimetype=mtype or 'application/pdf', as_attachment=False)
            # If path doesn't exist, try encoding as bytes
            data = data.encode('latin-1')
        else:
            data = data.encode('latin-1')

    # MAGIC NUMBER DETECTION for robustness
    is_pdf = data.startswith(b'%PDF-')
    is_zip = data.startswith(b'PK\x03\x04')
    
    mimetype = 'application/pdf' if is_pdf else 'application/octet-stream'
    
    if filename:
        filename = filename.strip()
        ext = filename.lower().split('.')[-1]
        if ext == 'pdf':
            mimetype = 'application/pdf'
        elif ext in ['doc', 'docx']:
            mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' if ext == 'docx' else 'application/msword'
        elif ext == 'zip':
            mimetype = 'application/zip'
        elif ext == 'png':
            mimetype = 'image/png'
        elif ext in ['jpg', 'jpeg']:
            mimetype = 'image/jpeg'
    else:
        # Fallback filename based on magic numbers
        if is_pdf:
            filename = f"{f_name}.pdf"
            mimetype = 'application/pdf'
        elif is_zip:
            filename = f"{f_name}.docx"
            mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        else:
            filename = f"{f_name}.dat"
            mimetype = 'application/octet-stream'

    return send_file(
        io.BytesIO(data),
        mimetype=mimetype,
        download_name=filename,
        as_attachment=False
    )

db = SQLAlchemy(app)
migrate = Migrate(app, db)

web_url='https://jbs-ethics.onrender.com'

##import dummy_data

##dummy_data


# Health check endpoint for Render
@app.route('/health', methods=['GET'])
def health():
    return 'OK', 200

    
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc'}

def get_upload_folder():
    """
    Get the upload folder path - uses persistent disk on production, local folder for development
    """
    # Check if we're on Render.com production with persistent disk
    persistent_path = os.getenv('UPLOAD_PATH')
    if persistent_path and os.path.exists(os.path.dirname(persistent_path)):
        upload_folder = os.path.join(persistent_path, 'form')
        os.makedirs(upload_folder, exist_ok=True)
        return upload_folder
    
    # Fallback to local static folder for development
    local_path = os.path.join('static', 'uploads', 'form')
    os.makedirs(local_path, exist_ok=True)
    return local_path

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def read_file_blob(file_obj_or_key):
    """Read and validate a PDF/DOCX upload."""
    if isinstance(file_obj_or_key, str):
        file = request.files.get(file_obj_or_key)
    else:
        file = file_obj_or_key
        
    return read_validated_upload(file, app.config.get('MAX_FILE_LENGTH', 524288000))


def assign_private_permission_upload(record, file_storage):
    """Validate and assign a private-permission document to a form record."""
    data, filename = read_file_blob(file_storage)
    record.private_permission_file = data
    record.private_permission_filename = filename


@app.route('/delete_form_b_private_permission/<string:form_id>', methods=['POST'])
@login_required
def delete_form_b_private_permission(form_id):
    """Delete the student's section 2.7 upload during corrections."""
    user_id = session.get('id')
    form = db_session.query(FormB).filter_by(form_id=form_id).first()
    if not form:
        return "Form not found", 404
    if getattr(form, 'user_id', None) != user_id:
        abort(403)
    if not is_student_correction_state(form):
        flash(
            'This document can only be deleted while the form is with you for resubmission.',
            'warning',
        )
        return redirect(url_for('student_form_pdf', form_id=form_id, form_type='B'))

    form.private_permission_file = None
    form.private_permission_filename = None
    db_session.commit()
    flash('The 2.7 permission document was deleted.', 'success')
    return redirect(request.referrer or url_for('student_edit_formb'))


# DEBUG ENDPOINT REMOVED FOR PRODUCTION SECURITY

# --- Send Back for Corrections endpoint for FormA ---
from sqlalchemy.exc import SQLAlchemyError


# --- Send Back for Corrections endpoint for FormA ---
@app.route('/send_back_for_corrections/<id>', methods=['POST'], endpoint='send_back_for_corrections')
@app.route('/send_back_for_corrections_a/<id>', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def send_back_for_corrections_a(id):
    try:
        form = db_session.query(FormA).filter_by(form_id=id).first()
        if not form:
            flash('Form not found.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_a'))
        if not has_all_required_reviews(form):
            flash('Form A can only be sent back after all assigned reviewers have submitted their reviews.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_a'))
        if get_admin_reviewer_outcome(form) != 'approved_with_minor_changes':
            flash('Only a form approved with minor changes can be returned for corrections.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_a'))
        # Mark the form as needing corrections and add feedback
        feedback = (request.form.get('corrections_feedback') or '').strip()
        if not feedback:
            flash('A comment is required before sending Form A back for corrections.', 'danger')
            return redirect(url_for('ethics_reviewer_committee_form_a'))
        form.status = 'Submitted to Student for Corrections'
        form.review_form_status = feedback
        form.visible_to_student = True  # Ensure student sees it in dashboard
        db_session.commit()
        flash('Form sent back for corrections.', 'warning')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash('Database error: {}'.format(str(e)), 'danger')
    return redirect(url_for('ethics_reviewer_committee_form_a'))

# --- Student resubmits corrected FormA ---
@app.route('/resubmit_forma/<id>', methods=['POST'])
def resubmit_forma(id):
    try:
        form = db_session.query(FormA).filter_by(form_id=id).first()
        if not form:
            flash('Form not found.', 'danger')
            return redirect(url_for('student_dashboard'))
        # Only allow resubmission if corrections were required
        if form.status not in ('Corrections Required', 'Submitted to Student for Corrections'):
            flash('Form is not awaiting corrections.', 'danger')
            return redirect(url_for('student_dashboard'))
        # Ensure reviewers are assigned before resubmission
        if assigned_reviewer_count(form) < 1:
            flash('You cannot resubmit Form A until at least one reviewer is assigned.', 'danger')
            return redirect(url_for('student_dashboard'))
        form.status = 'Resubmitted'
        form.ethics_status = None
        form.form_supervisor_status = 'Resubmitted'
        form.submitted_at = get_local_time()
        form.visible_to_student = False
        reset_form_review_feedback(form)
        db_session.commit()
        flash('Form resubmitted to admin and supervisor.', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash('Database error: {}'.format(str(e)), 'danger')
    return redirect(url_for('student_dashboard'))
# Use proper logging and monitoring tools instead


# Adding an exception handler

from sqlalchemy.exc import SQLAlchemyError, OperationalError
import time

def is_transient_db_connection_error(exc):
    message = str(exc).lower()
    transient_markers = (
        "ssl error",
        "ssl connection has been closed unexpectedly",
        "server closed the connection unexpectedly",
        "connection not open",
        "could not receive data from server",
    )
    return any(marker in message for marker in transient_markers)

def run_db_query_with_retry(query_factory, *, retries=1, retry_delay=0.25):
    last_exc = None
    for attempt in range(retries + 1):
        try:
            return query_factory()
        except OperationalError as exc:
            last_exc = exc
            try:
                db_session.rollback()
            except Exception:
                pass
            try:
                db_session.remove()
            except Exception:
                pass
            if attempt >= retries or not is_transient_db_connection_error(exc):
                raise
            time.sleep(retry_delay)
    raise last_exc

@app.errorhandler(SQLAlchemyError)
def handle_db_errors(e):
    # Safely rollback the session, handling cases where session is in a bad state
    try:
        db_session.rollback()
    except Exception as rollback_error:
        # If rollback fails, try to remove the session entirely
        try:
            db_session.remove()
        except Exception:
            pass
    
    # Handle SSL connection errors specifically
    if isinstance(e, OperationalError) and is_transient_db_connection_error(e):
        print(f"Database SSL connection error: {e}")
        # For SSL errors, we could try to reconnect or show a user-friendly message
        return render_template('error.html', 
                             error_message="Database connection temporarily unavailable. Please try again in a moment."), 503
    
    raise e



@app.route('/api')
def index():
    response={
         "message": "Welocme",
    }
    return jsonify(response), 200

def check_user_has_submitted_forms(user_id):
    """
    Check if user has submitted any form (FormA, FormB, or FormC)
    Returns True if user has submitted any form, False otherwise
    """
    try:
        # Check FormA submissions
        form_a_exists = db_session.query(FormA).filter_by(user_id=user_id).first()
        if form_a_exists:
            return True
            
        # Check FormB submissions
        form_b_exists = db_session.query(FormB).options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        ).filter_by(user_id=user_id).first()
        if form_b_exists:
            return True
            
        # Check FormC submissions
        form_c_exists = db_session.query(FormC).filter_by(user_id=user_id).first()
        if form_c_exists:
            return True
            
        return False
    except Exception as e:
        print(f"Error checking user form submissions: {e}")
        return False


@app.route("/checks_if_student_submitted_forms",methods=['POST','GET'] )
def checks_if_student_submitted_forms():
    user_id=session.get('id')
    has_submitted = check_user_has_submitted_forms(user_id)
    if has_submitted:
        return redirect(url_for('student_dashboard'))
        




@app.route('/student_ethics_pack_to_dashboards', methods=['GET','POST'])
def student_ethics_pack_to_dashboards():
    user_id=session.get('id')
    
    # Check if user has already submitted any form
    if check_user_has_submitted_forms(user_id):
        return redirect(url_for('student_dashboard'))
    
    if request.method == 'GET':
        return redirect(url_for('ethics_pack'))

    selected_form = (request.form.get('selected_form') or '').strip().upper()

    if selected_form == 'A':

        return render_template('form-a-upload.html')

    elif selected_form == 'B':

        return render_template('form-b-upload.html')

    elif selected_form == 'C':

        return render_template('form-c-upload.html')

    return redirect(url_for('ethics_pack'))
    


    
@app.route('/student-dashboard', methods=['GET'])
def student_dashboard():
    try:
        user_id = session.get('id')

        if not user_id:
            return redirect(url_for('login_page'))

        user = db_session.query(User).filter_by(user_id=user_id).first()

        formA = _find_all_forms_for_user(FormA, user_id)
        formB = _find_all_forms_for_user(
            FormB,
            user_id,
            options=[
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file),
            ],
        )
        formC = _find_all_forms_for_user(FormC, user_id)
        formD = db_session.query(FormD).filter_by(user_id=user_id).first()

        full_name = user.full_name if user else "Unknown User"

        # Check if student has FormARequirements and if files still exist
        form_requirements = db_session.query(FormARequirements).filter_by(user_id=user_id).first()
        files_missing = False
        missing_files_info = {}
        
        if not form_requirements and not (formA or formB or formC):
            # A student with no form has not started the workflow yet. Existing
            # (including restored) forms must always take priority over this gate.
            return redirect(url_for('ethics_pack'))
            
            
        return render_template(
            'dashboard.html',
            full_name=full_name,
            formA=formA,
            formB=formB,
            formC=formC,
            formD=formD,
            current_form_a_id=formA[0].form_id if formA else None,
            current_form_b_id=formB[0].form_id if formB else None,
            current_form_c_id=formC[0].form_id if formC else None,
            form_requirements=form_requirements,
            files_missing=files_missing,
            missing_files_info=missing_files_info
        )
    except Exception as e:
        # log error safely
        app.logger.error(f"Error loading student dashboard: {e}")
        return "An unexpected error occurred. Please try again later.", 500


@app.route('/quiz', methods=['GET'])
def quiz():
    return render_template('quiz.html')


@app.route('/logout', methods=['GET'])
def logout():
    # Do not carry alerts from the authenticated area back to the login page.
    clear_auth_session(clear_flashes=True)
    return redirect('/login?system=ethics')

# Auto-logout after 45 minutes of inactivity
@app.before_request
def auto_logout():
    # Only check for logged-in students
    if 'id' in session:
        user_role = session.get('role')
        # Only apply auto-logout to students
        if user_role and str(user_role).upper() == 'STUDENT':
            now = datetime.utcnow()
            last_active = session.get('last_active')
            # If never set, set now
            if not last_active:
                session['last_active'] = now.isoformat()
            else:
                try:
                    last_active_dt = datetime.fromisoformat(last_active)
                except Exception:
                    # fallback if format is wrong
                    last_active_dt = now
                # If inactive for more than 45 minutes
                if now - last_active_dt > timedelta(minutes=60):
                    clear_auth_session(clear_flashes=True)
                    return redirect('/login?system=ethics')
            # Always update last_active if not logging out
            session['last_active'] = now.isoformat()


@app.before_request
def start_request_activity_timer():
    g.request_started_at = time.time()


@app.after_request
def capture_logged_in_user_activity(response):
    if request.endpoint == 'static' or not session.get('id'):
        return response

    if getattr(g, 'activity_already_logged', False):
        return response

    request_started_at = getattr(g, 'request_started_at', None)
    duration_seconds = None
    if request_started_at is not None:
        duration_seconds = max(0, int(round(time.time() - request_started_at)))

    action = 'request_error' if response.status_code >= 400 else 'page_visit'
    details = {
        'method': request.method,
        'path': request.path,
        'endpoint': request.endpoint,
        'status_code': response.status_code,
        'query_params': request.args.to_dict(flat=True),
    }

    try:
        create_user_activity_entry(
            user_id=session.get('id'),
            action=action,
            page=request.path,
            details=details,
            duration_seconds=duration_seconds
        )
        g.activity_already_logged = True
    except Exception:
        try:
            db_session.rollback()
        except Exception:
            pass

    return response


@app.teardown_request
def capture_logged_in_user_exceptions(exception=None):
    if not exception or request.endpoint == 'static' or not session.get('id'):
        return None

    details = {
        'method': request.method,
        'path': request.path,
        'endpoint': request.endpoint,
        'error_type': type(exception).__name__,
        'error_message': str(exception),
    }

    try:
        create_user_activity_entry(
            user_id=session.get('id'),
            action='request_exception',
            page=request.path,
            details=details,
            duration_seconds=None
        )
    except Exception:
        try:
            db_session.rollback()
        except Exception:
            pass
    return None


def is_student_account_activated(user):
    """Accept activation performed through either integrated or legacy admin UI."""
    legacy_value = str(getattr(user, 'authenticate_student', '') or '').strip().lower()
    return legacy_value in {'true', '1', 'yes'} or bool(getattr(user, 'authenticated_student', False))


###
###
### this is the function to focus on when intergrating MBA and Ethics
###
@app.route('/')
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'GET' and 'id' in session:
        clear_auth_session()
    if request.method == 'GET':
        return redirect_to_shared_login()

    if request.method == 'POST':
        email = request.form.get('email')
        user_password = request.form.get('password')
        email = request.form.get('email')
        user_password = request.form.get('password')
        user = db_session.query(User).filter_by(email=email).first()
        
        
        if user:
            if user.verify_password(user_password):
                # Block unauthenticated user from logging in
                # Check if authenticate_student is falsy (False, "False", "false", None, empty string, etc.)
                if not is_student_account_activated(user):
                        clear_auth_session()
                        return redirect_to_shared_login("You are authenticated. Please wait for admin approval.")
                return _complete_ethics_login(user)
            else:
                return redirect_to_shared_login('Incorrect email or password')
        else:
            return redirect_to_shared_login('Incorrect email or password')
    return redirect_to_shared_login()


@app.route('/sso-login', methods=['GET'])
def sso_login():
    token = (request.args.get('token') or '').strip()
    if not token:
        clear_auth_session()
        return redirect_to_shared_login("Access denied.")

    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        payload = serializer.loads(token, salt=ETHICS_SSO_SALT, max_age=300)
    except SignatureExpired:
        clear_auth_session()
        return redirect_to_shared_login("Access denied. Your ethics sign-in link expired.")
    except BadSignature:
        clear_auth_session()
        return redirect_to_shared_login("Access denied.")

    email = (payload.get('email') or '').strip().lower()
    if not email:
        clear_auth_session()
        return redirect_to_shared_login("Access denied.")

    user = db_session.query(User).filter(func.lower(User.email) == email).first()
    if not user:
        clear_auth_session()
        return redirect_to_shared_login("Access denied.")

    if getattr(user, 'role', None) is None:
        clear_auth_session()
        return redirect_to_shared_login("Access denied.")

    user_role = str(getattr(user.role, 'value', user.role) or '').upper()
    if user_role == 'STUDENT' and not is_student_account_activated(user):
        clear_auth_session()
        return redirect_to_shared_login("Access denied.")

    return _complete_ethics_login(
        user,
        audit_action='sso_login',
        audit_page='sso-login',
        destination=(payload.get('destination') or '').strip(),
    )



###
###
### this is the function to focus on when intergrating MBA and Ethics
###
@app.route('/api/register', methods=['GET', 'POST'])
def register():
    try:
        msg = {}

        if request.method == 'POST':
            full_name = request.form.get('full_name', '').strip()
            student_number_raw = request.form.get('student_number', '').strip()
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()

            if password != confirm_password:
                msg = 'Passwords do not match.'
                return render_template('register.html', messages=[msg])

            # users.student_number is stored as Integer in the DB.
            # Validate early so bad input does not trigger DB rollback errors.
            if not student_number_raw.isdigit():
                msg = 'Student number must contain digits only.'
                return render_template('register.html', messages=[msg])
            student_number = int(student_number_raw)

            # Validate UJ email
            #if not email.endswith('student.uj.ac.za'):
                #msg = "Only University of Johannesburg email allowed"
                #return render_template('register.html', messages=[msg])

            # Validate password
            try:
                is_valid, message = validate_password(password)
                if not is_valid:
                    msg['message'] = message
                    return render_template('register.html', messages=[msg])
            except Exception as e:
                app.logger.error(f"Password validation error: {e}")
                return render_template('register.html', messages=["Invalid password. Please try again."])

            # Check if user exists
            try:
                user = db_session.query(User).filter_by(email=email).first()
                if user:
                    msg = 'Email already registered!'
                    return render_template('register.html', messages=[msg])
            except Exception as e:
                app.logger.error(f"DB lookup error: {e}")
                return render_template('register.html', messages=["Database error. Please try again later."])

            try:
                # Hash the password before storing (replace with proper hash function)
                # Example: password = generate_password_hash(password)

                # Create new user
                new_user = User(
                    full_name=full_name,
                    student_number=student_number,
                    email=email,
                    password=password,  # ⚠️ Replace with hashed version
                    role=UserRole.STUDENT
                )

                db_session.add(new_user)
                db_session.commit()

                stored_user = db_session.query(User).filter_by(email=email).first()
                session['pending_registration_declaration_user_id'] = stored_user.user_id
                return redirect(url_for('applicant_registration_declaration'))

            except Exception as e:
                db_session.rollback()
                app.logger.error(f"User registration error: {e}")
                msg = 'Registration failed. Please try again.'
                return render_template('register.html', messages=[msg])

        return render_template('register.html', messages=[])

    except Exception as e:
        app.logger.error(f"Unexpected error in register route: {e}")
        return render_template('register.html', messages=["Unexpected error occurred. Please try again later."])


@app.route('/registration/applicant-declaration', methods=['GET', 'POST'])
def applicant_registration_declaration():
    user_id = session.get('pending_registration_declaration_user_id')
    if not user_id:
        flash('Please register an account before completing the applicant declaration.', 'warning')
        return redirect(url_for('register'))

    user = db_session.query(User).filter_by(user_id=user_id).first()
    if not user:
        session.pop('pending_registration_declaration_user_id', None)
        flash('Your registration could not be found. Please register again.', 'danger')
        return redirect(url_for('register'))

    if request.method == 'POST':
        if request.form.get('applicant_declaration_agreed') != 'yes':
            flash('You must read and agree to the declaration before continuing.', 'danger')
            return render_template('applicant_registration_declaration.html', user=user, completed=False)

        create_user_activity_entry(
            user.user_id,
            'applicant_registration_declaration_agreed',
            page='registration/applicant-declaration',
            details={
                'declaration_version': '2026-08-07',
                'electronic_signature': True,
                'student_number': user.student_number,
            },
        )
        try:
            message = (
                f'An account was created for student number {user.student_number}. '
                'Please wait for an administrator to activate your account before logging in.'
            )
            send_email(app, mail, message, [user.email])
        except Exception as error:
            app.logger.error('Registration confirmation email failed for %s: %s', user.email, error)

        session.pop('pending_registration_declaration_user_id', None)
        return render_template('applicant_registration_declaration.html', user=user, completed=True)

    return render_template('applicant_registration_declaration.html', user=user, completed=False)


@app.route("/student_choose_supervisor",methods=['POST','GET'])
def student_choose_supervisor():
    user_id = session.get('id')
    if not user_id:
        flash("Your session has expired. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    supervisors = db_session.query(User).filter(
        or_(
            User.role == UserRole.SUPERVISOR,
            User.role == UserRole.REVIEWER
        )
    ).all()

    user = db_session.query(User).filter_by(user_id=user_id).first()
    if not user:
        flash("Your account could not be found. Please log in again.", "danger")
        session.clear()
        return redirect(url_for('login_page'))

    if request.method == 'POST':
        supervisor_id = (request.form.get('supervisors') or '').strip()
        if not supervisor_id:
            flash("Please choose a supervisor before submitting.", "danger")
            return render_template('student_choose_supervisor.html', supervisors=supervisors)

        supervisor = db_session.query(User).filter_by(user_id=supervisor_id).first()
        if not supervisor:
            flash("The selected supervisor could not be found. Please choose again.", "danger")
            return render_template('student_choose_supervisor.html', supervisors=supervisors)

        user.supervisor_id = supervisor_id
        db_session.commit()

        if user.supervisor_id:
            # Sending email should not break supervisor assignment.
            try:
                message = (
                    f'You have been assigned to be the supervisor of the student with the '
                    f'name {user.full_name} and student number- {user.student_number}.'
                )
                send_email(app, mail, message, [supervisor.email])
            except Exception as e:
                print("Email sending error:", str(e))

            return render_template('video.html', name=session.get('name', user.full_name))

    return render_template('student_choose_supervisor.html', supervisors=supervisors)


def sync_student_supervisor_forms(student, supervisor):
    updated_counts = {"form_a": 0, "form_b": 0, "form_c": 0}

    form_a_records = db_session.query(FormA).filter_by(user_id=student.user_id).all()
    for form in form_a_records:
        form.supervisor = supervisor.full_name
        form.supervisor_email = supervisor.email
        updated_counts["form_a"] += 1

    form_b_records = db_session.query(FormB).filter_by(user_id=student.user_id).all()
    for form in form_b_records:
        form.supervisor = supervisor.full_name
        form.supervisor_email = supervisor.email
        updated_counts["form_b"] += 1

    form_c_records = db_session.query(FormC).filter_by(user_id=student.user_id).all()
    for form in form_c_records:
        form.supervisor_name = supervisor.full_name
        form.supervisor_email = supervisor.email
        updated_counts["form_c"] += 1

    return updated_counts


def has_review_feedback_started(form):
    return bool(
        getattr(form, 'form_reviewed_by', None)
        or getattr(form, 'form_reviewed_by1', None)
        or getattr(form, 'review_status', None)
        or getattr(form, 'review_status1', None)
    )


def can_reassign_reviewers(form):
    return True


def get_form_submission_timestamp(form, form_type):
    if form_type == 'FORM A':
        return getattr(form, 'submitted_at', None) or getattr(form, 'created_at', None)
    if form_type == 'FORM B':
        return getattr(form, 'submitted_at', None) or getattr(form, 'created_at', None)
    return getattr(form, 'submission_date', None) or getattr(form, 'created_at', None)


def get_latest_forms_for_reviewer_reassignment():
    latest_forms = []

    form_configs = [
        ('FORM A', FormA, func.coalesce(FormA.submitted_at, FormA.created_at), FormA.submitted_at),
        ('FORM B', FormB, func.coalesce(FormB.submitted_at, FormB.created_at), FormB.submitted_at),
        ('FORM C', FormC, func.coalesce(FormC.submission_date, FormC.created_at), FormC.submission_date),
    ]

    for form_type, model, order_column, submission_column in form_configs:
        records = (
            db_session.query(model)
            .filter(
                submission_column.isnot(None),
                or_(
                    getattr(model, 'submitted_to_admin') == True,
                    getattr(model, 'submitted_to_reviewers') == True,
                )
            )
            .order_by(desc(order_column))
            .all()
        )

        latest_by_user = {}
        for record in records:
            if record.user_id not in latest_by_user:
                latest_by_user[record.user_id] = record

        for record in latest_by_user.values():
            latest_forms.append((form_type, record))

    latest_forms.sort(
        key=lambda item: get_form_submission_timestamp(item[1], item[0]) or datetime.min,
        reverse=True
    )
    return latest_forms


def build_reviewer_reassignment_rows(form_records, reviewer_lookup, student_lookup):
    rows = []

    for form_type, form in form_records:
        reviewer_ids = [
            reviewer_id for reviewer_id in
            [getattr(form, 'reviewer_name1', None), getattr(form, 'reviewer_name2', None)]
            if reviewer_id
        ]
        assigned_reviewers = [
            reviewer_lookup[reviewer_id]
            for reviewer_id in reviewer_ids
            if reviewer_id in reviewer_lookup
        ]
        student = student_lookup.get(form.user_id)
        rows.append({
            'form_id': form.form_id,
            'form_type': form_type,
            'user_id': form.user_id,
            'student_name': getattr(form, 'applicant_name', None) or (student.full_name if student else 'Unknown Student'),
            'student_number': getattr(form, 'student_number', None) or (student.student_number if student else None),
            'student_email': (
                getattr(form, 'email', None)
                or getattr(form, 'email_address', None)
                or (student.email if student else None)
            ),
            'submitted_at': get_form_submission_timestamp(form, form_type),
            'assigned_reviewers': assigned_reviewers,
            'selected_reviewer_ids': reviewer_ids,
            'can_reassign': can_reassign_reviewers(form),
            'review_started': has_review_feedback_started(form),
            'view_url': url_for('chair_form_view', id=form.form_id, form_name=form_type),
        })

    return rows


def generate_temporary_password(length=12):
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    specials = "!@#$%^&*"

    while True:
        core_length = max(length - 1, 5)
        password_chars = [secrets.choice(alphabet) for _ in range(core_length)]
        password_chars.append(secrets.choice(specials))
        secrets.SystemRandom().shuffle(password_chars)
        candidate = ''.join(password_chars)
        is_valid, _ = validate_password(candidate)
        if is_valid:
            return candidate


def _pdf_escape(text):
    return (
        str(text or "")
        .replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
    )


def build_simple_text_pdf(lines, title="Document"):
    page_width = 612
    page_height = 792
    margin_left = 50
    top_y = 760
    line_height = 18
    lines_per_page = 36

    pages = []
    for start in range(0, len(lines), lines_per_page):
        pages.append(lines[start:start + lines_per_page])

    if not pages:
        pages = [[""]]

    objects = []

    def add_object(payload):
        objects.append(payload)
        return len(objects)

    font_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    title_font_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    page_ids = []
    content_ids = []
    pages_placeholder_index = len(objects)
    objects.append(None)

    for page_lines in pages:
        content_stream_lines = [
            "BT",
            f"/F2 18 Tf",
            f"1 0 0 1 {margin_left} {top_y} Tm",
            f"({_pdf_escape(title)}) Tj",
        ]

        current_y = top_y - 34
        for line in page_lines:
            content_stream_lines.extend([
                f"/F1 12 Tf",
                f"1 0 0 1 {margin_left} {current_y} Tm",
                f"({_pdf_escape(line)}) Tj",
            ])
            current_y -= line_height

        content_stream_lines.append("ET")
        content_stream = "\n".join(content_stream_lines)
        content_id = add_object(
            f"<< /Length {len(content_stream.encode('latin-1', errors='replace'))} >>\nstream\n"
            f"{content_stream}\nendstream"
        )
        content_ids.append(content_id)

        page_id = add_object(
            "<< /Type /Page /Parent {parent} 0 R "
            f"/MediaBox [0 0 {page_width} {page_height}] "
            f"/Resources << /Font << /F1 {font_id} 0 R /F2 {title_font_id} 0 R >> >> "
            f"/Contents {content_id} 0 R >>"
        )
        page_ids.append(page_id)

    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects[pages_placeholder_index] = (
        f"<< /Type /Pages /Count {len(page_ids)} /Kids [{kids}] >>"
    )
    pages_id = pages_placeholder_index + 1

    for page_id in page_ids:
        objects[page_id - 1] = objects[page_id - 1].replace("{parent}", str(pages_id))

    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

    pdf_parts = [b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"]
    offsets = [0]

    for index, obj in enumerate(objects, start=1):
        offsets.append(sum(len(part) for part in pdf_parts))
        pdf_parts.append(f"{index} 0 obj\n{obj}\nendobj\n".encode("latin-1", errors="replace"))

    xref_position = sum(len(part) for part in pdf_parts)
    pdf_parts.append(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    pdf_parts.append(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf_parts.append(f"{offset:010d} 00000 n \n".encode("latin-1"))

    trailer = (
        f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n"
        f"startxref\n{xref_position}\n%%EOF"
    )
    pdf_parts.append(trailer.encode("latin-1"))
    return b"".join(pdf_parts)


def build_student_password_pdf(student, password, admin_user):
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    lines = [
        "Student Password Reset Confirmation",
        "",
        f"Student Name: {student.full_name}",
        f"Student Number: {student.student_number or 'N/A'}",
        f"Student Email: {student.email}",
        "",
        f"New Password: {password}",
        "",
        f"Generated By: {admin_user.full_name}",
        f"Generated At: {generated_at}",
        "",
        f"Login Link: {web_url}",
        "",
        "Please keep this document secure and share it only with the intended student.",
    ]
    return build_simple_text_pdf(lines, title="UJ Ethics Password Reset")


def get_supervisor_reassignment_state(student_id, forms=None):
    """Allow supervisor changes only while every existing form is with the student."""
    if forms is None:
        forms = []
        for model, label in ((FormA, 'Form A'), (FormB, 'Form B'), (FormC, 'Form C')):
            for form in db_session.query(model).filter_by(user_id=student_id).all():
                forms.append((label, form))

    blocking_forms = []
    for label, form in forms:
        stage = get_workflow_stage(form)
        if stage not in {'draft', 'with-student-revisions'}:
            blocking_forms.append(f"{label}: {get_workflow_location(form)}")

    if blocking_forms:
        return {
            'allowed': False,
            'message': (
                'The supervisor cannot be changed after submission. Wait until the form '
                'is returned to the student. Current location: ' + '; '.join(blocking_forms)
            ),
        }
    return {
        'allowed': True,
        'message': 'Supervisor reassignment is available because the form is with the student.',
    }


def build_supervisor_reassignment_states(students):
    """Build reassignment states with one query per form type instead of per student."""
    student_ids = [student.user_id for student in students]
    if not student_ids:
        return {}

    forms_by_student = defaultdict(list)
    try:
        for model, label in ((FormA, 'Form A'), (FormB, 'Form B'), (FormC, 'Form C')):
            for form in db_session.query(model).filter(model.user_id.in_(student_ids)).all():
                forms_by_student[form.user_id].append((label, form))
    except SQLAlchemyError:
        db_session.rollback()
        app.logger.exception('Could not calculate supervisor reassignment workflow states')
        return {
            student_id: {
                'allowed': False,
                'message': 'Workflow status is temporarily unavailable. No supervisor change was permitted.',
            }
            for student_id in student_ids
        }

    return {
        student_id: get_supervisor_reassignment_state(
            student_id,
            forms=forms_by_student.get(student_id, []),
        )
        for student_id in student_ids
    }


@app.route('/admin/reassign_supervisors', methods=['GET', 'POST'])
def admin_reassign_supervisors():
    admin_id = session.get('id')
    if not admin_id:
        flash("Your session has expired. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    user_profile = db_session.query(User).filter_by(user_id=admin_id).first()
    if not user_profile:
        session.clear()
        flash("Your account could not be found. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    role = user_profile.role.value if user_profile.role else ''
    if role not in ['ADMIN', 'SUPER_ADMIN']:
        flash("You are not authorized to access that page.", "danger")
        return redirect(url_for('login_page'))

    supervisors = db_session.query(User).filter(
        or_(
            User.role == UserRole.SUPERVISOR,
            User.role == UserRole.REVIEWER
        )
    ).order_by(User.full_name.asc()).all()

    if request.method == 'POST':
        page = request.form.get('page', 1, type=int)

        filter_student = (request.form.get('filter_student') or '').strip()
        filter_student_number = (request.form.get('filter_student_number') or '').strip()
        filter_email = (request.form.get('filter_email') or '').strip()

        student_id = (request.form.get('student_id') or '').strip()
        supervisor_id = (request.form.get('supervisor_id') or '').strip()

        redirect_filters = {
            "page": page,
            "filter_student": filter_student,
            "filter_student_number": filter_student_number,
            "filter_email": filter_email
        }

        student = db_session.query(User).filter_by(user_id=student_id).first()
        if not student or student.role != UserRole.STUDENT:
            flash("The selected student could not be found.", "danger")
            return redirect(url_for('admin_reassign_supervisors', **redirect_filters))

        if not supervisor_id:
            flash("Please choose a supervisor before saving.", "danger")
            return redirect(url_for('admin_reassign_supervisors', **redirect_filters))

        supervisor = db_session.query(User).filter_by(user_id=supervisor_id).first()
        if not supervisor or supervisor.role not in [UserRole.SUPERVISOR, UserRole.REVIEWER]:
            flash("The selected supervisor could not be found.", "danger")
            return redirect(url_for('admin_reassign_supervisors', **redirect_filters))

        reassignment_state = get_supervisor_reassignment_state(student.user_id)
        if not reassignment_state['allowed']:
            flash(reassignment_state['message'], 'danger')
            return redirect(url_for('admin_reassign_supervisors', **redirect_filters))

        student.supervisor_id = supervisor.user_id
        updated_counts = sync_student_supervisor_forms(student, supervisor)

        db_session.add(UserActivityLog(
            user_id=admin_id,
            action='admin_reassign_supervisor',
            page=request.path,
            target_user_id=student.user_id,
            timestamp=datetime.now(),
            details=(
                f"Reassigned supervisor for {student.full_name} ({student.email}) "
                f"to {supervisor.full_name} ({supervisor.email}). "
                f"Updated forms: A={updated_counts['form_a']}, "
                f"B={updated_counts['form_b']}, C={updated_counts['form_c']}"
            )
        ))

        db_session.commit()

        try:
            message = (
                f'You have been assigned to be the supervisor of the student with the '
                f'name {student.full_name} and student number- {student.student_number}.'
            )
            send_email(app, mail, message, [supervisor.email])
        except Exception as e:
            print("Email sending error:", str(e))

        flash(
            f"Supervisor updated for {student.full_name}. "
            f"Student forms synced: Form A ({updated_counts['form_a']}), "
            f"Form B ({updated_counts['form_b']}), Form C ({updated_counts['form_c']}).",
            "success"
        )

        return redirect(url_for('admin_reassign_supervisors', **redirect_filters))

    page = request.args.get('page', 1, type=int)
    per_page = 15

    filter_student = (request.args.get('filter_student') or '').strip()
    filter_student_number = (request.args.get('filter_student_number') or '').strip()
    filter_email = (request.args.get('filter_email') or '').strip()

    students_query = db_session.query(User).filter(
        User.role == UserRole.STUDENT
    )

    if filter_student:
        students_query = students_query.filter(
            func.lower(User.full_name).like(f"%{filter_student.lower()}%")
        )

    if filter_student_number:
        students_query = students_query.filter(
            func.cast(User.student_number, db.String).like(f"%{filter_student_number}%")
        )

    if filter_email:
        students_query = students_query.filter(
            func.lower(User.email).like(f"%{filter_email.lower()}%")
        )

    students_query = students_query.order_by(User.full_name.asc())

    total_students = students_query.count()
    students = students_query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = (total_students + per_page - 1) // per_page

    supervisor_lookup = {supervisor.user_id: supervisor for supervisor in supervisors}
    reassignment_states = build_supervisor_reassignment_states(students)

    filter_students_list = db_session.query(User).filter(
        User.role == UserRole.STUDENT
    ).order_by(User.full_name.asc()).all()


    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": admin_id,
            },
        ).scalar()
    )

    return render_template(
        'admin_reassign_supervisors.html',
        role=role,
        user_profile=user_profile,
        students=students,
        supervisors=supervisors,
        supervisor_lookup=supervisor_lookup,
        reassignment_states=reassignment_states,
        page=page,
        total_pages=total_pages,
        total_students=total_students,
        filter_students_list=filter_students_list,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )





@app.route('/admin/reassign_reviewers', methods=['GET', 'POST'])
def admin_reassign_reviewers():
    admin_id = session.get('id')
    if not admin_id:
        flash("Your session has expired. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    user_profile = db_session.query(User).filter_by(user_id=admin_id).first()
    if not user_profile:
        session.clear()
        flash("Your account could not be found. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    role = user_profile.role.value if user_profile.role else ''
    if role not in ['ADMIN', 'SUPER_ADMIN']:
        flash("You are not authorized to access that page.", "danger")
        return redirect(url_for('login_page'))

    reviewers = (
        db_session.query(User)
        .filter(User.role == UserRole.REVIEWER)
        .order_by(User.full_name.asc())
        .all()
    )
    reviewer_lookup = {reviewer.user_id: reviewer for reviewer in reviewers}

    if request.method == 'POST':
        page = request.form.get('page', 1, type=int)
        form_id = (request.form.get('form_id') or '').strip()
        form_type = (request.form.get('form_type') or '').strip().upper()
        reviewer1_id = (request.form.get('reviewer_1_id') or '').strip()
        reviewer2_id = (request.form.get('reviewer_2_id') or '').strip()

        form_model_map = {
            'FORM A': FormA,
            'FORM B': FormB,
            'FORM C': FormC,
        }
        model = form_model_map.get(form_type)
        if not model:
            flash("The selected form type is invalid.", "danger")
            return redirect(url_for('admin_reassign_reviewers', page=page))

        form = db_session.query(model).filter_by(form_id=form_id).first()
        if not form:
            flash("The selected form could not be found.", "danger")
            return redirect(url_for('admin_reassign_reviewers', page=page))

        missing_submission_redirect = redirect_if_missing_student_submission(
            form,
            'reviewers',
            'admin_reassign_reviewers',
            page=page,
        )
        if missing_submission_redirect:
            return missing_submission_redirect

        selected_ids = []
        for reviewer_id in [reviewer1_id, reviewer2_id]:
            if reviewer_id and reviewer_id not in selected_ids:
                selected_ids.append(reviewer_id)

        if len(selected_ids) != 1:
            flash("Please choose exactly one reviewer before saving.", "danger")
            return redirect(url_for('admin_reassign_reviewers', page=page))

        selected_reviewers = (
            db_session.query(User)
            .filter(
                User.user_id.in_(selected_ids),
                User.role == UserRole.REVIEWER
            )
            .order_by(User.full_name.asc())
            .all()
        )
        if len(selected_reviewers) != len(selected_ids):
            flash("One or more selected reviewers could not be found.", "danger")
            return redirect(url_for('admin_reassign_reviewers', page=page))

        form.reviewer_name1 = selected_ids[0]
        form.reviewer_name2 = None
        form.submitted_to_reviewers = True

        db_session.add(UserActivityLog(
            user_id=admin_id,
            action='admin_reassign_reviewers',
            page=request.path,
            target_user_id=form.user_id,
            timestamp=datetime.now(),
            details=(
                f"Reassigned reviewers for {form_type} {form.form_id} to "
                f"{', '.join(selected_ids)}"
            )
        ))
        db_session.commit()

        try:
            reviewer_emails = [reviewer.email for reviewer in selected_reviewers if reviewer.email]
            if reviewer_emails:
                message = (
                    f'You have been assigned as a reviewer for {form_type} '
                    f'belonging to {getattr(form, "applicant_name", "a student")}. '
                    f'Please log into the ethics application system and review the form.'
                )
                send_email(app, mail, message, reviewer_emails)
        except Exception as e:
            print("Email sending error:", str(e))

        flash(
            f"Reviewers updated for {getattr(form, 'applicant_name', 'the selected student')} ({form_type}).",
            "success"
        )
        return redirect(url_for('admin_reassign_reviewers', page=page))

    page = request.args.get('page', 1, type=int)
    per_page = 15
    form_records = get_latest_forms_for_reviewer_reassignment()
    student_ids = list({form.user_id for _, form in form_records})
    students = db_session.query(User).filter(User.user_id.in_(student_ids)).all() if student_ids else []
    student_lookup = {student.user_id: student for student in students}
    rows = build_reviewer_reassignment_rows(form_records, reviewer_lookup, student_lookup)

    total_forms = len(rows)
    total_pages = (total_forms + per_page - 1) // per_page if total_forms else 1
    start_index = (page - 1) * per_page
    paginated_rows = rows[start_index:start_index + per_page]


    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": admin_id,
            },
        ).scalar()
    )

    return render_template(
        'admin_reassign_reviewers.html',
        role=role,
        user_profile=user_profile,
        reviewer_rows=paginated_rows,
        reviewers=reviewers,
        page=page,
        total_pages=total_pages,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )


@app.route('/admin/student-password-resets', methods=['GET', 'POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def admin_student_password_resets():
    admin_id = session.get('id')
    user_profile = db_session.query(User).filter_by(user_id=admin_id).first() if admin_id else None

    if not user_profile:
        flash("Session expired. Please login again.", "danger")
        return redirect(url_for('login_page'))

    role = user_profile.role.value if user_profile.role else ''

    if request.method == 'POST':
        page = request.form.get('page', 1, type=int)
        search_query = (request.form.get('search') or '').strip()
        student_id = (request.form.get('student_id') or '').strip()
        provided_password = (request.form.get('new_password') or '').strip()

        student = db_session.query(User).filter_by(user_id=student_id).first()
        if not student or student.role != UserRole.STUDENT:
            flash("The selected student could not be found.", "danger")
            return redirect(url_for('admin_student_password_resets', page=page, search=search_query))

        new_password = provided_password or generate_temporary_password()
        is_valid, message = validate_password(new_password)
        if not is_valid:
            flash(message, "danger")
            return redirect(url_for('admin_student_password_resets', page=page, search=search_query))

        try:
            student.password = new_password
            db_session.add(UserActivityLog(
                user_id=admin_id,
                action='admin_reset_student_password',
                page=request.path,
                target_user_id=student.user_id,
                timestamp=datetime.now(),
                details=(
                    f"Admin reset password for {student.full_name} ({student.email}). "
                    f"PDF credential sheet generated."
                )
            ))
            db_session.commit()
        except Exception as e:
            db_session.rollback()
            app.logger.error(f"Student password reset failed: {e}")
            flash("Password reset failed. Please try again.", "danger")
            return redirect(url_for('admin_student_password_resets', page=page, search=search_query))

        pdf_bytes = build_student_password_pdf(student, new_password, user_profile)
        safe_name = secure_filename(student.full_name or 'student')
        filename = f"{safe_name or 'student'}_new_password.pdf"

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    page = request.args.get('page', 1, type=int)
    per_page = 15
    search_query = (request.args.get('search') or '').strip()

    students_query = db_session.query(User).filter(User.role == UserRole.STUDENT)
    if search_query:
        search_term = f"%{search_query}%"
        students_query = students_query.filter(
            or_(
                User.full_name.ilike(search_term),
                User.email.ilike(search_term),
                cast(User.student_number, String).ilike(search_term)
            )
        )

    students_query = students_query.order_by(User.full_name.asc())
    total_students = students_query.count()
    students = students_query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = max(1, (total_students + per_page - 1) // per_page)

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": admin_id,
            },
        ).scalar()
    )


    return render_template(
        'admin_student_password_resets.html',
        role=role,
        user_profile=user_profile,
        students=students,
        page=page,
        total_pages=total_pages,
        search_query=search_query,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )


@app.route("/authenticate_student/<string:id>",methods=['POST','GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def authenticate_student(id):
    if request.method=='POST':
        try:
            user=db_session.query(User).filter_by(user_id=id).first()
            admin_id = session.get('id')
            if user:
                print(f"Authenticating student: {user.full_name}")
                user.authenticate_student='true'
                user.authenticated_student=True
                # Log admin action
                db_session.add(UserActivityLog(
                    user_id=admin_id,
                    action='admin_authenticate_user',
                    page=request.path,
                    target_user_id=user.user_id,
                    timestamp=datetime.now(),
                    details=f"Admin authenticated user: {user.full_name} ({user.email})"
                ))
                # Commit the authentication change to database
                db_session.commit()
                print(f"Student {user.full_name} authenticated successfully in database.")
                # Send email notification (separate from database transaction)
                message=(f'Your account has been authenticated. Please follow the link to login '
                f'{web_url}')
                try:
                    send_email(app,mail, message,[user.email])
                    print(f"Authentication email sent to {user.email}")
                except Exception as e:
                    print("Email sending error:", str(e))
                
                return redirect(url_for('all_users'))
            else:
                flash("no such student on our data")
                return redirect(url_for("all_users"))
        except Exception as e:
            print(f"Error authenticating student: {str(e)}")
            db_session.rollback()
            flash("Error authenticating student")
            return redirect(url_for("all_users"))
    return redirect(url_for('all_users')) 



@app.route("/diactivate_user_account/<string:id>",methods=['POST','GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def diactivate_user_account(id):
    if request.method=='POST':
        try:
            user=db_session.query(User).filter_by(user_id=id).first()
            admin_id = session.get('id')
            if user:
                print(f"Deactivating user account: {user.full_name}")
                user.authenticate_student='false'
                user.authenticated_student=False
                # Log admin action
                db_session.add(UserActivityLog(
                    user_id=admin_id,
                    action='admin_deactivate_user',
                    page=request.path,
                    target_user_id=user.user_id,
                    timestamp=datetime.now(),
                    details=f"Admin deactivated user: {user.full_name} ({user.email})"
                ))
                # Commit the authentication change to database
                db_session.commit()
                print(f"Student {user.full_name} deactivated successfully in database.")
                # Send email notification (separate from database transaction)
                message=(f'Your account has been deactivated. To stay active please contact Ethics support team'
                f'{web_url}')
                try:
                    send_email(app,mail, message,[user.email])
                    print(f"Deactivation email sent to {user.email}")
                except Exception as e:
                    print("Email sending error:", str(e))
                
                return redirect(url_for('all_users'))
            else:
                flash("no such student on our data")
                return redirect(url_for("all_users"))
        except Exception as e:
            print(f"Error deactivating student: {str(e)}")
            db_session.rollback()
            flash("Error deactivating student")
            return redirect(url_for("all_users"))
    return redirect(url_for('all_users'))


@app.route('/register_reviewer', methods=['GET', 'POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def register_reviewer():
    current_actor = get_current_user()
    assignable_roles = sorted(allowed_assignable_roles(current_actor))

    messages=''
    if request.method == 'POST':
        full_name = request.form.get('full_name', '')
        staff_number = request.form.get('staff_number', '')
        email = request.form.get('email', '').lower()
        password = request.form.get('password', '')
        password2=request.form.get('password2')
        specialisation = request.form.get('specialisation')
        role=request.form.get('role')
        if role not in assignable_roles:
            return render_template('register_reviewer.html', messages=['You do not have permission to assign that role.'], assignable_roles=assignable_roles)
        if password == password2:

            # Validate password
            is_valid, message = validate_password(password)
            if not is_valid:
                return render_template('register_reviewer.html', messages=[message], assignable_roles=assignable_roles)

            # Check if user exists
            user = db_session.query(User).filter_by(email=email).first()
            if user:
                messages = 'Email already registered!'
                return render_template('register_reviewer.html', messages=[messages], assignable_roles=assignable_roles)
            
            try:
                # Hash the password properly
                
                # Create new user
                new_user = User(
                    full_name=full_name,
                    staff_number=staff_number,
                    email=email,
                    password=password,  # Make sure this is the hashed version
                    specialisation=specialisation,
                    role=role
                )
                
                db_session.add(new_user)
                db_session.commit()
                #sending email to the reviewers
                ###
                ### uncomment the code bellow for real testing

                
                try:
                    message=(f'An account was created on your behalf. ' 
                    f'Please follow the link {web_url} use your '
                    f' email as username and password is = {password}')
                    send_email(app,mail, message,[email])
                    messages = 'You have successfully registered!'
                except Exception as e:
                    print("Email sending error:", str(e))
                return redirect(url_for('reviewer_list'))
                
            except Exception as e:
                db_session.rollback()
                print("Registration error:", str(e))
                messages = 'Registration failed. Please try again.'
                return render_template('register_reviewer.html', messages=[messages], assignable_roles=assignable_roles)
        else:
            messages="Passwords mismatch"
            return render_template('register_reviewer.html', messages=[messages], assignable_roles=assignable_roles)
    messages= 'Please fill out the form completely!'
    return render_template('register_reviewer.html', messages=[messages], assignable_roles=assignable_roles)


@app.route('/super_admin_registration', methods=['GET', 'POST'])
@role_required('SUPER_ADMIN')
def super_admin_registration():
    messages = ''

    if request.method == 'POST':
        full_name = request.form.get('full_name', '')
        staff_number = request.form.get('staff_number', '')
        email = request.form.get('email', '').lower()
        password = request.form.get('password', '')
        password2 = request.form.get('password2')
        specialisation = request.form.get('specialisation')
        role = request.form.get('super_admin')

        if password != password2:
            messages = "Passwords mismatch"
            return render_template(
                'super_admin_registration.html',
                messages=[messages]
            )

        # Validate password
        is_valid, message = validate_password(password)
        if not is_valid:
            return render_template(
                'super_admin_registration.html',
                messages=[message]
            )

        # Check if user exists
        user = db_session.query(User).filter_by(email=email).first()
        if user:
            messages = 'Email already registered!'
            return render_template(
                'super_admin_registration.html',
                messages=[messages]
            )

        try:
            # Create new user
            new_user = User(
                full_name=full_name,
                staff_number=staff_number,
                email=email,
                password=password,  # 🔒 TODO: hash properly with bcrypt
                specialisation=specialisation,
                role=role
            )

            db_session.add(new_user)
            db_session.commit()

            # Prepare and send email
            message = (
                f'You have successfully created an account. '
                f'Please follow the link {web_url} '
                f'use your email as username and password is= {password}'
            )
            try:
                send_email(app, mail, message, [email])
            except Exception as e:
                print("Email sending error:", str(e))

            return redirect(url_for('login_page'))

        except Exception as e:
            db_session.rollback()
            error_trace = traceback.format_exc()  # full traceback
            print("Registration error:", str(e))
            print(error_trace)

            # Instead of generic template, return JSON so you see exact error
            return jsonify({
                "status": "error",
                "message": str(e),
                "trace": error_trace
            }), 500

    return render_template(
        'super_admin_registration.html',
        messages=['Please fill out the form completely!']
    )



###
###
### this is the function to focus on when intergrating MBA and Ethics
###
@app.route('/edit_user/<string:id>', methods=['POST','GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def edit_user(id):
    """
    Edit a user's profile. Only apply changes for fields that are provided and non-empty.
    - Leave existing DB values unchanged when inputs are blank on the frontend.
    - Change password only if both password and password2 are provided, match, and pass validation.
    - Avoid clobbering template 'role' with the posted role by using separate variables.
    """
    user = db_session.query(User).filter_by(user_id=id).first()
    user_id = session.get('id')
    user_profile = db_session.query(User).filter_by(user_id=user_id).first() if user_id else None
    current_role = user_profile.role.value if user_profile and user_profile.role else None

    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('all_users'))

    if not can_manage_ethics_user(user_profile, user):
        flash('You do not have permission to manage this user.', 'danger')
        return redirect(url_for('all_users'))

    if request.method == "POST":
        # Read form values and strip whitespace; treat empty strings as "not provided"
        full_name = (request.form.get('full_name') or '').strip()
        staff_number = (request.form.get('staff_number') or '').strip()
        email = (request.form.get('email') or '').strip().lower()
        new_password = (request.form.get('password') or '').strip()
        confirm_password = (request.form.get('password2') or '').strip()
        specialisation = (request.form.get('specialisation') or '').strip()
        new_role_raw = (request.form.get('role') or '').strip()

        try:
            changed = False
            password_was_changed = False

            # Apply non-empty field updates only
            if full_name:
                if user.full_name != full_name:
                    user.full_name = full_name
                    changed = True

            if staff_number:
                if getattr(user, 'staff_number', None) != staff_number:
                    user.staff_number = staff_number
                    changed = True

            if email:
                if user.email != email:
                    user.email = email
                    changed = True

            if specialisation:
                if getattr(user, 'specialisation', None) != specialisation:
                    user.specialisation = specialisation
                    changed = True

            # Role mapping (optional). Only update if a value was provided.
            if new_role_raw:
                new_role_value = None
                try:
                    # Try mapping by enum name (e.g., 'ADMIN')
                    new_role_value = UserRole[new_role_raw.upper()]
                except Exception:
                    # Try mapping by enum value string (case-insensitive)
                    try:
                        for m in UserRole:
                            if str(m.value).lower() == new_role_raw.lower():
                                new_role_value = m
                                break
                    except Exception:
                        new_role_value = None

                if new_role_value is not None:
                    if new_role_value.value not in allowed_assignable_roles(user_profile):
                        flash('You do not have permission to assign that role.', 'danger')
                        return redirect(url_for('edit_user', id=user.user_id))
                    if user.role != new_role_value:
                        user.role = new_role_value
                        changed = True
                else:
                    # If we can't map to enum, as a conservative choice do not change role
                    pass

            # Password behavior:
            # - If no new password is provided, keep the current password unchanged.
            # - If new password is provided, update password.
            # - If confirm field is also provided, it must match.
            if new_password:
                if confirm_password and new_password != confirm_password:
                    flash('Passwords do not match', 'danger')
                    return redirect(url_for('edit_user', id=user.user_id))

                old_password_hash = user.password
                print(f"[EDIT_USER DEBUG] User {user.user_id} old password-column hash: {old_password_hash}")

                # Use model setter: hashes and stores into DB column `password`
                user.password = new_password
                print(f"[EDIT_USER DEBUG] User {user.user_id} new password-column hash (pre-flush): {user.password}")
                changed = True

                # Ensure the new hash is written and immediately verifiable before final commit
                db_session.flush()
                print(f"[EDIT_USER DEBUG] User {user.user_id} password-column hash after flush: {user.password}")
                if not user.verify_password(new_password):
                    db_session.rollback()
                    flash('Password update failed. Please try again.', 'danger')
                    return redirect(url_for('edit_user', id=user.user_id))
                print(f"[EDIT_USER DEBUG] User {user.user_id} verify new password: True")
                password_was_changed = True
            elif confirm_password:
                flash('Enter a new password before confirming it.', 'danger')
                return redirect(url_for('edit_user', id=user.user_id))

            if changed:
                db_session.commit()
                print(f"[EDIT_USER DEBUG] User {user.user_id} commit successful")

                if password_was_changed:
                    try:
                        notify_msg = (
                            f'Your profile was changed on your behalf. '
                            f'Please follow the link {web_url} use your '
                            f'email as username and password is = {new_password}'
                        )
                        send_email(app, mail, notify_msg, [user.email])
                    except Exception as e:
                        app.logger.error(f"Failed to send email to {user.email}: {e}")

                flash('User details updated successfully.', 'success')
                return redirect(url_for('edit_user', id=user.user_id))
            else:
                flash('No changes to update.', 'info')
                return redirect(url_for('edit_user', id=user.user_id))

        except Exception as e:
            db_session.rollback()
            app.logger.error(f"Update error: {e}")
            flash('Update failed. Please try again.', 'danger')
            return redirect(url_for('edit_user', id=user.user_id))

    
    if not user_profile:
        flash("Session expired. Please login again.")
        return redirect(url_for('login_page'))

    role = user_profile.role.value
    return render_template(
        'edit_user.html',
        role=role,
        user_profile=user_profile,
        user=user,
        current_role=current_role,
        assignable_roles=allowed_assignable_roles(user_profile)
    )


@app.route('/admin/upload_student_docs', methods=['GET', 'POST'])
@app.route('/admin/upload_student_docs/<string:id>', methods=['GET', 'POST'])
@csrf.exempt
@role_required('ADMIN', 'SUPER_ADMIN')
def admin_upload_student_docs(id=None):
    # Optional: fetch user info for the sidebar if needed by the layout
    current_uid = session.get('id')
    current_user = db_session.query(User).filter_by(user_id=current_uid).first() if current_uid else None

    if request.method == 'POST':
        record_id = (request.form.get('id') or id or '').strip()
        form_type = (request.form.get('form_type') or '').strip()
        allowed_form_types = {'FormA', 'FormB', 'FormC'}
        
        if not record_id:
            flash("No Requirement ID provided. Please select one from the dropdown.", "admin-warning")
            return redirect(url_for('admin_upload_student_docs'))

        if form_type not in allowed_form_types:
            flash("Please select a valid form type.", "admin-warning")
            return redirect(url_for('admin_upload_student_docs'))

        deferred_documents = [
            defer(getattr(FormARequirements, field_name))
            for field_name in REQUIREMENT_FILE_FIELDS
            if hasattr(FormARequirements, field_name)
        ]
        req = (
            db_session.query(FormARequirements)
            .options(*deferred_documents)
            .filter_by(id=record_id)
            .first()
        )
        if not req:
            flash(f"Requirement record with ID '{record_id}' does not exist.", "admin-danger")
            return redirect(url_for('admin_upload_student_docs'))
        
        # Accept only the fields displayed for the selected form so hidden
        # sections cannot overwrite unrelated documents.
        form_file_fields = {
            'FormA': [
                'pending_note', 'permission_letter', 'prior_clearance',
                'prior_clearance1', 'research_tools_path', 'proposal_path',
                'impact_assessment_path', 'participation_info_sheet',
            ],
            'FormB': [
                'pending_note', 'permission_letter', 'prior_clearance_path',
                'proposal_path', 'prior_clearance1',
            ],
            'FormC': ['files'],
        }
        filename_mapping = {
            'research_tools_path': 'research_tools_filename',
            'proposal_path': 'proposal_filename',
            'impact_assessment_path': 'impact_assessment_filename',
            'participation_info_sheet': 'participation_info_filename',
            'prior_clearance_path': 'prior_clearance_path_filename',
            'ethics_evidence_path': 'ethics_evidence_path_filename'
        }

        uploaded_any = False
        try:
            req.form_type = form_type

            bool_fields = ['needs_permission', 'has_clearance', 'company_requires_jbs']
            for bool_field in bool_fields:
                setattr(req, bool_field, form_type != 'FormC' and bool_field in request.form)

            for field in form_file_fields[form_type]:
                file_obj = request.files.get(field)
                if not file_obj or not file_obj.filename:
                    continue

                file_data, safe_filename = read_file_blob(file_obj)

                setattr(req, field, file_data)
                filename_column = filename_mapping.get(field, f"{field}_filename")
                setattr(req, filename_column, safe_filename)
                uploaded_any = True

            req.updated_at = datetime.now()
            db_session.commit()
            if uploaded_any:
                flash(f"Successfully uploaded documents and updated flags for ID: {record_id}", "admin-success")
            else:
                flash(f"Updated status flags for ID: {record_id}", "admin-success")
        except ValueError as e:
            db_session.rollback()
            flash(str(e), "admin-warning")
        except Exception as e:
            db_session.rollback()
            app.logger.exception("Admin document upload failed for requirement %s", record_id)
            flash(f"Error saving database changes: {str(e)}", "admin-danger")

        return redirect(url_for('admin_upload_student_docs'))

    requirement_rows = (
        db_session.query(
            FormARequirements.id,
            FormARequirements.user_id,
            FormARequirements.form_type,
        )
        .distinct(FormARequirements.user_id)
        .all()
    )
    user_ids = [row.user_id for row in requirement_rows if row.user_id]
    student_names = {}
    if user_ids:
        for model in (FormA, FormB, FormC):
            for user_id, applicant_name in (
                db_session.query(model.user_id, model.applicant_name)
                .filter(model.user_id.in_(user_ids))
                .all()
            ):
                if applicant_name and user_id not in student_names:
                    student_names[user_id] = applicant_name

    all_requirements = [
        {
            'id': row.id,
            'user_id': row.user_id,
            'form_type': row.form_type,
            'student_name': student_names.get(
                row.user_id, f"Unknown (ID: {row.user_id})"
            ),
        }
        for row in requirement_rows
    ]
    admin_messages = [
        (category.removeprefix('admin-'), message)
        for category, message in get_flashed_messages(with_categories=True)
        if category.startswith('admin-')
    ]
    return render_template(
        "admin_upload_docs.html",
        requirements=all_requirements,
        user_profile=current_user,
        role='super_admin',
        admin_messages=admin_messages,
    )


@app.route('/super_admin', methods=['GET', 'POST'])
@role_required('SUPER_ADMIN')
def super_admin():
    user_id=session.get('id')
   
    
    user=db_session.query(User).filter(User.user_id==user_id).first()
    user_profile=db_session.query(User).filter_by(user_id=user_id).first()
    all_users = db_session.query(User).all()

    role=user.role.value
    return render_template("superadmin_dashboard.html",role=role,user_profile=user_profile,all_users=all_users,current_year=datetime.now().year)



### Enhanced Analytics Dashboard
### Form A Analytics with Professional Visualizations
@role_required('SUPER_ADMIN')
# =====================================================================================================
# ADMIN/SUPERADMIN STATUS MONITORING PAGE
# =====================================================================================================
@app.route('/admin_status_monitor', methods=['GET'])
def admin_status_monitor():
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))
    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user or user.role.value.lower() not in ['admin', 'super_admin']:
        return redirect(url_for('login_page'))
    role = user.role.value

    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20

    def normalize_datetime(dt_value):
        """Normalize DB datetimes so sorting and template math do not mix aware/naive values."""
        if not dt_value or not isinstance(dt_value, datetime):
            return dt_value
        return dt_value.astimezone(timezone.utc).replace(tzinfo=None) if dt_value.tzinfo else dt_value

    def apply_name_search(records, search_value):
        if not search_value:
            return records
        normalized_search = search_value.casefold()
        return [
            r for r in records
            if (r["applicant_name"] or "").casefold().find(normalized_search) != -1
        ]

    # Query FormA, FormB, FormC and combine
    def get_form_records(model, form_type):
        # Get the most recent form per user
        if hasattr(model, 'submitted_at'):
            latest_subq = (
                db_session.query(
                    model.user_id,
                    func.max(model.submitted_at).label("latest_date")
                )
                .group_by(model.user_id)
                .subquery()
            )
            query = db_session.query(model)
            query = query.join(latest_subq,
                (model.user_id == latest_subq.c.user_id) &
                (model.submitted_at == latest_subq.c.latest_date))
            results = run_db_query_with_retry(
                lambda: query.order_by(model.submitted_at.desc()).all()
            )
        elif hasattr(model, 'submission_date'):
            latest_subq = (
                db_session.query(
                    model.user_id,
                    func.max(model.submission_date).label("latest_date")
                )
                .group_by(model.user_id)
                .subquery()
            )
            query = db_session.query(model)
            query = query.join(latest_subq,
                (model.user_id == latest_subq.c.user_id) &
                (model.submission_date == latest_subq.c.latest_date))
            results = run_db_query_with_retry(
                lambda: query.order_by(model.submission_date.desc()).all()
            )
        else:
            query = db_session.query(model)
            results = run_db_query_with_retry(lambda: query.all())

        reviewer_ids = {
            reviewer_id
            for form in results
            for reviewer_id in [getattr(form, 'reviewer_name1', None), getattr(form, 'reviewer_name2', None)]
            if reviewer_id
        }
        reviewer_lookup = {}
        if reviewer_ids:
            try:
                reviewer_lookup = {
                    reviewer.user_id: reviewer.full_name
                    for reviewer in run_db_query_with_retry(
                        lambda: db_session.query(User)
                        .filter(User.user_id.in_(reviewer_ids))
                        .all()
                    )
                }
            except Exception:
                db_session.rollback()
                reviewer_lookup = {}

        records = []
        for form in results:
            rev1_name = None
            rev2_name = None
            if getattr(form, 'reviewer_name1', None):
                rev1_name = reviewer_lookup.get(form.reviewer_name1, form.reviewer_name1)
            if getattr(form, 'reviewer_name2', None):
                rev2_name = reviewer_lookup.get(form.reviewer_name2, form.reviewer_name2)

            rev1_data = {"date": None, "recommendation": None, "comments": None}
            rev2_data = {"date": None, "recommendation": None, "comments": None}

            slot0_reviewer = getattr(form, 'form_reviewed_by', None)
            slot1_reviewer = getattr(form, 'form_reviewed_by1', None)
            slot0_date = normalize_datetime(getattr(form, 'review_signature_date', None) or getattr(form, 'reviewer_date', None))
            slot1_date = normalize_datetime(getattr(form, 'review_signature_date1', None) or getattr(form, 'reviewer_date1', None))

            if slot0_reviewer == getattr(form, 'reviewer_name1', None):
                rev1_data["date"] = slot0_date
                rev1_data["recommendation"] = getattr(form, 'review_recommendation', None)
                rev1_data["comments"] = getattr(form, 'review_additional_comments', None)
            elif slot0_reviewer == getattr(form, 'reviewer_name2', None):
                rev2_data["date"] = slot0_date
                rev2_data["recommendation"] = getattr(form, 'review_recommendation', None)
                rev2_data["comments"] = getattr(form, 'review_additional_comments', None)

            if slot1_reviewer == getattr(form, 'reviewer_name1', None):
                rev1_data["date"] = slot1_date
                rev1_data["recommendation"] = getattr(form, 'review_recommendation1', None)
                rev1_data["comments"] = getattr(form, 'review_additional_comments1', None)
            elif slot1_reviewer == getattr(form, 'reviewer_name2', None):
                rev2_data["date"] = slot1_date
                rev2_data["recommendation"] = getattr(form, 'review_recommendation1', None)
                rev2_data["comments"] = getattr(form, 'review_additional_comments1', None)
            record = {
                "id": form.form_id,
                "form_type": form_type,
                "user_id":form.user_id,
                "applicant_name": getattr(form, 'applicant_name', None),
                "submitted_at": normalize_datetime(getattr(form, 'submitted_at', None) or getattr(form, 'submission_date', None)),
                "submitted": getattr(form, 'submitted', None),
                "risk_rating": getattr(form, 'risk_rating', None) or getattr(form, 'risk_level', None),
                "supervisor": getattr(form, 'supervisor', None) if hasattr(form, 'supervisor') else getattr(form, 'supervisor_name', None),
                "supervisor_date": normalize_datetime(getattr(form, 'supervisor_date', None)),
                "ethics_signature_date": normalize_datetime(getattr(form, 'ethics_signature_date', None)),
                "supervisor_recommendation": getattr(form, 'recommendation', None),
                "status": getattr(form, 'status', None),
                "first_reviewer_name": rev1_name,
                "second_reviewer_name": rev2_name,
                "first_reviewer": rev1_name,
                "second_reviewer": rev2_name,
                "signature_date": normalize_datetime(getattr(form, 'signature_date', None)),
                "recommendation": rev1_data["recommendation"],
                "first_reviewer_recommendation": rev1_data["recommendation"],
                "second_reviewer_recommendation": rev2_data["recommendation"],
                "first_reviewer_date": rev1_data["date"],
                "second_reviewer_date": rev2_data["date"],
                "first_reviewer_comment": rev1_data["comments"],
                "second_reviewer_comment": rev2_data["comments"],
                "certificate_issued": getattr(form, 'certificate_issued', None) if getattr(form, 'certificate_issued', None) is not None else 'Not Issued',
                "certificate_received": normalize_datetime(getattr(form, 'certificate_received', None)),
                "submitted_to_reviewers": getattr(form, 'submitted_to_reviewers', None),
                "submitted_to_rec": getattr(form, 'submitted_to_rec', None),
                "rec_status": getattr(form, 'rec_status', None),
                "rejected_or_accepted": getattr(form, 'rejected_or_accepted', None),
                "form_supervisor_status": getattr(form, 'form_supervisor_status', None),
                "ethics_status": getattr(form, 'ethics_status', None),
                "form_review_comment": getattr(form, 'form_review_comment', None),
                "form_review_comment1": getattr(form, 'form_review_comment1', None),
            }
            records.append(record)
        return records

    all_records = []
    for model, ftype in [(FormA, 'FORM A'), (FormB, 'FORM B'), (FormC, 'FORM C')]:
        all_records.extend(get_form_records(model, ftype))

    name_suggestions = sorted({
        record["applicant_name"]
        for record in all_records
        if record["applicant_name"]
    }, key=lambda name: name.casefold())
    filtered_records = apply_name_search(all_records, search_query)
    filtered_records.sort(
        key=lambda item: item["submitted_at"] or datetime.min,
        reverse=True
    )

    total = len(filtered_records)
    start_idx = max((page - 1) * per_page, 0)
    end_idx = start_idx + per_page
    forms_list = filtered_records[start_idx:end_idx]
    current = datetime.utcnow()

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )
    return render_template(
        'admin_status_monitor.html',
        role=role,
        current_time=current,
        forms_list=forms_list,
        page=page,
        per_page=per_page,
        total=total,
        search_query=search_query,
        name_suggestions=name_suggestions,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )

@app.route('/delete_user/<string:id>', methods=['GET','POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def delete_user(id):
    if request.method != 'POST':
        flash("Use the delete button to remove a user.", "warning")
        return redirect(url_for('all_users'))

    try:
        user_to_delete = db_session.query(User).filter_by(user_id=id).first()
        admin_id = session.get('id')
        if user_to_delete:
            actor = db_session.query(User).filter_by(user_id=admin_id).first() if admin_id else None
            if not can_manage_ethics_user(actor, user_to_delete):
                flash("You do not have permission to delete this user.", "danger")
                return redirect(url_for('all_users'))
            deleted_user_label = user_to_delete.email or user_to_delete.full_name or id

            # Remove dependent rows first to avoid FK constraint failures.
            db_session.query(LoginLog).filter_by(user_id=id).delete(synchronize_session=False)
            db_session.query(UserActivityLog).filter(
                or_(
                    UserActivityLog.user_id == id,
                    UserActivityLog.target_user_id == id
                )
            ).delete(synchronize_session=False)
            db_session.query(FormARequirements).filter_by(user_id=id).delete(synchronize_session=False)
            db_session.query(FormA).filter_by(user_id=id).delete(synchronize_session=False)
            db_session.query(FormB).filter_by(user_id=id).delete(synchronize_session=False)
            db_session.query(FormC).filter_by(user_id=id).delete(synchronize_session=False)

            # Use bulk delete for User to avoid ORM relationship synchronization
            # that can attempt to set child FKs to NULL (logs.user_id is NOT NULL).
            db_session.query(User).filter_by(user_id=id).delete(synchronize_session=False)
            # Log admin action
            if admin_id:
                db_session.add(UserActivityLog(
                    user_id=admin_id,
                    action='admin_delete_user',
                    page=request.path,
                    timestamp=datetime.now(),
                    details=f"Admin deleted user with user_id: {id} ({deleted_user_label})"
                ))
            db_session.commit()
            flash("User deleted successfully", "success")
            return redirect(url_for('all_users'))
        else:
            flash("User not found", "warning")
            return redirect(url_for('all_users'))

    except Exception as e:
        db_session.rollback()  # rollback to recover from DB error
        app.logger.exception("Delete user error for user_id=%s", id)
        flash("An error occurred while deleting the user. Check server logs for details.", "danger")
        return redirect(url_for('all_users'))


@app.route('/all_users', methods=['GET', 'POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def all_users():
    user_id=session.get('id')
    user_profile=db_session.query(User).filter_by(user_id=user_id).first()
    
    # If user_profile is None, redirect to login
    if not user_profile:
        flash("Session expired. Please login again.")
        return redirect(url_for('login_page'))
    
    # Pagination logic
    page = request.args.get('page', 1, type=int)
    per_page = 20
    users_query = db_session.query(User)
    if is_admin(user_profile):
        users_query = users_query.filter(User.role != UserRole.SUPER_ADMIN)
    search_query = (request.args.get('search') or '').strip()
    auth_status = (request.args.get('auth_status') or '').strip().lower()
    role_filter = (request.args.get('role') or '').strip().upper()

    allowed_role_filters = {role.name for role in UserRole}
    if is_admin(user_profile):
        allowed_role_filters.discard(UserRole.SUPER_ADMIN.name)
    if role_filter not in allowed_role_filters:
        role_filter = ''

    if search_query:
        search_term = f"%{search_query}%"
        users_query = users_query.filter(
            or_(
                User.full_name.ilike(search_term),
                User.email.ilike(search_term),
                User.role.cast(String).ilike(search_term)
            )
        )

    if auth_status == 'authenticated':
        users_query = users_query.filter(
            func.lower(func.coalesce(cast(User.authenticate_student, String), 'false')).in_(['true', '1'])
        )
    elif auth_status == 'not_authenticated':
        users_query = users_query.filter(
            func.lower(func.coalesce(cast(User.authenticate_student, String), 'false')).in_(['false', '0', 'none', ''])
        )

    if role_filter:
        users_query = users_query.filter(User.role == UserRole[role_filter])

    total_users = users_query.count()
    all_users = users_query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = (total_users + per_page - 1) // per_page
    role = user_profile.role.value

    # Backward-compatible fallback for legacy redirects using ?messages=...
    query_messages = []
    query_msg = request.args.get('messages')
    if query_msg:
        query_messages.append(query_msg)

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )

    return render_template(
        "user-list.html",
        role=role,
        user_profile=user_profile,
        all_users=all_users,
        messages=query_messages,
        search_query=search_query,
        auth_status=auth_status,
        role_filter=role_filter,
        page=page,
        total_pages=total_pages,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,

    )
@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    email = data.get('email')
    if not email:
        return jsonify({'message': 'Email is required'}), 400

    user = db_session.query(User).filter_by(email=email).first()
    if not user:
        # For security, don't reveal if the email exists
        return jsonify({'message': 'If that email exists, a reset link has been sent.'}), 200

    # Generate token and expiry
    token = generate_reset_token()
    user.reset_token = token
    user.reset_token_expiry = datetime.now(timezone.utc) + timedelta(minutes=15)
    session.commit()

    # Send email to user with the token
    try:
        send_email("motsietsepang@gmail.com", "UJ Ethics System: Password Resset", token)
    except:
        print("Error in ['/api/forgot-password'] ", "password reset token email sending failed.")
        return jsonify({'message': 'Server failed to send email. Contact admin.'}), 500

    return jsonify({'message': 'If that email exists, a reset code has been sent.'}), 200



@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    token = data.get('token')
    new_password = data.get('new_password')

    if not token or not new_password:
        return jsonify({'message': 'Token and new password are required'}), 400

    user = session.query(User).filter_by(reset_token=token).first()
    if not user or not user.reset_token_expiry or user.reset_token_expiry < datetime.now(timezone.utc):
        return jsonify({'message': 'Invalid or expired token'}), 400
    
    is_valid, message = validate_password(new_password)
    if not is_valid:
        return jsonify({'message': message}), 400

    # Update password
    user.password = new_password
    user.reset_token = None
    user.reset_token_expiry = None
    session.commit()

    return jsonify({'message': 'Password has been reset successfully.'}), 200

@app.route('/api/supervisors', methods=['GET'])
def get_supervisors():
    supervisors = db_session.query(User).filter(User.role == UserRole.SUPERVISOR.value).all()
    try:
        # Convert to list of dicts
        result = [
            {
                "user_id": sup.user_id,
                "full_name": sup.full_name,
                "email": sup.email
            }
            for sup in supervisors
        ]
        
        return jsonify(result), 200
    except Exception as e:
            print("Error fetching supervisors:", str(e))  # visible in Render logs
            return jsonify({"error": "Failed to fetch supervisors"}), 500


@app.route('/ethics_pack', methods=['GET'])
def ethics_pack():
    try:
        user_id = session.get('id')
        if not user_id:
            return redirect(url_for('login_page'))

        watched_video = db_session.query(Watched).filter_by(user_id=user_id).first()

        if watched_video is None:
            watched_video = Watched(user_id=user_id, watched=True)
            db_session.add(watched_video)
            db_session.commit()

        return render_template('ethics_pack.html')

    except Exception as e:
        db_session.rollback()
        print("Error in /ethics_pack:", str(e))
        msg = "An error occurred while loading the ethics pack."
        return render_template('ethics_pack.html', messages=[msg])



@app.route('/dashboard', methods=['GET'])
def dashboard ():
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))
    
    # Get form submission info
    form_a = db_session.query(FormA).filter_by(user_id=user_id).first()
    date_str = ''
    if form_a:
        date = form_a.submitted_at
        date_str = date.strftime('%Y-%m-%d')
    
    # Check if student has FormARequirements and if files still exist
    form_requirements = db_session.query(FormARequirements).filter_by(user_id=user_id).first()
    files_missing = False
    missing_files_info = {}
    
    if form_requirements:
        # Check if uploaded files still exist on the filesystem
        upload_folder = get_upload_folder()
        file_fields = [
            ('permission_letter', 'Permission Letters'),
            ('research_tools_path', 'Research Tools'),
            ('proposal_path', 'Research Proposal'),
            ('impact_assessment_path', 'Impact Assessment'),
            ('prior_clearance', 'Prior Clearance'),
            ('ethics_evidence_path', 'Ethics Evidence'),
            ('participation_info_sheet', 'Participant Info Sheet'),
            ('pending_note', 'Pending Note')
        ]
        
        for field_name, display_name in file_fields:
            file_data = getattr(form_requirements, field_name)
            if file_data:
                # If it's bytes or memoryview, it's a blob and definitely exists
                if isinstance(file_data, (bytes, memoryview)):
                    continue
                    
                # If it's a string, it's a legacy path, check if it exists on disk
                if isinstance(file_data, str):
                    if file_data.startswith('uploads/form/'):
                        clean_path = file_data.replace('uploads/form/', '')
                    elif file_data.startswith('form/'):
                        clean_path = file_data.replace('form/', '')
                    else:
                        clean_path = file_data
                    
                    full_path = os.path.join(upload_folder, clean_path)
                    if not os.path.exists(full_path):
                        files_missing = True
                        missing_files_info[field_name] = {
                            'display_name': display_name,
                            'original_path': file_data
                        }
    
    return render_template('dashboard.html', 
                         date=date_str, 
                         form_requirements=form_requirements,
                         files_missing=files_missing,
                         missing_files_info=missing_files_info)



# =====================================================================================================
# THIS SECTION IS FOR HANDLING FORMS
# =====================================================================================================

# FORM A =====================================================================================================

_REQUIREMENT_DOCUMENTS = {
    "FORM A": [
        ("permission_letter", "permission_letter_filename", "Company Permission Letter", "permission"),
        ("pending_note", "pending_note_filename", "Pending Note", "pending"),
        ("prior_clearance", "prior_clearance_filename", "Prior Ethical Clearance Document", "clearance"),
        ("prior_clearance1", "prior_clearance1_filename", "POPIA Information Letter and Consent Form", "personal"),
        ("research_tools_path", "research_tools_filename", "Survey / Questionnaire / Focus Group Questions", "always"),
        ("proposal_path", "proposal_filename", "Research Proposal", "always"),
        ("impact_assessment_path", "impact_assessment_filename", "Personal Impact Assessment Form", "always"),
        ("participation_info_sheet", "participation_info_filename", "Participation Information Sheet & Consent Form", "always"),
    ],
    "FORM B": [
        ("permission_letter", "permission_letter_filename", "Permission Letter", "permission"),
        ("pending_note", "pending_note_filename", "Pending Note", "pending"),
        ("prior_clearance_path", "prior_clearance_path_filename", "Prior Ethical Clearance Document", "clearance"),
        ("proposal_path", "proposal_filename", "Approved Proposal", "always"),
        ("prior_clearance1", "prior_clearance1_filename", "POPIA Information Letter and Consent Form", "personal"),
    ],
    "FORM C": [("files", "files_filename", "Research Proposal", "always")],
}


def _normalize_requirement_form_type(value):
    """Normalize legacy values such as FormA, FORM_A, and FORM A."""
    compact = ''.join(
        character for character in str(value or '').upper() if character.isalnum()
    )
    return {
        'FORMA': 'FORM A',
        'FORMB': 'FORM B',
        'FORMC': 'FORM C',
    }.get(compact, str(value or '').strip().upper())


def _student_requirement_documents(form, form_type):
    """Build the upload checklist without loading document bytes in the template."""
    documents = []
    for field, filename_field, label, condition in _REQUIREMENT_DOCUMENTS[form_type]:
        required = condition == "always"
        if form:
            required = required or (condition == "permission" and form.needs_permission is True)
            required = required or (condition == "pending" and form.needs_permission is None)
            required = required or (condition == "clearance" and form.has_clearance is True)
            required = required or (condition == "personal" and form.company_requires_jbs is True)
        filename = getattr(form, filename_field, None) if form else None
        has_file = bool(filename and getattr(form, field, None)) if form else False
        documents.append({"field": field, "label": label, "filename": filename,
                          "uploaded": has_file, "required": required})
    return documents


@app.route('/student/requirements/document/<string:field>', methods=['GET'])
def download_student_requirement_document(field):
    user_id = session.get('id')
    form = db_session.query(FormARequirements).filter_by(user_id=user_id).first() if user_id else None
    allowed = {item[0]: item[1] for items in _REQUIREMENT_DOCUMENTS.values() for item in items}
    if not form or field not in allowed:
        abort(404)
    data = getattr(form, field, None)
    filename = getattr(form, allowed[field], None)
    if not data or not filename:
        abort(404)
    if isinstance(data, memoryview):
        data = data.tobytes()
    if isinstance(data, bytes):
        data = decode_legacy_binary(data)
        mimetype, filename, _ = response_document_metadata(data, filename, field)
        return send_file(
            io.BytesIO(data), mimetype=mimetype, as_attachment=True,
            download_name=filename
        )
    safe_name = os.path.basename(str(data).replace('\\', '/'))
    path = os.path.join(get_upload_folder(), safe_name)
    if not os.path.isfile(path):
        abort(404)
    mimetype = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
    return send_file(path, mimetype=mimetype, as_attachment=True, download_name=filename)


@app.route('/student/requirements/document/<string:field>/delete', methods=['GET', 'POST'])
def delete_student_requirement_document(field):
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))
    form = db_session.query(FormARequirements).filter_by(user_id=user_id).first() if user_id else None
    allowed = {item[0]: item[1] for items in _REQUIREMENT_DOCUMENTS.values() for item in items}
    form_type = _normalize_requirement_form_type(form.form_type) if form else None
    if not form or field not in allowed or field not in {
        item[0] for item in _REQUIREMENT_DOCUMENTS.get(form_type, [])
    }:
        abort(404)
    endpoint = {
        "FORM A": "submit_form_a_requirements",
        "FORM B": "submit_form_b_requirements",
        "FORM C": "submit_form_c_requirements",
    }.get(form_type, "student_dashboard")
    if request.method == 'GET':
        flash("Use the Delete button and confirm the action to remove a document.", "info")
        return redirect(url_for(endpoint))
    setattr(form, field, None)
    setattr(form, allowed[field], None)
    form.updated_at = datetime.utcnow()
    db_session.commit()
    flash("Document deleted. It is now marked for re-upload.", "success")
    return redirect(url_for(endpoint))

@app.route('/submit_form_a_requirements', methods=['GET', 'POST'])
def submit_form_a_requirements():

    # Get user ID
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))
    
    form = db_session.query(FormARequirements).filter_by(user_id=user_id).first()

    if request.method == 'POST':
        try:
            UPLOAD_FOLDER = get_upload_folder()
            # Ensure upload directory exists before saving any files
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)

            # Get form data
            needs_permission = request.form.get('need_permission') == 'Yes'
            has_clearance = request.form.get('has_clearance') == 'Yes'
            company_requires_jbs = request.form.get('company_requires_jbs') == 'Yes'

            

            # Restrict only when the student has a meaningful active/submitted
            # record in another form type.
            if has_blocking_student_form(FormB, user_id, options=[
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file),
            ]) or has_blocking_student_form(FormC, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))
            formA=db_session.query(FormA).filter_by(user_id=user_id).first()
            
            # Handle permission letters (multiple files)
            permission_letter_data = None
            permission_letter_fname = None
            pending_note_data = None
            pending_note_fname = None
            
            needs_permission = request.form.get('need_permission')
            if needs_permission == 'Yes':
                uploaded_files = request.files.getlist('permission_letter[]')
                valid_files = [f for f in uploaded_files if f and f.filename != '']
                validated_files = [read_file_blob(f) for f in valid_files]
                
                if len(validated_files) > 1:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zf:
                        for file_data, safe_name in validated_files:
                            zf.writestr(safe_name, file_data)
                    permission_letter_data = zip_buffer.getvalue()
                    permission_letter_fname = "permission_letters.zip"
                    needs_permission = True
                elif len(validated_files) == 1:
                    permission_letter_data, permission_letter_fname = validated_files[0]
                    needs_permission = True
            else:
                if needs_permission == 'No':
                    needs_permission = False
                elif needs_permission == "Pending":
                    needs_permission = None
                    pending_note_data, pending_note_fname = read_file_blob(request.files.get('pending_note'))
            
            # Handle other file uploads (single files)
            research_tools_data, research_tools_fname = read_file_blob(request.files.get('research_tools_path'))
            prior_clearance_data, prior_clearance_fname = read_file_blob(request.files.get('prior_clearance')) if has_clearance else (None, None)
            prior_clearance1_data, prior_clearance1_fname = read_file_blob(request.files.get('prior_clearance1')) if company_requires_jbs else (None, None)
            proposal_data, proposal_fname = read_file_blob(request.files.get('proposal_path'))
            impact_data, impact_fname = read_file_blob(request.files.get('impact_assessment_path'))
            participation_data, participation_fname = read_file_blob(request.files.get('participation_info_sheet'))


            # Save or update DB record
            form = db_session.query(FormARequirements).filter_by(user_id=user_id).first()

            # Validate required/conditional uploads using incoming files OR existing stored values
            missing_files = []

            def has_file(new_data, existing_data):
                return bool(new_data) or bool(existing_data)

            if not has_file(research_tools_data, form.research_tools_path if form else None):
                missing_files.append('Survey / Questionnaire / Focus Group Questions')
            if not has_file(proposal_data, form.proposal_path if form else None):
                missing_files.append('Research Proposal')
            if not has_file(impact_data, form.impact_assessment_path if form else None):
                missing_files.append('Personal Impact Assessment Form')
            if not has_file(participation_data, form.participation_info_sheet if form else None):
                missing_files.append('Participation Information Sheet & Consent Form')

            if needs_permission is True and not has_file(permission_letter_data, form.permission_letter if form else None):
                missing_files.append('Company Permission Letter')
            if needs_permission is None and not has_file(pending_note_data, form.pending_note if form else None):
                missing_files.append('Pending Note')
            if has_clearance and not has_file(prior_clearance_data, form.prior_clearance if form else None):
                missing_files.append('Prior Ethical Clearance Document')
            if company_requires_jbs and not has_file(prior_clearance1_data, form.prior_clearance1 if form else None):
                missing_files.append('JBS Ethical Clearance Document')

            if missing_files:
                flash('Missing required upload(s): ' + ', '.join(missing_files), 'danger')
                return redirect(url_for('submit_form_a_requirements'))
            
            if form:
                form.needs_permission = needs_permission
                form.has_clearance = has_clearance
                form.company_requires_jbs = company_requires_jbs
                form.form_type = "FORM A"
                
                if permission_letter_data:
                    form.permission_letter = permission_letter_data
                    form.permission_letter_filename = permission_letter_fname
                
                if research_tools_data:
                    form.research_tools_path = research_tools_data
                    form.research_tools_filename = research_tools_fname
                    
                if prior_clearance_data:
                    form.prior_clearance = prior_clearance_data
                    form.prior_clearance_filename = prior_clearance_fname
                    
                if prior_clearance1_data:
                    form.prior_clearance1 = prior_clearance1_data
                    form.prior_clearance1_filename = prior_clearance1_fname
                    
                if proposal_data:
                    form.proposal_path = proposal_data
                    form.proposal_filename = proposal_fname
                    
                if pending_note_data:
                    form.pending_note = pending_note_data
                    form.pending_note_filename = pending_note_fname
                    
                if impact_data:
                    form.impact_assessment_path = impact_data
                    form.impact_assessment_filename = impact_fname
                    
                if participation_data:
                    form.participation_info_sheet = participation_data
                    form.participation_info_filename = participation_fname
                
                db_session.add(form)
                db_session.commit()
                if form and not formA:
                    return redirect(url_for('form_a_sec1'))
                else:
                    return redirect(url_for('student_dashboard'))
                
            else:
                form = FormARequirements(
                    user_id=user_id,
                    form_type="FORM A",
                    needs_permission=needs_permission,
                    permission_letter=permission_letter_data,
                    permission_letter_filename=permission_letter_fname,
                    has_clearance=has_clearance,
                    pending_note=pending_note_data,
                    pending_note_filename=pending_note_fname,
                    company_requires_jbs=company_requires_jbs,
                    research_tools_path=research_tools_data,
                    research_tools_filename=research_tools_fname,
                    proposal_path=proposal_data,
                    proposal_filename=proposal_fname,
                    impact_assessment_path=impact_data,
                    impact_assessment_filename=impact_fname,
                    prior_clearance=prior_clearance_data,
                    prior_clearance_filename=prior_clearance_fname,
                    prior_clearance1=prior_clearance1_data,
                    prior_clearance1_filename=prior_clearance1_fname,
                    participation_info_sheet=participation_data,
                    participation_info_filename=participation_fname
                )
                db_session.add(form)
                db_session.commit()

                # New form flow continues to section 1 as before
                return redirect(url_for('form_a_sec1'))

        except UploadValidationError as e:
            db_session.rollback()
            flash(str(e), 'danger')
            return redirect(url_for('submit_form_a_requirements'))
        except Exception as e:
            db_session.rollback()
            return jsonify({'error': str(e)}), 500

    return render_template('form-a-upload.html', from_dashboard=form,
                           uploaded_documents=_student_requirement_documents(form, "FORM A"))

@app.route('/submit_form_c_requirements', methods=['GET', 'POST'])
def submit_form_c_requirements():


    # Get user ID from session (adjust based on your auth system)
    user_id = session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
            
    
    # Check if form exists for this user
    form = db_session.query(FormARequirements).filter_by(user_id=user_id).first()
    if request.method=='POST':
        try:
            UPLOAD_FOLDER = get_upload_folder()
            

            if has_blocking_student_form(FormB, user_id, options=[
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file),
            ]):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))
                
            if has_blocking_student_form(FormA, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            formC = get_latest_student_form_record(FormC, user_id)
            
            # Save files based on form field names (corrected from request.form to request.files)
            proposal_data, proposal_filename = read_file_blob('proposal')
          
            # Validate required files
            if not proposal_data and not (form and form.files):
                return jsonify({'error': 'Missing required files'}), 400
                
            
         
            if form:
                # Update existing form
                form.user_id=user_id
                form.form_type="FORM C"
                form.updated_at=datetime.now()
                
                if proposal_data:
                    form.files = proposal_data
                    form.files_filename = proposal_filename

                db_session.add(form)
                db_session.commit()
                if form and not formC:
                    return redirect(url_for('form_c_sec1'))
                else:
                    return redirect(url_for('student_dashboard'))
                
            else:
                # Create new record
                form = FormARequirements(
                    user_id=user_id,
                    form_type="FORM C",
                    updated_at=datetime.now(),
                    files = proposal_data,
                    files_filename = proposal_filename
                )
            
                db_session.add(form)
                db_session.commit()
            
                return redirect(url_for('form_c_sec1'))
            
        except UploadValidationError as e:
            db_session.rollback()
            flash(str(e), 'danger')
            return redirect(url_for('submit_form_c_requirements'))
        except Exception as e:
            db_session.rollback()
            return jsonify({'error': str(e)}), 500
  
    return render_template('form-c-upload.html', from_dashboard=form,
                           uploaded_documents=_student_requirement_documents(form, "FORM C"))


@app.route('/submit_form_b_requirements', methods=['GET', 'POST'])
def submit_form_b_requirements():
    # Get user ID from session
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))
        
    # Check if form exists for this user
    form = db_session.query(FormARequirements).filter_by(user_id=user_id).first()

    if request.method == 'POST':
        try:
            # Basic status flags
            needs_permission_val = request.form.get('need_permission')
            has_clearance = request.form.get('has_clearance') == 'Yes'
            has_personal_information = request.form.get('company_requires_jbs') == 'Yes'
            # Note: has_ethics_evidence might not be in Form B template, but let's be safe
            has_ethics_evidence = request.form.get('has_ethics_evidence') == 'Yes'
            
            # Check for conflict with other forms
            if has_blocking_student_form(FormA, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))
                
            if has_blocking_student_form(FormC, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))
            
            formB = db_session.query(FormB).options(
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file)
            ).filter_by(user_id=user_id).first()

            # Handle permission letters (multiple files possible)
            permission_letter_data = None
            permission_letter_fname = None
            
            if needs_permission_val == 'Yes':
                uploaded_files = request.files.getlist('permission_letter_path[]')
                valid_files = [f for f in uploaded_files if f and f.filename != '']
                validated_files = [read_file_blob(f) for f in valid_files]
                
                if len(validated_files) > 1:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zf:
                        for file_data, safe_name in validated_files:
                            zf.writestr(safe_name, file_data)
                    permission_letter_data = zip_buffer.getvalue()
                    permission_letter_fname = "permission_letters.zip"
                elif len(validated_files) == 1:
                    permission_letter_data, permission_letter_fname = validated_files[0]
            
            # Convert needs_permission to boolean/None for storage
            needs_permission_bool = None
            if needs_permission_val == 'Yes':
                needs_permission_bool = True
            elif needs_permission_val == 'No':
                needs_permission_bool = False
            # 'Pending' leaves it as None
            
            # Handle other files
            prior_clearance_data, prior_clearance_fname = read_file_blob('prior_clearance_path') if has_clearance else (None, None)
            ethics_evidence_data, ethics_evidence_fname = read_file_blob('ethics_evidence') if has_ethics_evidence else (None, None)
            proposal_data, proposal_fname = read_file_blob('proposal_path')
            pending_note_data, pending_note_filename = read_file_blob('pending_note_path')
            popia_data, popia_filename = read_file_blob('prior_clearance1') if has_personal_information else (None, None)

            if has_personal_information and not popia_data and not (form and form.prior_clearance1):
                flash("POPIA Information Letter and Consent Form is required", "error")
                return redirect(url_for('submit_form_b_requirements'))
            
            if not form and not proposal_data:
                flash("Proposal is required", "error")
                return redirect(url_for('submit_form_b_requirements'))

            if not form:
                form = FormARequirements(user_id=user_id, form_type="FORM B")
                db_session.add(form)
            
            # Update fields
            form.form_type = "FORM B"
            form.needs_permission = needs_permission_bool
            form.has_clearance = has_clearance
            form.company_requires_jbs = has_personal_information
            form.has_ethics_evidence = has_ethics_evidence
            form.updated_at = datetime.utcnow()
            
            if permission_letter_data:
                form.permission_letter = permission_letter_data
                form.permission_letter_filename = permission_letter_fname
            
            if prior_clearance_data:
                form.prior_clearance_path = prior_clearance_data
                form.prior_clearance_path_filename = prior_clearance_fname
                
            if ethics_evidence_data:
                form.ethics_evidence_path = ethics_evidence_data
                form.ethics_evidence_path_filename = ethics_evidence_fname
                
            if proposal_data:
                form.proposal_path = proposal_data
                form.proposal_filename = proposal_fname
                
            if pending_note_data:
                form.pending_note = pending_note_data
                form.pending_note_filename = pending_note_filename

            if popia_data:
                form.prior_clearance1 = popia_data
                form.prior_clearance1_filename = popia_filename

            db_session.commit()
            
            if not formB:
                return redirect(url_for('form_b_sec1'))
            return redirect(url_for('student_dashboard'))
            
        except UploadValidationError as e:
            db_session.rollback()
            flash(str(e), "danger")
            return redirect(url_for('submit_form_b_requirements'))
        except Exception as e:
            traceback.print_exc()
            db_session.rollback()
            flash(f"Error submitting requirements: {str(e)}", "danger")
            return redirect(url_for('submit_form_b_requirements'))
            
    return render_template('form-b-upload.html', from_dashboard=form,
                           uploaded_documents=_student_requirement_documents(form, "FORM B"))

            
@app.route('/edit-form-a/<form_id>', methods=['GET'])
def edit_form_a(form_id):
    data = getFormAData(form_id)
    if data:
        session['active_forma_id'] = data.form_id
    return render_template('form-a-section1.html', form_data=data)



# ---------------- Section 1 ------------------
@app.route('/form_a_sec1', methods=['GET', 'POST'])
def form_a_sec1 ():

    sup_list = getSupervisorsList()

    try:
        if request.method == 'POST':
            # Verify user is logged in
            user_id = session.get('id')
            if not user_id:
                return jsonify({'error': 'Unauthorized'}), 401

            # Check if user already submitted other forms
            formB_options = [
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file)
            ]
            if has_blocking_student_form(FormB, user_id, options=formB_options):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            if has_blocking_student_form(FormC, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            # Get form data
            form_data = request.form
            user = db_session.query(User).filter(User.user_id == user_id).first()
            if not user:
                flash("User not found. Please log in again.", "danger")
                return redirect(url_for("login"))

            supervisor = get_student_supervisor_or_flash(user)
            if not supervisor:
                return redirect(url_for("student_dashboard"))

            form_requirements = db_session.query(FormARequirements).filter(FormARequirements.user_id == user_id).first()
            if not form_requirements:
                flash("Form A requirements not found. Please complete them first.", "warning")
                return redirect(url_for("student_dashboard"))

            # Create or update record
            form, error_response = _get_or_create_forma_draft(user_id, form_data)
            if error_response:
                return error_response

            form.attachment_id = form_requirements.id
            _apply_forma_autosave_payload(form, form_data, section='sec1', include_declaration=False)
            form.email = user.email
            form.supervisor = supervisor.full_name
            form.supervisor_email = supervisor.email

            db_session.add(form)
            db_session.commit()
            flash("Form submitted successfully.", "success")
            return redirect(url_for("form_a_sec2"))

        # GET request: always load/create a single active draft for autosave.
        current_user_id = session.get('id')
        form_data, error_response = _get_or_create_forma_draft(current_user_id, {})
        if error_response:
            return error_response
        db_session.commit()
        return render_template('form-a-section1.html', supervisors=sup_list, form_data=form_data)

    except Exception as e:
        db_session.rollback()
        print(f"⚠️ Error in form_a_section1: {str(e)}")  # Optional: use logging
        flash("An unexpected error occurred while submitting the form. Please try again.", "danger")
        return redirect(url_for("student_dashboard"))

    finally:
        db_session.close()

   
@app.route('/submit_form_a_sec1', methods=['GET', 'POST'])
def submit_form_a_sec1():
    try:
        # Ensure user is logged in
        user_id = session.get('id')
        if not user_id:
            flash("You must be logged in to submit this form.", "warning")
            return redirect(url_for("login"))

        # Ensure attachments are set
        attachment_id = session.get('formA-attachments_id')
        if not attachment_id:
            flash("Attachment information is missing. Please upload attachments first.", "warning")
            return redirect(url_for("form_a_sec1"))

        # Reuse the active draft when present; create only if no draft exists.
        formA_record, error_response = _get_or_create_forma_draft(user_id, request.form)
        if error_response:
            return error_response

        formA_record.user_id = user_id
        formA_record.attachment_id = attachment_id
        for key, value in request.form.items():
            if key in {'csrf_token', 'forma_id'}:
                continue
            if hasattr(FormA, key):
                setattr(formA_record, key, value)

        db_session.add(formA_record)
        db_session.commit()

        flash("Form A (Section 1) submitted successfully!", "success")
        return redirect(url_for("form_a_sec2"))

    except KeyError as ke:
        db_session.rollback()
        flash(f"Missing required field: {ke}", "danger")
        print(f"⚠️ Missing key in form submission: {ke}")
        return redirect(url_for("form_a_sec1"))

    except Exception as e:
        db_session.rollback()
        flash("An unexpected error occurred while submitting your form. Please try again.", "danger")
        print(f"⚠️ Error in /submit_form_a_sec1: {str(e)}")

        return redirect(url_for("form_a_sec1"))

    finally:
        db_session.close()


# ---------------- Section 2 ------------------
@app.route('/form_a_sec2', methods=['GET', 'POST'])
def form_a_sec2():
    try:
        data = request.form
        if request.method == 'POST':
            def to_bool(val):
                if isinstance(val, bool):
                    return val
                if val in [None, '']:
                    return False
                return str(val).lower() in ['yes', 'true', '1', 'on', 'checked']

            # ✅ Ensure user session exists
            user_id = session.get('id')
            if not user_id:
                flash("You must be logged in to fill this form.", "warning")
                return redirect(url_for("login"))

            # ✅ Check if the user already filled other forms
            formB_options = [
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file)
            ]
            if has_blocking_student_form(FormB, user_id, options=formB_options):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            if has_blocking_student_form(FormC, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            # ✅ Fetch the existing Form A record
            form = _get_latest_forma_for_user(user_id)
            if not form:
                flash("No existing Form A record found for this user.", "danger")
                return redirect(url_for("form_a_sec1"))

            _apply_forma_autosave_payload(form, data, section='sec2', include_declaration=False)

            # ✅ Commit to DB
            db_session.add(form)
            db_session.commit()

            flash("Form A Section 2 submitted successfully.", "success")
            return redirect(url_for("form_a_sec3"))

        # GET method fallback - preload existing values for checkbox/text persistence
        user_id = session.get('id')
        form_data = _get_latest_forma_for_user(user_id)
        return render_template('form-a-section2.html', form_data=form_data)

    except KeyError as ke:
        db_session.rollback()
        flash(f"Missing required session or field: {ke}", "danger")
        print(f"⚠️ Missing key in /form_a_sec2: {ke}")
        return redirect(url_for("form_a_sec2"))

    except Exception as e:
        db_session.rollback()
        flash("An unexpected error occurred while submitting Form A Section 2. Please try again.", "danger")
        print(f"⚠️ Error in /form_a_sec2: {str(e)}")
        return redirect(url_for("form_a_sec2"))

    finally:
        db_session.close()

@app.route('/submit_form_a_sec2', methods=['GET', 'POST'])
def submit_form_a_sec2():
    try:
        if request.method == 'POST':
            user_id = session.get('id')
            if not user_id:
                flash("Unauthorized access. Please log in to continue.", "warning")
                return redirect(url_for("login"))

            form_record = _get_latest_forma_for_user(user_id)
            if not form_record:
                flash("No existing Form A record found for this user.", "danger")
                return redirect(url_for("form_a_sec1"))

            data = _build_autosave_data(request.form)

            bool_fields = [
                'survey', 'focus_groups', 'observations', 'interviews', 'documents',
                'vulnerable_communities', 'age_range', 'uj_employees', 'vulnerable',
                'non_english', 'own_students', 'poverty', 'no_education',
                'disclosure', 'discomfiture', 'deception',
                'sensitive', 'prejudice', 'intrusive_techniques', 'illegal_activities',
                'personal', 'available_records', 'inventories', 'risk_activities',
                'incentives', 'financial_costs', 'reward', 'conflict',
                'uj_premises', 'uj_facilities', 'uj_funding'
            ]

            for field in bool_fields:
                setattr(form_record, field, _autosave_str_to_bool(data[field]) if field in data else False)

            form_record.assessment_other_specify = request.form.get('assessment_other_specify', '')
            form_record.vulnerable_other_specify = request.form.get('vulnerable_other_specify', '')
            form_record.activity_specify = request.form.get('activity_specify', '')
            form_record.vulnerable_comments_1 = request.form.get('vulnerable_comments_1', '')
            form_record.vulnerable_comments_2 = request.form.get('vulnerable_comments_2', '')
            form_record.vulnerable_comments_3 = request.form.get('vulnerable_comments_3', '')
            form_record.risk_rating = request.form.get('risk_rating', '')
            form_record.risk_justification = request.form.get('risk_justification', '')
            form_record.benefits_description = request.form.get('benefits_description', '')
            form_record.risk_mitigation = request.form.get('risk_mitigation', '')
            form_record.apply_comments = request.form.get('apply_comments', '')
            form_record.other_sec2 = request.form.get('other_sec2', '')

            form_record.interviews_one = _autosave_str_to_bool(data.get('interviews')) if 'interviews' in data else False
            form_record.documents_one = _autosave_str_to_bool(data.get('documents')) if 'documents' in data else False

            db_session.commit()
            flash("Form A Section 2 submitted successfully!", "success")
            return redirect(url_for("form_a_sec3"))

        user_id = session.get('id')
        form_data = _get_latest_forma_for_user(user_id)
        return render_template('form-a-section2.html', form_data=form_data)

    except KeyError as ke:
        db_session.rollback()
        flash(f"Missing required data: {ke}", "danger")
        print(f"⚠️ KeyError in /submit_form_a_sec2: {ke}")
        return redirect(url_for("form_a_sec2"))

    except Exception as e:
        db_session.rollback()
        flash("An unexpected error occurred while submitting the form. Please try again.", "danger")
        print(f"⚠️ Exception in /submit_form_a_sec2: {e}")
        return redirect(url_for("form_a_sec2"))

    finally:
        db_session.close()

# ---------------- Section 3 ------------------
@app.route('/form_a_sec3', methods=['GET', 'POST'])
def form_a_sec3():
    try:
        if request.method == 'POST':
            data = request.form
            def to_yes_no(val):
                if isinstance(val, bool):
                    return 'Yes' if val else 'No'
                if val in [None, '']:
                    return ''
                return 'Yes' if str(val).lower() in ['yes', 'true', '1', 'on', 'checked'] else 'No'

            user_id = session.get('id')

            # ✅ Check user authentication
            if not user_id:
                flash("Unauthorized access. Please log in to continue.", "warning")
                return redirect(url_for("login"))

            # ✅ Retrieve existing Form A record
            form = _get_latest_forma_for_user(user_id)
            if not form:
                flash("No existing Form A record found for this user.", "danger")
                return redirect(url_for("form_a_sec1"))

            _apply_forma_autosave_payload(form, data, section='sec3', include_declaration=False)

            # ✅ Save updates
            db_session.add(form)
            db_session.commit()

            flash("Form A Section 3 submitted successfully!", "success")
            return redirect(url_for("form_a_sec4"))

        # ✅ Handle GET request
        user_id = session.get('id')
        form_data = _get_latest_forma_for_user(user_id)
        return render_template('form-a-section3.html', form_data=form_data)

    except KeyError as ke:
        db_session.rollback()
        flash(f"Missing field or session key: {ke}", "danger")
        print(f"⚠️ KeyError in /form_a_sec3: {ke}")
        return redirect(url_for("form_a_sec3"))

    except Exception as e:
        db_session.rollback()
        flash("An unexpected error occurred while submitting Form A Section 3. Please try again.", "danger")
        print(f"⚠️ Exception in /form_a_sec3: {e}")
        return redirect(url_for("form_a_sec3"))

    finally:
        db_session.close()

@app.route('/form_a_upload', methods=['GET'])
def form_a_upload ():
    return render_template('form-a-upload.html')


###
###
### this is the function to focus on when intergrating MBA and Ethics
###
@app.route('/back_end/monitor', methods=['GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def monitor():
    user_id = session.get('id')

    if not user_id:
        flash("Your session has expired. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    user_profile = db_session.query(User).filter_by(user_id=user_id).first()

    if not user_profile:
        session.clear()
        flash("Your account could not be found. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    role = user_profile.role.value if user_profile.role else ''

    # -----------------------------
    # FILTER VALUES FROM TEMPLATE
    # -----------------------------
    filter_full_name = (request.args.get('filter_full_name') or '').strip()
    filter_email = (request.args.get('filter_email') or '').strip()
    filter_role = (request.args.get('filter_role') or '').strip()

    page = request.args.get('page', 1, type=int)
    per_page = 15

    # ------------------------------------------------
    # MAIN QUERY
    # Default view shows STUDENTS.
    # If a role is selected, it filters by that role.
    # ------------------------------------------------
    users_query = db_session.query(User)

    if filter_role:
        try:
            users_query = users_query.filter(User.role == UserRole[filter_role])
        except KeyError:
            users_query = users_query.filter(User.role == UserRole.STUDENT)
    else:
        users_query = users_query.filter(User.role == UserRole.STUDENT)

    if filter_full_name:
        users_query = users_query.filter(
            func.lower(User.full_name).like(f"%{filter_full_name.lower()}%")
        )

    if filter_email:
        users_query = users_query.filter(
            func.lower(User.email).like(f"%{filter_email.lower()}%")
        )

    users_query = users_query.order_by(User.full_name.asc())

    total_users = users_query.count()

    paginated_users = users_query.offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    total_pages = (total_users + per_page - 1) // per_page

    users_list = []
    for user in paginated_users:
        users_list.append({
            "user_id": user.user_id,
            "full_name": user.full_name,
            "email": user.email,
            "role": user.role.value if user.role else '',
        })

    # ------------------------------------------------
    # LIST FOR DATALIST OPTIONS
    # This is separate from pagination so the datalist
    # can still show all names/emails.
    # ------------------------------------------------
    filter_users = db_session.query(User).order_by(User.full_name.asc()).all()

    filter_users_list = []
    for user in filter_users:
        filter_users_list.append({
            "full_name": user.full_name,
            "email": user.email,
        })

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )

    return render_template(
        'monitor.html',
        role=role,
        users_list=users_list,
        filter_users_list=filter_users_list,
        page=page,
        total_pages=total_pages,
        total_users=total_users,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )





@app.route('/back_end/monitor_forms/view_forms/<string:user_id>', methods=['GET','POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def monitor_forms(user_id):
    current_user_id = session.get('id')
    user_profile = db_session.query(User).filter_by(user_id=current_user_id).first()
    student = db_session.query(User).filter_by(user_id=user_id).first()

    if not user_profile or not student or student.role != UserRole.STUDENT:
        flash("The selected student could not be found.", "danger")
        return redirect(url_for('all_users', role='STUDENT'))

    form_models = {"A": FormA, "B": FormB, "C": FormC}

    if request.method == "POST":
        selected_forms = request.form.getlist('selected_forms')
        if not selected_forms:
            flash("Select at least one form to archive.", "warning")
            return redirect(url_for('monitor_forms', user_id=user_id))

        archived_count = 0
        try:
            for selection in dict.fromkeys(selected_forms):
                try:
                    form_type, form_id = selection.split(':', 1)
                except ValueError:
                    continue
                model = form_models.get(form_type.upper())
                if model is None:
                    continue

                form_record = db_session.query(model).filter_by(
                    form_id=form_id,
                    user_id=user_id,
                ).first()
                if not form_record:
                    continue

                snapshot = {}
                for column in form_record.__table__.columns:
                    value = getattr(form_record, column.name)
                    if isinstance(value, (bytes, bytearray, memoryview)):
                        raw_value = bytes(value)
                        value = {
                            "__archive_type__": "bytes",
                            "base64": base64.b64encode(raw_value).decode("ascii"),
                            "size": len(raw_value),
                        }
                    elif isinstance(value, (datetime, date)):
                        value = value.isoformat()
                    elif hasattr(value, 'value'):
                        value = value.value
                    snapshot[column.name] = value

                submitted_at = (
                    getattr(form_record, 'submitted_at', None)
                    or getattr(form_record, 'submission_date', None)
                )
                db_session.add(ArchivedEthicsForm(
                    form_type=f"Form {form_type.upper()}",
                    original_form_id=form_record.form_id,
                    student_user_id=user_id,
                    student_name=student.full_name,
                    student_email=student.email,
                    original_created_at=getattr(form_record, 'created_at', None),
                    original_submitted_at=submitted_at,
                    archived_by_user_id=current_user_id,
                    archive_reason=(request.form.get('archive_reason') or '').strip() or None,
                    snapshot_json=json.dumps(snapshot, ensure_ascii=False, default=str),
                ))
                db_session.delete(form_record)
                archived_count += 1

            if not archived_count:
                flash("None of the selected forms could be archived.", "warning")
                return redirect(url_for('monitor_forms', user_id=user_id))

            db_session.commit()
            flash(
                f"{archived_count} form{'s' if archived_count != 1 else ''} moved to the archive.",
                "success",
            )
        except Exception:
            db_session.rollback()
            app.logger.exception("Unable to archive forms for student %s", user_id)
            flash("The selected forms could not be archived. No changes were saved.", "danger")

        return redirect(url_for('monitor_forms', user_id=user_id))

    forms = []
    for form_type, model in form_models.items():
        for record in db_session.query(model).filter_by(user_id=user_id).all():
            forms.append({
                "selection_value": f"{form_type}:{record.form_id}",
                "form_type": f"Form {form_type}",
                "form_id": record.form_id,
                "applicant_name": getattr(record, 'applicant_name', None) or student.full_name,
                "email": (
                    getattr(record, 'email_address', None)
                    or getattr(record, 'email', None)
                    or student.email
                ),
                "submitted_at": (
                    getattr(record, 'submitted_at', None)
                    or getattr(record, 'submission_date', None)
                ),
            })

    return render_template(
        'monitor.html',
        role=user_profile.role.value,
        user_profile=user_profile,
        student=student,
        users_list=[],
        form=forms,
    )


@app.route('/back_end/archived_forms', methods=['GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def archived_forms():
    current_user = db_session.query(User).filter_by(user_id=session.get('id')).first()
    archives = db_session.query(ArchivedEthicsForm).order_by(
        ArchivedEthicsForm.archived_at.desc()
    ).all()
    return render_template(
        'archived_forms.html',
        role=current_user.role.value,
        user_profile=current_user,
        archives=archives,
    )


@app.route('/back_end/archived_forms/<string:form_type_key>/<string:archive_id>', methods=['GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def archived_form_detail(form_type_key, archive_id):
    current_user = db_session.query(User).filter_by(user_id=session.get('id')).first()
    archive = db_session.query(ArchivedEthicsForm).filter_by(archive_id=archive_id).first()
    if not archive or archive.form_type_key != form_type_key.lower():
        abort(404)

    snapshot = json.loads(archive.snapshot_json or '{}')
    archived_by = db_session.query(User).filter_by(
        user_id=archive.archived_by_user_id
    ).first()
    form_models = {"form-a": FormA, "form-b": FormB, "form-c": FormC}
    archive_model = form_models.get(archive.form_type_key)
    active_form_types = [
        label
        for label, model in (("Form A", FormA), ("Form B", FormB), ("Form C", FormC))
        if db_session.query(model).filter_by(user_id=archive.student_user_id).first()
    ]
    restore_block_reason = None
    if archive_model is None:
        restore_block_reason = "This archive has an unsupported form type."
    if not restore_block_reason and active_form_types:
        restore_block_reason = (
            f"The student has already started an active {', '.join(active_form_types)}. "
            f"Archive that active form before restoring this {archive.form_type}."
        )
    if (
        not restore_block_reason
        and db_session.query(archive_model).filter_by(
            form_id=archive.original_form_id
        ).first()
    ):
        restore_block_reason = "This exact form record is already active."

    def archived_display_value(value):
        if isinstance(value, dict) and value.get("__archive_type__") == "bytes":
            return f"[Archived binary data: {value.get('size', 0)} bytes]"
        return '' if value is None else str(value)

    form_fields = [
        {
            'label': key.replace('_', ' ').title(),
            'display_value': archived_display_value(value),
        }
        for key, value in snapshot.items()
    ]
    return render_template(
        'archived_form_detail.html',
        role=current_user.role.value,
        user_profile=current_user,
        archive=archive,
        archived_by_name=archived_by.full_name if archived_by else 'Former or unavailable user',
        restore_block_reason=restore_block_reason,
        form_type=archive.form_type,
        form_fields=form_fields,
    )


@app.route('/back_end/archived_forms/<string:archive_id>/restore', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def restore_archived_form(archive_id):
    archive = db_session.query(ArchivedEthicsForm).filter_by(archive_id=archive_id).first()
    if not archive:
        abort(404)

    form_models = {"form-a": FormA, "form-b": FormB, "form-c": FormC}
    archive_model = form_models.get(archive.form_type_key)
    if archive_model is None:
        flash("This archived form type cannot be restored.", "danger")
        return redirect(url_for(
            'archived_form_detail',
            form_type_key=archive.form_type_key,
            archive_id=archive.archive_id,
        ))

    active_forms = [
        label
        for label, model in (("Form A", FormA), ("Form B", FormB), ("Form C", FormC))
        if db_session.query(model).filter_by(user_id=archive.student_user_id).first()
    ]
    if active_forms:
        flash(
            f"Restore blocked: the student has already started an active {', '.join(active_forms)}. "
            f"Archive the active form before restoring {archive.form_type}.",
            "warning",
        )
        return redirect(url_for(
            'archived_form_detail',
            form_type_key=archive.form_type_key,
            archive_id=archive.archive_id,
        ))
    if db_session.query(archive_model).filter_by(
        form_id=archive.original_form_id
    ).first():
        flash("Restore blocked: this exact form record is already active.", "warning")
        return redirect(url_for(
            'archived_form_detail',
            form_type_key=archive.form_type_key,
            archive_id=archive.archive_id,
        ))

    try:
        snapshot = json.loads(archive.snapshot_json or '{}')
        restored_values = {}
        for column in archive_model.__table__.columns:
            if column.name not in snapshot:
                continue
            value = snapshot[column.name]
            if isinstance(value, dict) and value.get("__archive_type__") == "bytes":
                value = base64.b64decode(value.get("base64") or "")
            elif isinstance(value, str) and value.startswith("[binary data archived:"):
                value = None
            elif isinstance(value, str):
                try:
                    if column.type.python_type is datetime:
                        value = datetime.fromisoformat(value)
                except (AttributeError, NotImplementedError, TypeError, ValueError):
                    pass
            restored_values[column.name] = value

        restored_values["user_id"] = archive.student_user_id
        restored_values["form_id"] = archive.original_form_id
        restored_form_type = archive.form_type
        restored_student_name = archive.student_name
        restored_student_id = archive.student_user_id
        db_session.add(archive_model(**restored_values))
        db_session.delete(archive)
        db_session.commit()
        flash(
            f"{restored_form_type} was restored successfully for {restored_student_name}.",
            "success",
        )
        return redirect(url_for('monitor_forms', user_id=restored_student_id))
    except Exception:
        db_session.rollback()
        app.logger.exception("Unable to restore archived ethics form %s", archive_id)
        flash("The archived form could not be restored. No changes were saved.", "danger")
        return redirect(url_for(
            'archived_form_detail',
            form_type_key=archive.form_type_key,
            archive_id=archive.archive_id,
        ))

# ---------------- Section 4 ------------------
@app.route('/form_a_sec4', methods=['GET', 'POST'])
def form_a_sec4():
    if request.method == 'POST':
        try:
            cleaned_sample_sizes, sample_size_error = normalize_forma_sample_sizes(
                request.form.getlist('sample_size[]')
            )
            if sample_size_error:
                flash(sample_size_error, "danger")
                return redirect(request.referrer or url_for('form_a_sec4'))
            def to_bool(val):
                if isinstance(val, bool):
                    return val
                if val in [None, '']:
                    return False
                return str(val).lower() in ['yes', 'true', '1', 'on', 'checked']

            user_id = session.get('id')
            if not user_id:
                return "Unauthorized access. Please log in.", 401

            # --- Fetch the existing FormA record ---
            form = _get_latest_forma_for_user(user_id)
            if not form:
                return "No existing Form A record found for this user.", 404

            _apply_forma_autosave_payload(form, request.form, section='sec4', include_declaration=False)
            form.sampling_size = ','.join(cleaned_sample_sizes)

            # --- Commit to Database ---
            db_session.commit()

            flash("Form A Section 4 submitted successfully.", "success")
            return redirect(url_for("form_a_sec5"))

        except SQLAlchemyError as e:
            db_session.rollback()
            print("Database Error:", e)
            traceback.print_exc()
            return render_template('form-a-section4.html', messages=["Database error occurred. Please try again."])

        except Exception as e:
            db_session.rollback()
            print("Unexpected Error:", e)
            traceback.print_exc()
            return render_template('form-a-section4.html', messages=["An unexpected error occurred. Please try again."])

    # If GET request → render Section 4 with existing values
    user_id = session.get('id')
    form_data = _get_latest_forma_for_user(user_id)
    return render_template('form-a-section4.html', form_data=form_data)
@app.route('/submit_form_a_sec4', methods=['GET', 'POST'])
def submit_form_a_sec4 ():
    try:
        if request.method == 'POST':
            cleaned_sample_sizes, sample_size_error = normalize_forma_sample_sizes(
                request.form.getlist('sample_size[]')
            )
            if sample_size_error:
                flash(sample_size_error, "danger")
                return redirect(request.referrer or url_for('form_a_sec4'))
            user_id = session.get('id')
            if not user_id:
                return "Unauthorized access. Please log in.", 401

            form = _get_latest_forma_for_user(user_id)
            if not form:
                return "No existing Form A record found for this user.", 404

            data = _build_autosave_data(request.form)

            form.quantitative = _autosave_str_to_bool(data.get('quantitative')) if 'quantitative' in data else False
            form.qualitative = _autosave_str_to_bool(data.get('qualitative')) if 'qualitative' in data else False
            form.mixed_methods = _autosave_str_to_bool(data.get('mixed_methods')) if 'mixed_methods' in data else False
            form.paradigm_explanation = request.form.get('paradigm_explanation', '')
            form.design = request.form.get('design', '')
            form.participants_description = request.form.get('participants_description', '')
            form.population = data.get('population', '')
            form.sampling_method = data.get('sampling_method', '')
            form.sampling_size = ','.join(cleaned_sample_sizes)
            form.inclusion_criteria = data.get('inclusion_criteria', '')
            form.duration_timing = request.form.get('duration_timing', '')
            form.contact_details_method = request.form.get('contact_details_method', '')
            form.conflict_interest = _autosave_str_to_bool(data.get('conflict_interest')) if 'conflict_interest' in data else False
            form.conflict_explanation = request.form.get('conflict_explanation', '')

            form.data_methods = data.get('data_methods', '')
            form.questionnaire_type = request.form.get('questionnaire_type', '')
            questionnaire_permission = request.form.get('questionnaire_permission', '')
            if questionnaire_permission == 'Yes':
                form.permission_obtained = 'Yes'
                form.open_source = 'No'
            elif questionnaire_permission == 'Open Source':
                form.permission_obtained = 'No'
                form.open_source = 'Yes'
            else:
                form.permission_obtained = None
                form.open_source = None
            form.interview_type = request.form.get('interview_type', '')
            form.interview_recording = request.form.get('interview_recording', '')
            selected_methods = [method.strip() for method in form.data_methods.split(',') if method.strip()]
            form.use_focus_groups = 'focus' in selected_methods
            form.focus_recording = request.form.get('focus_recording', '')
            form.observation_details = request.form.get('observation_details', '')
            form.documents_details = request.form.get('documents_details', '')
            form.other_details = request.form.get('other_details', '')
            form.data_collection_procedure = request.form.get('data_collection_procedure', '')
            form.data_collectors = request.form.get('data_collectors', '')
            form.intervention = _autosave_str_to_bool(data.get('intervention')) if 'intervention' in data else False
            form.intervention_details = request.form.get('intervention_details', '')
            form.sensitive_data = request.form.get('sensitive_data', '')
            form.translator = _autosave_str_to_bool(data.get('translator')) if 'translator' in data else False
            form.translator_procedure = request.form.get('translator_procedure', '')
            form.instrument_attachment_reason = request.form.get('instrument_attachment_reason', '')

            secondary_data = data.get('uses_secondary_data', data.get('secondary_data', ''))
            form.uses_secondary_data = _autosave_str_to_bool(secondary_data) if secondary_data is not None else False

            shared_fields = [
                'data_nature', 'data_origin', 'access_conditions', 'personal_info',
                'personal_info_comment', 'data_anonymized', 'anonymization_comment',
                'shortcomings_reported',
                'limitations_reporting', 'methodology_alignment', 'data_acknowledgment'
            ]
            for field in shared_fields:
                if hasattr(form, field):
                    setattr(form, field, request.form.get(field, ''))

            if form.uses_secondary_data:
                form.secondary_data_type = request.form.get('data_type', '')
            else:
                form.secondary_data_type = ''

            db_session.commit()
            flash("Form A Section 4 submitted successfully.", "success")
            return redirect(url_for("form_a_sec5"))

        user_id = session.get('id')
        form_data = _get_latest_forma_for_user(user_id)
        return render_template('form-a-section4.html', form_data=form_data)

    except Exception as e:
        db_session.rollback()
        print(f"⚠️ Exception in /submit_form_a_sec4: {e}")
        return render_template('form-a-section4.html', messages=["An unexpected error occurred. Please try again."])

# ---------------- Section 5 ------------------
@app.route('/form_a_sec5', methods=['GET', 'POST'])
def form_a_sec5 ():
    if request.method == 'POST':
        def to_bool(val):
            if isinstance(val, bool):
                return val
            if val in [None, '']:
                return False
            return str(val).lower() in ['yes', 'true', '1', 'on', 'checked']

        user_id = session.get('id')
        

        if not user_id:
            return "Unauthorized access. Please log in.", 401

        # Fetch existing form entry for the user
        form = _get_latest_forma_for_user(user_id)
        if not form:
            return "No existing Form A record found for this user.", 404
        
        
        
        _apply_forma_autosave_payload(form, request.form, section='sec5', include_declaration=False)
    
        db_session.add(form)
        db_session.commit()
        return redirect(url_for("form_a_sec6"))
        

    user_id = session.get('id')
    form_data = _get_latest_forma_for_user(user_id)
    return render_template('form-a-section5.html', form_data=form_data)

@app.route('/submit_form_a_sec5', methods=['GET', 'POST'])
def submit_form_a_sec5 ():
    try:
        if request.method == 'POST':
            user_id = session.get('id')
            if not user_id:
                return "Unauthorized access. Please log in.", 401

            form = _get_latest_forma_for_user(user_id)
            if not form:
                return "No existing Form A record found for this user.", 404

            data = _build_autosave_data(request.form)

            form.informed_consent = request.form.get('informed_consent', '')
            form.data_storage = data.get('data_storage', '')
            form.study_benefits = request.form.get('study_benefits', '')
            form.participant_risks = request.form.get('participant_risks', '')
            form.adverse_steps = request.form.get('adverse_steps', '')
            form.community_participation = request.form.get('community_participation', '')
            form.community_effects = request.form.get('community_effects', '')
            form.privacy = data.get('privacy', '')

            for key in ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s']:
                field_name = f'q6_9{key}'
                if hasattr(form, field_name):
                    setattr(form, field_name, _autosave_str_to_bool(data.get(field_name)) if field_name in data else False)

            form.results_feedback = request.form.get('results_feedback', '')
            form.products_access = request.form.get('products_access', '')
            form.publication_plans = request.form.get('publication_plans', '')
            form.participant_comp = request.form.get('participant_comp', '')
            form.participant_costs = request.form.get('participant_costs', '')
            form.ethics_reporting = request.form.get('ethics_reporting', '')

            db_session.commit()
            return redirect(url_for("form_a_sec6"))

        user_id = session.get('id')
        form_data = _get_latest_forma_for_user(user_id)
        return render_template('form-a-section5.html', form_data=form_data)

    except Exception as e:
        db_session.rollback()
        print(f"⚠️ Exception in /submit_form_a_sec5: {e}")
        return render_template('form-a-section5.html')

# ---------------- Section 6 ------------------
@app.route('/form_a_sec6', methods=['GET', 'POST'])
def form_a_sec6():
    if request.method == 'POST':
        try:
            # --- Check user session ---
            user_id = session.get('id')
            if not user_id:
                return "Unauthorized access. Please log in.", 401

            # --- Fetch form and user records ---
            form = _get_latest_forma_for_user(user_id)
            user = db_session.query(User).filter_by(user_id=user_id).first()

            if not form:
                return "No existing Form A record found for this user.", 404

            # --- Update Declaration Section ---
            form.submitted_at = datetime.now()
            form.declaration_name = request.form.get('declaration_name')
            form.applicant_signature = request.form.get('applicant_signature')
            form.declaration_date = datetime.now()
            
            # --- Commit to database ---
            db_session.add(form)
            db_session.commit()

            # --- Try sending confirmation email ---
            try:
                message = f"{form.declaration_name or 'Applicant'} has submitted Form A for review."
                send_email(app, mail, message, [form.supervisor_email])
            except Exception as e:
                app.logger.error(f"⚠️ Failed to send email to {form.supervisor_email}: {e}")
                traceback.print_exc()

            flash("✅ Form A submitted successfully.", "success")
            return redirect(url_for('student_dashboard'))

        except SQLAlchemyError as e:
            db_session.rollback()
            app.logger.error(f"Database error during Form A submission: {e}")
            traceback.print_exc()
            flash("❌ Database error occurred. Please try again.", "danger")
            return render_template('form-a-section6.html')

        except Exception as e:
            db_session.rollback()
            app.logger.error(f"Unexpected error in Form A submission: {e}")
            traceback.print_exc()
            flash("⚠️ An unexpected error occurred. Please try again.", "danger")
            return render_template('form-a-section6.html')

    # GET request → Render the section
    user_id = session.get('id')
    form_data = _get_latest_forma_for_user(user_id)
    return render_template('form-a-section6.html', form_data=form_data)



@app.route('/submit_form_a_sec6', methods=['GET', 'POST'])
def submit_form_a_sec6():
    # This route is handled by form_a_sec6 function above
    # Redirect to avoid duplicate handling
    return redirect(url_for('form_a_sec6'))




# FORM B =====================================================================================================
@app.route('/api/form-b/<form_id>', methods=['GET'])
@login_required
def get_form_b(form_id):
    form_b = db_session.query(FormB).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).filter_by(form_id=form_id).first()
    if not form_b:
        return jsonify({"message": "Form not found"}), 404
    if not can_access_form(get_current_user(), form_b):
        return jsonify({"message": "Forbidden"}), 403
    return jsonify(form_b.to_dict()), 200


@app.route('/api/form-b/<form_id>/reassign-reviewers', methods=['POST'])
def api_reassign_form_b_reviewers(form_id):
    user_id = session.get('id')
    user_role = (session.get('role') or '').upper()
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    if user_role not in ['ADMIN', 'SUPER_ADMIN']:
        return jsonify({'success': False, 'error': 'Forbidden'}), 403

    form = db_session.query(FormB).filter_by(form_id=form_id).first()
    if not form:
        return jsonify({'success': False, 'error': 'Form B not found'}), 404

    data = request.get_json(silent=True) or {}
    reviewer_ids = data.get('reviewer_ids') or data.get('reviewers') or []
    if not isinstance(reviewer_ids, list):
        return jsonify({'success': False, 'error': 'reviewer_ids must be a list'}), 400

    selected_ids = []
    for reviewer_id in reviewer_ids:
        reviewer_id = str(reviewer_id or '').strip()
        if reviewer_id and reviewer_id not in selected_ids:
            selected_ids.append(reviewer_id)

    if len(selected_ids) != 1:
        return jsonify({'success': False, 'error': 'Please provide exactly one reviewer'}), 400

    selected_reviewers = (
        db_session.query(User)
        .filter(
            User.user_id.in_(selected_ids),
            User.role == UserRole.REVIEWER
        )
        .all()
    )

    if len(selected_reviewers) != len(selected_ids):
        return jsonify({'success': False, 'error': 'One or more selected reviewers are invalid'}), 400

    form.reviewer_name1 = selected_ids[0]
    form.reviewer_name2 = None
    form.submitted_to_reviewers = True

    db_session.add(UserActivityLog(
        user_id=user_id,
        action='reassign_reviewers_api',
        page='API Form B Reassignment',
        target_user_id=form.user_id,
        details=f"Reassigned reviewers via API: {form.reviewer_name1}, {form.reviewer_name2} for FORM B {form.form_id}"
    ))
    db_session.commit()

    return jsonify({
        'success': True,
        'message': 'Form B reviewers reassigned successfully',
        'form_id': form.form_id,
        'reviewer_ids': selected_ids
    }), 200




@app.route('/form_b_upload', methods=['GET','POST'])
def form_b_upload():
    UPLOAD_FOLDER = 'static/uploads/form_b'
    user_id = session.get('id')

    if not user_id:
        return "Unauthorized", 401
    
    if has_blocking_student_form(FormA, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
        
    if has_blocking_student_form(FormC, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))

    if request.method=='POST':
        form = db_session.query(FormB).options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        ).filter_by(user_id=user_id).first()
        # Get form data
        need_permission = request.form.get('need_permission')
        has_clearance = request.form.get('has_clearance')
        has_ethics_evidence = request.form.get('has_ethics_evidence')

        # Create upload folder if it doesn't exist
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        def get_file_blob(field_name):
            return read_file_blob(field_name)

        perm_data, perm_fname = get_file_blob('permission_letter_path')
        prior_data, prior_fname = get_file_blob('prior_clearance_path')
        pending_data, pending_fname = get_file_blob('pending_note_path')
        prop_data, prop_fname = get_file_blob('proposal_path')

        if form:
            form.user_id = user_id
            form.need_permission = need_permission == 'Yes'
            if perm_data:
                form.permission_letter = perm_data
                form.permission_letter_filename = perm_fname
            
            if pending_data:
                form.pending_note = pending_data
                form.pending_note_filename = pending_fname

            form.has_clearance = has_clearance == 'Yes'
            if prior_data:
                form.prior_clearance = prior_data
                form.prior_clearance_filename = prior_fname
            
            if prop_data:
                form.proposal_path = prop_data
                form.proposal_filename = prop_fname
        else:
            # Save to database
            form = FormB(
                user_id=user_id,
                need_permission=need_permission == 'Yes',
                permission_letter=perm_data,
                permission_letter_filename=perm_fname,
                pending_note=pending_data,
                pending_note_filename=pending_fname,
                has_clearance=has_clearance == 'Yes',
                prior_clearance=prior_data,
                prior_clearance_filename=prior_fname,
                proposal_path=prop_data,
                proposal_filename=prop_fname
            )

        db_session.add(form)
        db_session.commit()
        message="form submited succesffuly"
        return render_template("form-b-section1.html",messages=[message])

    return render_template("form-b-upload.html")


@app.route('/form_b_sec1', methods=['GET', 'POST'])
def form_b_sec1():
    try:
        if request.method == 'POST':
            user_id = session.get('id')
            if not user_id:
                return jsonify({'error': 'unauthorized'}), 401

            if has_blocking_student_form(FormA, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            if has_blocking_student_form(FormC, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            form_data = request.form
            user = db_session.query(User).filter(User.user_id == user_id).first()
            supervisor = get_student_supervisor_or_flash(user)

            if not user or not supervisor:
                return redirect(url_for("student_dashboard"))

            form = db_session.query(FormB).filter_by(user_id=user_id).first()

            if form:
                # Update existing record
                form.user_id = user_id
                form.applicant_name = form_data.get('applicant_name')
                form.student_number = form_data.get('student_number')
                form.institution = form_data.get('institution')
                form.department = form_data.get('department')
                form.degree = form_data.get('degree')
                form.study_title = form_data.get('study_title')
                form.mobile = form_data.get('mobile')
                form.email = user.email
                form.supervisor = supervisor.full_name
                form.supervisor_email = supervisor.email
            else:
                # Create new record
                form = FormB(
                    user_id=user_id,
                    applicant_name=form_data.get('applicant_name'),
                    student_number=form_data.get('student_number'),
                    institution=form_data.get('institution'),
                    department=form_data.get('department'),
                    degree=form_data.get('degree'),
                    study_title=form_data.get('study_title'),
                    mobile=form_data.get('mobile'),
                    email=user.email,
                    supervisor=supervisor.full_name,
                    supervisor_email=supervisor.email,
                )

            db_session.add(form)
            db_session.commit()

            message = 'Form submitted successfully'
            flash(message, 'success')
            return render_template("form-b-section2.html", messages=[message])

        # GET request
        return render_template('form-b-section1.html')

    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Error in form_b_sec1: {e}")
        flash("An error occurred while submitting the form. Please try again.", "danger")
        return render_template('form-b-section1.html'), 500




@app.route('/form_b_sec2', methods=['GET', 'POST'])
def form_b_sec2():
    try:
        user_id = session.get('id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401

        # Restrict access if Form A or Form C already exists
        if has_blocking_student_form(FormA, user_id):
            flash("You are not permitted to fill this form", "warning")
            return redirect(url_for("student_dashboard"))
        
        if has_blocking_student_form(FormC, user_id):
            flash("You are not permitted to fill this form", "warning")
            return redirect(url_for("student_dashboard"))
        
        # Handle form submission
        if request.method == 'POST':
            form_data = request.form
            form = db_session.query(FormB).filter_by(user_id=user_id).first()

            if not form:
                form = FormB(user_id=user_id)

            # ✅ Assign form fields safely
            form.project_description = form_data.get('project_description')
            form.data_nature = form_data.get('data_nature')
            form.data_origin = form_data.get('data_origin')
            _apply_formb_public_domain_answers(form, form_data)
            form.personal_info = form_data.get('personal_info') == 'Yes'
            form.data_anonymized = form_data.get('data_anonymized')
            form.anonymization_comment = form_data.get('anonymization_comment')
            form.private_permission = form_data.get('private_permission') == 'Yes'
            form.permission_details = form_data.get('permission_details')
            form.shortcomings_reported = form_data.get('shortcomings_reported') == 'Yes'
            form.limitations_reporting = form_data.get('limitations_reporting')
            form.methodology_alignment = form_data.get('methodology_alignment') == 'Yes'
            form.data_acknowledgment = form_data.get('data_acknowledgment')
            form.rejected_or_accepted = False

            # Handle file upload
            file = request.files.get('private_permission_file')
            if file and file.filename:
                assign_private_permission_upload(form, file)

            # ✅ Commit to database
            db_session.add(form)
            db_session.commit()

            message = "Section 2 saved successfully."
            flash(message, "success")
            return render_template("form-b-section3.html", messages=[message])

        # Handle GET request
        return render_template('form-b-section2.html')

    except Exception as e:
        # Roll back and log the error
        db_session.rollback()
        app.logger.error(f"Error in form_b_sec2: {e}")
        flash("An error occurred while saving your data. Please try again.", "danger")
        return render_template('form-b-section2.html'), 500

@app.route('/form_b_sec3', methods=['GET', 'POST'])
def form_b_sec3():
    try:
        user_id = session.get('id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401

        if request.method == 'POST':
            form = db_session.query(FormB).filter_by(user_id=user_id).first()
            user = db_session.query(User).filter_by(user_id=user_id).first()

            if not form:
                form = FormB(user_id=user_id)

            # --- Assign form fields ---
            form.original_clearance = request.form.get('original_clearance')
            form.participant_permission = request.form.get('participant_permission')
            form.data_safekeeping = request.form.get('data_safekeeping')
            form.risk_level = request.form.get('risk_level')
            form.submitted = True
            form.risk_comments = request.form.get('risk_comments')
            form.declaration_name = request.form.get('declaration_name')
            form.full_name = request.form.get('full_name')
            form.declaration_date = datetime.now()
            form.submitted_at = datetime.now()
            form.rejected_or_accepted = False

            db_session.add(form)
            db_session.commit()

            # --- Email notification (nested try/catch) ---
            try:
                message = f"{form.applicant_name} has submitted the form and it needs to be reviewed."
                send_email(app, mail, message, [form.supervisor_email])
            except Exception as e:
                app.logger.error(f"⚠️ Failed to send email to {form.supervisor_email}: {e}")
                traceback.print_exc()

            flash("✅ Form submitted successfully.", "success")
            return redirect(url_for('student_dashboard'))

        # GET request
        return render_template('form_b_section3.html', messages=[], show_modal=False)

    except SQLAlchemyError as e:
        db_session.rollback()
        app.logger.error(f"Database error in form_b_sec3: {e}")
        traceback.print_exc()
        flash("❌ Database error occurred. Please try again.", "danger")
        return render_template('form_b_section3.html', messages=[], show_modal=False), 500

    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Unexpected error in form_b_sec3: {e}")
        traceback.print_exc()
        flash("⚠️ An unexpected error occurred. Please try again.", "danger")
        return render_template('form_b_section3.html', messages=[], show_modal=False), 500





@app.route('/form_c_upload', methods=['GET'])
def form_c_upload ():
    return render_template('form-c-upload.html')

@app.route('/form_c_sec1', methods=['GET', 'POST'])
def form_c_sec1():
    try:
        if request.method == 'POST':
            user_id = session.get('id')
            if not user_id:
                return jsonify({'error': 'Unauthorized'}), 401

            # Prevent filling Form C if Form A or Form B exists
            if has_blocking_student_form(FormA, user_id):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            formB_options = [
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file)
            ]
            if has_blocking_student_form(FormB, user_id, options=formB_options):
                flash("You are not permitted to fill this form", "warning")
                return redirect(url_for("student_dashboard"))

            user = db_session.query(User).filter(User.user_id == user_id).first()
            supervisor = get_student_supervisor_or_flash(user)

            if not user or not supervisor:
                return redirect(url_for("student_dashboard"))

            form = db_session.query(FormC).filter_by(user_id=user_id).first()
            if not form:
                form = FormC(user_id=user_id)

            # --- Assign form fields ---
            form.applicant_name = request.form.get('applicant_name')
            form.student_number = request.form.get('student_number')
            form.institution = request.form.get('institution')
            form.department = request.form.get('department')
            form.degree = request.form.get('degree')
            form.project_title = request.form.get('project_title')
            form.mobile_number = request.form.get('mobile_number')
            form.email_address = user.email
            form.supervisor_name = supervisor.full_name
            form.supervisor_email = supervisor.email
            form.rejected_or_accepted = False
            form.supervisor_comments = ""

            db_session.add(form)
            db_session.commit()

            message = "Form submitted successfully"
            flash(message, "success")
            return render_template("form-c-section2.html", messages=[message])

        # GET request
        return render_template("form-c-section1.html")

    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Error in form_c_sec1: {e}")
        flash("An error occurred while submitting the form. Please try again.", "danger")
        return render_template("form-c-section1.html"), 500


@app.route('/form_a_supervisor/<string:form_id>',methods=['GET','POST'])
@login_required
def form_a_supervisor(form_id):
    form = db_session.query(FormA).filter_by(form_id=form_id).order_by(FormA.submitted_at.desc()).first()
    if not form:
        return "Form not found", 404
    if not can_act_as_assigned_supervisor(get_current_user(), form):
        abort(403)
    if is_with_ethics(form):
        flash('You have already reviewed this form. It is currently with Ethics.', 'info')
        return redirect(url_for('supervisor_dashboard'))
    
    data={
        "org_name" : parse_field(form.org_name),
        "org_contact" : parse_field(form.org_contact),
        "org_role": parse_field(form.org_role),
        "org_permission" : parse_field(form.org_permission),

        "fund_org" : parse_field(form.fund_org),
        "fund_contact" :parse_field(form.fund_contact),
        "fund_role": parse_field(form.fund_role),
        "fund_amount": parse_field(form.fund_amount),

        "population" :parse_field(form.population),
        "sampling_method" : parse_field(form.sampling_method),
        "sampling_size": parse_field(form.sampling_size),
        "inclusion_criteria": parse_field(form.inclusion_criteria)
    }

    return render_template("form_a_supervisor.html",formA=form,data=data)


    
    
@app.route('/form_b_supervisor/<string:form_id>',methods=['GET','POST'])
@login_required
def form_b_supervisor(form_id):
    form = db_session.query(FormB).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).filter_by(form_id=form_id).order_by(FormB.submitted_at.desc()).first()
    if not form:
        return "Form not found", 404
    if not can_act_as_assigned_supervisor(get_current_user(), form):
        abort(403)
    if is_with_ethics(form):
        flash('You have already reviewed this form. It is currently with Ethics.', 'info')
        return redirect(url_for('supervisor_dashboard'))

    return render_template("form_b_supervisor.html",formB=form)

@app.route('/form_c_supervisor/<string:form_id>',methods=['GET','POST'])
@login_required
def form_c_supervisor(form_id):
    ##
    ## check the user id and form id trace back the error
    ##
    form = db_session.query(FormC).filter_by(form_id=form_id).order_by(FormC.submission_date.desc()).first()
    if not form:
        return "Form not found", 404
    if not can_act_as_assigned_supervisor(get_current_user(), form):
        abort(403)
    if is_with_ethics(form):
        flash('You have already reviewed this form. It is currently with Ethics.', 'info')
        return redirect(url_for('supervisor_dashboard'))
    return render_template("form_c_supervisor.html",formc=form)


@app.route('/reject_or_Accept_form_a/<string:id>',methods=['GET','POST'])
@role_required('SUPERVISOR', 'REVIEWER')
def reject_or_Accept_form_a(id):

    user_id=session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        forma = db_session.query(FormA).filter_by(form_id=id).order_by(FormA.submitted_at.desc()).first()
    except ValueError as e:
        # Handle corrupted date data in database
        app.logger.error(f"Date parsing error for form {id}: {e}")
        forma = db_session.query(FormA).filter_by(form_id=id).first()

    if not forma:
        return "Form not found", 404
    if not can_act_as_assigned_supervisor(get_current_user(), forma):
        abort(403)
    if is_with_ethics(forma):
        flash('You have already reviewed this form. It is currently with Ethics.', 'info')
        return redirect(url_for('supervisor_dashboard'))
    
    #admin=db_session.query(User).filter_by(role="Admin").all()
    data={
        "org_name" : parse_field(forma.org_name) if forma else [],
        "org_contact" : parse_field(forma.org_contact) if forma else [],
        "org_role": parse_field(forma.org_role) if forma else [],
        "org_permission" : parse_field(forma.org_permission) if forma else [],

        "fund_org" : parse_field(forma.fund_org) if forma else [],
        "fund_contact" :parse_field(forma.fund_contact) if forma else [],
        "fund_role": parse_field(forma.fund_role) if forma else [],
        "fund_amount": parse_field(forma.fund_amount) if forma else [],

        "population" :parse_field(forma.population) if forma else [],
        "sampling_method" : parse_field(forma.sampling_method) if forma else [],
        "sampling_size": parse_field(forma.sampling_size) if forma else [],
        "inclusion_criteria": parse_field(forma.inclusion_criteria) if forma else []
    }

    if request.method=="POST":
        forma = (
            db_session.query(FormA)
            .filter_by(form_id=id)
            .with_for_update()
            .first()
        )
        if has_supervisor_submitted_feedback(forma):
            flash('You have already submitted feedback for this application. It was not submitted again.', 'warning')
            return redirect(url_for('supervisor_dashboard'))
        org_permission_comment=(request.form.get('org_permission_comment') or '').strip()
        waiver_comment=(request.form.get('waiver_comment') or '').strip()
        form_a_comment=(request.form.get('form_a_comment') or '').strip()
        questions_comment=(request.form.get('questions_comment') or '').strip()
        consent_comment=(request.form.get('consent_comment') or '').strip()
        proposal_comment=(request.form.get('proposal_comment') or '').strip()
        supervisor_feedback=(request.form.get('supervisor_feedback') or '').strip()
        recommendation=request.form.get('recommendation')
        supervisor_signature=(request.form.get('supervisor_signature') or '').strip()
        signature_date=datetime.now()

        if not supervisor_feedback:
            flash('Supervisor feedback is required before submitting your recommendation.', 'danger')
            return redirect(url_for('form_a_supervisor', form_id=id))
        
        
        if request.form.get('recommendation')=='Ready for submission':
            if not supervisor_signature:
                flash('Supervisor signature is required when approving for submission.', 'danger')
                return redirect(url_for('form_a_supervisor', form_id=id))
            missing_submission_redirect = redirect_if_missing_student_submission(
                forma,
                'ethics admin',
                'form_a_supervisor',
                form_id=id,
            )
            if missing_submission_redirect:
                return missing_submission_redirect
            forma.supervisor_date=datetime.now()
            forma.org_permission_comment=org_permission_comment
            forma.waiver_comment=waiver_comment
            forma.form_a_comment=form_a_comment
            forma.questions_comment=questions_comment
            forma.consent_comment=consent_comment
            forma.proposal_comment=proposal_comment
            forma.supervisor_feedback=supervisor_feedback
            forma.recommendation=recommendation
            forma.status='Submitted to Ethics'
            forma.ethics_status=None
            forma.form_supervisor_status='Ready for submission'
            forma.supervisor_signature=supervisor_signature
            forma.signature_date=signature_date
            forma.rejected_or_accepted=True
            forma.submitted_to_admin=True
            #Uncomment the code bellow for testing
            ##
            try:
                message=f' Form belongning to {forma.applicant_name} has been reviewed by the supervisor and has been accepted for further reviewing. ' 
            
                send_email(app,mail, message,[forma.email])
            except Exception as e:
                app.logger.error(f"Failed to send email to {forma.email}: {e}")
        else:
            forma.supervisor_date=datetime.now()
            forma.org_permission_comment=org_permission_comment
            forma.waiver_comment=waiver_comment
            forma.form_a_comment=form_a_comment
            forma.questions_comment=questions_comment
            forma.consent_comment=consent_comment
            forma.proposal_comment=proposal_comment
            forma.supervisor_feedback=supervisor_feedback
            forma.recommendation=recommendation
            forma.status='Revisions required'
            forma.ethics_status=None
            forma.form_supervisor_status='Revisions required'
            forma.submitted_to_admin=False
            forma.supervisor_signature=None
            forma.signature_date=None
            forma.rejected_or_accepted=False
            try:
                message=f' Form belonging to {forma.applicant_name} has been reviewed and returned back to you. Please login to view the feedback.' 
            
                send_email(app,mail, message,[forma.email])
                
            except Exception as e:
                app.logger.error(f"Failed to send email to {forma.email}: {e}")

                

                # New form flow: respect dashboard origin when present
                from_dashboard = request.form.get('from_dashboard') or request.args.get('from_dashboard')
                if from_dashboard:
                    return redirect(url_for('dashboard'))
        db_session.add(forma)
        db_session.commit()
    return redirect(url_for('supervisor_dashboard'))

@app.route('/reject_or_Accept_form_b/<string:id>',methods=['GET','POST'])
@role_required('SUPERVISOR', 'REVIEWER')
def reject_or_Accept_form_b(id):
    user_id=session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        formb = db_session.query(FormB).options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        ).filter_by(form_id=id).order_by(FormB.submitted_at.desc()).first()
    except ValueError as e:
        # Handle corrupted date data in database
        app.logger.error(f"Date parsing error for form {id}: {e}")
        formb = db_session.query(FormB).options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        ).filter_by(form_id=id).first()
    #admin=db_session.query(User).filter_by(role="Admin").all()
    if not formb:
        return "Form not found", 404
    if not can_act_as_assigned_supervisor(get_current_user(), formb):
        abort(403)
    if is_with_ethics(formb):
        flash('You have already reviewed this form. It is currently with Ethics.', 'info')
        return redirect(url_for('supervisor_dashboard'))
    if request.method=="POST":
        formb = (
            db_session.query(FormB)
            .filter_by(form_id=id)
            .with_for_update()
            .first()
        )
        if has_supervisor_submitted_feedback(formb):
            flash('You have already submitted feedback for this application. It was not submitted again.', 'warning')
            return redirect(url_for('supervisor_dashboard'))

        org_permission_comment=(request.form.get('org_permission_comment') or '').strip()
        waiver_comment=(request.form.get('waiver_comment') or '').strip()
        form_a_comment=(request.form.get('form_a_comment') or '').strip()
        questions_comment=(request.form.get('questions_comment') or '').strip()
        consent_comment=(request.form.get('consent_comment') or '').strip()
        proposal_comment=(request.form.get('proposal_comment') or '').strip()
        supervisor_feedback=(request.form.get('supervisor_feedback') or '').strip()
        recommendation=request.form.get('recommendation')
        supervisor_signature=(request.form.get('supervisor_signature') or '').strip()
        signature_date=datetime.now()
        if not supervisor_feedback:
            flash('Supervisor feedback is required before submitting your recommendation.', 'danger')
            return redirect(url_for('form_b_supervisor', form_id=id))

        if request.form.get('recommendation')=='Ready for submission':
            if not supervisor_signature:
                flash('Supervisor signature is required when approving for submission.', 'danger')
                return redirect(url_for('form_b_supervisor', form_id=id))
            missing_submission_redirect = redirect_if_missing_student_submission(
                formb,
                'ethics admin',
                'form_b_supervisor',
                form_id=id,
            )
            if missing_submission_redirect:
                return missing_submission_redirect
            formb.supervisor_date=datetime.now()
            formb.org_permission_comment=org_permission_comment
            formb.waiver_comment=waiver_comment
            formb.form_a_comment=form_a_comment
            formb.questions_comment=questions_comment
            formb.consent_comment=consent_comment
            formb.proposal_comment=proposal_comment
            formb.supervisor_feedback=supervisor_feedback
            formb.recommendation=recommendation
            formb.status='Submitted to Ethics'
            formb.ethics_status=None
            formb.form_supervisor_status='Ready for submission'
            formb.supervisor_signature=supervisor_signature
            formb.signature_date=signature_date
            formb.rejected_or_accepted=True
            formb.submitted_to_admin=True
            #Uncomment the code bellow for testing
            ##
            try:
                message=f' Form belongning to {formb.applicant_name} has been reviewed by the supervisor and has been accepted for further reviewing.' 
            
                send_email(app,mail, message,[formb.email])
            except Exception as e:
                app.logger.error(f"Failed to send email to {formb.email}: {e}")
            
        else:
            formb.supervisor_date=datetime.now()
            formb.org_permission_comment=org_permission_comment
            formb.waiver_comment=waiver_comment
            formb.form_a_comment=form_a_comment
            formb.questions_comment=questions_comment
            formb.consent_comment=consent_comment
            formb.proposal_comment=proposal_comment
            formb.supervisor_feedback=supervisor_feedback
            formb.recommendation=recommendation
            formb.status='Revisions required'
            formb.ethics_status=None
            formb.form_supervisor_status='Revisions required'
            formb.submitted_to_admin=False
            formb.supervisor_signature=None
            formb.signature_date=None
            formb.rejected_or_accepted=False
            try:
                message=f'Form belonging to {formb.applicant_name} has been reviewed and returned back to you. Please login to view the feedback.' 
            
                send_email(app,mail, message,[formb.email])
            except Exception as e:
                app.logger.error(f"Failed to send email to {formb.email}: {e}")
            

        
        db_session.add(formb)
        db_session.commit()
    return redirect(url_for('supervisor_dashboard'))


@app.route('/reject_or_Accept_form_c/<string:id>',methods=['GET','POST'])
@role_required('SUPERVISOR', 'REVIEWER')
def reject_or_Accept_form_c(id):
    user_id=session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        formc = db_session.query(FormC).filter_by(form_id=id).order_by(FormC.submission_date.desc()).first()
    except ValueError as e:
        # Handle corrupted date data in database
        app.logger.error(f"Date parsing error for form {id}: {e}")
        formc = db_session.query(FormC).filter_by(form_id=id).first()
    #admin=db_session.query(User).filter_by(role="Admin").all()
    if not formc:
        return "Form not found", 404
    if not can_act_as_assigned_supervisor(get_current_user(), formc):
        abort(403)
    if is_with_ethics(formc):
        flash('You have already reviewed this form. It is currently with Ethics.', 'info')
        return redirect(url_for('supervisor_dashboard'))
    if request.method=="POST":
        formc = (
            db_session.query(FormC)
            .filter_by(form_id=id)
            .with_for_update()
            .first()
        )
        if has_supervisor_submitted_feedback(formc):
            flash('You have already submitted feedback for this application. It was not submitted again.', 'warning')
            return redirect(url_for('supervisor_dashboard'))

        org_permission_comment=(request.form.get('org_permission_comment') or '').strip()
        waiver_comment=(request.form.get('waiver_comment') or '').strip()
        form_a_comment=(request.form.get('form_a_comment') or '').strip()
        questions_comment=(request.form.get('questions_comment') or '').strip()
        consent_comment=(request.form.get('consent_comment') or '').strip()
        proposal_comment=(request.form.get('proposal_comment') or '').strip()
        supervisor_feedback=(request.form.get('supervisor_feedback') or '').strip()
        recommendation=request.form.get('recommendation')
        supervisor_signature=(request.form.get('supervisor_signature') or '').strip()
        signature_date=datetime.now()

        if not supervisor_feedback:
            flash('Supervisor feedback is required before submitting your recommendation.', 'danger')
            return redirect(url_for('form_c_supervisor', form_id=id))
        
        if request.form.get('recommendation')=='Ready for submission':
            if not supervisor_signature:
                flash('Supervisor signature is required when approving for submission.', 'danger')
                return redirect(url_for('form_c_supervisor', form_id=id))
            missing_submission_redirect = redirect_if_missing_student_submission(
                formc,
                'ethics admin',
                'form_c_supervisor',
                form_id=id,
            )
            if missing_submission_redirect:
                return missing_submission_redirect
            formc.supervisor_date=datetime.now()
            formc.org_permission_comment=org_permission_comment
            formc.waiver_comment=waiver_comment
            formc.form_a_comment=form_a_comment
            formc.questions_comment=questions_comment
            formc.consent_comment=consent_comment
            formc.proposal_comment=proposal_comment
            formc.supervisor_feedback=supervisor_feedback
            formc.recommendation=recommendation
            formc.status='Submitted to Ethics'
            formc.ethics_status=None
            formc.form_supervisor_status='Ready for submission'
            formc.supervisor_signature=supervisor_signature
            formc.signature_date=signature_date
            formc.rejected_or_accepted=True
            formc.submitted_to_admin=True
            #Uncomment the code bellow for testing
            ##
            try:
                message=f'Form belongning to {formc.applicant_name} has been reviewed by the supervisor and has been accepted for further reviewing.' 
            
                send_email(app,mail, message,[formc.email_address])
            except Exception as e:
                app.logger.error(f"Failed to send email to {formc.email_address}: {e}")
             
        else:
            formc.supervisor_date=datetime.now()
            formc.org_permission_comment=org_permission_comment
            formc.waiver_comment=waiver_comment
            formc.form_a_comment=form_a_comment
            formc.questions_comment=questions_comment
            formc.consent_comment=consent_comment
            formc.proposal_comment=proposal_comment
            formc.supervisor_feedback=supervisor_feedback
            formc.recommendation=recommendation
            formc.status='Revisions required'
            formc.ethics_status=None
            formc.form_supervisor_status='Revisions required'
            formc.submitted_to_admin=False
            formc.supervisor_signature=None
            formc.signature_date=None
            formc.rejected_or_accepted=False
            try:
                message=f' Form belonging to {formc.applicant_name} has been reviewed and returned back to you. Please login to view the feedback.' 
            
                send_email(app,mail, message,[formc.email_address])
            except Exception as e:
                app.logger.error(f"Failed to send email to {formc.email_address}: {e}")
                
        db_session.add(formc)
        db_session.commit()
    return redirect(url_for('supervisor_dashboard'))

@app.route('/form_c_sec2', methods=['GET', 'POST'])
def form_c_sec2():
    try:
        user_id = session.get('id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401

        if request.method == "POST":
            form = db_session.query(FormC).filter_by(user_id=user_id).first()
            if not form:
                form = FormC(user_id=user_id)

            # --- Map boolean fields ---
            boolean_fields = {
                'vulnerable': 'vulnerable',
                'age_under_18_or_over_65': 'age_under_18_or_over_65',
                'uj_employee': 'uj_employees',
                'non_vulnerable_context': 'non_vulnerable_context',
                'non_english': 'non_english',
                'own_student': 'own_students',
                'poverty': 'poverty',
                'non_education': 'no_education',
                'consent_violation': 'consent_violation',
                'discomfiture': 'discomfiture',
                'deception': 'deception',
                'sensitive_issues': 'sensitive_issues',
                'prejuditial_info': 'prejudicial_info',
                'intrusive': 'intrusive',
                'illegal': 'illegal',
                'direct_social_info': 'direct_social_info',
                'identifiable_records': 'identifiable_records',
                'psychology_tests': 'psychology_tests',
                'reseacher_risk': 'researcher_risk',
                'incentives': 'incentives',
                'participant_costs': 'participant_costs',
                'researcher_interest': 'researcher_interest',
                'conflict_of_interest': 'conflict_of_interest',
                'uj_premises': 'uj_premises',
                'uj_facilities': 'uj_facilities',
                'uj_funding': 'uj_funding'
            }

            for form_field, model_field in boolean_fields.items():
                setattr(form, model_field, request.form.get(form_field) == "Yes")

            # --- Map text fields ---
            text_fields = [
                'vulnerable_other_description',
                'vulnerable_comments',
                'activity_other_description',
                'activity_comments',
                'consideration_comments',
                'risk_level',
                'risk_justification',
                'risk_benefits',
                'risk_mitigation'
            ]

            for field in text_fields:
                setattr(form, field, request.form.get(field))

            db_session.add(form)
            db_session.commit()

            message = "Form submitted successfully"
            flash(message, "success")
            return render_template("form-c-section3.html", messages=[message])

        # GET request
        return render_template("form-c-section2.html")

    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Error in form_c_sec2: {e}")
        flash("An error occurred while submitting the form. Please try again.", "danger")
        return render_template("form-c-section2.html"), 500


@app.route('/form_c_sec3', methods=['GET', 'POST'])
def form_c_sec3():
    try:
        user_id = session.get('id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401

        if request.method == "POST":
            form = db_session.query(FormC).filter_by(user_id=user_id).first()
            if not form:
                form = FormC(user_id=user_id)

            # --- Map text fields dynamically ---
            text_fields = [
                'summary_title',
                'executive_summary',
                'research_questions',
                'research_purpose',
                'secondary_data_info',
                'exemption_reason'
            ]

            for field in text_fields:
                setattr(form, field, request.form.get(field))

            db_session.add(form)
            db_session.commit()

            message = "Form submitted successfully"
            flash(message, "success")
            return render_template("form-c-section4.html", messages=[message])

        # GET request
        return render_template("form-c-section3.html")

    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Error in form_c_sec3: {e}")
        flash("An error occurred while submitting the form. Please try again.", "danger")
        return render_template("form-c-section3.html"), 500


@app.route('/form_c_sec4', methods=['GET', 'POST'])
def form_c_sec4():
    try:
        user_id = session.get('id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401

        if request.method == "POST":
            form = db_session.query(FormC).filter_by(user_id=user_id).first()
            user = db_session.query(User).filter_by(user_id=user_id).first()

            if not form:
                form = FormC(user_id=user_id)

            # Update form declaration fields
            form.declaration_name = request.form.get('declaration_name')
            form.full_name = request.form.get('full_name')
            form.submitted = True
            form.submission_date = datetime.now()

            db_session.add(form)
            db_session.commit()

            # Try sending notification email
            try:
                message = f'{form.applicant_name} has submitted a form that needs to be reviewed.'
                send_email(app, mail, message, [form.supervisor_email])
            except Exception as e:
                app.logger.error(f"Failed to send email to {form.supervisor_email}: {e}")

            flash("Form submitted successfully", "success")
            return redirect(url_for('student_dashboard'))

        # GET request
        return render_template("form-c-section4.html")

    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Error in form_c_sec4: {e}")
        flash("An error occurred while submitting the form. Please try again.", "danger")
        return render_template("form-c-section4.html"), 500


@app.route('/form_a_answers', methods=['GET','POST'])
def form_a_answers():
    user_id=session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    form = db_session.query(FormA).filter_by(user_id=user_id).first()
    
    return render_template("form_a_answers.html",formA=form)


@app.route('/student_edit_forma', methods=['GET', 'POST'])
def student_edit_forma():
    user_id = session.get('id')

    formB_options = [
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ]

    if has_blocking_student_form(FormB, user_id, options=formB_options):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))

    if has_blocking_student_form(FormC, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))

    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401

    user = db_session.query(User).filter(User.user_id == user_id).first()

    supervisor = get_student_supervisor_or_flash(user)
    if not supervisor:
        return redirect(url_for("student_dashboard"))

    form = _get_latest_forma_for_user(user_id)
    if form is not None and getattr(form, 'submitted_at', None) is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_forma_resubmission_draft(user_id, form)
        if error_response:
            return error_response
    elif form is None or (getattr(form, 'submitted_at', None) is not None and not is_student_correction_state(form)):
        form, error_response = _get_or_create_forma_draft(user_id, {})
        if error_response:
            return error_response

    form_requirements = db_session.query(FormARequirements).filter(FormARequirements.user_id == user_id).first()
    source_form = form

    def get_context_data(current_form):
        if current_form:
            data_org = {
                "org_name": parse_field(current_form.org_name),
                "org_contact": parse_field(current_form.org_contact),
                "org_role": parse_field(current_form.org_role),
                "org_permission": parse_field(current_form.org_permission),
            }

            data_fund = {
                "fund_org": parse_field(current_form.fund_org),
                "fund_contact": parse_field(current_form.fund_contact),
                "fund_role": parse_field(current_form.fund_role),
                "fund_amount": parse_field(current_form.fund_amount),
            }

            data_sampling = {
                "population": parse_field(current_form.population),
                "sampling_method": parse_field(current_form.sampling_method),
                "sampling_size": parse_field(current_form.sampling_size),
                "inclusion_criteria": parse_field(current_form.inclusion_criteria)
            }

            data_store = {
                "data": parse_field(current_form.data_storage)
            }

            privacy_data = {
                "privacy": parse_field(current_form.privacy)
            }

        else:
            data_org = {"org_name": [], "org_contact": [], "org_role": [], "org_permission": []}
            data_fund = {"fund_org": [], "fund_contact": [], "fund_role": [], "fund_amount": []}
            data_sampling = {"population": [], "sampling_method": [], "sampling_size": [], "inclusion_criteria": []}
            data_store = {"data": []}
            privacy_data = {"privacy": []}

        return data_org, data_fund, data_sampling, data_store, privacy_data

    data_org, data_fund, data_sampling, data_store, _privacy = get_context_data(form)

    if not form:
        form = FormA(user_id=user_id)
        inherit_previous_reviewers(form, FormA, user_id, FormA.submitted_at)

    def is_yes(field_name):
        return request.form.get(field_name) == 'Yes'

    def join_list(field_name):
        return ','.join(request.form.getlist(field_name))

    def populate_form_from_request(target_form):
        public_data_description = ""

        # Section 1
        if form_requirements:
            target_form.attachment_id = form_requirements.id

        target_form.applicant_name = request.form.get('applicant_name')
        target_form.student_number = request.form.get('student_number')
        target_form.institution = request.form.get('institution')
        target_form.department = request.form.get('department')
        target_form.degree = request.form.get('degree')
        target_form.study_title = request.form.get('study_title')
        target_form.mobile = request.form.get('mobile')
        target_form.email = user.email
        target_form.supervisor = supervisor.full_name
        target_form.supervisor_email = supervisor.email

        # Section 2
        target_form.survey = is_yes('survey')
        target_form.observations = is_yes('observations')
        target_form.focus_groups = is_yes('focus_groups')
        target_form.interviews = is_yes('interviews')
        target_form.documents = is_yes('documents')

        target_form.vulnerable_other_specify = request.form.get('vulnerable_other_specify')
        target_form.vulnerable_communities = is_yes('vulnerable_communities')
        target_form.age_range = is_yes('age_range')
        target_form.uj_employees = is_yes('uj_employees')
        target_form.vulnerable = is_yes('vulnerable')
        target_form.non_english = is_yes('non_english')
        target_form.own_students = is_yes('own_students')
        target_form.poverty = is_yes('poverty')
        target_form.no_education = is_yes('no_education')
        target_form.assessment_other_specify = request.form.get('assessment_other_specify')

        target_form.disclosure = is_yes('disclosure')
        target_form.discomfiture = is_yes('discomfiture')
        target_form.deception = is_yes('deception')
        target_form.sensitive = is_yes('sensitive')
        target_form.prejudice = is_yes('prejudice')
        target_form.intrusive_techniques = is_yes('intrusive_techniques')
        target_form.illegal_activities = is_yes('illegal_activities')
        target_form.personal = is_yes('personal')
        target_form.available_records = is_yes('available_records')
        target_form.inventories = is_yes('inventories')
        target_form.risk_activities = is_yes('risk_activities')
        target_form.activity_specify = request.form.get('activity_specify')
        target_form.vulnerable_comments_2 = request.form.get('vulnerable_comments_2')

        target_form.incentives = is_yes('incentives')
        target_form.financial_costs = is_yes('financial_costs')
        target_form.reward = is_yes('reward')
        target_form.conflict = is_yes('conflict')
        target_form.uj_premises = is_yes('uj_premises')
        target_form.uj_facilities = is_yes('uj_facilities')
        target_form.uj_funding = is_yes('uj_funding')
        target_form.vulnerable_comments_3 = request.form.get('vulnerable_comments_3')

        target_form.risk_rating = request.form.get('risk_rating')
        target_form.risk_justification = request.form.get('risk_justification')
        target_form.benefits_description = request.form.get('benefits_description')
        target_form.risk_mitigation = request.form.get('risk_mitigation')

        target_form.interviews_one = is_yes('interviews')
        target_form.documents_one = is_yes('documents')
        target_form.other_sec2 = request.form.get('other_sec2')

        # Section 3
        target_form.title_provision = request.form.get('title_provision')
        target_form.abstract = request.form.get('abstract')
        target_form.questions = request.form.get('questions')
        target_form.purpose_objectives = request.form.get('purpose_objectives')

        # Section 4
        target_form.grant_permission = request.form.get('grant_permission')
        if (target_form.grant_permission or '').strip().lower() == 'yes':
            target_form.org_name = join_list('org_name[]')
            target_form.org_contact = join_list('org_contact[]')
            target_form.org_role = join_list('org_role[]')
            target_form.org_permission = join_list('org_permission[]')
        else:
            target_form.org_name = ''
            target_form.org_contact = ''
            target_form.org_role = ''
            target_form.org_permission = ''

        target_form.researcher_affiliation = request.form.get('researcher_affiliation')
        target_form.affiliation_details = request.form.get('affiliation_details') if (target_form.researcher_affiliation or '').lower() == 'yes' else ''

        target_form.collective_involvement = request.form.get('collective_involvement')
        target_form.collective_details = request.form.get('collective_details') if (target_form.collective_involvement or '').lower() == 'yes' else ''

        target_form.is_funded = request.form.get('is_funded')
        target_form.fund_org = join_list('fund_org[]')
        target_form.fund_contact = join_list('fund_contact[]')
        target_form.fund_role = join_list('fund_role[]')
        target_form.fund_amount = join_list('fund_amount[]')

        target_form.indemnity_arrangements = request.form.get('indemnity_arrangements')
        target_form.other_committee = request.form.get('other_committee')

        # Section 5
        target_form.quantitative = "Yes" in request.form.getlist('quantitative[]')
        target_form.qualitative = "Yes" in request.form.getlist('qualitative[]')
        target_form.mixed_methods = "Yes" in request.form.getlist('mixed_methods[]')
        target_form.paradigm_explanation = request.form.get('paradigm_explanation')

        target_form.design = request.form.get('design')

        target_form.participants_description = request.form.get('participants_description')
        target_form.population = join_list('population[]')
        target_form.sampling_method = join_list('sampling_method[]')
        target_form.sampling_size = ','.join(cleaned_sample_sizes)
        target_form.inclusion_criteria = join_list('inclusion_criteria[]')
        target_form.duration_timing = request.form.get('duration_timing')
        target_form.contact_details_method = request.form.get('contact_details_method')
        target_form.conflict_interest = is_yes('conflict_interest')
        target_form.conflict_explanation = request.form.get('conflict_explanation')

        target_form.questionnaire_type = request.form.get('questionnaire_type')

        questionnaire_permission = request.form.get('questionnaire_permission')
        if questionnaire_permission is not None:
            questionnaire_permission = questionnaire_permission.strip()
            if questionnaire_permission == 'Yes':
                target_form.permission_obtained = 'Yes'
                target_form.open_source = 'No'
            elif questionnaire_permission == 'Open Source':
                target_form.permission_obtained = 'No'
                target_form.open_source = 'Yes'
            else:
                target_form.permission_obtained = None
                target_form.open_source = None
        else:
            target_form.permission_obtained = request.form.get('permission_obtained')
            target_form.open_source = request.form.get('open_source')

        target_form.instrument_attachment_reason = request.form.get('instrument_attachment_reason')
        target_form.data_collection_procedure = request.form.get('data_collection_procedure')
        target_form.interview_type = request.form.get('interview_type')
        target_form.interview_recording = request.form.get('interview_recording')
        target_form.focus_recording = request.form.get('focus_recording')
        target_form.observation_details = request.form.get('observation_details')
        target_form.documents_details = request.form.get('documents_details')
        target_form.other_details = request.form.get('other_details')
        target_form.data_collectors = request.form.get('data_collectors')
        target_form.data_methods = join_list('data_methods[]')
        target_form.intervention = is_yes('intervention')
        target_form.intervention_details = request.form.get('intervention_details')
        target_form.sensitive_data = request.form.get('sensitive_data')
        target_form.translator = is_yes('translator')
        target_form.translator_procedure = request.form.get('translator_procedure')

        # 5.5 Secondary Data Usage
        uses_secondary_data = request.form.get('uses_secondary_data')
        if uses_secondary_data == 'Yes':
            target_form.uses_secondary_data = True
        elif uses_secondary_data == 'No':
            target_form.uses_secondary_data = False

        target_form.secondary_data_type = request.form.get('data_type')
        target_form.data_nature = request.form.get('data_nature')
        target_form.data_origin = request.form.get('data_origin')
        target_form.access_conditions = request.form.get('access_conditions')
        target_form.personal_info = request.form.get('personal_info')
        target_form.personal_info_comment = request.form.get('personal_info_comment')
        target_form.data_anonymized = request.form.get('data_anonymized')
        target_form.anonymization_comment = request.form.get('anonymization_comment')
        target_form.permission_details = request.form.get('permission_details')
        target_form.shortcomings_reported = request.form.get('shortcomings_reported')
        target_form.limitations_reporting = request.form.get('limitations_reporting')
        target_form.methodology_alignment = request.form.get('methodology_alignment')
        target_form.data_acknowledgment = request.form.get('data_acknowledgment')
        target_form.private_permission = request.form.get('privatePermission')
        target_form.public_data_description = request.form.get('public_data_description')

        if not target_form.uses_secondary_data:
            _clear_forma_secondary_data_details(target_form)

        # Section 6
        target_form.informed_consent = request.form.get('informed_consent')
        target_form.secure_location = join_list('secure_location[]')
        target_form.password_protected = join_list('password_protected[]')
        target_form.protected_place = join_list('protected_place[]')
        target_form.retention = join_list('retention[]')
        target_form.data_storage = join_list('data_storage[]')
        target_form.study_benefits = request.form.get('study_benefits')
        target_form.participant_risks = request.form.get('participant_risks')
        target_form.adverse_steps = request.form.get('adverse_steps')
        target_form.community_participation = request.form.get('community_participation')
        target_form.community_effects = request.form.get('community_effects')

        target_form.privacy = join_list('privacy[]')
        for key in ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's']:
            field_name = f'q6_9{key}'
            setattr(target_form, field_name, bool(_autosave_str_to_bool(request.form.get(field_name))))

        target_form.results_feedback = request.form.get('results_feedback')
        target_form.products_access = request.form.get('products_access')
        target_form.publication_plans = request.form.get('publication_plans')
        target_form.participant_comp = request.form.get('participant_comp')
        target_form.participant_costs = request.form.get('participant_costs')
        target_form.ethics_reporting = request.form.get('ethics_reporting')

        return target_form

    if request.method == 'POST':
        autosave_only = request.form.get('autosave_only') == '1'
        cleaned_sample_sizes, sample_size_error = normalize_forma_sample_sizes(
            request.form.getlist('sample_size[]')
        )
        if sample_size_error:
            if autosave_only:
                return jsonify({'success': False, 'error': sample_size_error}), 422
            flash(sample_size_error, "danger")
            return redirect(request.referrer or url_for('student_edit_forma'))

        if autosave_only:
            autosave_form, error_response = _get_or_create_forma_draft(user_id, request.form)
            if error_response:
                return error_response

            populate_form_from_request(autosave_form)
            db_session.commit()

            return jsonify({
                "success": True,
                "message": "Autosaved current Form A successfully.",
                "form_id": autosave_form.form_id
            })

        was_in_corrections = is_student_correction_state(source_form)

        form = FormA(user_id=user_id)
        inherit_previous_reviewers(form, FormA, user_id, FormA.submitted_at)
        db_session.add(form)

        populate_form_from_request(form)

        declaration_name = request.form.get('declaration_name', '').strip()
        applicant_signature = request.form.get('applicant_signature', '').strip()
        declaration_date_raw = datetime.now()

        if declaration_name and applicant_signature:
            try:
                declaration_date = parse_html_date(declaration_date_raw) or get_local_time()
            except ValueError:
                flash('Please provide a valid declaration date.', 'warning')
                return redirect(url_for('student_edit_forma'))

            form.declaration_name = declaration_name
            form.applicant_signature = applicant_signature
            form.declaration_date = declaration_date
            form.submitted = True
            form.submitted_at = datetime.now()
            form.rejected_or_accepted = False
            form.status = 'Resubmitted' if was_in_corrections else 'Submitted'
            form.visible_to_student = False
            form.ethics_status = None
            form.form_supervisor_status = 'Resubmitted' if was_in_corrections else 'Submitted'
            reset_form_review_feedback(form)
        else:
            flash('Declaration section must be completed to submit the form', 'warning')
            return redirect(url_for('student_edit_forma'))

        db_session.commit()

        try:
            message = (
                f'you have successfully edited and submitted your form. '
                f'Please wait while its under review.'
            )

            send_email(app, mail, message, [user.email])

            messages = (
                f'{form.applicant_name} has submitted a form that needs to be reviewed.'
            )

            send_email(app, mail, messages, [form.supervisor_email])

        except Exception as e:
            print("Email sending error:", str(e))

        return redirect(url_for('student_dashboard'))

    return render_template(
        "student_edit_forma.html",
        formA=form,
        data_org=data_org,
        data_fund=data_fund,
        data_sampling=data_sampling,
        data_storage=data_store,
        privacy=_privacy
    )







@app.route('/student_continue_forma', methods=['GET','POST'])
def student_continue_forma():
    user_id=session.get('id')

    formB_options = [
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ]
    if has_blocking_student_form(FormB, user_id, options=formB_options):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
        
    if has_blocking_student_form(FormC, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
    

    public_data_description=""
    private_permission_file=None
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor = get_student_supervisor_or_flash(user)
    if not supervisor:
        return redirect(url_for("student_dashboard"))
    
    # Check if form exists in database - if yes UPDATE, if no CREATE
    form = _get_latest_forma_for_user(user_id)
    if form is not None and getattr(form, 'submitted_at', None) is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_forma_resubmission_draft(user_id, form)
        if error_response:
            return error_response
    elif form is None or (form.submitted_at is not None and not is_student_correction_state(form)):
        form, error_response = _get_or_create_forma_draft(user_id, {})
        if error_response:
            return error_response

    form_requirements = db_session.query(FormARequirements).filter(FormARequirements.user_id == user_id).first()
    if not form_requirements:
        flash("Form A requirements not found. Please complete them first.", "warning")
        return redirect(url_for("submit_form_a_requirements"))

    data_org={
            "org_name" : parse_field(form.org_name),
            "org_contact" : parse_field(form.org_contact),
            "org_role": parse_field(form.org_role),
            "org_permission" : parse_field(form.org_permission),
        }
    data_fund={
            "fund_org" : parse_field(form.fund_org),
            "fund_contact" :parse_field(form.fund_contact),
            "fund_role": parse_field(form.fund_role),
            "fund_amount": parse_field(form.fund_amount),
        }
    data_sampling={
            "population" :parse_field(form.population),
            "sampling_method" : parse_field(form.sampling_method),
            "sampling_size": parse_field(form.sampling_size),
            "inclusion_criteria": parse_field(form.inclusion_criteria)
        }
        
    if request.method == 'POST':
        try:
            cleaned_sample_sizes, sample_size_error = normalize_forma_sample_sizes(
                request.form.getlist('sample_size[]')
            )
            if sample_size_error:
                flash(sample_size_error, "danger")
                return redirect(request.referrer or url_for('student_continue_forma'))

            if request.form.get('survey')=='Yes':
                survey=True
            else:
                survey=False

            if request.form.get('focus_groups')=='Yes':
                focus_groups=True
            else:
                focus_groups=False

            if request.form.get('observations')=='Yes':
                observations=True
            else:
                observations=False

            if request.form.get('interviews')=='Yes':
                interviews=True
            else:
                interviews=False

            if request.form.get('documents')=='Yes':
                documents=True
            else:
                documents=False
            
            if request.form.get('vulnerable_other_specify')=='Yes':
                vulnerable_other_specify=True
            else:
                vulnerable_other_specify=False

            # section 2.1
            if request.form.get('vulnerable_communities')=='Yes':
                vulnerable_communities=True
            else:
                vulnerable_communities=False

            if request.form.get('age_range')=='Yes':
                age_range=True
            else:
                age_range=False

            if request.form.get('uj_employees')=='Yes':
                uj_employees=True
            else:
                uj_employees=False

            if request.form.get('vulnerable')=='Yes':
                vulnerable=True
            else:
                vulnerable=False

            if request.form.get('non_english')=='Yes':
                non_english=True
            else:
                non_english=False

            if request.form.get('own_students')=='Yes':
                own_students=True
            else:
                own_students=False

            if request.form.get('poverty')=='Yes':
                poverty=True
            else:
                poverty=False
            
            if request.form.get('no_education')=='Yes':
                no_education=True
            else:
                no_education=False
            
            form.assessment_other_specify=request.form.get('assessment_other_specify')

            if request.form.get('vulnerable_comments_1')=='Yes':
                vulnerable_comments_1=True
            else:
                vulnerable_comments_1=False

            # 2.2
            if request.form.get('disclosure')=='Yes':
                disclosure=True
            else:
                disclosure=False

            if request.form.get('discomfiture')=='Yes':
                discomfiture=True
            else:
                discomfiture=False

            if request.form.get('deception')=='Yes':
                deception=True
            else:
                deception=False
            
            if request.form.get('sensitive')=='Yes':
                sensitive=True
            else:
                sensitive=False

            if request.form.get('prejudice')=='Yes':
                prejudice=True
            else:
                prejudice=False

            
            if request.form.get('intrusive_techniques')=='Yes':
                intrusive_techniques=True
            else:
                intrusive_techniques=False

            if request.form.get('illegal_activities')=='Yes':
                illegal_activities=True
            else:
                illegal_activities=False

            if request.form.get('personal')=='Yes':
                personal=True
            else:
                personal=False
                
            if request.form.get('available_records')=='Yes':
                available_records=True
            else:
                available_records=False

            if request.form.get('inventories')=='Yes':
                inventories=True
            else:
                inventories=False

            if request.form.get('risk_activities')=='Yes':
                risk_activities=True
            else:
                risk_activities=False

            activity_specify=request.form.get('activity_specify')
           
            vulnerable_comments_2=request.form.get('vulnerable_comments_2')
            
            
            # Risk Assessment 2.3
            if request.form.get('incentives')=='Yes':
                incentives=True
            else:
                incentives=False

            if request.form.get('financial_costs')=='Yes':
                financial_costs=True
            else:
                financial_costs=False

            if request.form.get('reward')=='Yes':
                reward=True
            else:
                reward=False
            
            if request.form.get('conflict')=='Yes':
                conflict=True
            else:
                conflict=False

            if request.form.get('uj_premises')=='Yes':
                uj_premises=True
            else:
                uj_premises=False
    
            if request.form.get('uj_facilities')=='Yes':
                uj_facilities=True
            else:
                uj_facilities=False

            if request.form.get('uj_funding')=='Yes':
                uj_funding=True
            else:
                uj_funding=False
            
            form.vulnerable_comments_3=request.form.get('vulnerable_comments_3')
            
            if request.form.get('dataType') == 'public':
                    public_data_description = request.form.get('public_data_description')

            if request.form.get('researcher_affiliation')=='Yes':
                researcher_affiliation=True
            else:
                researcher_affiliation=False

            if request.form.get('collective_involvement')=='Yes':
                collective_involvement=True
            else:
                collective_involvement=False
        
            secondary_data = request.form.get('uses_secondary_data', request.form.get('secondary_data'))
            form.uses_secondary_data = _autosave_str_to_bool(secondary_data)
            if form.uses_secondary_data:
                form.secondary_data_type = request.form.get('data_type')
                if form.secondary_data_type == 'private':
                    form.private_permission = request.form.get('privatePermission') == 'Yes'
                    # Handle file upload for permission if required
                    # Add logic for saving file securely if uploaded
                elif form.secondary_data_type == 'public':
                    form.public_data_description = request.form.get('public_data_description')
            else:
                _clear_forma_secondary_data_details(form)
                
            
                
            if request.form.get('translator')=='Yes':
                translator=True
            else:
                translator=False

            if request.form.get('intervention')=='Yes':
                intervention=True
            else:
                intervention=False


            if request.form.get('use_focus_groups')=='Yes':
                use_focus_groups=True

            else:
                use_focus_groups=False


            if request.form.get('conflict_interest')=='Yes':
                conflict_interest=True
            else:
                conflict_interest=False

            
            data_nature=request.form.get('data_nature')
            data_origin=request.form.get('data_origin')
            access_conditions=request.form.get('access_conditions')
            personal_info=request.form.get('personal_info')
            personal_info_comment=request.form.get('personal_info_comment')
            data_anonymized=request.form.get('data_anonymized')
            anonymization_comment=request.form.get('anonymization_comment')
        
            permission_details=request.form.get('permission_details')
            public_data_description=request.form.get('public_data_description')
            shortcomings_reported=request.form.get('shortcomings_reported')
            limitations_reporting=request.form.get('limitations_reporting')
            methodology_alignment=request.form.get('methodology_alignment')
            data_acknowledgment=request.form.get('data_acknowledgment')

            interviews_one = request.form.get('interviews') == 'Yes'
            documents_one = request.form.get('documents') == 'Yes'
            
            form.user_id=user_id
            form.attachment_id=form_requirements.id
            form.applicant_name=request.form.get('applicant_name')
            form.student_number=request.form.get('student_number')
            form.institution=request.form.get('institution')
            form.department=request.form.get('department')
            form.degree=request.form.get('degree')
            form.study_title=request.form.get('study_title')
            form.mobile=request.form.get('mobile')
            form.email=user.email
            form.supervisor=supervisor.full_name
            form.supervisor_email=supervisor.email
            form.survey=survey
            form.observations=observations
            form.focus_groups=focus_groups
            form.interviews=interviews
            form.documents=documents
            form.vulnerable_other_specify=vulnerable_other_specify
            form.vulnerable_communities=vulnerable_communities
            form.age_range=age_range
            form.uj_employees=uj_employees
            form.vulnerable=vulnerable
            form.non_english=non_english
            form.own_students=own_students
            form.poverty=poverty
            form.no_education=no_education
            form.assessment_other_specify=request.form.get('assessment_other_specify')
            form.vulnerable_comments_1=vulnerable_comments_1
            form.disclosure=disclosure
            form.discomfiture=discomfiture
            form.deception=deception
            form.sensitive=sensitive
            form.prejudice=prejudice
            form.intrusive_techniques=intrusive_techniques
            form.illegal_activities=illegal_activities
            form.personal=personal
            form.available_records=available_records
            form.inventories=inventories
            form.risk_activities=risk_activities
            form.activity_specify=activity_specify
            form.vulnerable_comments_2=vulnerable_comments_2
            form.incentives=incentives
            form.financial_costs=financial_costs
            form.reward=reward
            form.conflict=conflict
            form.uj_premises=uj_premises
            form.uj_facilities=uj_facilities
            form.uj_funding=uj_funding
            form.vulnerable_comments_3=request.form.get('vulnerable_comments_3')
            form.risk_rating = request.form.get('risk_rating')
            form.risk_justification = request.form.get('risk_justification')
            form.benefits_description = request.form.get('benefits_description')
            form.risk_mitigation = request.form.get('risk_mitigation')

            form.interviews_one = interviews_one
            form.documents_one = documents_one
            form.other_sec2 = request.form.get('other_sec2')
                # Section 3: Project Information
            form.title_provision = request.form.get('title_provision')
            form.abstract = request.form.get('abstract')
            form.questions = request.form.get('questions')
            form.purpose_objectives = request.form.get('purpose_objectives')

                # Section 4: Organisational Permissions and Affiliations
            form.grant_permission=request.form.get('grant_permission')
            if (form.grant_permission or '').strip().lower() == 'yes':
                form.org_name = ','.join(request.form.getlist('org_name[]'))
                form.org_contact = ','.join(request.form.getlist('org_contact[]'))
                form.org_role = ','.join(request.form.getlist('org_role[]'))
                form.org_permission=','.join(request.form.getlist('org_permission[]'))
            else:
                form.org_name = ''
                form.org_contact = ''
                form.org_role = ''
                form.org_permission = ''
                
            form.researcher_affiliation = 'Yes' if researcher_affiliation else 'No'
            form.affiliation_details = request.form.get('affiliation_details') if researcher_affiliation else ''

            form.collective_involvement = 'Yes' if collective_involvement else 'No'
                

            form.collective_details = request.form.get('collective_details') if collective_involvement else ''
                # Funding Information
            form.is_funded = request.form.get('is_funded')
            form.fund_org = ','.join(request.form.getlist('fund_org[]'))
            form.fund_contact = ','.join(request.form.getlist('fund_contact[]'))
            form.fund_role = ','.join(request.form.getlist('fund_role[]'))
            form.fund_amount =','.join(request.form.getlist('fund_amount[]'))

                # Indemnity & Other Committee Info
            form.indemnity_arrangements = request.form.get('indemnity_arrangements')
            form.other_committee = request.form.get('other_committee')

        
                # 5.1 Research Paradigm
            form.quantitative = "Yes" in request.form.getlist('quantitative[]')
            form.qualitative ="Yes" in request.form.getlist('qualitative[]')
            form.mixed_methods = "Yes" in request.form.getlist('mixed_methods[]')
            form.paradigm_explanation = request.form.get('paradigm_explanation')

                # 5.2 Research Design
            form.design = request.form.get('design')

                # 5.3 Participant Details
            form.participants_description = request.form.get('participants_description')
            form.population = ','.join(request.form.getlist('population[]'))
            form.sampling_method = ','.join(request.form.getlist('sampling_method[]'))
            form.sampling_size = ','.join(cleaned_sample_sizes)
            form.inclusion_criteria =','.join(request.form.getlist('inclusion_criteria[]'))
            form.duration_timing = request.form.get('duration_timing')
            form.contact_details_method = request.form.get('contact_details_method')
            form.conflict_interest = conflict_interest
            form.conflict_explanation = request.form.get('conflict_explanation')

                # 5.4 Instruments
            form.questionnaire_type = request.form.get('questionnaire_type')
            questionnaire_permission = request.form.get('questionnaire_permission')
            if questionnaire_permission is not None:
                questionnaire_permission = questionnaire_permission.strip()
                if questionnaire_permission == 'Yes':
                    form.permission_obtained = 'Yes'
                    form.open_source = 'No'
                elif questionnaire_permission == 'Open Source':
                    form.permission_obtained = 'No'
                    form.open_source = 'Yes'
                else:
                    form.permission_obtained = None
                    form.open_source = None
            else:
                form.permission_obtained = request.form.get('permission_obtained')
                form.open_source = request.form.get('open_source')
            form.instrument_attachment_reason = request.form.get('instrument_attachment_reason')
            form.data_collection_procedure = request.form.get('data_collection_procedure')
            form.interview_type = request.form.get('interview_type')
            form.interview_recording = request.form.get('interview_recording')
            form.use_focus_groups = use_focus_groups
            form.focus_recording = request.form.get('focus_recording')
            form.observation_details = request.form.get('observation_details')
            form.documents_details = request.form.get('documents_details')
            form.other_details = request.form.get('other_details')
            form.data_collectors = request.form.get('data_collectors')
            form.data_methods=','.join(request.form.getlist('data_methods[]'))
                #in_depth=request.form.get("in_depth"),
                #semi_structured=request.form.get("semi_structured"),
                #unstructured=request.form.get("unstructured"),
            form.intervention =intervention
            form.intervention_details = request.form.get('intervention_details')
            form.sensitive_data = request.form.get('sensitive_data')
            form.translator = translator
            form.translator_procedure = request.form.get('translator_procedure')

                # 5.5 Secondary Data Usage
            form.data_nature=data_nature
            form.data_origin=data_origin
            form.access_conditions=access_conditions
            form.personal_info=personal_info
            form.personal_info_comment=personal_info_comment
            form.data_anonymized=data_anonymized
            form.anonymization_comment=anonymization_comment
            form.permission_details=permission_details
            form.shortcomings_reported=shortcomings_reported
            form.limitations_reporting=limitations_reporting
            form.methodology_alignment=methodology_alignment
            form.data_acknowledgment=data_acknowledgment
            form.private_permission= request.form.get('privatePermission')
            form.public_data_description=public_data_description
            form.informed_consent=request.form.get('informed_consent')
            form.secure_location=','.join(request.form.getlist('secure_location[]'))
            form.password_protected=','.join(request.form.getlist('password_protected[]'))
            form.protected_place=','.join(request.form.getlist('protected_place[]'))
            form.retention=','.join(request.form.getlist('retention[]'))
            form.data_storage=','.join(request.form.getlist('data_storage[]'))
            form.study_benefits=request.form.get('study_benefits')
            form.participant_risks=request.form.get('participant_risks')
            form.adverse_steps=request.form.get('adverse_steps')
            form.community_participation=request.form.get('community_participation')
            form.community_effects=request.form.get('community_effects')
                #remove_identifiers=request.form.getlist("remove_identifiers"),
                #encryption=request.form.getlist("encryption"),
                #pseudonyms=request.form.getlist("pseudonyms"),
                #focus_group_warning=request.form.getlist("focus_group_warning"),
            form.privacy=','.join(request.form.getlist('privacy[]'))
            form.q6_9a= request.form.get("q6_9a")=='yes'
            form.q6_9b=request.form.get("q6_9b")=='yes'
            form.q6_9c=request.form.get("q6_9c")=='yes'
            form.q6_9d=request.form.get("q6_9d")=='yes'
            form.q6_9e=request.form.get("q6_9e")=='yes'
            form.q6_9f=request.form.get("q6_9f")=='yes'
            form.q6_9g=request.form.get("q6_9g")=='yes'
            form.q6_9h=request.form.get("q6_9h")=='yes'
            form.q6_9i=request.form.get("q6_9i")=='yes'
            form.q6_9j=request.form.get("q6_9j")=='yes'
            form.q6_9k=request.form.get("q6_9k")=='yes'
            form.q6_9l=request.form.get("q6_9l")=='yes'
            form.q6_9m=request.form.get("q6_9m")=='yes'
            form.q6_9n=request.form.get("q6_9n")=='yes'
            form.q6_9o=request.form.get("q6_9o")=='yes'
            form.q6_9p=request.form.get("q6_9p")=='yes'
            form.q6_9q=request.form.get("q6_9q")=='yes'
            form.q6_9r=request.form.get("q6_9r")=='yes'
            form.q6_9s=request.form.get("q6_9s")=='yes'
            form.results_feedback=request.form.get('results_feedback')
            form.products_access=request.form.get('products_access')
            form.publication_plans=request.form.get('publication_plans')
            form.participant_comp=request.form.get('participant_comp')
            form.participant_costs=request.form.get('participant_costs')
            form.ethics_reporting=request.form.get('ethics_reporting')
            

            if not form.uses_secondary_data:
                _clear_forma_secondary_data_details(form)
    
            db_session.add(form)
            db_session.commit()
         
            return redirect(url_for('student_dashboard'))
        except Exception as e:
            db_session.rollback()
            flash(f"Error saving Form A: {str(e)}", "danger")
            print("Exception in FormA save:", e)
            import traceback; traceback.print_exc()
            return render_template("student_continue_forma.html",formA=form,data_org=data_org,data_fund=data_fund,data_sampling=data_sampling)
    return render_template("student_continue_forma.html",formA=form,data_org=data_org,data_fund=data_fund,data_sampling=data_sampling)


@app.route('/submit_form_a/<string:form_id>',methods=['GET','POST'])
def submit_form_a(form_id):
    """
    Submit form to supervisor - saves ALL sections including Section 7 and marks as submitted
    """
    print(f"DEBUG submit_form_a called: form_id={form_id}, method={request.method}")
    
    user_id=session.get('id')

    public_data_description=""
    private_permission_file=None
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor=db_session.query(User).filter(User.user_id == user.supervisor_id).first()
    
    # Resolve the draft to submit. Never create a new row at submit time,
    # otherwise a stale/missing form_id can produce duplicate records.
    form = db_session.query(FormA).filter(FormA.user_id==user_id, FormA.form_id==form_id).first()
    form_exists = form is not None

    if not form_exists:
        # Fallback to the latest unsubmitted draft for this user.
        form = (
            db_session.query(FormA)
            .filter(FormA.user_id == user_id, FormA.submitted_at.is_(None))
            .order_by(FormA.created_at.desc())
            .first()
        )
        if not form:
            flash("No draft Form A found to submit. Please open your draft and try again.", "warning")
            return redirect(url_for('student_continue_forma'))
    elif getattr(form, 'submitted_at', None) is not None:
        if is_student_correction_state(form):
            form, error_response = _get_or_create_forma_resubmission_draft(user_id, form)
            if error_response:
                return error_response
        else:
            flash("This submitted Form A is locked. Please open your current draft and try again.", "warning")
            return redirect(url_for('student_continue_forma'))
   
    form_requirements = db_session.query(FormARequirements).filter(FormARequirements.user_id == user_id).first()
    
    if request.method=='POST':
        cleaned_sample_sizes, sample_size_error = normalize_forma_sample_sizes(
            request.form.getlist('sample_size[]')
        )
        if sample_size_error:
            flash(sample_size_error, "danger")
            return redirect(request.referrer or url_for('student_continue_forma', form_id=form_id))

        submitted_at=get_local_time()
        declaration_name = request.form.get('declaration_name')
        applicant_signature = request.form.get('applicant_signature')
        declaration_date_raw = datetime.now()
        was_in_corrections = is_student_correction_state(form)
        
        print(f"DEBUG submit_form_a: declaration_name={declaration_name}, applicant_signature={applicant_signature}")
        print(f"DEBUG submit_form_a: form_id={form_id}, form_exists={form_exists}")
        
        # Validate Section 7 is complete
        if not (declaration_name and applicant_signature and declaration_date_raw):
            flash("Please complete all Section 7 (Declaration) fields before submitting.", "error")
            return redirect(url_for('student_continue_forma', form_id=form_id))

        try:
            declaration_date = parse_html_date(declaration_date_raw)
        except ValueError:
            flash("Please provide a valid declaration date.", "error")
            return redirect(url_for('student_continue_forma', form_id=form_id))
        
        # First, call the same processing logic as student_continue_forma to save all sections
        # This ensures all data from the form is saved before marking as submitted
        
        # Helper function to convert checkbox values to boolean
        def to_bool(value):
            if isinstance(value, bool):
                return value
            if value is None:
                return False
            if isinstance(value, str):
                return value.lower() in ('yes', 'on', 'true', '1')
            return bool(value)
        
        try:
            # Process all the form data (reuse logic from student_continue_forma)
            # Convert checkboxes
            survey = to_bool(request.form.get('survey'))
            focus_groups = to_bool(request.form.get('focus_groups'))
            observations = to_bool(request.form.get('observations'))
            interviews = to_bool(request.form.get('interviews'))
            documents = to_bool(request.form.get('documents'))
            vulnerable_other_specify = to_bool(request.form.get('vulnerable_other_specify'))
            vulnerable_communities = to_bool(request.form.get('vulnerable_communities'))
            age_range = to_bool(request.form.get('age_range'))
            uj_employees = to_bool(request.form.get('uj_employees'))
            vulnerable = to_bool(request.form.get('vulnerable'))
            non_english = to_bool(request.form.get('non_english'))
            own_students = to_bool(request.form.get('own_students'))
            poverty = to_bool(request.form.get('poverty'))
            no_education = to_bool(request.form.get('no_education'))
            vulnerable_comments_1 = to_bool(request.form.get('vulnerable_comments_1'))
            disclosure = to_bool(request.form.get('disclosure'))
            discomfiture = to_bool(request.form.get('discomfiture'))
            deception = to_bool(request.form.get('deception'))
            sensitive = to_bool(request.form.get('sensitive'))
            prejudice = to_bool(request.form.get('prejudice'))
            intrusive_techniques = to_bool(request.form.get('intrusive_techniques'))
            illegal_activities = to_bool(request.form.get('illegal_activities'))
            personal = to_bool(request.form.get('personal'))
            available_records = to_bool(request.form.get('available_records'))
            inventories = to_bool(request.form.get('inventories'))
            risk_activities = to_bool(request.form.get('risk_activities'))
            activity_specify =request.form.get('activity_specify')
            vulnerable_comments_2 =request.form.get('vulnerable_comments_2')
            incentives = to_bool(request.form.get('incentives'))
            financial_costs = to_bool(request.form.get('financial_costs'))
            reward = to_bool(request.form.get('reward'))
            conflict = to_bool(request.form.get('conflict'))
            uj_premises = to_bool(request.form.get('uj_premises'))
            uj_facilities = to_bool(request.form.get('uj_facilities'))
            uj_funding = to_bool(request.form.get('uj_funding'))
            researcher_affiliation = to_bool(request.form.get('researcher_affiliation'))
            collective_involvement = to_bool(request.form.get('collective_involvement'))
            
            secondary_data = request.form.get('uses_secondary_data', request.form.get('secondary_data'))
            form.uses_secondary_data = _autosave_str_to_bool(secondary_data)
            if form.uses_secondary_data:
                form.secondary_data_type = request.form.get('data_type')
                if form.secondary_data_type == 'private':
                    form.private_permission = to_bool(request.form.get('privatePermission'))
            else:
                _clear_forma_secondary_data_details(form)

            # Assign all form fields
            form.user_id = user_id
            if form_requirements:
                form.attachment_id = form_requirements.id
            elif not form.attachment_id:
                form.attachment_id = 'AUTOSAVE_PENDING'
            form.applicant_name = request.form.get('applicant_name')
            form.student_number = request.form.get('student_number')
            form.institution = request.form.get('institution')
            form.department = request.form.get('department')
            form.degree = request.form.get('degree')
            form.study_title = request.form.get('study_title')
            form.mobile = request.form.get('mobile')
            form.email = user.email
            form.supervisor = supervisor.full_name if supervisor else ''
            form.supervisor_email = supervisor.email if supervisor else ''
            
            # Section 2 fields
            form.survey = survey
            form.observations = observations
            form.focus_groups = focus_groups
            form.interviews = interviews
            form.documents = documents
            form.vulnerable_other_specify = vulnerable_other_specify
            form.vulnerable_communities = vulnerable_communities
            form.age_range = age_range
            form.uj_employees = uj_employees
            form.vulnerable = vulnerable
            form.non_english = non_english
            form.own_students = own_students
            form.poverty = poverty
            form.no_education = no_education
            form.assessment_other_specify = request.form.get('assessment_other_specify')
            form.vulnerable_comments_1 = vulnerable_comments_1
            form.disclosure = disclosure
            form.discomfiture = discomfiture
            form.deception = deception
            form.sensitive = sensitive
            form.prejudice = prejudice
            form.intrusive_techniques = intrusive_techniques
            form.illegal_activities = illegal_activities
            form.personal = personal
            form.available_records = available_records
            form.inventories = inventories
            form.risk_activities = risk_activities
            form.activity_specify = activity_specify
            form.vulnerable_comments_2 = vulnerable_comments_2
            form.incentives = incentives
            form.financial_costs = financial_costs
            form.reward = reward
            form.conflict = conflict
            form.uj_premises = uj_premises
            form.uj_facilities = uj_facilities
            form.uj_funding = uj_funding
            form.vulnerable_comments_3 = request.form.get('vulnerable_comments_3')
            form.risk_rating = request.form.get('risk_rating')
            form.risk_justification = request.form.get('risk_justification')
            form.benefits_description = request.form.get('benefits_description')
            form.risk_mitigation = request.form.get('risk_mitigation')
            form.interviews_one = to_bool(request.form.get('interviews'))
            form.documents_one = to_bool(request.form.get('documents'))
            form.other_sec2 = request.form.get('other_sec2')
            
            # Section 3
            form.title_provision = request.form.get('title_provision')
            form.abstract = request.form.get('abstract')
            form.questions = request.form.get('questions')
            form.purpose_objectives = request.form.get('purpose_objectives')

            # Section 4
            form.grant_permission = request.form.get('grant_permission')
            if (form.grant_permission or '').strip().lower() == 'yes':
                form.org_name = ','.join(request.form.getlist('org_name[]'))
                form.org_contact = ','.join(request.form.getlist('org_contact[]'))
                form.org_role = ','.join(request.form.getlist('org_role[]'))
                form.org_permission = ','.join(request.form.getlist('org_permission[]'))
            else:
                form.org_name = ''
                form.org_contact = ''
                form.org_role = ''
                form.org_permission = ''
            form.researcher_affiliation = 'Yes' if researcher_affiliation else 'No'
            form.affiliation_details = request.form.get('affiliation_details') if researcher_affiliation else ''
            form.collective_involvement = 'Yes' if collective_involvement else 'No'
            form.collective_details = request.form.get('collective_details') if collective_involvement else ''
            form.is_funded = request.form.get('is_funded')
            form.fund_org = ','.join(request.form.getlist('fund_org[]'))
            form.fund_contact = ','.join(request.form.getlist('fund_contact[]'))
            form.fund_role = ','.join(request.form.getlist('fund_role[]'))
            form.fund_amount = ','.join(request.form.getlist('fund_amount[]'))
            form.indemnity_arrangements = request.form.get('indemnity_arrangements')
            form.other_committee = request.form.get('other_committee')

            # Section 5
            form.quantitative = "Yes" in request.form.getlist('quantitative[]') or to_bool(request.form.get('quantitative'))
            form.qualitative = "Yes" in request.form.getlist('qualitative[]') or to_bool(request.form.get('qualitative'))
            form.mixed_methods = "Yes" in request.form.getlist('mixed_methods[]') or to_bool(request.form.get('mixed_methods'))
            form.paradigm_explanation = request.form.get('paradigm_explanation')
            form.population = ','.join(request.form.getlist('population[]'))
            form.sampling_method = ','.join(request.form.getlist('sampling_method[]'))
            form.sampling_size = ','.join(cleaned_sample_sizes)
            form.inclusion_criteria = ','.join(request.form.getlist('inclusion_criteria[]'))
            form.translator = to_bool(request.form.get('translator'))
            form.translator_details = request.form.get('translator_details')
            form.intervention = to_bool(request.form.get('intervention'))
            form.intervention_details = request.form.get('intervention_details')
            form.personal_info = to_bool(request.form.get('personal_info'))
            form.personal_info_comment = request.form.get('personal_info_comment')
            form.conflict_interest = to_bool(request.form.get('conflict_interest'))
            form.conflict_explanation = request.form.get('conflict_explanation')
            
            # Method checkboxes
            form.questionnaire = 'questionnaire' in request.form.getlist('method[]')
            form.interview = 'interview' in request.form.getlist('method[]')
            form.focus = 'focus' in request.form.getlist('method[]')
            form.observation = 'observation' in request.form.getlist('method[]')
            form.documents_method = 'documents' in request.form.getlist('method[]')
            form.other_method = 'other' in request.form.getlist('method[]')
            form.questionnaire_explain = request.form.get('questionnaire_explain')
            form.interview_explain = request.form.get('interview_explain')
            form.focus_explain = request.form.get('focus_explain')
            form.observation_explain = request.form.get('observation_explain')
            form.documents_explain = request.form.get('documents_explain')
            form.other_explain = request.form.get('other_explain')
            form.data_public = request.form.get('data_public')
            form.public_data_link = request.form.get('public_data_link')
            
            # Section 6
            form.procedures = request.form.get('procedures')
            form.anonymity_confidentiality = request.form.get('anonymity_confidentiality')
            form.data_storage_security = request.form.get('data_storage_security')
            form.informed_consent = request.form.get('informed_consent')
            form.secure_location = ','.join(request.form.getlist('secure_location[]'))
            form.password_protected = ','.join(request.form.getlist('password_protected[]'))
            form.protected_place = ','.join(request.form.getlist('protected_place[]'))
            form.retention = ','.join(request.form.getlist('retention[]'))
            form.data_storage = ','.join(request.form.getlist('data_storage[]'))
            form.study_benefits = request.form.get('study_benefits')
            form.participant_risks = request.form.get('participant_risks')
            form.adverse_steps = request.form.get('adverse_steps')
            form.community_participation = request.form.get('community_participation')
            form.community_effects = request.form.get('community_effects')
            form.privacy = ','.join(request.form.getlist('privacy[]'))

            for key in ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's']:
                field_name = f'q6_9{key}'
                if hasattr(form, field_name):
                    setattr(form, field_name, to_bool(request.form.get(field_name)))

            form.voluntary_participation = request.form.get('voluntary_participation')
            form.results_feedback = request.form.get('results_feedback')
            form.products_access = request.form.get('products_access')
            form.publication_plans = request.form.get('publication_plans')
            form.participant_comp = request.form.get('participant_comp')
            form.participant_costs = request.form.get('participant_costs')
            form.ethics_reporting = request.form.get('ethics_reporting')

            # Section 7 - Mark as submitted
            form.submitted_at = submitted_at
            form.declaration_name = declaration_name
            form.applicant_signature = applicant_signature
            form.declaration_date = declaration_date
            form.submitted = True
            form.rejected_or_accepted = False
            form.status = 'Resubmitted' if was_in_corrections else 'Submitted'
            form.visible_to_student = False
            form.ethics_status = None
            form.form_supervisor_status = 'Resubmitted' if was_in_corrections else 'Submitted'
            reset_form_review_feedback(form)

            if not form.uses_secondary_data:
                _clear_forma_secondary_data_details(form)
            
            print(f"DEBUG: About to commit form - declaration_name={form.declaration_name}, submitted={form.submitted}")

            db_session.add(form)
            db_session.commit()
            
            print(f"DEBUG: Form committed successfully - form_id={form.form_id}")

            # Send notifications
            try:
                message = 'You have successfully edited and submitted your form. Please wait while its under review.'
                send_email(app, mail, message, [user.email])
                messages = f'{form.applicant_name} has submitted a form that needs to be reviewed.'
                send_email(app, mail, messages, [form.supervisor_email])
            except Exception as e:
                app.logger.error(f"Failed to send email to {form.supervisor_email}: {e}")
            
            flash("Form successfully submitted to supervisor!", "success")
            return redirect(url_for('student_dashboard'))
            
        except Exception as e:
            db_session.rollback()
            print(f"ERROR in submit_form_a: {str(e)}")
            print(f"ERROR traceback: {traceback.format_exc()}")
            app.logger.error(f"Error submitting form: {e}")
            flash(f"An error occurred while submitting the form: {str(e)}", "error")
            return redirect(url_for('student_continue_forma', form_id=form_id))


    data_org={
        "org_name" : parse_field(form.org_name),
        "org_contact" : parse_field(form.org_contact),
        "org_role": parse_field(form.org_role),
        "org_permission" : parse_field(form.org_permission),
    }
    data_fund={
        "fund_org" : parse_field(form.fund_org),
        "fund_contact" :parse_field(form.fund_contact),
        "fund_role": parse_field(form.fund_role),
        "fund_amount": parse_field(form.fund_amount),
    }
    data_sampling={
        "population" :parse_field(form.population),
        "sampling_method" : parse_field(form.sampling_method),
        "sampling_size": parse_field(form.sampling_size),
        "inclusion_criteria": parse_field(form.inclusion_criteria)
    } 

    return render_template("student_continue_forma.html",formA=form,data_org=data_org,data_fund=data_fund,data_sampling=data_sampling)



@app.route('/student_edit_formb', methods=['GET','POST'])
def student_edit_formb():
    user_id=session.get('id')

    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor=db_session.query(User).filter(User.user_id == user.supervisor_id).first()
    form = _find_latest_editable_form_for_user(
        FormB,
        user_id,
        'submitted_at',
        options=[
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file),
        ],
    )
    if form is not None and form.submitted_at is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_formb_resubmission_draft(user_id, form)
        if error_response:
            return error_response

    locked_response = redirect_if_student_form_locked(form, 'Form B')
    if locked_response:
        return locked_response
    
    if request.method=="POST":
        print("[DEBUG] FormB POST request received")
        data_public= request.form.get('data_public')=='Yes'
        personal_info=request.form.get('personal_info')== 'Yes'
        private_permission=request.form.get('private_permission')=='Yes'
        shortcomings_reported=request.form.get('shortcomings_reported')=="Yes"
        methodology_alignment=request.form.get('methodology_alignment')=="Yes"
        declaration_name = (request.form.get('declaration_name') or request.form.get('applicant_name')).strip()
        applicant_signature = (request.form.get('applicant_signature') or '').strip()
        print(f"[DEBUG] declaration_name: {declaration_name}, applicant_signature: {applicant_signature}")
        if not declaration_name or not applicant_signature:
            print("[DEBUG] Declaration fields missing, redirecting.")
            flash('Declaration section must be completed to submit the form', 'warning')
            return redirect(url_for('student_edit_formb'))
        if not form:
            form = FormB(user_id=user_id)
            inherit_previous_reviewers(form, FormB, user_id, FormB.submitted_at)
            db_session.add(form)
            print("[DEBUG] FormB instance created and added to session.")
        form.applicant_name=request.form.get('applicant_name')
        form.student_number=request.form.get('student_number')
        form.institution=request.form.get('institution')
        form.department=request.form.get('department')
        form.degree=request.form.get('degree')
        form.study_title=request.form.get('study_title')
        form.mobile=request.form.get('mobile')
        form.email=user.email
        form.supervisor=supervisor.full_name
        form.supervisor_email=supervisor.email
        form.project_description = request.form.get('project_description')
        form.data_nature = request.form.get('data_nature')
        form.data_origin = request.form.get('data_origin')
        form.data_public=data_public
        form.personal_info=personal_info
        _apply_formb_public_domain_answers(form, request.form)
        form.private_permission=private_permission
        form.data_anonymized = request.form.get('data_anonymized')
        form.anonymization_comment = request.form.get('anonymization_comment')
        form.shortcomings_reported=shortcomings_reported
        form.methodology_alignment=methodology_alignment
        form.permission_details = request.form.get('permission_details')
        form.limitations_reporting = request.form.get('limitations_reporting')
        form.original_clearance=request.form.get('original_clearance')
        form.participant_permission=request.form.get('participant_permission')
        form.data_safekeeping=request.form.get('data_safekeeping')
        form.risk_level=request.form.get('risk_level')
        form.risk_comments=request.form.get('risk_comments')
        form.declaration_name=declaration_name
        form.applicant_signature=applicant_signature
        form.full_name=(request.form.get('full_name') or request.form.get('applicant_name')).strip()
        form.submitted_at = get_local_time()
        form.declaration_date = get_local_time()
        form.submitted = True
        form.rejected_or_accepted = False
        was_in_corrections = is_student_correction_state(form)
        form.status = 'Resubmitted' if was_in_corrections else (form.status or 'Submitted')
        form.visible_to_student = False
        form.ethics_status = None
        form.form_supervisor_status = 'Resubmitted' if was_in_corrections else (form.form_supervisor_status or 'Submitted')
        reset_form_review_feedback(form)
        # Handle file upload
        file = request.files.get('private_permission_file')
        if file and file.filename:
            print(f"[DEBUG] File uploaded: {file.filename}")
            form.private_permission_file = file.read()
            form.private_permission_filename = file.filename
        import traceback
        try:
            db_session.commit()
            print("[DEBUG] FormB committed successfully.")
        except Exception as e:
            db_session.rollback()
            print(f"[DEBUG] FormB commit failed: {e}")
            traceback.print_exc()
            flash("An error occurred while submitting Form B. Please try again.", "danger")
            return redirect(url_for('student_edit_formb'))
        flash("Form B submitted successfully.", "success")
        return redirect(url_for('student_dashboard'))
    return render_template("student_edit_formb.html",formB=form)


@app.route('/student_continue_formb', methods=['GET','POST'])
def student_continue_formb():
    user_id=session.get('id')
    
    if has_blocking_student_form(FormA, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
        
    if has_blocking_student_form(FormC, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
    

    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor = get_student_supervisor_or_flash(user)
    if not supervisor:
        return redirect(url_for("student_dashboard"))
    
    form = db_session.query(FormB).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).filter(
        FormB.user_id == user_id,
        FormB.submitted_at.is_(None)
    ).order_by(FormB.created_at.desc().nullslast()).first()

    if form is None:
        form = db_session.query(FormB).options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        ).filter_by(user_id=user_id).order_by(FormB.submitted_at.desc().nullslast(), FormB.created_at.desc().nullslast()).first()

    if form is not None and form.submitted_at is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_formb_resubmission_draft(user_id, form)
        if error_response:
            return error_response

    # Ensure there is a draft instance before processing POST payload.
    if form is None:
        form = FormB(user_id=user_id)
        inherit_previous_reviewers(form, FormB, user_id, FormB.submitted_at)
        db_session.add(form)

    locked_response = redirect_if_student_form_locked(form, 'Form B')
    if locked_response:
        return locked_response

    if request.method=="POST":
        try:
            data_public= request.form.get('data_public')=='Yes'
            personal_info=request.form.get('personal_info')== 'Yes'
            private_permission=request.form.get('private_permission')=='Yes'
            shortcomings_reported=request.form.get('shortcomings_reported')=="Yes"
            methodology_alignment=request.form.get('methodology_alignment')=="Yes"
                        
            
            form.user_id=user_id
            form.applicant_name=request.form.get('applicant_name')
            form.student_number=request.form.get('student_number')
            form.institution=request.form.get('institution')
            form.department=request.form.get('department')
            form.degree=request.form.get('degree')
            form.study_title=request.form.get('study_title')
            form.mobile=request.form.get('mobile')
            form.email=user.email
            form.supervisor=supervisor.full_name
            form.supervisor_email=supervisor.email
            form.project_description = request.form.get('project_description')
            form.data_nature = request.form.get('data_nature')
            form.data_origin = request.form.get('data_origin')
            form.data_public=data_public
            form.personal_info=personal_info
            _apply_formb_public_domain_answers(form, request.form)
            form.data_acknowledgment = request.form.get('data_acknowledgment')
            form.private_permission=private_permission

            form.data_anonymized = request.form.get('data_anonymized')
            form.anonymization_comment = request.form.get('anonymization_comment')

            form.shortcomings_reported=shortcomings_reported
            form.methodology_alignment=methodology_alignment

            form.permission_details = request.form.get('permission_details')
            
            # Handle file upload
            file = request.files.get('private_permission_file')
            if file and file.filename:
                form.private_permission_file = file.read()
                form.private_permission_filename = file.filename 

            form.limitations_reporting = request.form.get('limitations_reporting')

                    
            form.original_clearance=request.form.get('original_clearance')
            form.participant_permission=request.form.get('participant_permission')
            form.data_safekeeping=request.form.get('data_safekeeping')
            form.risk_level=request.form.get('risk_level')
            form.risk_comments=request.form.get('risk_comments')
            
            db_session.add(form)
            db_session.commit()
            return redirect(url_for('student_continue_formb'))
        except Exception as e:
            db_session.rollback()
            flash(f"Error saving Form A: {str(e)}", "danger")
            print("Exception in FormA save:", e)
            import traceback; traceback.print_exc()
            return render_template("student_continue_formb.html",formB=form)
        
    return render_template("student_continue_formb.html",formB=form)


@app.route('/submit_form_b/<string:form_id>',methods=['GET','POST'])
def submit_form_b(form_id):
    """
    Submit Form B to supervisor - saves ALL sections including Section 5 and marks as submitted
    """
    user_id=session.get('id')
    
    if has_blocking_student_form(FormA, user_id):
        flash("You are not permitted to fill this form (FormA exists)", "warning")
        print(f"[ERROR] User {user_id} attempted to submit FormB but FormA exists.")
        return redirect(url_for("student_dashboard"))
    if has_blocking_student_form(FormC, user_id):
        flash("You are not permitted to fill this form (FormC exists)", "warning")
        print(f"[ERROR] User {user_id} attempted to submit FormB but FormC exists.")
        return redirect(url_for("student_dashboard"))
    

    if not user_id:
        print(f"[ERROR] Unauthorized access to submit_form_b. No user_id in session.")
        flash("Unauthorized access. Please log in again.", "danger")
        return jsonify({'error': 'Unauthorized'}), 401
    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor = get_student_supervisor_or_flash(user)
    if not supervisor:
        return redirect(url_for("student_dashboard"))
    
    # Check if form exists in database - if yes UPDATE (add declaration), if no CREATE with declaration
    form = db_session.query(FormB).filter(FormB.user_id==user_id, FormB.form_id==form_id).first()
    if form is not None and form.submitted_at is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_formb_resubmission_draft(user_id, form)
        if error_response:
            return error_response
        form_id = form.form_id
    if not form:
        form = (
            db_session.query(FormB)
            .filter(FormB.user_id == user_id, FormB.submitted_at.is_(None))
            .order_by(FormB.created_at.desc())
            .first()
        )
        if not form:
            print(f"[ERROR] FormB draft not found for user_id={user_id}, form_id={form_id}")
            flash("No draft Form B found to submit. Please open your draft and try again.", "warning")
            return redirect(url_for('student_continue_formb'))

    locked_response = redirect_if_student_form_locked(form, 'Form B')
    if locked_response:
        return locked_response
   
    
    if request.method=='POST':
        print(f"[DEBUG] POST request received for submit_form_b by user_id={user_id}, form_id={form_id}")
        declaration_name=request.form.get('declaration_name')
        full_name=request.form.get('full_name')
        declaration_date = get_local_time()
        submitted_at=get_local_time()
        print(f"[DEBUG] Declaration: name={declaration_name}, full_name={full_name}, date={declaration_date}, submitted_at={submitted_at}")
        # Validate Section 5 (Declaration) is complete
        if not (declaration_name and full_name):
            print(f"[ERROR] Declaration fields missing for user_id={user_id}, form_id={form_id}. declaration_name={declaration_name}, full_name={full_name}")
            flash("Please complete all Section 5 (Declaration) fields before submitting.", "error")
            return redirect(url_for('student_continue_formb', form_id=form_id))
        
        # Helper function to convert checkbox values to boolean
        def to_bool(value):
            if isinstance(value, bool):
                return value
            if value is None:
                return False
            if isinstance(value, str):
                return value.lower() in ('yes', 'on', 'true', '1')
            return bool(value)
        
        try:
            print("[DEBUG] Processing form submission...")
            # Process all form data (same as student_continue_formb)
            data_public = to_bool(request.form.get('data_public'))
            personal_info = to_bool(request.form.get('personal_info'))
            private_permission = to_bool(request.form.get('private_permission'))
            shortcomings_reported = to_bool(request.form.get('shortcomings_reported'))
            methodology_alignment = to_bool(request.form.get('methodology_alignment'))
            
            # Assign all form fields
            form.user_id = user_id
            form.applicant_name = request.form.get('applicant_name')
            form.student_number = request.form.get('student_number')
            form.institution = request.form.get('institution')
            form.department = request.form.get('department')
            form.degree = request.form.get('degree')
            form.study_title = request.form.get('study_title')
            form.mobile = request.form.get('mobile')
            form.email = user.email
            form.supervisor = supervisor.full_name if supervisor else ''
            form.supervisor_email = supervisor.email if supervisor else ''
            form.project_description = request.form.get('project_description')
            form.data_nature = request.form.get('data_nature')
            form.data_origin = request.form.get('data_origin')
            form.data_public = data_public
            form.personal_info = personal_info
            _apply_formb_public_domain_answers(form, request.form)
            form.data_acknowledgment = request.form.get('data_acknowledgment')
            form.private_permission = private_permission
            form.data_anonymized = request.form.get('data_anonymized')
            form.anonymization_comment = request.form.get('anonymization_comment')
            form.shortcomings_reported = shortcomings_reported
            form.methodology_alignment = methodology_alignment
            form.permission_details = request.form.get('permission_details')
            
            # Handle file upload
            file = request.files.get('private_permission_file')
            if file and file.filename:
                assign_private_permission_upload(form, file)
            
            form.limitations_reporting = request.form.get('limitations_reporting')
            form.original_clearance = request.form.get('original_clearance')
            form.participant_permission = request.form.get('participant_permission')
            form.data_safekeeping = request.form.get('data_safekeeping')
            form.risk_level = request.form.get('risk_level')
            form.risk_comments = request.form.get('risk_comments')
            
            # Section 5 - Mark as submitted
            form.declaration_name = (request.form.get('declaration_name') or request.form.get('applicant_name')).strip()
            form.full_name = (request.form.get('full_name') or request.form.get('applicant_name')).strip()
            form.declaration_date = declaration_date
            form.submitted_at = submitted_at
            form.submitted = True
            form.rejected_or_accepted = False
            was_in_corrections = is_student_correction_state(form)
            form.status = 'Resubmitted' if was_in_corrections else 'Submitted'
            form.ethics_status = None
            form.form_supervisor_status = 'Resubmitted' if was_in_corrections else 'Submitted'
            reset_form_review_feedback(form)
            print(f"[DEBUG] FormB updated: declaration_name={form.declaration_name}, full_name={form.full_name}, declaration_date={form.declaration_date}, submitted_at={form.submitted_at}")
            
            db_session.add(form)
            db_session.commit()
            print(f"[DEBUG] FormB committed successfully: form_id={form.form_id}")
            
            # Send notifications
            try:
                message = 'You have successfully edited and submitted your form. Please wait while its under review.'
                print(f"[DEBUG] Sending email to user: {user.email}")
                send_email(app, mail, message, [user.email])
                messages = f'{form.applicant_name} has submitted a form that needs to be reviewed.'
                print(f"[DEBUG] Sending email to supervisor: {form.supervisor_email}")
                send_email(app, mail, messages, [form.supervisor_email])
            except Exception as e:
                print(f"[DEBUG] Failed to send email to {form.supervisor_email}: {e}")
                app.logger.error(f"Failed to send email to {form.supervisor_email}: {e}")
            
            print("[DEBUG] Form successfully submitted to supervisor!")
            flash("Form successfully submitted to supervisor!", "success")
            print(f"[REDIRECT] Redirecting user_id={user_id} to student_dashboard after successful submission.")
            return redirect(url_for('student_dashboard'))
            
        except Exception as e:
            db_session.rollback()
            print(f"[ERROR] Exception during FormB submission for user_id={user_id}, form_id={form_id}: {e}")
            import traceback; traceback.print_exc()
            app.logger.error(f"Error submitting Form B: {e}")
            flash(f"An error occurred while submitting the form: {e}", "error")
            return redirect(url_for('submit_form_b', form_id=form_id))
    print(f"[DEBUG] Non-POST request to submit_form_b for user_id={user_id}, form_id={form_id}")
    return redirect(url_for('student_continue_formb'))



@app.route('/student_edit_formc', methods=['GET','POST'])
def student_edit_formc():
    user_id=session.get('id')
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401

    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor=db_session.query(User).filter(User.user_id == user.supervisor_id).first()
    form = _find_latest_editable_form_for_user(FormC, user_id, 'submission_date')
    if form is not None and form.submission_date is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_formc_resubmission_draft(user_id, form)
        if error_response:
            return error_response

    locked_response = redirect_if_student_form_locked(form, 'Form C')
    if locked_response:
        return locked_response
    if request.method=="POST":
        print("[DEBUG] FormC POST request received")
        declaration_name = (request.form.get('declaration_name') or request.form.get('applicant_name')).strip()
        full_name = (request.form.get('full_name') or request.form.get('applicant_name')).strip()
        submission_date = request.form.get('submission_date')
        print(f"[DEBUG] declaration_name: {declaration_name}, full_name: {full_name}, submission_date: {submission_date}")
        if not declaration_name or not full_name:
            print("[DEBUG] Declaration fields missing, redirecting.")
            flash('Declaration section must be completed to submit the form', 'warning')
            return redirect(url_for('student_edit_formc'))
        # Update existing form instance
        if not form:
            form = FormC(user_id=user_id)
            inherit_previous_reviewers(form, FormC, user_id, FormC.submission_date)
            db_session.add(form)
            print("[DEBUG] FormC instance created and added to session.")
        try:
            form.applicant_name=request.form.get('applicant_name')
            form.student_number=request.form.get('student_number')
            form.institution=request.form.get('institution')
            form.department=request.form.get('department')
            form.degree=request.form.get('degree')
            form.project_title=request.form.get('project_title')
            form.mobile_number=request.form.get('mobile_number')
            form.email_address=user.email
            form.supervisor_name=supervisor.full_name
            form.supervisor_email=supervisor.email
            form.vulnerable=True if request.form.get('vulnerable') else False
            form.age_under_18_or_over_65=True if request.form.get('age_under_18_or_over_65') else False
            form.uj_employees=True if request.form.get('uj_employees') else False
            form.non_vulnerable_context=True if request.form.get('non_vulnerable_context') else False
            form.non_english=True if request.form.get('non_english') else False
            form.own_students=True if request.form.get('own_students') else False
            form.poverty=True if request.form.get('poverty') else False
            form.no_education=True if request.form.get('no_education') else False
            form.vulnerable_other_description=True if request.form.get('vulnerable_other_description') else False
            form.vulnerable_comments=request.form.get('vulnerable_comments')
            form.consent_violation=True if request.form.get('consent_violation') else False
            form.discomfiture=True if request.form.get('discomfiture') else False
            form.deception=True if request.form.get('deception') else False
            form.sensitive_issues=True if request.form.get('sensitive_issues') else False
            form.prejudicial_info=True if request.form.get('prejudicial_info') else False
            form.intrusive=True if request.form.get('intrusive') else False
            form.illegal=True if request.form.get('illegal') else False
            form.direct_social_info=True if request.form.get('direct_social_info') else False
            form.identifiable_records=True if request.form.get('identifiable_records') else False
            form.psychology_tests=True if request.form.get('psychology_tests') else False
            form.researcher_risk=True if request.form.get('researcher_risk') else False
            form.activity_other_description=request.form.get('activity_other_description')
            form.activity_comments=request.form.get('activity_comments')
            form.incentives=True if request.form.get('incentives') else False
            form.participant_costs=True if request.form.get('participant_costs') else False
            form.researcher_interest=True if request.form.get('researcher_interest') else False
            form.conflict_of_interest=True if request.form.get('conflict_of_interest') else False
            form.uj_premises=True if request.form.get('uj_premises') else False
            form.uj_facilities=True if request.form.get('uj_facilities') else False
            form.uj_funding=True if request.form.get('uj_funding') else False
            form.consideration_comments=request.form.get('consideration_comments')
            form.risk_level=request.form.get('risk_level')
            form.risk_justification=request.form.get('risk_justification')
            form.risk_benefits=request.form.get('risk_benefits')
            form.risk_mitigation=request.form.get('risk_mitigation')
            form.summary_title=request.form.get('summary_title')
            form.executive_summary=request.form.get('executive_summary')
            form.research_questions=request.form.get('research_questions')
            form.research_purpose=request.form.get('research_purpose')
            form.secondary_data_info=request.form.get('secondary_data_info')
            form.exemption_reason=request.form.get('exemption_reason')
            form.declaration_name=declaration_name
            form.full_name=full_name
            form.submission_date=get_local_time()
            form.submitted = True
            form.rejected_or_accepted = False
            was_in_corrections = is_student_correction_state(form)
            form.status = 'Resubmitted' if was_in_corrections else (form.status or 'Submitted')
            form.visible_to_student = False
            form.ethics_status = None
            form.form_supervisor_status = 'Resubmitted' if was_in_corrections else (form.form_supervisor_status or 'Submitted')
            reset_form_review_feedback(form)
            db_session.commit()
            print("[DEBUG] FormC committed successfully.")
            flash("Form C submitted successfully.", "success")
            return redirect(url_for('student_dashboard'))
        except Exception as e:
            db_session.rollback()
            print(f"[DEBUG] FormC commit failed: {e}")
            app.logger.error(f"Error submitting Form C: {e}")
            print(f"[DEBUG] Outer exception: {e}")
            flash("An error occurred while submitting the form. Please try again.", "error")
            return redirect(url_for('student_dashboard'))
    return render_template('student_edit_formc.html',formc=form)


@app.route('/student_continue_formc', methods=['GET','POST'])
def student_continue_formc():
    user_id=session.get('id')
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401

    if has_blocking_student_form(FormA, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
        
    if has_blocking_student_form(FormB, user_id, options=[
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file),
    ]):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
    

    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor = get_student_supervisor_or_flash(user)

    if not user or not supervisor:
        return redirect(url_for("student_dashboard"))
    
    form = db_session.query(FormC).filter(
        FormC.user_id == user_id,
        FormC.submission_date.is_(None)
    ).order_by(FormC.created_at.desc().nullslast()).first()

    if form is None:
        form = db_session.query(FormC).filter_by(user_id=user_id).order_by(FormC.submission_date.desc().nullslast(), FormC.created_at.desc().nullslast()).first()

    if form is not None and form.submission_date is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_formc_resubmission_draft(user_id, form)
        if error_response:
            return error_response

    if form is None:
        form = FormC(user_id=user_id)
        inherit_previous_reviewers(form, FormC, user_id, FormC.submission_date)
        db_session.add(form)

    locked_response = redirect_if_student_form_locked(form, 'Form C')
    if locked_response:
        return locked_response
    
    if request.method=="POST":
        
        try:
            form.user_id=user_id
            form.applicant_name=request.form.get('applicant_name')
            form.student_number=request.form.get('student_number')
            form.institution=request.form.get('institution')
            form.department=request.form.get('department')
            form.degree=request.form.get('degree')
            form.project_title=request.form.get('project_title')
            form.mobile_number=request.form.get('mobile_number')
            form.email_address=user.email
            form.supervisor_name=supervisor.full_name
            form.supervisor_email=supervisor.email
            form.vulnerable=True if request.form.get('vulnerable') else False
                
            form.age_under_18_or_over_65=True if request.form.get('age_under_18_or_over_65') else False
            form.uj_employees=True if request.form.get('uj_employees') else False

            form.non_vulnerable_context=True if request.form.get('non_vulnerable_context') else False
            form.non_english=True if request.form.get('non_english')else False
            form.own_students=True if request.form.get('own_students') else False

            form.poverty=True if request.form.get('poverty') else False
            form.no_education=True if request.form.get('no_education') else False
            form.vulnerable_other_description=True if request.form.get('vulnerable_other_description') else False
            form.vulnerable_comments=request.form.get('vulnerable_comments')

            form.consent_violation=True if request.form.get('consent_violation') else False
            form.discomfiture=True if request.form.get('discomfiture') else False
            form.deception=True if request.form.get('deception') else False
            form.sensitive_issues=True if request.form.get('sensitive_issues') else False
            form.prejudicial_info=True if request.form.get('prejudicial_info') else False
            form.intrusive=True if request.form.get('intrusive') else False
            form.illegal=True if request.form.get('illegal') else False
            form.direct_social_info=True if request.form.get('direct_social_info') else False
            form.identifiable_records=True if request.form.get('identifiable_records') else False
            form.psychology_tests=True if request.form.get('psychology_tests') else False
            form.researcher_risk=True if request.form.get('researcher_risk') else False
            form.activity_other_description=request.form.get('activity_other_description')

            form.activity_comments=request.form.get('activity_comments')

            form.incentives=True if request.form.get('incentives') else False
            form.participant_costs=True if request.form.get('participant_costs') else False
            form.researcher_interest=True if request.form.get('researcher_interest') else False
            form.conflict_of_interest=True if request.form.get('conflict_of_interest') else False
            form.uj_premises=True if request.form.get('uj_premises') else False
            form.uj_facilities=True if request.form.get('uj_facilities') else False
            form.uj_funding=True if request.form.get('uj_funding') else False
            form.consideration_comments=request.form.get('consideration_comments')
                
            form.risk_level=request.form.get('risk_level')
            form.risk_justification=request.form.get('risk_justification')
            form.risk_benefits=request.form.get('risk_benefits')
            form.risk_mitigation=request.form.get('risk_mitigation')

            form.summary_title=request.form.get('summary_title')
            form.executive_summary=request.form.get('executive_summary')
            form.research_questions=request.form.get('research_questions')
            form.research_purpose=request.form.get('research_purpose')
            form.secondary_data_info=request.form.get('secondary_data_info')
            form.exemption_reason=request.form.get('exemption_reason')
            
            db_session.add(form)
            db_session.commit()
        except Exception as e:
            db_session.rollback()
            app.logger.error(f"Error submitting Form B: {e}")
            flash("An error occurred while submitting the form. Please try again.", "error")
            return redirect(url_for('student_dashboard'))
    return render_template('student_continue_formc.html',formc=form)




@app.route('/submit_form_c/<string:form_id>',methods=['GET','POST'])
def submit_form_c(form_id):
    """
    Submit Form C to supervisor - saves ALL sections including Section 4 and marks as submitted
    """
    user_id=session.get('id')
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401

    if has_blocking_student_form(FormA, user_id):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
        
    if has_blocking_student_form(FormB, user_id, options=[
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file),
    ]):
        flash("You are not permitted to fill this form", "warning")
        return redirect(url_for("student_dashboard"))
    

    user = db_session.query(User).filter(User.user_id == user_id).first()
    supervisor = get_student_supervisor_or_flash(user)

    if not user or not supervisor:
        return redirect(url_for("student_dashboard"))
    
    # Resolve the draft to submit. Never create a new row at submit time,
    # otherwise a stale/missing form_id can produce duplicate records.
    form = db_session.query(FormC).filter(FormC.user_id==user_id, FormC.form_id==form_id).first()
    if form is not None and form.submission_date is not None and is_student_correction_state(form):
        form, error_response = _get_or_create_formc_resubmission_draft(user_id, form)
        if error_response:
            return error_response
        form_id = form.form_id
    form_exists = form is not None
    
    if not form_exists:
        # Fallback to the latest unsubmitted draft for this user.
        form = (
            db_session.query(FormC)
            .filter(FormC.user_id == user_id, FormC.submission_date.is_(None))
            .order_by(FormC.created_at.desc())
            .first()
        )
        if not form:
            flash("No draft Form C found to submit. Please open your draft and try again.", "warning")
            return redirect(url_for('student_continue_formc'))

    locked_response = redirect_if_student_form_locked(form, 'Form C')
    if locked_response:
        return locked_response
    
    if request.method=="POST":
        declaration_name=request.form.get('declaration_name')
        full_name=request.form.get('full_name')
        submission_date=get_local_time()
        
        # Validate Section 4 (Declaration) is complete
        if not (declaration_name and full_name):
            flash("Please complete all Section 4 (Declaration) fields before submitting.", "error")
            return redirect(url_for('student_continue_formc', form_id=form_id))
        
        # Helper function to convert checkbox values to boolean
        def to_bool(value):
            if isinstance(value, bool):
                return value
            if value is None:
                return False
            if isinstance(value, str):
                return value.lower() in ('yes', 'on', 'true', '1')
            return bool(value)
        
        try:
            # Process all form data (same as student_continue_formc)
            form.user_id = user_id
            form.applicant_name = request.form.get('applicant_name')
            form.student_number = request.form.get('student_number')
            form.institution = request.form.get('institution')
            form.department = request.form.get('department')
            form.degree = request.form.get('degree')
            form.project_title = request.form.get('project_title')
            form.mobile_number = request.form.get('mobile_number')
            form.email_address = user.email
            form.supervisor_name = supervisor.full_name if supervisor else ''
            form.supervisor_email = supervisor.email if supervisor else ''
            form.vulnerable = to_bool(request.form.get('vulnerable'))
            form.age_under_18_or_over_65 = to_bool(request.form.get('age_under_18_or_over_65'))
            form.uj_employees = to_bool(request.form.get('uj_employees'))
            form.non_vulnerable_context = to_bool(request.form.get('non_vulnerable_context'))
            form.non_english = to_bool(request.form.get('non_english'))
            form.own_students = to_bool(request.form.get('own_students'))
            form.poverty = to_bool(request.form.get('poverty'))
            form.no_education = to_bool(request.form.get('no_education'))
            form.vulnerable_other_description = to_bool(request.form.get('vulnerable_other_description'))
            form.vulnerable_comments = request.form.get('vulnerable_comments')
            form.consent_violation = to_bool(request.form.get('consent_violation'))
            form.discomfiture = to_bool(request.form.get('discomfiture'))
            form.deception = to_bool(request.form.get('deception'))
            form.sensitive_issues = to_bool(request.form.get('sensitive_issues'))
            form.prejudicial_info = to_bool(request.form.get('prejudicial_info'))
            form.intrusive = to_bool(request.form.get('intrusive'))
            form.illegal = to_bool(request.form.get('illegal'))
            form.direct_social_info = to_bool(request.form.get('direct_social_info'))
            form.identifiable_records = to_bool(request.form.get('identifiable_records'))
            form.psychology_tests = to_bool(request.form.get('psychology_tests'))
            form.researcher_risk = to_bool(request.form.get('researcher_risk'))
            form.activity_other_description = request.form.get('activity_other_description')
            form.activity_comments = request.form.get('activity_comments')
            form.incentives = to_bool(request.form.get('incentives'))
            form.participant_costs = to_bool(request.form.get('participant_costs'))
            form.researcher_interest = to_bool(request.form.get('researcher_interest'))
            form.conflict_of_interest = to_bool(request.form.get('conflict_of_interest'))
            form.uj_premises = to_bool(request.form.get('uj_premises'))
            form.uj_facilities = to_bool(request.form.get('uj_facilities'))
            form.uj_funding = to_bool(request.form.get('uj_funding'))
            form.consideration_comments = request.form.get('consideration_comments')
            form.risk_level = request.form.get('risk_level')
            form.risk_justification = request.form.get('risk_justification')
            form.risk_benefits = request.form.get('risk_benefits')
            form.risk_mitigation = request.form.get('risk_mitigation')
            form.summary_title = request.form.get('summary_title')
            form.executive_summary = request.form.get('executive_summary')
            form.research_questions = request.form.get('research_questions')
            form.research_purpose = request.form.get('research_purpose')
            form.secondary_data_info = request.form.get('secondary_data_info')
            form.exemption_reason = request.form.get('exemption_reason')
            
            # Section 4 - Mark as submitted
            form.declaration_name = declaration_name
            form.full_name = full_name
            form.submission_date = submission_date
            form.submitted = True
            form.rejected_or_accepted = False
            was_in_corrections = is_student_correction_state(form)
            form.status = 'Resubmitted' if was_in_corrections else 'Submitted'
            form.visible_to_student = False
            form.ethics_status = None
            form.form_supervisor_status = 'Resubmitted' if was_in_corrections else 'Submitted'
            reset_form_review_feedback(form)
            
            db_session.add(form)
            db_session.commit()

            # Send notifications
            try:
                message = 'You have successfully edited and submitted your form. Please wait while its under review.'
                send_email(app, mail, message, [user.email])
                messages = f'{form.applicant_name} has submitted a form that needs to be reviewed.'
                send_email(app, mail, messages, [form.supervisor_email])
            except Exception as e:
                app.logger.error(f"Failed to send email to {form.supervisor_email}: {e}")
            
            flash("Form successfully submitted to supervisor!", "success")
            return redirect(url_for('student_dashboard'))
            
        except Exception as e:
            db_session.rollback()
            app.logger.error(f"Error submitting Form C: {e}")
            flash("An error occurred while submitting the form. Please try again.", "error")
            return redirect(url_for('student_continue_formc', form_id=form_id))



@app.route('/form_c_answers', methods=['GET','POST'])
def form_c_answers():
    user_id=session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    form = db_session.query(FormC).filter_by(user_id=user_id).first()
    return render_template("form_c_answers.html",formc=form)

@app.route('/form_d_answers', methods=['GET','POST'])
def form_d_answers():
    user_id=session.get('id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    form = db_session.query(FormD).filter_by(user_id=user_id).first()
    return render_template("form_d_answers.html",form=form)



@app.route('/api/form-c/<form_id>', methods=['GET'])
@login_required
def get_form_c(form_id):
    form_c = db_session.query(FormC).filter_by(form_id=form_id).first()
    if not form_c:
        return jsonify({"message": "Form not found"}), 404
    if not can_access_form(get_current_user(), form_c):
        return jsonify({"message": "Forbidden"}), 403
    return jsonify(form_c.to_dict()), 200


@app.route('/api/form-c/<form_id>/reassign-reviewers', methods=['POST'])
def api_reassign_form_c_reviewers(form_id):
    user_id = session.get('id')
    user_role = (session.get('role') or '').upper()
    if not user_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    if user_role not in ['ADMIN', 'SUPER_ADMIN']:
        return jsonify({'success': False, 'error': 'Forbidden'}), 403

    form = db_session.query(FormC).filter_by(form_id=form_id).first()
    if not form:
        return jsonify({'success': False, 'error': 'Form C not found'}), 404

    data = request.get_json(silent=True) or {}
    reviewer_ids = data.get('reviewer_ids') or data.get('reviewers') or []
    if not isinstance(reviewer_ids, list):
        return jsonify({'success': False, 'error': 'reviewer_ids must be a list'}), 400

    selected_ids = []
    for reviewer_id in reviewer_ids:
        reviewer_id = str(reviewer_id or '').strip()
        if reviewer_id and reviewer_id not in selected_ids:
            selected_ids.append(reviewer_id)

    if len(selected_ids) != 1:
        return jsonify({'success': False, 'error': 'Please provide exactly one reviewer'}), 400

    selected_reviewers = (
        db_session.query(User)
        .filter(
            User.user_id.in_(selected_ids),
            User.role == UserRole.REVIEWER
        )
        .all()
    )

    if len(selected_reviewers) != len(selected_ids):
        return jsonify({'success': False, 'error': 'One or more selected reviewers are invalid'}), 400

    form.reviewer_name1 = selected_ids[0]
    form.reviewer_name2 = None
    form.submitted_to_reviewers = True

    db_session.add(UserActivityLog(
        user_id=user_id,
        action='reassign_reviewers_api',
        page='API Form C Reassignment',
        target_user_id=form.user_id,
        details=f"Reassigned reviewers via API: {form.reviewer_name1}, {form.reviewer_name2} for FORM C {form.form_id}"
    ))
    db_session.commit()

    return jsonify({
        'success': True,
        'message': 'Form C reviewers reassigned successfully',
        'form_id': form.form_id,
        'reviewer_ids': selected_ids
    }), 200



@app.route('/chair_dashboard', methods=['GET','POST'])
def chair_dashboard():
    submitted_form_a = (db_session.query(FormA)
    .filter(FormA.submitted_at != None)
    .all())
    submitted_form_b = (db_session.query(FormB)
    .options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    )
    .filter(FormB.submitted_at != None)
    .all())
    submitted_form_c = (db_session.query(FormC)
    .filter(FormC.submission_date != None)
    .all())

    today = date.today()
    return render_template('chair-dashboard.html',today=today,submitted_form_a=submitted_form_a,submitted_form_b=submitted_form_b,submitted_form_c=submitted_form_c)


def parse_field(field):
        if field and isinstance(field, str):
            return field.strip("{}").split(",")
        return []


@app.route('/chair_forma_view/<string:id>', methods=['GET'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def chair_forma_view(id):
    form = (
        db_session.query(FormA)
        .filter(
            FormA.user_id == id,
            or_(
                FormA.submitted_at.isnot(None),
                FormA.created_at.isnot(None),
            )
        )
        .order_by(
            FormA.submitted_at.is_(None).asc(),
            FormA.submitted_at.desc().nullslast(),
            FormA.created_at.desc().nullslast(),
            FormA.form_id.desc(),
        )
        .all()
    )

    # The first row is always the latest submitted version.
    current_form = form[0] if form else None

    form_name="FORM A"
    data={}
    if current_form:
        data={
        "org_name" : parse_field(current_form.org_name),
        "org_contact" : parse_field(current_form.org_contact),
        "org_role": parse_field(current_form.org_role),
        "org_permission" : parse_field(current_form.org_permission),

        "fund_org" : parse_field(current_form.fund_org),
        "fund_contact" :parse_field(current_form.fund_contact),
        "fund_role": parse_field(current_form.fund_role),
        "fund_amount": parse_field(current_form.fund_amount),

        "population" :parse_field(current_form.population),
        "sampling_method" : parse_field(current_form.sampling_method),
        "sampling_size": parse_field(current_form.sampling_size),
        "inclusion_criteria": parse_field(current_form.inclusion_criteria)
    }
    
    today = date.today()
    user_role = session.get('role', '')
    return render_template(
        "chair-forms-dashboard.html",
        data=data,
        today=today,
        form_name=form_name,
        submitted_form=form,
        current_form_id=current_form.form_id if current_form else None,
        role=user_role
    )


@app.route('/chair_forma_reassign_reviewers/<string:form_id>', methods=['GET', 'POST'])
def chair_forma_reassign_reviewers(form_id):
    form = db_session.query(FormA).filter_by(form_id=form_id).first()
    if not form:
        flash("Form A record not found.", "danger")
        return redirect(url_for('chair_dashboard'))

    reviewers = (
        db_session.query(User)
        .filter(User.role == UserRole.REVIEWER)
        .order_by(User.full_name.asc())
        .all()
    )

    current_reviewer_ids = [rid for rid in [form.reviewer_name1, form.reviewer_name2] if rid]
    current_reviewers = []
    if current_reviewer_ids:
        current_reviewers = (
            db_session.query(User)
            .filter(User.user_id.in_(current_reviewer_ids))
            .all()
        )

    if request.method == 'POST':
        selected_ids = []
        for reviewer_id in request.form.getlist('reviewer_names[]'):
            reviewer_id = (reviewer_id or '').strip()
            if reviewer_id and reviewer_id not in selected_ids:
                selected_ids.append(reviewer_id)

        if len(selected_ids) != 1:
            flash("Please select exactly one reviewer.", "danger")
            return render_template(
                'chair-forma-reassign-reviewers.html',
                form=form,
                reviewers=reviewers,
                current_reviewers=current_reviewers,
                selected_reviewer_ids=selected_ids
            )

        selected_reviewers = (
            db_session.query(User)
            .filter(
                User.user_id.in_(selected_ids),
                User.role == UserRole.REVIEWER
            )
            .all()
        )

        if len(selected_reviewers) != len(selected_ids):
            flash("One or more selected reviewers are invalid.", "danger")
            return render_template(
                'chair-forma-reassign-reviewers.html',
                form=form,
                reviewers=reviewers,
                current_reviewers=current_reviewers,
                selected_reviewer_ids=selected_ids
            )

        form.reviewer_name1 = selected_ids[0]
        form.reviewer_name2 = None
        form.submitted_to_reviewers = True

        db_session.add(UserActivityLog(
            user_id=session.get('id'),
            action='reassign_reviewers',
            page='Chair Form A View',
            target_user_id=form.user_id,
            details=f"Reassigned reviewers: {form.reviewer_name1}, {form.reviewer_name2} for FORM A {form.form_id}"
        ))
        db_session.commit()

        flash("Reviewers reassigned successfully.", "success")
        return redirect(url_for('chair_forma_view', id=form.user_id))

    return render_template(
        'chair-forma-reassign-reviewers.html',
        form=form,
        form_name='FORM A',
        reviewers=reviewers,
        current_reviewers=current_reviewers,
        selected_reviewer_ids=current_reviewer_ids
    )


@app.route('/chair_formb_reassign_reviewers/<string:form_id>', methods=['GET', 'POST'])
def chair_formb_reassign_reviewers(form_id):
    form = db_session.query(FormB).filter_by(form_id=form_id).first()
    if not form:
        flash("Form B record not found.", "danger")
        return redirect(url_for('chair_dashboard'))

    reviewers = (
        db_session.query(User)
        .filter(User.role == UserRole.REVIEWER)
        .order_by(User.full_name.asc())
        .all()
    )

    current_reviewer_ids = [rid for rid in [form.reviewer_name1, form.reviewer_name2] if rid]
    current_reviewers = []
    if current_reviewer_ids:
        current_reviewers = (
            db_session.query(User)
            .filter(User.user_id.in_(current_reviewer_ids))
            .all()
        )

    if request.method == 'POST':
        selected_ids = []
        for reviewer_id in request.form.getlist('reviewer_names[]'):
            reviewer_id = (reviewer_id or '').strip()
            if reviewer_id and reviewer_id not in selected_ids:
                selected_ids.append(reviewer_id)

        if len(selected_ids) != 1:
            flash("Please select exactly one reviewer.", "danger")
            return render_template(
                'chair-forma-reassign-reviewers.html',
                form=form,
                form_name='FORM B',
                reviewers=reviewers,
                current_reviewers=current_reviewers,
                selected_reviewer_ids=selected_ids
            )

        selected_reviewers = (
            db_session.query(User)
            .filter(
                User.user_id.in_(selected_ids),
                User.role == UserRole.REVIEWER
            )
            .all()
        )

        if len(selected_reviewers) != len(selected_ids):
            flash("One or more selected reviewers are invalid.", "danger")
            return render_template(
                'chair-forma-reassign-reviewers.html',
                form=form,
                form_name='FORM B',
                reviewers=reviewers,
                current_reviewers=current_reviewers,
                selected_reviewer_ids=selected_ids
            )

        form.reviewer_name1 = selected_ids[0]
        form.reviewer_name2 = None
        form.submitted_to_reviewers = True

        db_session.add(UserActivityLog(
            user_id=session.get('id'),
            action='reassign_reviewers',
            page='Chair Form B View',
            target_user_id=form.user_id,
            details=f"Reassigned reviewers: {form.reviewer_name1}, {form.reviewer_name2} for FORM B {form.form_id}"
        ))
        db_session.commit()

        flash("Reviewers reassigned successfully.", "success")
        return redirect(url_for('chair_formb_view', id=form.user_id))

    return render_template(
        'chair-forma-reassign-reviewers.html',
        form=form,
        form_name='FORM B',
        reviewers=reviewers,
        current_reviewers=current_reviewers,
        selected_reviewer_ids=current_reviewer_ids
    )


@app.route('/chair_formc_reassign_reviewers/<string:form_id>', methods=['GET', 'POST'])
def chair_formc_reassign_reviewers(form_id):
    form = db_session.query(FormC).filter_by(form_id=form_id).first()
    if not form:
        flash("Form C record not found.", "danger")
        return redirect(url_for('chair_dashboard'))

    reviewers = (
        db_session.query(User)
        .filter(User.role == UserRole.REVIEWER)
        .order_by(User.full_name.asc())
        .all()
    )

    current_reviewer_ids = [rid for rid in [form.reviewer_name1, form.reviewer_name2] if rid]
    current_reviewers = []
    if current_reviewer_ids:
        current_reviewers = (
            db_session.query(User)
            .filter(User.user_id.in_(current_reviewer_ids))
            .all()
        )

    if request.method == 'POST':
        selected_ids = []
        for reviewer_id in request.form.getlist('reviewer_names[]'):
            reviewer_id = (reviewer_id or '').strip()
            if reviewer_id and reviewer_id not in selected_ids:
                selected_ids.append(reviewer_id)

        if len(selected_ids) != 1:
            flash("Please select exactly one reviewer.", "danger")
            return render_template(
                'chair-forma-reassign-reviewers.html',
                form=form,
                form_name='FORM C',
                reviewers=reviewers,
                current_reviewers=current_reviewers,
                selected_reviewer_ids=selected_ids
            )

        selected_reviewers = (
            db_session.query(User)
            .filter(
                User.user_id.in_(selected_ids),
                User.role == UserRole.REVIEWER
            )
            .all()
        )

        if len(selected_reviewers) != len(selected_ids):
            flash("One or more selected reviewers are invalid.", "danger")
            return render_template(
                'chair-forma-reassign-reviewers.html',
                form=form,
                form_name='FORM C',
                reviewers=reviewers,
                current_reviewers=current_reviewers,
                selected_reviewer_ids=selected_ids
            )

        form.reviewer_name1 = selected_ids[0]
        form.reviewer_name2 = None
        form.submitted_to_reviewers = True

        db_session.add(UserActivityLog(
            user_id=session.get('id'),
            action='reassign_reviewers',
            page='Chair Form C View',
            target_user_id=form.user_id,
            details=f"Reassigned reviewers: {form.reviewer_name1}, {form.reviewer_name2} for FORM C {form.form_id}"
        ))
        db_session.commit()

        flash("Reviewers reassigned successfully.", "success")
        return redirect(url_for('chair_formc_view', id=form.user_id))

    return render_template(
        'chair-forma-reassign-reviewers.html',
        form=form,
        form_name='FORM C',
        reviewers=reviewers,
        current_reviewers=current_reviewers,
        selected_reviewer_ids=current_reviewer_ids
    )


@app.route('/chair_formb_view/<string:id>', methods=['GET'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def chair_formb_view(id):
    form = (
        db_session.query(FormB)
        .options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        )
        .filter(
            FormB.user_id == id,
            or_(
                FormB.submitted_at.isnot(None),
                FormB.created_at.isnot(None),
            )
        )
        .order_by(
            FormB.submitted_at.is_(None).asc(),
            FormB.submitted_at.desc().nullslast(),
            FormB.created_at.desc().nullslast(),
            FormB.form_id.desc(),
        )
        .all()
    )

    # The first row is always the latest submitted version.
    current_form = form[0] if form else None

    form_name="FORM B"
    today = date.today()
    user_role = session.get('role', '')
    return render_template(
        "chair-forms-dashboard.html",
        today=today,
        form_name=form_name,
        submitted_form=form,
        current_form_id=current_form.form_id if current_form else None,
        role=user_role
    )

@app.route('/chair_formc_view/<string:id>', methods=['GET'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def chair_formc_view(id):
    form = (
        db_session.query(FormC)
        .filter(
            FormC.user_id == id,
            or_(
                FormC.submission_date.isnot(None),
                FormC.created_at.isnot(None),
            )
        )
        .order_by(
            FormC.submission_date.is_(None).asc(),
            FormC.submission_date.desc().nullslast(),
            FormC.created_at.desc().nullslast(),
            FormC.form_id.desc(),
        )
        .all()
    )

    # The first row is always the latest submitted version.
    current_form = form[0] if form else None

    form_name="FORM C"
    
    today = date.today()
    user_role = session.get('role', '')
    return render_template(
        "chair-forms-dashboard.html",
        today=today,
        form_name=form_name,
        submitted_form=form,
        current_form_id=current_form.form_id if current_form else None,
        role=user_role
    )

# Unified endpoint to view all forms for a user
@app.route('/chair_form_view_fixed/<string:user_id>', methods=['GET'])
def chair_form_view_fixed(user_id):
    # Query all forms for the user
    form_a_list = (
        db_session.query(FormA)
        .filter(FormA.user_id == user_id, FormA.submitted_at.isnot(None))
        .order_by(FormA.submitted_at.desc())
        .all()
    )
    form_b_list = (
        db_session.query(FormB)
        .filter(FormB.user_id == user_id, FormB.submitted_at.isnot(None))
        .order_by(FormB.submitted_at.desc())
        .all()
    )
    form_c_list = (
        db_session.query(FormC)
        .filter(FormC.user_id == user_id, FormC.submission_date.isnot(None))
        .order_by(FormC.submission_date.desc())
        .all()
    )

    user = db_session.query(User).filter_by(user_id=user_id).first()
    user_name = user.full_name if user else "Unknown"

    return render_template(
        "chair-forms-view-fixed.html",
        user_id=user_id,
        user_name=user_name,
        form_a_list=form_a_list,
        form_b_list=form_b_list,
        form_c_list=form_c_list
    )


@app.route("/send_certificate/<string:id>",methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def send_certificate(id):
    certificate_details = get_certificate_form(id)
    if not certificate_details:
        flash('Certificate record not found.', 'danger')
        return redirect(url_for('chair_landing'))
    if not certificate_details.certificate_issued:
        flash('Issue the certificate before sending it to the student.', 'danger')
        return redirect(url_for('modify_certificate', id=id))

    certificate_details.certificate_received = True
    certificate_details.certificate_modified = False
    db_session.commit()
    try:
        message = (
            'You have been issued with the Ethical Clearance Certificate. '
            f'Please follow the link {web_url} to view your certificate.'
        )
        recipient = getattr(certificate_details, 'email', None) or getattr(certificate_details, 'email_address', None)
        if recipient:
            send_email(app, mail, message, [recipient])
    except Exception as error:
        app.logger.error('Failed to send certificate email: %s', error)

    flash('Certificate sent to the student.', 'success')
    return redirect(url_for('modify_certificate', id=id))



@app.route("/faq",methods=['GET','POST'])
def faq():
    return render_template("faq.html")



@app.route('/student_view_feedback/<string:id>', methods=['GET'])
def student_view_feedback(id):
    form = None
    for model in [FormA, FormB, FormC]:
        form = db_session.query(model).filter_by(form_id=id).first()
        if form:
            break  # Stop once the form is found
    
    if form:
        form = merge_reviewer_feedback_from_related_draft(form)
        return render_template("student-view-feedback.html", view_form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('dashboard'))


@app.route('/supervisor_view_feedback/<string:id>', methods=['GET'])
@role_required('SUPERVISOR', 'REVIEWER', 'ADMIN', 'SUPER_ADMIN')
def supervisor_view_feedback(id):
    form = None
    for model in [FormA, FormB, FormC]:
        order_field = getattr(model, "submitted_at", getattr(model, "submission_date", None))
        if not order_field:
            continue
        form = (
            db_session.query(model)
            .filter_by(form_id=id)
            .order_by(order_field.desc())
            .first()
        )
        if form:
            break

    if form:
        current_user = get_current_user()
        current_role = str(role_value(current_user) or '').upper()
        form_owner = db_session.query(User).filter_by(
            user_id=getattr(form, 'user_id', None)
        ).first()
        if (
            current_role not in {'ADMIN', 'SUPER_ADMIN'}
            and (
                not form_owner
                or getattr(form_owner, 'supervisor_id', None) != current_user.user_id
            )
        ):
            abort(403)

    if form:
        form = merge_reviewer_feedback_from_related_draft(form)
        return render_template("supervisor-view-feedback.html", view_form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('supervisor_dashboard'))


@app.route('/ethics_view_feedback/<string:id>', methods=['GET'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def ethics_view_feedback(id):
    form = None
    for model in [FormA, FormB, FormC]:
        # pick whichever timestamp column the model has
        order_field = getattr(model, "submitted_at", getattr(model, "submission_date", None))
        if not order_field:
            continue  # skip if model doesn't have a timestamp field

        form = (
            db_session.query(model)
            .filter_by(form_id=id)
            .order_by(order_field.desc())
            .first()
        )
        if form:
            break

    if form and not can_access_form(get_current_user(), form):
        abort(403)

    if form:
        form = merge_reviewer_feedback_from_related_draft(form)
        return render_template("ethics-view-feedback.html", view_form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('chair_landing'))


###
###
### this is the function to focus on when intergrating MBA and Ethics
###
@app.route('/reviewer_list/', methods=['GET'])
def reviewer_list():
    user_id = session.get('id')

    if not user_id:
        flash("Your session has expired. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    user_profile = db_session.query(User).filter_by(user_id=user_id).first()

    if not user_profile:
        session.clear()
        flash("Your account could not be found. Please log in again.", "danger")
        return redirect(url_for('login_page'))

    filter_full_name = (request.args.get('filter_full_name') or '').strip()
    filter_email = (request.args.get('filter_email') or '').strip()
    filter_specialisation = (request.args.get('filter_specialisation') or '').strip()

    page = request.args.get('page', 1, type=int)
    per_page = 9

    reviewers_base_query = db_session.query(User).filter(User.role == "REVIEWER")

    # This is only for datalist suggestions, so it must not be paginated
    filter_users_list = reviewers_base_query.order_by(User.full_name.asc()).all()

    reviewers_query = reviewers_base_query

    if filter_full_name:
        reviewers_query = reviewers_query.filter(User.full_name.ilike(f"%{filter_full_name}%"))

    if filter_email:
        reviewers_query = reviewers_query.filter(User.email.ilike(f"%{filter_email}%"))

    if filter_specialisation:
        reviewers_query = reviewers_query.filter(User.specialisation.ilike(f"%{filter_specialisation}%"))

    reviewers_query = reviewers_query.order_by(User.full_name.asc())

    total_reviewers = reviewers_query.count()
    form = reviewers_query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = (total_reviewers + per_page - 1) // per_page

    role = user_profile.role.value

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )

    return render_template(
        "reviewer-list.html",
        role=role,
        user_profile=user_profile,
        view_form=form,
        filter_users_list=filter_users_list,
        page=page,
        total_pages=total_pages,
        total_reviewers=total_reviewers,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )


   


@app.route('/review_feedback/<string:form_id>', methods=['GET','POST'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def review_feedback(form_id):
    user_id=session.get('id')
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    form = None
    for model in [FormA, FormB, FormC]:
        form = db_session.query(model).filter_by(form_id=form_id).first()
        if form:
            break  # Stop once the form is found
    if not form:
        return "Form not found", 404
    if not can_access_as_assigned_reviewer(get_current_user(), form):
        abort(403)

    if form and is_submitted_form_record(form):
        form = merge_reviewer_feedback_from_related_draft(form)
        first_reviewer=''
        second_reveiwer=''
        if user_id == form.reviewer_name1 and user_id != form.reviewer_name2:
            first_reviewer=user_id
        elif user_id == form.reviewer_name2 and user_id != form.reviewer_name1:
            second_reveiwer=user_id
        return render_template("reviewer_feedback.html",view_form=form,first=first_reviewer,second=second_reveiwer)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('review_version'))
    


@app.route('/chair_form_view/<string:id>/<string:form_name>', methods=['GET','POST'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def chair_form_view(id,form_name):
    user_id=session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))

    user_name=db_session.query(User).filter_by(user_id=user_id).first()
    if not user_name:
        session.clear()
        return redirect(url_for('login_page'))

    user_role = getattr(getattr(user_name, 'role', None), 'value', '')
    forma = db_session.query(FormA).filter_by(form_id=id).first()
    formb = db_session.query(FormB).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).filter_by(form_id=id).first()
    formc = db_session.query(FormC).filter_by(form_id=id).first()

    requested_form = {
        'FORM A': forma,
        'A': forma,
        'FORM B': formb,
        'B': formb,
        'FORM C': formc,
        'C': formc,
    }.get((form_name or '').strip().upper())
    if not requested_form:
        abort(404)
    if not can_access_form(user_name, requested_form):
        abort(403)

    selected_form = requested_form
    if user_role == 'REVIEWER' and not can_access_as_assigned_reviewer(
        user_name, selected_form
    ):
        abort(403)
    if user_role == 'REVIEWER' and is_reviewed_by(selected_form, user_id):
        flash('You have already reviewed and submitted feedback for this application.', 'info')
        return redirect(url_for('review_dashboard'))

    if request.method == 'POST' and user_role == 'REVIEWER':
        selected_model = (
            FormA if forma is not None else FormB if formb is not None else FormC
        )
        selected_form = (
            db_session.query(selected_model)
            .filter_by(form_id=id)
            .with_for_update()
            .first()
        )
        if not selected_form:
            return "Form not found", 404
        if selected_model is FormA:
            forma = selected_form
        elif selected_model is FormB:
            formb = selected_form
        else:
            formc = selected_form
        if has_current_reviewer_submitted_feedback(selected_form, user_id):
            flash('You have already submitted feedback for this application. It was not submitted again.', 'warning')
            return redirect(url_for('review_dashboard'))

    if forma:
        formReviewers = db_session.query(User).filter(
        and_(
            User.role == "REVIEWER",
            User.email != forma.supervisor_email,
            User.authenticate_student.in_(['true', '1'])
        )
        ).all()
    elif formb:
        formReviewers = db_session.query(User).filter(
        and_(
            User.role == "REVIEWER",
            User.email != formb.supervisor_email,
            User.authenticate_student.in_(['true', '1'])
        )
            ).all()
    elif formc:
        formReviewers = db_session.query(User).filter(
        and_(
            User.role == "REVIEWER",
            User.email != formc.supervisor_email,
            User.authenticate_student.in_(['true', '1'])
        )
        ).all()
    
    list_of_revewers=[]
    id_of_reviewers=[]
    data={}
    if forma:
        data={
        "org_name" : parse_field(forma.org_name),
        "org_contact" : parse_field(forma.org_contact),
        "org_role": parse_field(forma.org_role),
        "org_permission" : parse_field(forma.org_permission),

        "fund_org" : parse_field(forma.fund_org),
        "fund_contact" :parse_field(forma.fund_contact),
        "fund_role": parse_field(forma.fund_role),
        "fund_amount": parse_field(forma.fund_amount),

        "population" :parse_field(forma.population),
        "sampling_method" : parse_field(forma.sampling_method),
        "sampling_size": parse_field(forma.sampling_size),
        "inclusion_criteria": parse_field(forma.inclusion_criteria)
    }

        
       
        latest_forma = forma
        if backfill_reviewer_from_previous_version(latest_forma, FormA, FormA.submitted_at):
            db_session.commit()
        Assigned_reviewer=db_session.query(User).filter(User.user_id.in_([latest_forma.reviewer_name1, latest_forma.reviewer_name2])).all()
        if Assigned_reviewer:
            for item in Assigned_reviewer:
                list_of_revewers.append(item.email)
                id_of_reviewers.append(item.user_id)
   
    #admin=db_session.query(User).filter_by(role="Admin").all()
    if form_name=="FORM A":
        
   
        if request.method=="POST":
           
            review_date=datetime.now()
            status=request.form.get('status')
            review_org_permission_status=request.form.get('org_permission_status')
            review_org_permission_comments=request.form.get('org_permission_comments')
            review_waiver_status=request.form.get('waiver_status')
            form_status=request.form.get('form_status')
            review_waiver_comments=request.form.get('waiver_comments')
            review_form_status=request.form.get('reject')
            review_form_comments=request.form.get('form_comments')
            review_questions_status=request.form.get('questions_status')
            review_questions_comments=request.form.get('questions_comments')
            review_consent_status=request.form.get('consent_status')
            review_consent_comments=request.form.get('consent_comments')
            review_proposal_status=request.form.get('proposal_status')
            review_proposal_comments=request.form.get('proposal_comments')
            review_additional_comments=request.form.get('additional_comments')
            review_recommendation=request.form.get('status')
            review_supervisor_signature=request.form.get('supervisor_signature')
            review_signature_date=datetime.now()
            form_review_comment=request.form.get('status')
            form_reviewed_by=user_id
 
            if request.form.get('status') in ['Approved','Approved with Minor Changes']:
                if str(user_id) == str(forma.reviewer_name1):
                    forma.review_date=review_date
                    forma.status=status
                    forma.review_org_permission_status=review_org_permission_status
                    forma.review_org_permission_comments=review_org_permission_comments
                    forma.review_waiver_status=review_waiver_status
                    forma.form_status=form_status
                    forma.review_waiver_comments=review_waiver_comments
                    forma.review_form_status=review_form_status
                    forma.review_form_comments=review_form_comments
                    forma.review_questions_status=review_questions_status
                    forma.review_questions_comments=review_questions_comments
                    forma.review_consent_status=review_consent_status
                    forma.review_consent_comments=review_consent_comments
                    forma.review_proposal_status=review_proposal_status
                    forma.review_proposal_comments=review_proposal_comments
                    forma.review_additional_comments=review_additional_comments
                    forma.review_recommendation=review_recommendation
                    forma.review_supervisor_signature=review_supervisor_signature
                    forma.review_signature_date=review_signature_date
                    forma.form_review_comment=form_review_comment
                    forma.form_reviewed_by=form_reviewed_by
                    forma.review_status=True
                    forma.rejected_or_accepted=True

                    #Uncomment the code bellow for testing
                    ##
                  
                    #add coments to Rec table
                    if user_role=='REVIEWER':
                       
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
                   
                else:
                    forma.review_date1=review_date
                    forma.status=status
                    forma.review_org_permission_status1=review_org_permission_status
                    forma.review_org_permission_comments1=review_org_permission_comments
                    forma.review_waiver_status1=review_waiver_status
                    forma.form_status1=form_status
                    forma.review_waiver_comments1=review_waiver_comments
                    forma.review_form_status1=review_form_status
                    forma.review_form_comments1=review_form_comments
                    forma.review_questions_status1=review_questions_status
                    forma.review_questions_comments1=review_questions_comments
                    forma.review_consent_status1=review_consent_status
                    forma.review_consent_comments1=review_consent_comments
                    forma.review_proposal_status1=review_proposal_status
                    forma.review_proposal_comments1=review_proposal_comments
                    forma.review_additional_comments1=review_additional_comments
                    forma.review_recommendation1=review_recommendation
                    forma.review_supervisor_signature1=review_supervisor_signature
                    forma.review_signature_date1=review_signature_date
                    forma.form_review_comment1=form_review_comment
                    forma.form_reviewed_by1=form_reviewed_by
                    forma.review_status1=True
                    forma.rejected_or_accepted=True
                  

                    #Uncomment the code bellow for testing
                    ##
                    
                    #add coments to Rec table
                    if user_role=='REVIEWER':
                     
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
                   
            else:
                if str(user_id) == str(forma.reviewer_name1):
                    forma.review_date=review_date
                    forma.status=status
                    forma.review_org_permission_status=review_org_permission_status
                    forma.review_org_permission_comments=review_org_permission_comments
                    forma.review_waiver_status=review_waiver_status
                    forma.form_status=form_status
                    forma.review_waiver_comments=review_waiver_comments
                    forma.review_form_status=review_form_status
                    forma.review_form_comments=review_form_comments
                    forma.review_questions_status=review_questions_status
                    forma.review_questions_comments=review_questions_comments
                    forma.review_consent_status=review_consent_status
                    forma.review_consent_comments=review_consent_comments
                    forma.review_proposal_status=review_proposal_status
                    forma.review_proposal_comments=review_proposal_comments
                    forma.review_additional_comments=review_additional_comments
                    forma.review_recommendation=review_recommendation
                    forma.form_review_comment=form_review_comment
                    forma.form_reviewed_by=form_reviewed_by
                    forma.review_status=False
                    forma.rejected_or_accepted=False

                    #Uncomment the code bellow for testing
                    ##
                    
                    try:
                        emails_of_student_and_supervisor=[]
                        emails_of_student_and_supervisor.append(forma.email)
                        emails_of_student_and_supervisor.append(forma.supervisor_email)
                        message=f' Form belonging to {forma.applicant_name} was sent back. Please view feedback.'
            
                        send_email(app,mail, message,emails_of_student_and_supervisor)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")
                        
                else:
                    forma.review_date1=review_date
                    forma.status=status
                    forma.review_org_permission_status1=review_org_permission_status
                    forma.review_org_permission_comments1=review_org_permission_comments
                    forma.review_waiver_status1=review_waiver_status
                    forma.form_status1=form_status
                    forma.review_waiver_comments1=review_waiver_comments
                    forma.review_form_status1=review_form_status
                    forma.review_form_comments1=review_form_comments
                    forma.review_questions_status1=review_questions_status
                    forma.review_questions_comments1=review_questions_comments
                    forma.review_consent_status1=review_consent_status
                    forma.review_consent_comments1=review_consent_comments
                    forma.review_proposal_status1=review_proposal_status
                    forma.review_proposal_comments1=review_proposal_comments
                    forma.review_additional_comments1=review_additional_comments
                    forma.review_recommendation1=review_recommendation
                    forma.review_supervisor_signature1=review_supervisor_signature
                    forma.review_signature_date1=review_signature_date
                    forma.form_review_comment1=form_review_comment
                    forma.form_reviewed_by1=form_reviewed_by
                    forma.review_status=False
                    forma.rejected_or_accepted=False

                    #Uncomment the code bellow for testing
                    ##
                    
                   
                    try:
                        emails_of_student_and_supervisor=[]
                        emails_of_student_and_supervisor.append(forma.email)
                        emails_of_student_and_supervisor.append(forma.supervisor_email)
                        message=f' Form belonging to {forma.applicant_name} was sent back. Please view feedback.'
            
                        send_email(app,mail, message,emails_of_student_and_supervisor)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")   
                

                #add coments to Rec table
                if user_role=='REVIEWER':
                       
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
            apply_reviewer_recommendation_routing(forma, review_recommendation)
            apply_admin_correction_routing(forma, review_recommendation, user_role)
            db_session.add(forma)
            db_session.commit()

            
            
            return redirect(url_for('review_dashboard'))
       
        return render_template("form_a_ethics.html",user_id=user_id,formA=forma,data=data,formReviewers=formReviewers,latest_forma=latest_forma)
    elif form_name=="FORM B":
        
        if formb:
            latest_formb = formb
            if backfill_reviewer_from_previous_version(latest_formb, FormB, FormB.submitted_at):
                db_session.commit()
            Assigned_reviewer=db_session.query(User).filter(User.user_id.in_([latest_formb.reviewer_name1, latest_formb.reviewer_name2])).all()
            if Assigned_reviewer:
                for item in Assigned_reviewer:
                    list_of_revewers.append(item.email)
                    id_of_reviewers.append(item.user_id)
        if request.method=="POST":
            review_date=datetime.now()
            status=request.form.get('status')
            review_org_permission_status=request.form.get('org_permission_status')
            review_org_permission_comments=request.form.get('org_permission_comments')
            review_waiver_status=request.form.get('waiver_status')
            form_status=request.form.get('form_status')
            review_waiver_comments=request.form.get('waiver_comments')
            review_form_status=request.form.get('reject')
            review_form_comments=request.form.get('form_comments')
            review_questions_status=request.form.get('questions_status')
            review_questions_comments=request.form.get('questions_comments')
            review_consent_status=request.form.get('consent_status')
            review_consent_comments=request.form.get('consent_comments')
            review_proposal_status=request.form.get('proposal_status')
            review_proposal_comments=request.form.get('proposal_comments')
            review_additional_comments=request.form.get('additional_comments')
            review_recommendation=request.form.get('status')
            review_supervisor_signature=request.form.get('supervisor_signature')
            review_signature_date=datetime.now()
            form_review_comment=request.form.get('status')
            form_reviewed_by=user_id
            if request.form.get('status') in ['Approved','Approved with Minor Changes']:
                if str(user_id) == str(formb.reviewer_name1):
                    formb.review_date=review_date
                    formb.status=status
                    formb.review_org_permission_status=review_org_permission_status
                    formb.review_org_permission_comments=review_org_permission_comments
                    formb.review_waiver_status=review_waiver_status
                    formb.form_status=form_status
                    formb.review_waiver_comments=review_waiver_comments
                    formb.review_form_status=review_form_status
                    formb.review_form_comments=review_form_comments
                    formb.review_questions_status=review_questions_status
                    formb.review_questions_comments=review_questions_comments
                    formb.review_consent_status=review_consent_status
                    formb.review_consent_comments=review_consent_comments
                    formb.review_proposal_status=review_proposal_status
                    formb.review_proposal_comments=review_proposal_comments
                    formb.review_additional_comments=review_additional_comments
                    formb.review_recommendation=review_recommendation
                    formb.review_supervisor_signature=review_supervisor_signature
                    formb.review_signature_date=review_signature_date
                    formb.form_review_comment=form_review_comment
                    formb.form_reviewed_by=form_reviewed_by
                    formb.review_status=True
                    formb.rejected_or_accepted=True
                    #Uncomment the code bellow for testing
                    ##
                        
                    #add coments to Rec table
                    if user_role=='REVIEWER':
                       
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
                else:
                    formb.review_date1=review_date
                    formb.status=status
                    formb.review_org_permission_status1=review_org_permission_status
                    formb.review_org_permission_comments1=review_org_permission_comments
                    formb.review_waiver_status1=review_waiver_status
                    formb.form_status1=form_status
                    formb.review_waiver_comments1=review_waiver_comments
                    formb.review_form_status1=review_form_status
                    formb.review_form_comments1=review_form_comments
                    formb.review_questions_status1=review_questions_status
                    formb.review_questions_comments1=review_questions_comments
                    formb.review_consent_status1=review_consent_status
                    formb.review_consent_comments1=review_consent_comments
                    formb.review_proposal_status1=review_proposal_status
                    formb.review_proposal_comments1=review_proposal_comments
                    formb.review_additional_comments1=review_additional_comments
                    formb.review_recommendation1=review_recommendation
                    formb.review_supervisor_signature1=review_supervisor_signature
                    formb.review_signature_date1=review_signature_date
                    formb.form_review_comment1=form_review_comment
                    formb.form_reviewed_by1=form_reviewed_by
                    formb.review_status1=True
                    formb.rejected_or_accepted=True

                    #Uncomment the code bellow for testing
                    ##
                    
                    #add coments to Rec table
                    if user_role=='REVIEWER':
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
                   
            else:
                if str(user_id) == str(formb.reviewer_name1):
                    formb.review_date=review_date
                    formb.status=status
                    formb.review_org_permission_status=review_org_permission_status
                    formb.review_org_permission_comments=review_org_permission_comments
                    formb.review_waiver_status=review_waiver_status
                    formb.form_status=form_status
                    formb.review_waiver_comments=review_waiver_comments
                    formb.review_form_status=review_form_status
                    formb.review_form_comments=review_form_comments
                    formb.review_questions_status=review_questions_status
                    formb.review_questions_comments=review_questions_comments
                    formb.review_consent_status=review_consent_status
                    formb.review_consent_comments=review_consent_comments
                    formb.review_proposal_status=review_proposal_status
                    formb.review_proposal_comments=review_proposal_comments
                    formb.review_additional_comments=review_additional_comments
                    formb.review_recommendation=review_recommendation
                    formb.form_review_comment=form_review_comment
                    formb.form_reviewed_by=form_reviewed_by
                    formb.review_status=False
                    formb.rejected_or_accepted=False

                    #Uncomment the code bellow for testing
                    ##
                    
                    try:
                        emails_of_student_and_supervisor=[]
                        emails_of_student_and_supervisor.append(formb.email)
                        emails_of_student_and_supervisor.append(formb.supervisor_email)
                        message=f' Form belonging to {formb.applicant_name} was sent back. Please view feedback.'
            
                        send_email(app,mail, message,emails_of_student_and_supervisor)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")
                else:
                    formb.review_date1=review_date
                    formb.status=status
                    formb.review_org_permission_status1=review_org_permission_status
                    formb.review_org_permission_comments1=review_org_permission_comments
                    formb.review_waiver_status1=review_waiver_status
                    formb.form_status1=form_status
                    formb.review_waiver_comments1=review_waiver_comments
                    formb.review_form_status1=review_form_status
                    formb.review_form_comments1=review_form_comments
                    formb.review_questions_status1=review_questions_status
                    formb.review_questions_comments1=review_questions_comments
                    formb.review_consent_status1=review_consent_status
                    formb.review_consent_comments1=review_consent_comments
                    formb.review_proposal_status1=review_proposal_status
                    formb.review_proposal_comments1=review_proposal_comments
                    formb.review_additional_comments1=review_additional_comments
                    formb.review_recommendation1=review_recommendation
                    formb.review_supervisor_signature1=review_supervisor_signature
                    formb.review_signature_date1=review_signature_date
                    formb.form_review_comment1=form_review_comment
                    formb.form_reviewed_by1=form_reviewed_by
                    formb.review_status=False
                    formb.rejected_or_accepted=False

                    #Uncomment the code bellow for testing
                    ##
                    
                    try:
                        emails_of_student_and_supervisor=[]
                        emails_of_student_and_supervisor.append(formb.email)
                        emails_of_student_and_supervisor.append(formb.supervisor_email)
                        message=f' Form belonging to {formb.applicant_name} was sent back. Please view feedback.'
            
                        send_email(app,mail, message,emails_of_student_and_supervisor)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")
                #add coments to Rec table
                if user_role=='REVIEWER':
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
            apply_reviewer_recommendation_routing(formb, review_recommendation)
            apply_admin_correction_routing(formb, review_recommendation, user_role)
            db_session.add(formb)
            db_session.commit()
            return redirect(url_for('review_dashboard'))
        return render_template("form_b_ethics.html",user_id=user_id,formB=formb,formReviewers=formReviewers,latest_formb=latest_formb)
    elif form_name=="FORM C":
        
        list_of_revewers=[]
        id_of_reviewers=[]
        if formc:
            latest_formc = formc
            if backfill_reviewer_from_previous_version(latest_formc, FormC, FormC.submission_date):
                db_session.commit()
            Assigned_reviewer=db_session.query(User).filter(User.user_id.in_([latest_formc.reviewer_name1, latest_formc.reviewer_name2])).all()
            
            if Assigned_reviewer:
                for item in Assigned_reviewer:
                    list_of_revewers.append(item.email)
                    id_of_reviewers.append(item.user_id)
        
        if request.method=="POST":
            
            review_date=datetime.now()
            status=request.form.get('status')
            review_org_permission_status=request.form.get('org_permission_status')
            review_org_permission_comments=request.form.get('org_permission_comments')
            review_waiver_status=request.form.get('waiver_status')
            form_status=request.form.get('form_status')
            review_waiver_comments=request.form.get('waiver_comments')
            review_form_status=request.form.get('reject')
            review_form_comments=request.form.get('form_comments')
            review_questions_status=request.form.get('questions_status')
            review_questions_comments=request.form.get('questions_comments')
            review_consent_status=request.form.get('consent_status')
            review_consent_comments=request.form.get('consent_comments')
            review_proposal_status=request.form.get('proposal_status')
            review_proposal_comments=request.form.get('proposal_comments')
            review_additional_comments=request.form.get('additional_comments')
            review_recommendation=request.form.get('status')
            review_supervisor_signature=request.form.get('supervisor_signature')
           
            review_signature_date=datetime.now()
            form_review_comment=request.form.get('status')
            form_reviewed_by=user_id
            if request.form.get('status') in ['Approved', 'Approved with Minor Changes']:
                
                if str(user_id) == str(formc.reviewer_name1):
                    formc.review_date=review_date
                    formc.status=status
                    formc.review_org_permission_status=review_org_permission_status
                    formc.review_org_permission_comments=review_org_permission_comments
                    formc.review_waiver_status=review_waiver_status
                    formc.form_status=form_status
                    formc.review_waiver_comments=review_waiver_comments
                    formc.review_form_status=review_form_status
                    formc.review_form_comments=review_form_comments
                    formc.review_questions_status=review_questions_status
                    formc.review_questions_comments=review_questions_comments
                    formc.review_consent_status=review_consent_status
                    formc.review_consent_comments=review_consent_comments
                    formc.review_proposal_status=review_proposal_status
                    formc.review_proposal_comments=review_proposal_comments
                    formc.review_additional_comments=review_additional_comments
                    formc.review_recommendation=review_recommendation
                    formc.review_supervisor_signature=review_supervisor_signature
                    formc.review_signature_date=review_signature_date
                    formc.form_review_comment=form_review_comment
                    formc.form_reviewed_by=form_reviewed_by
                    formc.review_status=True
                    formc.rejected_or_accepted=True

                    #Uncomment the code bellow for testing
                    
 
                    #add coments to Rec table
                    if user_role=='REVIEWER':
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
                else:
                    formc.review_date1=review_date
                    formc.status=status
                    formc.review_org_permission_status1=review_org_permission_status
                    formc.review_org_permission_comments1=review_org_permission_comments
                    formc.review_waiver_status1=review_waiver_status
                    formc.form_status1=form_status
                    formc.review_waiver_comments1=review_waiver_comments
                    formc.review_form_status1=review_form_status
                    formc.review_form_comments1=review_form_comments
                    formc.review_questions_status1=review_questions_status
                    formc.review_questions_comments1=review_questions_comments
                    formc.review_consent_status1=review_consent_status
                    formc.review_consent_comments1=review_consent_comments
                    formc.review_proposal_status1=review_proposal_status
                    formc.review_proposal_comments1=review_proposal_comments
                    formc.review_additional_comments1=review_additional_comments
                    formc.review_recommendation1=review_recommendation
                    formc.review_supervisor_signature1=review_supervisor_signature
                    formc.review_signature_date1=review_signature_date
                    formc.form_review_comment1=form_review_comment
                    formc.form_reviewed_by1=form_reviewed_by
                    formc.review_status1=True
                    formc.rejected_or_accepted=True

                    #Uncomment the code bellow for testing
                    ##
                    
                    #add coments to Rec table
                    if user_role=='REVIEWER':
                        form=Rec(
                        rec_id=user_id,
                        form_id=id,
                        full_name=user_name.full_name,
                        rec_comments = review_additional_comments,
                        rec_status = status,
                        rec_date=datetime.now()
                        )
                        db_session.add(form)
                   
            else:
                if str(user_id) == str(formc.reviewer_name1):
                    formc.review_date=review_date
                    formc.status=status
                    formc.review_org_permission_status=review_org_permission_status
                    formc.review_org_permission_comments=review_org_permission_comments
                    formc.review_waiver_status=review_waiver_status
                    formc.form_status=form_status
                    formc.review_waiver_comments=review_waiver_comments
                    formc.review_form_status=review_form_status
                    formc.review_form_comments=review_form_comments
                    formc.review_questions_status=review_questions_status
                    formc.review_questions_comments=review_questions_comments
                    formc.review_consent_status=review_consent_status
                    formc.review_consent_comments=review_consent_comments
                    formc.review_proposal_status=review_proposal_status
                    formc.review_proposal_comments=review_proposal_comments
                    formc.review_additional_comments=review_additional_comments
                    formc.review_recommendation=review_recommendation
                    formc.form_review_comment=form_review_comment
                    formc.form_reviewed_by=form_reviewed_by
                    formc.review_status=False
                    formc.rejected_or_accepted=False      
                    


                    try:
                        emails_of_student_and_supervisor=[]
                        emails_of_student_and_supervisor.append(formc.email_address)
                        emails_of_student_and_supervisor.append(formc.supervisor_email)
                        message=f' Form belonging to {formc.applicant_name} was sent back. Please view feedback.'
            
                        send_email(app,mail, message,emails_of_student_and_supervisor)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")
                else:
                    formc.review_date1=review_date
                    formc.status=status
                    formc.review_org_permission_status1=review_org_permission_status
                    formc.review_org_permission_comments1=review_org_permission_comments
                    formc.review_waiver_status1=review_waiver_status
                    formc.form_status1=form_status
                    formc.review_waiver_comments1=review_waiver_comments
                    formc.review_form_status1=review_form_status
                    formc.review_form_comments1=review_form_comments
                    formc.review_questions_status1=review_questions_status
                    formc.review_questions_comments1=review_questions_comments
                    formc.review_consent_status1=review_consent_status
                    formc.review_consent_comments1=review_consent_comments
                    formc.review_proposal_status1=review_proposal_status
                    formc.review_proposal_comments1=review_proposal_comments
                    formc.review_additional_comments1=review_additional_comments
                    formc.review_recommendation1=review_recommendation
                    formc.form_review_comment1=form_review_comment
                    formc.form_reviewed_by1=form_reviewed_by
                    formc.review_status1=False
                    formc.rejected_or_accepted=False
                    #Uncomment the code bellow for testing
                    ##
                    


                    try:
                        emails_of_student_and_supervisor=[]
                        emails_of_student_and_supervisor.append(formc.email_address)
                        emails_of_student_and_supervisor.append(formc.supervisor_email)
                        message=f' Form belonging to {formc.applicant_name} was sent back. Please view feedback.'
            
                        send_email(app,mail, message,emails_of_student_and_supervisor)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")

                #add coments to Rec table
                if user_role=='REVIEWER':
                        form=Rec(
                            rec_id=user_id,
                            form_id=id,
                            full_name=user_name.full_name,
                            rec_comments = review_additional_comments,
                            rec_status = status,
                            rec_date=datetime.now()
                            )
               
                        db_session.add(form)
            apply_reviewer_recommendation_routing(formc, review_recommendation)
            apply_admin_correction_routing(formc, review_recommendation, user_role)
            db_session.add(formc)
            db_session.commit()
            return redirect(url_for('review_dashboard'))
    
        return render_template("form_c_ethics.html",user_id=user_id,formc=formc,formReviewers=formReviewers,latest_formc=latest_formc)



@app.route('/ethics_reviewer_committee_form', methods=['GET'])
def ethics_reviewer_committee_form():
    return redirect(url_for('ethics_reviewer_committee_form_a', **request.args.to_dict()))


@app.route('/ethics_reviewer_committee_form_a', methods=['GET','POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def ethics_reviewer_committee_form_a():
    user_id = session.get('id')
    user_profile = db_session.query(User).filter_by(user_id=user_id).first() if user_id else None
    role = user_profile.role.value if user_profile and user_profile.role else ''
    # Get year and month from query parameters
    year_param = request.args.get('year')
    month_param = request.args.get('month')
    filter_applied = request.args.get('filter_applied', 'false') == 'true'

    latest_forma_subq = (
        db_session.query(
            FormA.user_id,
            func.max(FormA.submitted_at).label('latest_submitted_at')
        )
        .filter(FormA.submitted_at.isnot(None))
    )

    if not filter_applied:
        latest_forma_subq = latest_forma_subq.filter(
            or_(
                FormA.submitted_to_admin == True,
                FormA.submitted_to_reviewers == True,
                FormA.submitted_to_rec == True
            )
        )

    latest_forma_subq = latest_forma_subq.group_by(FormA.user_id).subquery()

    query = db_session.query(FormA).join(
        latest_forma_subq,
        and_(
            FormA.user_id == latest_forma_subq.c.user_id,
            FormA.submitted_at == latest_forma_subq.c.latest_submitted_at
        )
    )
    if role == 'REC':
        query = query.filter(FormA.submitted_to_rec == True)

    if year_param and month_param:
        query = query.filter(
            extract('year', FormA.submitted_at) == int(year_param),
            func.to_char(FormA.submitted_at, 'YYYY-MM') == month_param
        )

    page = request.args.get('page', default=1, type=int)
    page_size = request.args.get('page_size', default=20, type=int)
    total_records = run_db_query_with_retry(lambda: query.count())
    total_pages = max(1, (total_records + page_size - 1) // page_size) if total_records else 1
    page = min(max(page, 1), total_pages)
    supervisor_forms = run_db_query_with_retry(
        lambda: query.order_by(FormA.submitted_at.desc()).offset((page-1)*page_size).limit(page_size).all()
    )
    form_requirements_lookup = {}
    form_user_ids = [form.user_id for form in supervisor_forms]
    if form_user_ids:
        requirement_rows = run_db_query_with_retry(
            lambda: db_session.query(
                FormARequirements.id,
                FormARequirements.user_id,
                FormARequirements.proposal_filename,
                FormARequirements.impact_assessment_filename,
                FormARequirements.permission_letter_filename,
                FormARequirements.prior_clearance_path_filename,
                FormARequirements.research_tools_filename,
                FormARequirements.prior_clearance_filename,
                FormARequirements.prior_clearance1_filename,
                FormARequirements.participation_info_filename,
                FormARequirements.pending_note_filename,
            )
            .filter(FormARequirements.user_id.in_(form_user_ids))
            .all()
        )

        for row in requirement_rows:
            proxy = type('RequirementProxy', (), {})()
            proxy.id = row.id
            proxy.user_id = row.user_id
            proxy.proposal_path = bool(row.proposal_filename)
            proxy.impact_assessment_path = bool(row.impact_assessment_filename)
            proxy.permission_letter = bool(row.permission_letter_filename)
            proxy.prior_clearance_path = bool(row.prior_clearance_path_filename)
            proxy.research_tools_path = bool(row.research_tools_filename)
            proxy.prior_clearance = bool(row.prior_clearance_filename)
            proxy.prior_clearance1 = bool(row.prior_clearance1_filename)
            proxy.participation_info_sheet = bool(row.participation_info_filename)
            proxy.pending_note = bool(row.pending_note_filename)
            form_requirements_lookup[row.user_id] = proxy
    supervisor_formA = [
        (form, form_requirements_lookup.get(form.user_id))
        for form in supervisor_forms
    ]
    today = date.today()
    return render_template(
        'ethics_reviewer_committee.html',
        role=role,
        today=today,
        submitted_form_a=supervisor_formA,
        filter_applied=filter_applied,
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages
    )


@app.route('/ethics_reviewer_committee_form_b', methods=['GET','POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def ethics_reviewer_committee_form_b():
    user_id = session.get('id')
    user_profile = db_session.query(User).filter_by(user_id=user_id).first() if user_id else None
    role = user_profile.role.value if user_profile and user_profile.role else ''
    # Get year and month from query parameters
    year_param = request.args.get('year')
    month_param = request.args.get('month')
    filter_applied = request.args.get('filter_applied', 'false') == 'true'
    
    # Show only the latest Form B per user (like Form A)
    base_query = db_session.query(FormB, FormARequirements)
    base_query = base_query.options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).join(User, FormB.user_id == User.user_id)
    base_query = base_query.outerjoin(FormARequirements, FormARequirements.user_id == FormB.user_id)

    # Subquery to get latest submitted_at per user
    latest_formB_subq = db_session.query(
        FormB.user_id,
        func.max(FormB.submitted_at).label('latest_submitted_at')
    )
    if filter_applied:
        latest_formB_subq = latest_formB_subq.filter(FormB.submitted_at.isnot(None))
    else:
        latest_formB_subq = latest_formB_subq.filter(
            FormB.submitted_at.isnot(None),
            or_(
                FormB.submitted_to_admin == True,
                FormB.submitted_to_reviewers == True,
                FormB.submitted_to_rec == True
            )
        )
    latest_formB_subq = latest_formB_subq.group_by(FormB.user_id).subquery()

    # Join to only get latest per user
    query = base_query.join(
        latest_formB_subq,
        (FormB.user_id == latest_formB_subq.c.user_id) & (FormB.submitted_at == latest_formB_subq.c.latest_submitted_at)
    )
    if role == 'REC':
        query = query.filter(FormB.submitted_to_rec == True)
    if year_param and month_param:
        query = query.filter(
            extract('year', FormB.submitted_at) == int(year_param),
            func.to_char(FormB.submitted_at, 'YYYY-MM') == month_param
        )
    page = request.args.get('page', default=1, type=int)
    page_size = request.args.get('page_size', default=20, type=int)
    total_records = run_db_query_with_retry(lambda: query.count())
    total_pages = max(1, (total_records + page_size - 1) // page_size) if total_records else 1
    page = min(max(page, 1), total_pages)
    supervisor_formB = run_db_query_with_retry(
        lambda: query.order_by(FormB.submitted_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    )
    today = date.today()
    return render_template(
        'ethics_reviewer_committee.html',
        role=role,
        today=today,
        submitted_form_b=supervisor_formB,
        filter_applied=filter_applied,
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages
    )


@app.route('/ethics_reviewer_committee_form_c', methods=['GET','POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def ethics_reviewer_committee_form_c():
    user_id = session.get('id')
    user_profile = db_session.query(User).filter_by(user_id=user_id).first() if user_id else None
    role = user_profile.role.value if user_profile and user_profile.role else ''
    # Get year and month from query parameters
    year_param = request.args.get('year')
    month_param = request.args.get('month')
    filter_applied = request.args.get('filter_applied', 'false') == 'true'
    
    # Show only the latest Form C per user (like Form A)
    base_query = db_session.query(FormC, FormARequirements)
    base_query = base_query.join(User, FormC.user_id == User.user_id)
    base_query = base_query.outerjoin(FormARequirements, FormARequirements.user_id == FormC.user_id)

    # Subquery to get latest submission per user
    latest_formC_subq = db_session.query(
        FormC.user_id,
        func.max(FormC.submission_date).label('latest_submission_date')
    )
    if filter_applied:
        latest_formC_subq = latest_formC_subq.filter(FormC.submission_date.isnot(None))
    else:
        latest_formC_subq = latest_formC_subq.filter(
            FormC.submission_date.isnot(None),
            or_(
                FormC.submitted_to_admin == True,
                FormC.submitted_to_reviewers == True,
                FormC.submitted_to_rec == True
            )
        )
    latest_formC_subq = latest_formC_subq.group_by(FormC.user_id).subquery()

    # Join to only get latest per user
    query = base_query.join(
        latest_formC_subq,
        (FormC.user_id == latest_formC_subq.c.user_id) & (FormC.submission_date == latest_formC_subq.c.latest_submission_date)
    )
    if role == 'REC':
        query = query.filter(FormC.submitted_to_rec == True)
    if year_param and month_param:
        query = query.filter(
            extract('year', FormC.submission_date) == int(year_param),
            func.to_char(FormC.submission_date, 'YYYY-MM') == month_param
        )
    page = request.args.get('page', default=1, type=int)
    page_size = request.args.get('page_size', default=20, type=int)
    total_records = run_db_query_with_retry(lambda: query.count())
    total_pages = max(1, (total_records + page_size - 1) // page_size) if total_records else 1
    page = min(max(page, 1), total_pages)
    supervisor_formC = run_db_query_with_retry(
        lambda: query.order_by(FormC.submission_date.desc()).offset((page - 1) * page_size).limit(page_size).all()
    )
    today = date.today()
    return render_template(
        'ethics_reviewer_committee.html',
        role=role,
        today=today,
        submitted_form_c=supervisor_formC,
        filter_applied=filter_applied,
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages
    )


@app.route('/student_form_pdf/<string:form_id>/<string:form_type>', methods=['GET','POST'])
@login_required
def student_form_pdf(form_id,form_type):
    form = None
    for model in [FormA, FormB, FormC]:
        form = db_session.query(model).filter_by(form_id=form_id).first()
        if form:
            break  # Stop once the form is found
    if not form:
        abort(404)
    if not can_access_form(get_current_user(), form):
        abort(403)
    data={}
    if form_type=='A':
        data={
        "org_name" : parse_field(form.org_name),
        "org_contact" : parse_field(form.org_contact),
        "org_role": parse_field(form.org_role),
        "org_permission" : parse_field(form.org_permission),

        "fund_org" : parse_field(form.fund_org),
        "fund_contact" :parse_field(form.fund_contact),
        "fund_role": parse_field(form.fund_role),
        "fund_amount": parse_field(form.fund_amount),

        "population" :parse_field(form.population),
        "sampling_method" : parse_field(form.sampling_method),
        "sampling_size": parse_field(form.sampling_size),
        "inclusion_criteria": parse_field(form.inclusion_criteria)
    }
        return render_template('student_form_a_answer_pdf.html',formA=form,data=data)
    elif form_type=='B':
        return render_template('student_form_b_answer_pdf.html',formB=form)
    elif form_type=='C':
        return render_template('student_form_c_answer_pdf.html',formc=form)
    



@app.route('/ethics_form_pdf/<string:form_id>/<string:form_type>', methods=['GET','POST'])
@login_required
def ethics_form_pdf(form_id,form_type):
    form_type = (form_type or '').strip().upper()

    if form_type == "FORM A":
        form = db_session.query(FormA).filter_by(form_id=form_id).first()
        if not form:
            return "Form A not found.", 404
        if not can_access_form(get_current_user(), form):
            abort(403)

        data = {
            "org_name": parse_field(form.org_name),
            "org_contact": parse_field(form.org_contact),
            "org_role": parse_field(form.org_role),
            "org_permission": parse_field(form.org_permission),
            "fund_org": parse_field(form.fund_org),
            "fund_contact": parse_field(form.fund_contact),
            "fund_role": parse_field(form.fund_role),
            "fund_amount": parse_field(form.fund_amount),
            "population": parse_field(form.population),
            "sampling_method": parse_field(form.sampling_method),
            "sampling_size": parse_field(form.sampling_size),
            "inclusion_criteria": parse_field(form.inclusion_criteria)
        }
        
        return render_template('student_form_a_answer_pdf.html',formA=form,data=data)

    if form_type == "FORM B":
        form = db_session.query(FormB).filter_by(form_id=form_id).first()
        if not form:
            return "Form B not found.", 404
        if not can_access_form(get_current_user(), form):
            abort(403)
        return render_template('student_form_b_answer_pdf.html',formB=form)

    if form_type == "FORM C":
        form = db_session.query(FormC).filter_by(form_id=form_id).first()
        if not form:
            return "Form C not found.", 404
        if not can_access_form(get_current_user(), form):
            abort(403)
        return render_template('student_form_c_answer_pdf.html',formc=form)

    return "Invalid form type.", 400


@app.route('/chair_landing', methods=['POST', 'GET'])
@role_required('ADMIN', 'SUPER_ADMIN')
def chair_landing():
    user_id = session.get('id')
    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user:
        return redirect(url_for('login_page'))

    # Get year from query parameters
    year_param = request.args.get('year')
    year_filter = None
    try:
        if year_param:
            year_filter = int(year_param)
    except (ValueError, TypeError):
        year_filter = None

    def paginate_folder_rows(folder_map, page, per_page=20):
        folder_rows = []
        for year, month_map in folder_map.items():
            for month, forms in month_map.items():
                if not forms:
                    continue
                first_form, count, status = forms[0]
                folder_rows.append({
                    'year': year,
                    'month': month,
                    'month_sort': datetime.strptime(month, "%Y-%m"),
                    'form': first_form,
                    'count': count,
                    'status': status,
                })

        folder_rows.sort(key=lambda item: item['month_sort'], reverse=True)
        total_records = len(folder_rows)
        total_pages = max(1, math.ceil(total_records / per_page)) if total_records else 1
        current_page = min(max(page, 1), total_pages)
        start_index = (current_page - 1) * per_page
        end_index = start_index + per_page
        return folder_rows[start_index:end_index], total_records, total_pages, current_page

    # ---------- FORM A ----------
    latest_forma_subq = (
        db_session.query(
            FormA.user_id.label('user_id'),
            func.max(func.coalesce(FormA.submitted_at, FormA.created_at)).label('latest_date')
        )
        .group_by(FormA.user_id)
        .subquery()
    )
    formA_query = db_session.query(FormA, func.count(FormA.form_id).over(partition_by=FormA.user_id).label("total_forms"))\
        .join(
            latest_forma_subq,
            and_(
                FormA.user_id == latest_forma_subq.c.user_id,
                func.coalesce(FormA.submitted_at, FormA.created_at) == latest_forma_subq.c.latest_date
            )
        ).filter(
            # Treat workflow location as the visibility gate for the admin landing page.
            or_(
                FormA.submitted_to_admin == True,
                FormA.submitted_to_reviewers == True,
                FormA.submitted_to_rec == True
            )
        )
    if year_filter:
        formA_query = formA_query.filter(func.extract('year', func.coalesce(FormA.submitted_at, FormA.created_at)) == year_filter)
    formAs = formA_query.order_by(FormA.user_id, func.coalesce(FormA.submitted_at, FormA.created_at).desc()).all()

    forms_by_yearA = defaultdict(lambda: defaultdict(list))
    for form, count in formAs:
        timestamp = form.submitted_at if form.submitted_at else form.created_at
        year = timestamp.year
        month = timestamp.strftime("%Y-%m")
        status = 'reviewers' if getattr(form, 'submitted_to_reviewers', False) else 'admin'
        forms_by_yearA[year][month].append((form, count, status))
    page_a = request.args.get('page_a', default=1, type=int)
    paginated_rows_a, total_records_a, total_pages_a, page_a = paginate_folder_rows(forms_by_yearA, page_a)

    # ---------- FORM B ----------
    latest_formb_subq = (
        db_session.query(
            FormB.user_id.label('user_id'),
            func.max(func.coalesce(FormB.submitted_at, FormB.created_at)).label('latest_date')
        )
        .group_by(FormB.user_id)
        .subquery()
    )
    formB_query = db_session.query(FormB, func.count(FormB.form_id).over(partition_by=FormB.user_id).label("total_forms")).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).join(
        latest_formb_subq,
        and_(
            FormB.user_id == latest_formb_subq.c.user_id,
            func.coalesce(FormB.submitted_at, FormB.created_at) == latest_formb_subq.c.latest_date
        )
    ).filter(
        or_(
            FormB.submitted_to_admin == True,
            FormB.submitted_to_reviewers == True,
            FormB.submitted_to_rec == True
        )
    )
    if year_filter:
        formB_query = formB_query.filter(func.extract('year', func.coalesce(FormB.submitted_at, FormB.created_at)) == year_filter)
    formBs = formB_query.order_by(FormB.user_id, func.coalesce(FormB.submitted_at, FormB.created_at).desc()).all()

    forms_by_yearB = defaultdict(lambda: defaultdict(list))
    for form, count in formBs:
        timestamp = form.submitted_at if form.submitted_at else form.created_at
        year = timestamp.year
        month = timestamp.strftime("%Y-%m")
        status = 'reviewers' if getattr(form, 'submitted_to_reviewers', False) else 'admin'
        forms_by_yearB[year][month].append((form, count, status))
    page_b = request.args.get('page_b', default=1, type=int)
    paginated_rows_b, total_records_b, total_pages_b, page_b = paginate_folder_rows(forms_by_yearB, page_b)

    # ---------- FORM C ----------
    latest_formc_subq = (
        db_session.query(
            FormC.user_id.label('user_id'),
            func.max(func.coalesce(FormC.submission_date, FormC.created_at)).label('latest_date')
        )
        .group_by(FormC.user_id)
        .subquery()
    )
    formC_query = db_session.query(FormC, func.count(FormC.form_id).over(partition_by=FormC.user_id).label("total_forms"))\
        .join(
            latest_formc_subq,
            and_(
                FormC.user_id == latest_formc_subq.c.user_id,
                func.coalesce(FormC.submission_date, FormC.created_at) == latest_formc_subq.c.latest_date
            )
        ).filter(
            or_(
                FormC.submitted_to_admin == True,
                FormC.submitted_to_reviewers == True,
                FormC.submitted_to_rec == True
            )
        )
    if year_filter:
        formC_query = formC_query.filter(func.extract('year', func.coalesce(FormC.submission_date, FormC.created_at)) == year_filter)
    formCs = formC_query.order_by(FormC.user_id, func.coalesce(FormC.submission_date, FormC.created_at).desc()).all()

    forms_by_yearC = defaultdict(lambda: defaultdict(list))
    for form, count in formCs:
        timestamp = form.submission_date if form.submission_date else form.created_at
        year = timestamp.year
        month = timestamp.strftime("%Y-%m")
        status = 'reviewers' if getattr(form, 'submitted_to_reviewers', False) else 'admin'
        forms_by_yearC[year][month].append((form, count, status))
    page_c = request.args.get('page_c', default=1, type=int)
    paginated_rows_c, total_records_c, total_pages_c, page_c = paginate_folder_rows(forms_by_yearC, page_c)

    role = user.role.value if user and user.role else None

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )
    return render_template(
        "chair-landing-dashboard.html",
        role=role,
        folder_rows_a=paginated_rows_a,
        folder_rows_b=paginated_rows_b,
        folder_rows_c=paginated_rows_c,
        page_a=page_a,
        page_b=page_b,
        page_c=page_c,
        total_pages_a=total_pages_a,
        total_pages_b=total_pages_b,
        total_pages_c=total_pages_c,
        total_records_a=total_records_a,
        total_records_b=total_records_b,
        total_records_c=total_records_c,
        current_year=datetime.now().year,
        year_filter=year_filter,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,

    )


@app.route('/admin/login_logs', methods=['GET'])
def admin_login_logs():
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))

    current_user = db_session.query(User).filter(User.user_id == user_id).first()
    if not current_user or not current_user.role or current_user.role.value not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You are not authorized to view login logs.', 'danger')
        return redirect(url_for('login_page'))

    user_search = (request.args.get('search') or request.args.get('email') or '').strip()
    start_date_raw = (request.args.get('start_date') or '').strip()
    end_date_raw = (request.args.get('end_date') or '').strip()

    # --- Pagination parameters ---
    iu_page = int(request.args.get('iu_page', 1))
    iu_per_page = int(request.args.get('iu_per_page', 20))
    log_page = int(request.args.get('log_page', 1))
    log_per_page = int(request.args.get('log_per_page', 20))
    iu_inactive_only = request.args.get('iu_inactive_only', '0') == '1'
    iu_years = request.args.get('iu_years', '')
    try:
        iu_years = int(iu_years)
    except (ValueError, TypeError):
        iu_years = None

    # --- Inactive Users Table Data ---
    all_users = db_session.query(User).order_by(User.full_name.asc()).all()
    user_name_suggestions = sorted({
        user.full_name for user in all_users if user.full_name
    }, key=lambda name: name.casefold())

    def matches_user_search(user):
        if not user_search:
            return True
        search_value = user_search.casefold()
        return any(
            search_value in str(value).casefold()
            for value in [
                getattr(user, 'full_name', None),
                getattr(user, 'email', None),
                getattr(user, 'student_number', None),
                getattr(user, 'staff_number', None),
                getattr(user, 'user_id', None),
            ]
            if value
        )

    subq = (
        db_session.query(
            UserActivityLog.user_id,
            func.max(UserActivityLog.timestamp).label('last_login')
        )
        .filter(UserActivityLog.action == 'login')
        .group_by(UserActivityLog.user_id)
        .subquery()
    )
    user_last_logins = dict(
        db_session.query(subq.c.user_id, subq.c.last_login)
    )
    now = datetime.now(timezone.utc)
    inactive_users = []
    for user in all_users:
        if not matches_user_search(user):
            continue
        last_login = user_last_logins.get(user.user_id)
        if last_login:
            delta = now - last_login
            days = delta.days
            if days < 1:
                inactivity = f"<1 day"
            elif days < 7:
                inactivity = f"{days} day{'s' if days > 1 else ''}"
            elif days < 365:
                inactivity = f"{days // 7} week{'s' if days // 7 > 1 else ''}"
            else:
                inactivity = f"{days // 365} year{'s' if days // 365 > 1 else ''}"
        else:
            last_login = None
            inactivity = "Never logged in"
        # Determine if user is active based on authenticate_student
        is_active = (str(user.authenticate_student).lower() == 'true')
        # Filtering logic
        show = True
        if iu_inactive_only:
            # Only show users who have never logged in or last login > 30 days ago
            if last_login:
                show = (now - last_login).days > 30
            else:
                show = True
        if iu_years is not None:
            # Only show users who have not logged in for at least iu_years
            if last_login:
                show = show and ((now - last_login).days >= iu_years * 365)
            else:
                show = show and True
        if show:
            inactive_users.append({
                'user': user,
                'last_login': last_login,
                'inactivity': inactivity,
                'is_active': is_active
            })
    inactive_users.sort(key=lambda x: (x['last_login'] is not None, x['last_login'] or datetime(1900,1,1)), reverse=False)
    iu_total = len(inactive_users)
    iu_pages = max(1, (iu_total + iu_per_page - 1) // iu_per_page)
    iu_start = (iu_page - 1) * iu_per_page
    iu_end = iu_start + iu_per_page
    inactive_users_paginated = inactive_users[iu_start:iu_end]

    # --- Existing login log table ---
    query = (
        db_session.query(UserActivityLog, User)
        .join(User, UserActivityLog.user_id == User.user_id)
        .filter(UserActivityLog.action == 'login')
    )
    if user_search:
        search_term = f"%{user_search}%"
        query = query.filter(
            or_(
                User.full_name.ilike(search_term),
                User.email.ilike(search_term),
                cast(User.student_number, String).ilike(search_term),
                cast(User.staff_number, String).ilike(search_term),
                User.user_id.ilike(search_term),
            )
        )
    if start_date_raw:
        try:
            start_dt = parse_admin_log_date(start_date_raw)
            query = query.filter(UserActivityLog.timestamp >= start_dt)
        except ValueError:
            flash('Invalid start date format. Please use dd/mm/yyyy or YYYY-MM-DD.', 'warning')
    if end_date_raw:
        try:
            end_dt = parse_admin_log_date(end_date_raw)
            end_dt = datetime.combine(end_dt.date(), dt_time.max)
            query = query.filter(UserActivityLog.timestamp <= end_dt)
        except ValueError:
            flash('Invalid end date format. Please use dd/mm/yyyy or YYYY-MM-DD.', 'warning')
    log_total = query.count()
    log_pages = max(1, (log_total + log_per_page - 1) // log_per_page)
    log_rows = (
        query
        .order_by(UserActivityLog.timestamp.desc())
        .offset((log_page - 1) * log_per_page)
        .limit(log_per_page)
        .all()
    )
    # --- User Activity Logs Table ---

    from sqlalchemy.orm import aliased
    TargetUser = aliased(User)
    activity_page = int(request.args.get('activity_page', 1))
    activity_per_page = int(request.args.get('activity_per_page', 20))
    activity_query = (
        db_session.query(UserActivityLog, User, TargetUser)
        .join(User, UserActivityLog.user_id == User.user_id)
        .outerjoin(TargetUser, UserActivityLog.target_user_id == TargetUser.user_id)
        .order_by(UserActivityLog.timestamp.desc())
    )
    if user_search:
        search_term = f"%{user_search}%"
        activity_query = activity_query.filter(
            or_(
                User.full_name.ilike(search_term),
                User.email.ilike(search_term),
                cast(User.student_number, String).ilike(search_term),
                cast(User.staff_number, String).ilike(search_term),
                User.user_id.ilike(search_term),
                TargetUser.full_name.ilike(search_term),
                TargetUser.email.ilike(search_term),
                cast(TargetUser.student_number, String).ilike(search_term),
                cast(TargetUser.staff_number, String).ilike(search_term),
                TargetUser.user_id.ilike(search_term),
            )
        )
    activity_total = activity_query.count()
    activity_pages = max(1, (activity_total + activity_per_page - 1) // activity_per_page)
    activity_rows = (
        activity_query
        .offset((activity_page - 1) * activity_per_page)
        .limit(activity_per_page)
        .all()
    )

    role = current_user.role.value if current_user and current_user.role else None


    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )
    return render_template(
        'admin_login_logs.html',
        role=role,
        log_rows=log_rows,
        log_page=log_page,
        log_pages=log_pages,
        log_per_page=log_per_page,
        log_total=log_total,
        inactive_users=inactive_users_paginated,
        iu_page=iu_page,
        iu_pages=iu_pages,
        iu_per_page=iu_per_page,
        iu_total=iu_total,
        iu_inactive_only=iu_inactive_only,
        iu_years=request.args.get('iu_years', ''),
        activity_rows=activity_rows,
        activity_page=activity_page,
        activity_pages=activity_pages,
        activity_per_page=activity_per_page,
        activity_total=activity_total,
        user_name_suggestions=user_name_suggestions,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,

        filters={
            'search': user_search,
            'email': user_search,
            'start_date': start_date_raw,
            'end_date': end_date_raw,
        }
    )


@app.route('/api/admin/login_logs', methods=['GET'])
def api_admin_login_logs():
    user_id = session.get('id')
    if not user_id:
        return jsonify({'message': 'Unauthorized'}), 401

    current_user = db_session.query(User).filter(User.user_id == user_id).first()
    if not current_user or not current_user.role or current_user.role.value not in ['ADMIN', 'SUPER_ADMIN']:
        return jsonify({'message': 'Forbidden'}), 403

    email_filter = (request.args.get('email') or '').strip()
    start_date_raw = (request.args.get('start_date') or '').strip()
    end_date_raw = (request.args.get('end_date') or '').strip()
    limit_raw = (request.args.get('limit') or '200').strip()

    try:
        limit = int(limit_raw)
    except ValueError:
        limit = 200
    limit = max(1, min(limit, 1000))

    query = (
        db_session.query(UserActivityLog, User)
        .join(User, UserActivityLog.user_id == User.user_id)
        .filter(UserActivityLog.action == 'login')
    )

    if email_filter:
        query = query.filter(User.email.ilike(f"%{email_filter}%"))

    if start_date_raw:
        try:
            start_dt = datetime.fromisoformat(start_date_raw)
            query = query.filter(UserActivityLog.timestamp >= start_dt)
        except ValueError:
            return jsonify({'message': 'Invalid start_date. Use YYYY-MM-DD.'}), 400

    if end_date_raw:
        try:
            end_dt = datetime.fromisoformat(end_date_raw)
            end_dt = datetime.combine(end_dt.date(), dt_time.max)
            query = query.filter(UserActivityLog.timestamp <= end_dt)
        except ValueError:
            return jsonify({'message': 'Invalid end_date. Use YYYY-MM-DD.'}), 400

    rows = (
        query
        .order_by(UserActivityLog.timestamp.desc())
        .limit(limit)
        .all()
    )

    data = [
        {
            'activity_id': log.activity_id,
            'user_id': user.user_id,
            'full_name': user.full_name,
            'email': user.email,
            'role': user.role.value if user.role else None,
            'login_at': log.timestamp.isoformat() if log.timestamp else None,
        }
        for log, user in rows
    ]

    return jsonify({'count': len(data), 'data': data}), 200


@app.route('/review_version/<string:user_id>/<string:form_name>', methods=['GET','POST'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def review_version(user_id,form_name):
    normalized_form_name = (form_name or '').strip().upper()
    model_map = {
        'A': FormA,
        'FORM A': FormA,
        'B': FormB,
        'FORM B': FormB,
        'C': FormC,
        'FORM C': FormC,
    }
    model = model_map.get(normalized_form_name)
    if not model:
        return redirect(url_for('review_dashboard'))

    query = (
        db_session.query(model, FormARequirements)
        .outerjoin(FormARequirements, FormARequirements.user_id == model.user_id)
        .filter(model.user_id == user_id)
    )
    current_user = get_current_user()
    current_role = str(role_value(current_user) or '').upper()
    if current_role == 'REVIEWER':
        query = query.filter(or_(
            model.reviewer_name1 == current_user.user_id,
            model.reviewer_name2 == current_user.user_id,
            model.user_id.in_(
                db_session.query(User.user_id).filter(
                    User.supervisor_id == current_user.user_id
                )
            ),
        ))
    elif current_role == 'REC':
        query = query.filter(model.submitted_to_rec == True)

    if model == FormB:
        query = query.options(
            defer(FormB.permission_letter),
            defer(FormB.prior_clearance),
            defer(FormB.ethics_evidence),
            defer(FormB.proposal_path),
            defer(FormB.pending_note),
            defer(FormB.private_permission_file)
        )

    submission_field = get_form_submission_field(model)
    if submission_field is not None:
        query = query.filter(submission_field.isnot(None)).order_by(
            submission_field.desc().nullslast(),
            model.created_at.desc().nullslast()
        )

    form = query.all()
    if current_role == 'REVIEWER' and not form:
        abort(403)
    if current_role == 'REVIEWER' and is_reviewed_by(form[0][0], current_user.user_id):
        flash('You have already reviewed and submitted feedback for this application.', 'info')
        return redirect(url_for('review_dashboard'))

   
    user_id=session.get('id')

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )
    return render_template('review_version.html',form=form,form_name=form_name,user_id=user_id,      # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,)


@app.route('/review_dashboard', methods=['GET','POST'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN')
def review_dashboard():
    user_id=session.get('id')

    latest_form_a_subq = (
        db_session.query(
            FormA.user_id.label('user_id'),
            func.max(FormA.submitted_at).label('latest_submitted_at')
        )
        .filter(FormA.submitted_at.isnot(None))
        .group_by(FormA.user_id)
        .subquery()
    )

    submitted_form_a = (
        db_session.query(FormA, FormARequirements)
        .join(
            latest_form_a_subq,
            and_(
                FormA.user_id == latest_form_a_subq.c.user_id,
                FormA.submitted_at == latest_form_a_subq.c.latest_submitted_at
            )
        )
        .join(User, FormA.user_id == User.user_id)
        .outerjoin(FormARequirements, FormARequirements.user_id == FormA.user_id)
        .filter(
            or_(
                FormA.reviewer_name1 == user_id,
                FormA.reviewer_name2 == user_id
            )
        )
        .order_by(FormA.submitted_at.desc())
        .all()
    )
    form_aa=[]
    form_bb=[]
    form_cc=[]
    if submitted_form_a:
        
        for form_a, requirementsa in submitted_form_a:
        
            form_aa.append(
                {"forma":form_a,
                 "requirementsa":requirementsa})
            
         
    else:
        form_a, requirementsa = None, None
   
    latest_form_b_subq = (
        db_session.query(
            FormB.user_id.label('user_id'),
            func.max(FormB.submitted_at).label('latest_submitted_at')
        )
        .filter(FormB.submitted_at.isnot(None))
        .group_by(FormB.user_id)
        .subquery()
    )

    submitted_form_b = (
        db_session.query(FormB, FormARequirements)
        .join(
            latest_form_b_subq,
            and_(
                FormB.user_id == latest_form_b_subq.c.user_id,
                FormB.submitted_at == latest_form_b_subq.c.latest_submitted_at
            )
        )
        .join(User, FormB.user_id == User.user_id)
        .outerjoin(FormARequirements, FormARequirements.user_id == FormB.user_id)
        .filter(
            or_(
                FormB.reviewer_name1 == user_id,
                FormB.reviewer_name2 == user_id
            )
        )
        .order_by(FormB.submitted_at.desc())
        .all()
    )

    if submitted_form_b:
        for form_b, requirementsb in submitted_form_b:

            form_bb.append(
                {"formb":form_b,
                 "requirementsb":requirementsb})
            
            
    else:
        form_b, requirementsb = None, None
    # Form C
    latest_form_c_subq = (
        db_session.query(
            FormC.user_id.label('user_id'),
            func.max(FormC.submission_date).label('latest_submission_date')
        )
        .filter(FormC.submission_date.isnot(None))
        .group_by(FormC.user_id)
        .subquery()
    )

    submitted_form_c = (
        db_session.query(FormC, FormARequirements)
        .join(
            latest_form_c_subq,
            and_(
                FormC.user_id == latest_form_c_subq.c.user_id,
                FormC.submission_date == latest_form_c_subq.c.latest_submission_date
            )
        )
        .join(User, FormC.user_id == User.user_id)
        .outerjoin(FormARequirements, FormARequirements.user_id == FormC.user_id)
        .filter(
            or_(
                FormC.reviewer_name1 == user_id,
                FormC.reviewer_name2 == user_id
            )
        )
        .order_by(FormC.submission_date.desc())
        .all()
    )

    if submitted_form_c:
        for form_c, requirementsc in submitted_form_c:
        
            form_cc.append(
                {"formc":form_c,
                 "requirementsc":requirementsc})
    else:
        form_c, requirementsc = None, None
   
    today = date.today()

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )
   
    return render_template('review-dashboard.html',
                user_id=user_id,today=today,
                submitted_form_a=form_aa,
                submitted_form_b=form_bb,
                submitted_form_c=form_cc,
                     # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,)



@app.route('/submit_to_rec/<string:id>', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def submit_to_rec(id):
    form = None
    for model in [FormA, FormB, FormC]:
        form = db_session.query(model).filter_by(form_id=id).first()
        if form:
            break  # Stop once the form is found
    
    if not form:
        flash('Form not found.', 'danger')
        if not certificate_details:
            flash('Certificate record not found.', 'danger')
            return redirect(url_for('chair_landing'))
        flash('Certificate sent to the student.', 'success')
        return redirect(url_for('modify_certificate', id=id))

    outcome = get_admin_reviewer_outcome(form)
    if outcome not in {'approved', 'approved_with_minor_changes'}:
        flash('This form cannot be sent to REC until all assigned reviewers have approved it.', 'danger')
        return redirect(request.referrer or url_for('chair_landing'))

    risk = str(getattr(form, 'risk_level', None) or getattr(form, 'risk_rating', None) or '').lower()
    if outcome == 'approved_with_minor_changes' and not any(level in risk for level in ('mid', 'medium', 'high')):
        flash('A low-risk form with minor changes should be returned to the student or approved for a certificate.', 'warning')
        return redirect(request.referrer or url_for('chair_landing'))

    missing_submission_redirect = redirect_if_missing_student_submission(form, 'REC', 'chair_landing')
    if missing_submission_redirect:
        return missing_submission_redirect
    form.submitted_to_rec = True
    form.submitted_to_admin = False
    form.visible_to_student = False
    form.status = 'Submitted to REC'
    db_session.commit()
    flash('The form was sent to REC.', 'success')
    return redirect(url_for('chair_landing'))
    


@app.route('/reviewer_form_a/<string:id>', methods=['GET'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def reviewer_form_a(id):
    form = db_session.query(FormA).filter_by(form_id=id).first()
    if not form:
        return "Form not found", 404
    if not can_access_as_assigned_reviewer(get_current_user(), form):
        abort(403)
    if has_current_reviewer_submitted_feedback(form, session.get('id')):
        flash('You have already reviewed and submitted feedback for this application.', 'warning')
        return redirect(url_for('review_dashboard'))
    data={}
    if form and is_submitted_form_record(form):
        data={
        "org_name" : parse_field(form.org_name),
        "org_contact" : parse_field(form.org_contact),
        "org_role": parse_field(form.org_role),
        "org_permission" : parse_field(form.org_permission),

        "fund_org" : parse_field(form.fund_org),
        "fund_contact" :parse_field(form.fund_contact),
        "fund_role": parse_field(form.fund_role),
        "fund_amount": parse_field(form.fund_amount),

        "population" :parse_field(form.population),
        "sampling_method" : parse_field(form.sampling_method),
        "sampling_size": parse_field(form.sampling_size),
        "inclusion_criteria": parse_field(form.inclusion_criteria)
    }

    if form and is_submitted_form_record(form):
        return render_template("review_form_a.html", form=form,data=data)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('review_dashboard'))


@app.route('/reviewer_form_b/<string:id>', methods=['GET'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def reviewer_form_b(id):
   
    form = db_session.query(FormB).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).filter_by(form_id=id).first()
    if not form:
        return "Form not found", 404
    if not can_access_as_assigned_reviewer(get_current_user(), form):
        abort(403)
    if has_current_reviewer_submitted_feedback(form, session.get('id')):
        flash('You have already reviewed and submitted feedback for this application.', 'warning')
        return redirect(url_for('review_dashboard'))

    if form and is_submitted_form_record(form):
        return render_template("review_form_b.html",form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('review_dashboard'))



@app.route('/reviewer_form_c/<string:id>', methods=['GET'])
@role_required('REVIEWER', 'ADMIN', 'SUPER_ADMIN', 'REC')
def reviewer_form_c(id):
    form = db_session.query(FormC).filter_by(form_id=id).first()
    if not form:
        return "Form not found", 404
    if not can_access_as_assigned_reviewer(get_current_user(), form):
        abort(403)
    if has_current_reviewer_submitted_feedback(form, session.get('id')):
        flash('You have already reviewed and submitted feedback for this application.', 'warning')
        return redirect(url_for('review_dashboard'))
   
    if form and is_submitted_form_record(form):
        return render_template("review_form_c.html", form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('review_dashboard'))
    


@app.route('/rec_dashboard', methods=['GET', 'POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def rec_dashboard():
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))

    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user:
        return redirect(url_for('login_page'))
    
    
    today = date.today()
    role = user.role.value if hasattr(user.role, 'value') else str(user.role)
    # Count reviews per form_id
    form_review_counts = dict(
        db_session.query(Rec.form_id, func.count(Rec.rec_id))
        .group_by(Rec.form_id)
        .all()
    )

    


    # Shared filter conditions
    def get_common_filters(FormModel):
        return [
            FormModel.rejected_or_accepted == True,
            FormModel.review_signature_date != None,
            ~func.lower(FormModel.risk_level if hasattr(FormModel, 'risk_level') else FormModel.risk_rating).like('%low%'),
            FormModel.review_status == True,
            FormModel.review_status1 == True,
            FormModel.submitted_to_rec == True,
            FormModel.reviewer_name1 != user_id,
            FormModel.reviewer_name2 != user_id,
        ]

    def rec_sort_timestamp(form):
        return (
            getattr(form, 'submitted_at', None)
            or getattr(form, 'submission_date', None)
            or getattr(form, 'review_signature_date', None)
            or getattr(form, 'signature_date', None)
            or getattr(form, 'created_at', None)
            or datetime.min
        )

    def paginate_records(records, page, per_page=20):
        total_records = len(records)
        total_pages = max(1, math.ceil(total_records / per_page)) if total_records else 1
        current_page = min(max(page, 1), total_pages)
        start_index = (current_page - 1) * per_page
        end_index = start_index + per_page
        return records[start_index:end_index], total_records, total_pages, current_page

    # Form A
    submitted_form_a = [
        (form, req, form_review_counts.get(form.form_id, 0))
        for form, req in db_session.query(FormA, FormARequirements)
            .outerjoin(FormARequirements, FormA.user_id == FormARequirements.user_id)
            .outerjoin(Rec, Rec.form_id == FormA.form_id)
            .filter(*get_common_filters(FormA))
            .all()
    ]
    submitted_form_a.sort(key=lambda item: rec_sort_timestamp(item[0]), reverse=True)

    # Form B
    submitted_form_b = [
        (form, req, form_review_counts.get(form.form_id, 0))
        for form, req in db_session.query(FormB, FormARequirements)
            .options(
                defer(FormB.permission_letter),
                defer(FormB.prior_clearance),
                defer(FormB.ethics_evidence),
                defer(FormB.proposal_path),
                defer(FormB.pending_note),
                defer(FormB.private_permission_file)
            )
            .outerjoin(FormARequirements, FormB.user_id == FormARequirements.user_id)
            .outerjoin(Rec, Rec.form_id == FormB.form_id)
            .filter(*get_common_filters(FormB))
            .all()
    ]
    submitted_form_b.sort(key=lambda item: rec_sort_timestamp(item[0]), reverse=True)

    # Form C
    submitted_form_c = [
        (form, req, form_review_counts.get(form.form_id, 0))
        for form, req in db_session.query(FormC, FormARequirements)
            .outerjoin(FormARequirements, FormC.user_id == FormARequirements.user_id)
            .outerjoin(Rec, Rec.form_id == FormC.form_id)
            .filter(*get_common_filters(FormC))
            .all()
    ]
    submitted_form_c.sort(key=lambda item: rec_sort_timestamp(item[0]), reverse=True)

    page_a = request.args.get('page_a', default=1, type=int)
    page_b = request.args.get('page_b', default=1, type=int)
    page_c = request.args.get('page_c', default=1, type=int)
    submitted_form_a, total_records_a, total_pages_a, page_a = paginate_records(submitted_form_a, page_a)
    submitted_form_b, total_records_b, total_pages_b, page_b = paginate_records(submitted_form_b, page_b)
    submitted_form_c, total_records_c, total_pages_c, page_c = paginate_records(submitted_form_c, page_c)

    # Count all reviewers
    all_reviewers_counter = db_session.query(User).filter(User.role == 'REVIEWER').count()

    # Supervisor-specific requirements
    supervisor_formA_req = db_session.query(FormARequirements).filter_by(user_id=user_id).all()

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )

    return render_template(
        'rec-dashboard.html',
        today=today,
        role=role,
        all_Reviewers_counter=all_reviewers_counter,
        submitted_form_a=submitted_form_a,
        submitted_form_b=submitted_form_b,
        submitted_form_c=submitted_form_c,
        supervisor_formA_req=supervisor_formA_req,
        page_a=page_a,
        page_b=page_b,
        page_c=page_c,
        total_pages_a=total_pages_a,
        total_pages_b=total_pages_b,
        total_pages_c=total_pages_c,
        total_records_a=total_records_a,
        total_records_b=total_records_b,
        total_records_c=total_records_c,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
        
    )


@app.route('/admin_rec_form/<string:form_id>',methods=['GET','POST'])
def admin_rec_form(form_id):
  
    user_id = session.get('id')
    if not user_id:
        return redirect(url_for('login_page'))

    user=db_session.query(User).filter(User.user_id==user_id).first()
    if not user:
        return redirect(url_for('login_page'))

    form = None
    page = request.args.get('page', default=1, type=int)
    per_page = 20
    rec_query = (
        db_session.query(Rec)
        .filter(Rec.form_id == form_id)
        .order_by(Rec.rec_date.desc().nullslast(), Rec.rec_id.desc())
    )
    total_records = rec_query.count()
    total_pages = max(1, math.ceil(total_records / per_page)) if total_records else 1
    page = min(max(page, 1), total_pages)
    Rec_team = rec_query.offset((page - 1) * per_page).limit(per_page).all()
    for model in [FormA, FormB, FormC]:
        form = db_session.query(model).filter_by(form_id=form_id).first()
        if form:
            break  # Stop once the form is found

    role = user.role.value if hasattr(user.role, 'value') else str(user.role)
    if not form:
        return redirect(url_for('rec_dashboard'))
    # Count unique reviewers for this form
    unique_reviewer_ids = set()
    unique_reviewer_names = set()
    for rec in Rec_team:
        if rec.rec_id:
            unique_reviewer_ids.add(rec.rec_id)
        if rec.full_name:
            unique_reviewer_names.add(rec.full_name)
    unique_reviewer_count = len(unique_reviewer_ids)

    # Get all reviewers in the system
    all_reviewers = db_session.query(User).filter(User.role == 'REVIEWER').all()
    all_reviewer_ids = set([r.user_id for r in all_reviewers])
    all_reviewer_names = set([r.full_name for r in all_reviewers])

    # Reviewers who have not yet reviewed this form
    not_reviewed_ids = all_reviewer_ids - unique_reviewer_ids
    not_reviewed_names = all_reviewer_names - unique_reviewer_names

    all_reviewers_counter = len(all_reviewers)

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )


    return render_template(
        'chair_rec_form.html',
        Rec_team=Rec_team,
        role=role,
        form=form,
        all_reviewers_counter=all_reviewers_counter,
        unique_reviewer_count=unique_reviewer_count,
        unique_reviewer_names=list(unique_reviewer_names),
        not_reviewed_ids=list(not_reviewed_ids),
        not_reviewed_names=list(not_reviewed_names),
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_records=total_records,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,

    )
@app.route('/rec_form_a/<string:id>', methods=['GET'])
def rec_form_a(id):
    
    form = db_session.query(FormA).filter(FormA.form_id==id).first()
    data={
        "org_name" : parse_field(form.org_name),
        "org_contact" : parse_field(form.org_contact),
        "org_role": parse_field(form.org_role),
        "org_permission" : parse_field(form.org_permission),

        "fund_org" : parse_field(form.fund_org),
        "fund_contact" :parse_field(form.fund_contact),
        "fund_role": parse_field(form.fund_role),
        "fund_amount": parse_field(form.fund_amount),

        "population" :parse_field(form.population),
        "sampling_method" : parse_field(form.sampling_method),
        "sampling_size": parse_field(form.sampling_size),
        "inclusion_criteria": parse_field(form.inclusion_criteria)
    }
    if form:
        return render_template("rec_form_a.html", form=form,data=data)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('rec_dashboard'))


@app.route('/rec_form_b/<string:id>', methods=['GET'])
def rec_form_b(id):
   
    form = db_session.query(FormB).options(
        defer(FormB.permission_letter),
        defer(FormB.prior_clearance),
        defer(FormB.ethics_evidence),
        defer(FormB.proposal_path),
        defer(FormB.pending_note),
        defer(FormB.private_permission_file)
    ).filter(FormB.form_id==id).first()

    if form:
        return render_template("rec_form_b.html",form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('rec_dashboard'))



@app.route('/rec_form_c/<string:id>', methods=['GET'])
def rec_form_c(id):
    form = db_session.query(FormC).filter(FormC.form_id==id).first()
   
    if form:
        return render_template("rec_form_c.html", form=form)
    else:
        # You can pass an error message or just load the dashboard
        return redirect(url_for('rec_dashboard'))
    


@app.route('/rec_response/<string:id>', methods=['GET', 'POST'])
def rec_response(id):
    user_id=session.get('id')
    if request.method == 'POST':
        status = request.form.get('status')
        comments = request.form.get('rec_comments')  # ✅ corrected from 'additional_comments'

        # Loop through models to find the correct form by ID
        
        user_name=db_session.query(User).filter_by(user_id=user_id).first()
       
        form=Rec(
                rec_id=user_id,
                form_id=id,
                full_name=user_name.full_name,
                rec_comments = comments,
                rec_status = status,
                rec_date=datetime.now()
                )
        db_session.add(form)
        db_session.commit()
        flash("Form updated successfully", "success")

        return redirect(url_for('rec_dashboard'))

    return "Invalid access", 405



def generate_clearance_code(committee_acronym, decision_date=None):
    if decision_date is None:
        decision_date = datetime.today()

    # Format the date as YYYYMMDD
    date_str = decision_date.strftime('%Y%m%d')

    total_count = 0  # Initialize total decision count

    for model in [FormA, FormB, FormC]:
        count = (
            db_session.query(func.count())
            .select_from(model)
            .filter(cast(model.rec_date, Date) == decision_date.date())
            .scalar()
        )
        total_count += count
    
    # Increment for the new decision
    decision_number = total_count + 1
    
    # Format the final clearance code
    clearance_code = f"{committee_acronym}{date_str}{decision_number:02d}"
    return clearance_code


DEFAULT_CERTIFICATE_CONDITIONS = [
    'The researcher ensures adherence to the POPI Act.',
    'The researcher ensures adherence to conditions of use of open source questionnaires.',
    'The project adheres to ethical research requirements.',
    'The study is conducted as set out in the approved application.',
    'The project adheres to applicable legislation, professional codes and scientific standards.',
    'Proposed changes, concerns and unexpected ethical issues are reported to JBSREC.',
    'Changes affecting study-related risks are reported to JBSREC in writing.',
    'All permissions required to access data and organisations have been obtained.',
    'No fieldwork may continue after clearance expires without an approved extension.',
]


def get_certificate_form(form_id):
    for model in (FormA, FormB, FormC):
        record = db_session.query(model).filter_by(form_id=form_id).first()
        if record:
            return record
    return None


def get_certificate_conditions(record):
    raw = getattr(record, 'certificate_heading', None)
    if not raw:
        return list(DEFAULT_CERTIFICATE_CONDITIONS)
    try:
        values = json.loads(raw)
        if isinstance(values, list):
            return [str(value).strip() for value in values if str(value).strip()]
    except (TypeError, ValueError):
        pass
    separator = '|||' if '|||' in str(raw) else ','
    values = [value.strip() for value in str(raw).split(separator) if value.strip()]
    return values or list(DEFAULT_CERTIFICATE_CONDITIONS)


def save_certificate_draft(record):
    try:
        valid_years = int(request.form.get('valid_years', ''))
    except (TypeError, ValueError):
        raise ValueError('Enter a valid number of years for the certificate.')
    if valid_years < 1 or valid_years > 10:
        raise ValueError('Certificate validity must be between 1 and 10 years.')

    end_date_value = (request.form.get('end_date') or '').strip()
    try:
        end_date = datetime.strptime(end_date_value, '%Y-%m-%d')
    except ValueError:
        raise ValueError('Select a valid certificate end date.')
    if end_date.date() <= date.today():
        raise ValueError('The certificate end date must be after today.')

    conditions = [value.strip() for value in request.form.getlist('conditions[]') if value.strip()]
    if not conditions:
        raise ValueError('Add at least one certificate condition.')

    if not getattr(record, 'certificate_code', None):
        record.certificate_code = generate_clearance_code('JBSREC')
    record.certificate_valid_years = valid_years
    record.certificate_end_date = end_date
    record.certificate_issuer = (request.form.get('certificate_issuer') or '').strip() or 'JBS Research Ethics Committee'
    record.certificate_email = (request.form.get('email') or '').strip() or 'jbssubmissions@uj.ac.za'
    record.certificate_heading = json.dumps(conditions)
    record.certificate_modified = True
    record.certificate_received = False


@app.route('/certificate/<string:id>', methods=['GET', 'POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def certificate(id):
    certificate_details = get_certificate_form(id)
    if not certificate_details:
        return "No certificate data found.", 404

    if request.method == 'POST':
        try:
            save_certificate_draft(certificate_details)
            db_session.commit()
            flash('Certificate draft saved. Review it below, then issue it when ready.', 'success')
            return redirect(url_for('modify_certificate', id=id))
        except ValueError as error:
            db_session.rollback()
            flash(str(error), 'danger')

    return render_template(
        'edit_certificate.html',
        certificate_details=certificate_details,
        certificate_conditions=get_certificate_conditions(certificate_details),
        certificate_is_draft=not bool(certificate_details.certificate_issued),
    )



@app.route('/modify_certificate/<string:id>', methods=['GET', 'POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def modify_certificate(id):
    certificate_details = get_certificate_form(id)
    if not certificate_details:
        return "No certificate data found.", 404

    if request.method == 'POST':
        try:
            save_certificate_draft(certificate_details)
            db_session.commit()
            flash('Certificate changes saved.', 'success')
            return redirect(url_for('modify_certificate', id=id))
        except ValueError as error:
            db_session.rollback()
            flash(str(error), 'danger')

    return render_template(
        'edit_certificate.html',
        certificate_details=certificate_details,
        certificate_conditions=get_certificate_conditions(certificate_details),
        certificate_is_draft=not bool(certificate_details.certificate_issued),
    )


@app.route('/issue_certificate/<string:id>', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def issue_certificate(id):
    certificate_details = get_certificate_form(id)
    if not certificate_details:
        flash('Certificate record not found.', 'danger')
        return redirect(url_for('chair_landing'))
    required = (
        certificate_details.certificate_code,
        certificate_details.certificate_valid_years,
        certificate_details.certificate_end_date,
        certificate_details.certificate_heading,
    )
    if not all(required):
        flash('Save the complete certificate draft before issuing it.', 'danger')
        return redirect(url_for('modify_certificate', id=id))
    if certificate_details.certificate_issued:
        flash('This certificate has already been issued. You can still edit or view it.', 'info')
        return redirect(url_for('modify_certificate', id=id))

    certificate_details.certificate_issued = datetime.now()
    certificate_details.certificate_received = False
    db_session.commit()
    flash('Certificate issued successfully. Review the final copy, then send it to the student.', 'success')
    return redirect(url_for('edited_certificate', id=id))


@app.route('/view_certificate/<string:id>',methods=['GET','POST'])
@login_required
def view_certificate(id):
    
    if 'id' not in session:
        flash("Unauthorized access", "error")
        return redirect(url_for('login_page'))
    
    certificate_details = None
    for model in [FormA, FormB, FormC]:
        certificate_details = db_session.query(model).filter_by(form_id=id).first()
        if certificate_details:
            break
    if certificate_details and not can_access_form(get_current_user(), certificate_details):
        abort(403)
    return render_template(
        'view_certificate.html',
        certificate_details=certificate_details
    )

@app.route('/update_certificate_status/<string:id>', methods=['POST'])
@role_required('ADMIN', 'SUPER_ADMIN', 'REC')
def update_certificate_status(id):
    current_role = str(session.get('role') or '').upper()
    redirect_endpoint = 'ethics_reviewer_committee_form' if current_role == 'REC' else 'chair_landing'

    certificate_details = None
    for model in [FormA, FormB, FormC]:
        certificate_details = db_session.query(model).filter_by(form_id=id).first()
        if certificate_details:
            break

    if not certificate_details:
        flash("Certificate record not found.", "error")
        return redirect(url_for(redirect_endpoint))

    try:
        if not getattr(certificate_details, 'certificate_issued', None):
            certificate_details.certificate_issued = datetime.now()
        certificate_details.certificate_received = True

        db_session.add(certificate_details)
        db_session.commit()
        flash("Certificate status updated successfully.", "success")
    except Exception as error:
        db_session.rollback()
        app.logger.exception("Certificate status update failed for form_id=%s: %s", id, error)
        flash("Certificate status could not be updated. Please try again.", "error")

    return redirect(url_for(redirect_endpoint))



@app.route('/edited_certificate/<string:id>', methods=['GET'])
@login_required
def edited_certificate(id):
    certificate_details = None
    for model in [FormA, FormB, FormC]:
        certificate_details = db_session.query(model).filter_by(form_id=id).first()
        if certificate_details:
            break
    if not certificate_details:
        return "No certificate data found.", 404
    if not can_access_form(get_current_user(), certificate_details):
        abort(403)
    return render_template('edited_certificate.html', certificate_details=certificate_details)



###
### Admin Review Submision
###
@app.route('/ethics_reviewer_committee_forms/<string:id>/<string:form_name>', methods=['GET','POST'])
@role_required('ADMIN', 'SUPER_ADMIN')
def ethics_reviewer_committee_forms(id,form_name):
    def resolve_reviewer_ids(selected_reviewers, existing_reviewer_ids):
        unique_selected = []
        for reviewer_id in selected_reviewers or []:
            reviewer_id = (reviewer_id or '').strip()
            if reviewer_id and reviewer_id not in unique_selected:
                unique_selected.append(reviewer_id)
        if unique_selected:
            return unique_selected[:1]
        return (existing_reviewer_ids or [])[:1]

    forma = db_session.query(FormA).filter_by(form_id=id).first()
    
    Assigned_reviewer=''
    if forma:
        id_of_reviewers = get_reviewers_for_ethics_assignment(forma, FormA, FormA.submitted_at)[:1]
        if id_of_reviewers:
            Assigned_reviewer = db_session.query(User).filter(User.user_id.in_(id_of_reviewers)).all()
 
    list_of_revewers=[]
    id_of_reviewers=[]
    if Assigned_reviewer:
        for item in Assigned_reviewer:
            id_of_reviewers.append(item.user_id)
            list_of_revewers.append(item.email)
    else:
        reviewers=request.form.getlist('reviewer_names[]')

        if len(reviewers) >= 1:
            Assigned_reviewer=db_session.query(User).filter(User.user_id == reviewers[0]).all()
            

    if form_name=="FORM A":
        
        if request.method=="POST":
            reviewers=request.form.getlist('reviewer_names[]')
            if len({reviewer_id for reviewer_id in reviewers if reviewer_id}) > 1:
                flash('Please select exactly one reviewer before submitting.', 'danger')
                return redirect(url_for('ethics_reviewer_committee_forms', id=id, form_name=form_name))

            effective_reviewer_ids = resolve_reviewer_ids(reviewers, id_of_reviewers)
            if request.form.get('accept') in ['Accept','Approved with Minor Changes'] and len(effective_reviewer_ids) < 1:
                flash('Please assign exactly one reviewer before submitting.', 'danger')
                return redirect(url_for('ethics_reviewer_committee_forms', id=id, form_name=form_name))

            _reviewers_emails=[]
            if reviewers:
                _reviewers_emails=db_session.query(User).filter(User.user_id.in_(effective_reviewer_ids)).all()

            forma.reviewer_name1 = effective_reviewer_ids[0] if len(effective_reviewer_ids) >= 1 else None
            forma.reviewer_name2 = None
            forma.ethics_signature_date=datetime.now()
            forma.review_form_comments=request.form.get('additional_comments')
            forma.ethics_status=request.form.get('recommendation')
            if request.form.get('accept') in ['Accept','Approved with Minor Changes']:
                missing_submission_redirect = redirect_if_missing_student_submission(
                    forma,
                    'reviewers',
                    'ethics_reviewer_committee_forms',
                    id=id,
                    form_name=form_name,
                )
                if missing_submission_redirect:
                    return missing_submission_redirect
                if Assigned_reviewer:
                    #Uncomment the code bellow for testing
                    ##
                    for item in Assigned_reviewer:
                        list_of_revewers.append(item.email)
                    
                    try:
                        message=f' You are assigned as a reviewer for a form belonging to {forma.applicant_name}. Please log into the ethics application system and review the form.' 

                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")  
                
                elif _reviewers_emails:
                    for item in _reviewers_emails:
                        list_of_revewers.append(item.email)
                    try:
                        message=f' You are assigned as a reviewer for a form belonging to {forma.applicant_name}. Please log into the ethics application system and review the form.' 

                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")
                else:
                    #Uncomment the code bellow for testing
                    ##
                    
                    try:
                        message=f'You are assigned as a reviewer for a form belonging to {forma.applicant_name}. Please log into the ethics application system and review the form.' 

                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")  
                forma.rejected_or_accepted=True
                forma.submitted_to_reviewers=True
                # Log admin action for FormA
                db_session.add(UserActivityLog(
                    user_id=session.get('id'),
                    action='assign_reviewers',
                    page='FORM A Ethics Assignment',
                    target_user_id=forma.user_id,
                    details=f"Assigned reviewers: {forma.reviewer_name1}, {forma.reviewer_name2} to FORM A for {forma.applicant_name}"
                ))
            else:
                forma.submitted_to_reviewers=False
                forma.rejected_or_accepted=False
                emails=[]
                emails.append(forma.email)
                emails.append(forma.supervisor_email)
                try:
                    message=f' Form belonging to {forma.applicant_name} was sent back. Please view feedback.' 
            
                    send_email(app,mail, message,emails)
                except Exception as e:
                    app.logger.error(f"Failed to send email")
            db_session.add(forma)
            db_session.commit()
            return redirect(url_for('chair_landing'))
        return render_template("form_a_ethics.html",formA=forma)
    elif form_name=="FORM B":
        formb = db_session.query(FormB).filter_by(form_id=id).first()
        Assigned_reviewer=''
        if formb:
            id_of_reviewers = get_reviewers_for_ethics_assignment(formb, FormB, FormB.submitted_at)[:1]
            if id_of_reviewers:
                Assigned_reviewer = db_session.query(User).filter(User.user_id.in_(id_of_reviewers)).all()
 
        list_of_revewers=[]
        id_of_reviewers=[]
        if Assigned_reviewer:
            for item in Assigned_reviewer:
                id_of_reviewers.append(item.user_id)
                list_of_revewers.append(item.email)
        else:
            reviewers=request.form.getlist('reviewer_names[]')
            
            if len(reviewers) >= 1:
                Assigned_reviewer=db_session.query(User).filter(User.user_id == reviewers[0]).all()
  
        if request.method=="POST":
            reviewers=request.form.getlist('reviewer_names[]')
            if len({reviewer_id for reviewer_id in reviewers if reviewer_id}) > 1:
                flash('Please select exactly one reviewer before submitting.', 'danger')
                return redirect(url_for('ethics_reviewer_committee_forms', id=id, form_name=form_name))

            effective_reviewer_ids = resolve_reviewer_ids(reviewers, id_of_reviewers)
            if request.form.get('accept') in ['Accept','Approved with Minor Changes'] and len(effective_reviewer_ids) < 1:
                flash('Please assign exactly one reviewer before submitting.', 'danger')
                return redirect(url_for('ethics_reviewer_committee_forms', id=id, form_name=form_name))

            formb.reviewer_name1 = effective_reviewer_ids[0] if len(effective_reviewer_ids) >= 1 else None
            formb.reviewer_name2 = None
            formb.ethics_signature_date=datetime.now()
            formb.review_form_comments=request.form.get('additional_comments')
            formb.ethics_status=request.form.get('recommendation')
            
            if request.form.get('accept') in ['Accept','Approved with Minor Changes']:
                missing_submission_redirect = redirect_if_missing_student_submission(
                    formb,
                    'reviewers',
                    'ethics_reviewer_committee_forms',
                    id=id,
                    form_name=form_name,
                )
                if missing_submission_redirect:
                    return missing_submission_redirect
                if Assigned_reviewer:
                    for item in Assigned_reviewer:
                        list_of_revewers.append(item.email)
                    
                    #Uncomment the code bellow for testing
                    ##
                   
                    try:
                        message=f' You areassigned as a reviewer for a form belonging to {formb.applicant_name}. Please log into the ethics application system and review the form' 

                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")  

                else:
                    #Uncomment the code bellow for testing
                    ##
                    try:
                        message=f'You areassigned as a reviewer for a form belonging to {formb.applicant_name}. Please log into the ethics application system and review the form' 
            
                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")
                formb.rejected_or_accepted=True
                formb.submitted_to_reviewers=True
                # Log admin action for FormB
                db_session.add(UserActivityLog(
                    user_id=session.get('id'),
                    action='assign_reviewers',
                    page='FORM B Ethics Assignment',
                    target_user_id=formb.user_id,
                    details=f"Assigned reviewers: {formb.reviewer_name1}, {formb.reviewer_name2} to FORM B for {formb.applicant_name}"
                ))
            else:
                    review_date_str = request.form.get('review_date')
                    if review_date_str:
                        try:
                            # Try parsing as DD/MM/YYYY
                            review_date_obj = datetime.strptime(review_date_str, "%d/%m/%Y")
                            formb.supervisor_date = review_date_obj
                        except ValueError:
                            try:
                                # Fallback: try parsing as YYYY-MM-DD
                                review_date_obj = datetime.strptime(review_date_str, "%Y-%m-%d")
                                formb.supervisor_date = review_date_obj
                            except ValueError:
                                formb.supervisor_date = None  # or handle error as needed
                    else:
                        formb.supervisor_date = None
                        formb.rejected_or_accepted = False
                        formb.submitted_to_reviewers = False
                        emails = []
                        emails.append(formb.email)
                        emails.append(formb.supervisor_email)
                    try:
                        message = f' Form belonging to {formb.applicant_name} was sent back. Please view feedback.'
                        send_email(app, mail, message, emails)
                    except Exception as e:
                        app.logger.error(f"Failed to send email")
            db_session.add(formb)
            db_session.commit()
            return redirect(url_for('chair_landing'))
        return render_template("form_b_ethics.html",formB=formb)
    elif form_name=="FORM C":
       
        formc = db_session.query(FormC).filter_by(form_id=id).first()
        
        Assigned_reviewer=''
        if formc:
            id_of_reviewers = get_reviewers_for_ethics_assignment(formc, FormC, FormC.submission_date)[:1]
            if id_of_reviewers:
                Assigned_reviewer = db_session.query(User).filter(User.user_id.in_(id_of_reviewers)).all()
 
        list_of_revewers=[]
        id_of_reviewers=[]
        if Assigned_reviewer:
            for item in Assigned_reviewer:
                id_of_reviewers.append(item.user_id)
                list_of_revewers.append(item.email)
            
        else:
            reviewers=request.form.getlist('reviewer_names[]')
            if len(reviewers) >= 1:
                Assigned_reviewer=db_session.query(User).filter(User.user_id == reviewers[0]).all()
            
        if request.method=="POST":
            reviewers=request.form.getlist('reviewer_names[]')
            if len({reviewer_id for reviewer_id in reviewers if reviewer_id}) > 1:
                flash('Please select exactly one reviewer before submitting.', 'danger')
                return redirect(url_for('ethics_reviewer_committee_forms', id=id, form_name=form_name))

            effective_reviewer_ids = resolve_reviewer_ids(reviewers, id_of_reviewers)
            if request.form.get('accept') in ['Accept','Approved with Minor Changes'] and len(effective_reviewer_ids) < 1:
                flash('Please assign exactly one reviewer before submitting.', 'danger')
                return redirect(url_for('ethics_reviewer_committee_forms', id=id, form_name=form_name))

            _reviewers_emails=[]
            if reviewers:
                _reviewers_emails=db_session.query(User).filter(User.user_id.in_(effective_reviewer_ids)).all()

            formc.reviewer_name1 = effective_reviewer_ids[0] if len(effective_reviewer_ids) >= 1 else None
            formc.reviewer_name2 = None
            formc.ethics_signature_date=datetime.now()
            formc.review_form_comments=request.form.get('additional_comments')
            formc.ethics_status=request.form.get('recommendation')
            
            if request.form.get('accept') in ['Accept','Approved with Minor Changes']:
                missing_submission_redirect = redirect_if_missing_student_submission(
                    formc,
                    'reviewers',
                    'ethics_reviewer_committee_forms',
                    id=id,
                    form_name=form_name,
                )
                if missing_submission_redirect:
                    return missing_submission_redirect
                if Assigned_reviewer:
                    #Uncomment the code bellow for testing
                    ##
                    for item in Assigned_reviewer:
                        list_of_revewers.append(item.email)
                    
                    try:
                        message=f' You areassigned as a reviewer for a form belonging to {formc.applicant_name}. Please log into the ethics application system and review the form' 
            
                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")

                else:
                    #Uncomment the code bellow for testing
                    ##
                    try:
                        message=f'You areassigned as a reviewer for a form belonging to {formc.applicant_name}. ,Please log into the ethic applications system and review the form' 
            
                        send_email(app,mail, message,list_of_revewers)
                    except Exception as e:
                        print(f"Error sending email: {e}")

                formc.rejected_or_accepted=True
                formc.submitted_to_reviewers=True
                # Log admin action for FormC
                db_session.add(UserActivityLog(
                    user_id=session.get('id'),
                    action='assign_reviewers',
                    page='FORM C Ethics Assignment',
                    target_user_id=formc.user_id,
                    details=f"Assigned reviewers: {formc.reviewer_name1}, {formc.reviewer_name2} to FORM C for {formc.applicant_name}"
                ))
            else:
                formc.supervisor_date=request.form.get('review_date')
                formc.rejected_or_accepted=False
                formc.submitted_to_reviewers=False
                emails=[]
                emails.append(formc.email_address)
                emails.append(formc.supervisor_email)
                try:
                    message=f' Form belonging to {formc.applicant_name} was sent back. Please view feedback.' 
            
                    send_email(app,mail, message,emails)
                except Exception as e:
                    app.logger.error(f"Failed to send email")
            db_session.add(formc)
            db_session.commit()
            return redirect(url_for('chair_landing'))
        return render_template("form_c_ethics.html",formc=formc)



@app.route('/request-reset', methods=['POST'])
def request_reset():
    email = request.form.get('email')
    token = generate_reset_token(email)
    if token:
        reset_link = url_for('reset_password', token=token, _external=True)
        send_email(
            to=email,
            subject="Password Reset",
            body=f"Click here to reset: {reset_link}"
        )
        return "Reset email sent!"
    return "Email not found", 404

from sqlalchemy.orm import aliased
@app.route('/supervisor_dashboard', methods=['GET','POST'])
def supervisor_dashboard():
    supervisor_id=session.get('id')
    
    role=session.get('supervisor_role')
    if not role:
        return redirect(url_for('login_page'))
    #supervisor_id="bea65156-03ff-45c8-bd41-9d07f4bc48d2"
    if not supervisor_id:
        return jsonify({'error': 'Unauthorized'}), 401
 
    # Global stats (limited to avoid performance issues) - these are for dashboard stats only
    formA = db_session.query(FormA).filter(FormA.submitted_at != None).order_by(FormA.submitted_at.desc()).limit(5).all()
    
    # Safe FormB query - only load needed columns to avoid binary column issues (LIMITED)
    formB_results = db_session.query(
        FormB.form_id, FormB.user_id, FormB.applicant_name, FormB.student_number,
        FormB.email, FormB.supervisor, FormB.supervisor_email, FormB.submitted_at,
        FormB.recommendation, FormB.supervisor_date, FormB.ethics_status,
        FormB.signature_date, FormB.review_supervisor_signature, FormB.review_date,
        FormB.review_supervisor_signature1, FormB.review_date1, FormB.created_at, FormB.declaration_date
    ).filter(FormB.submitted_at != None).order_by(FormB.submitted_at.desc()).limit(5).all()
    
    # Convert tuples to proxy objects
    formB = []
    for result in formB_results:
        class FormBProxy:
            pass
        proxy = FormBProxy()
        proxy.form_id = result.form_id
        proxy.user_id = result.user_id
        proxy.applicant_name = result.applicant_name
        proxy.student_number = result.student_number
        proxy.email = result.email
        proxy.supervisor = result.supervisor
        proxy.supervisor_email = result.supervisor_email
        proxy.submitted_at = result.submitted_at
        proxy.recommendation = result.recommendation
        proxy.supervisor_date = result.supervisor_date
        proxy.ethics_status = result.ethics_status
        proxy.signature_date = result.signature_date
        proxy.review_supervisor_signature = result.review_supervisor_signature
        proxy.review_date = result.review_date
        proxy.review_supervisor_signature1 = result.review_supervisor_signature1
        proxy.review_date1 = result.review_date1
        proxy.created_at = result.created_at
        proxy.declaration_date = result.declaration_date
        formB.append(proxy)
    
    formC = db_session.query(FormC).filter(FormC.submission_date != None).order_by(FormC.submission_date.desc()).limit(5).all()
    
 
    # Query FormA and its FormARequirements for supervisor's users - SIMPLIFIED
    supervisor_formA_raw = (
        db_session.query(FormA, FormARequirements)
        .join(User, FormA.user_id == User.user_id)
        .outerjoin(FormARequirements, FormARequirements.user_id == FormA.user_id)
        .filter(User.supervisor_id == supervisor_id, FormA.submitted_at != None)
        .order_by(FormA.user_id, FormA.submitted_at.desc())
        .all()
    )
    
    # Keep only the latest per user and add count
    seen_users_a = set()
    supervisor_formA = []
    forma_counts = {}
    
    # First pass: count submissions per user
    for forma, req in supervisor_formA_raw:
        if forma.user_id not in forma_counts:
            forma_counts[forma.user_id] = 0
        forma_counts[forma.user_id] += 1
    
    # Second pass: keep latest per user
    for forma, req in supervisor_formA_raw:
        if forma.user_id not in seen_users_a:
            count = forma_counts.get(forma.user_id, 0)
            supervisor_formA.append((forma, req, count))
            seen_users_a.add(forma.user_id)
    
    # Query FormB and its FormARequirements for supervisor's users - SIMPLIFIED
    supervisor_formB_raw = (
        db_session.query(FormB, FormARequirements)
        .join(User, FormB.user_id == User.user_id)
        .outerjoin(FormARequirements, FormARequirements.user_id == FormB.user_id)
        .filter(User.supervisor_id == supervisor_id, FormB.submitted_at != None)
        .order_by(FormB.user_id, FormB.submitted_at.desc())
        .all()
    )
    
    # Keep only the latest per user and add count
    seen_users = set()
    supervisor_formB = []
    formb_counts = {}
    
    # First pass: count submissions per user
    for formb, req in supervisor_formB_raw:
        if formb.user_id not in formb_counts:
            formb_counts[formb.user_id] = 0
        formb_counts[formb.user_id] += 1
    
    # Second pass: keep latest per user
    for formb, req in supervisor_formB_raw:
        if formb.user_id not in seen_users:
            count = formb_counts.get(formb.user_id, 0)
            supervisor_formB.append((formb, req, count))
            seen_users.add(formb.user_id)
    
    # Subquery to count all FormC submissions per user and get latest - SIMPLIFIED
    supervisor_formC_raw = (
        db_session.query(FormC, FormARequirements)
        .join(User, FormC.user_id == User.user_id)
        .outerjoin(FormARequirements, FormARequirements.user_id == FormC.user_id)
        .filter(User.supervisor_id == supervisor_id, FormC.submission_date != None)
        .order_by(FormC.user_id, FormC.submission_date.desc())
        .all()
    )
    
    # Keep only the latest per user and add count
    seen_users_c = set()
    supervisor_formC = []
    formc_counts = {}
    
    # First pass: count submissions per user
    for formc, req in supervisor_formC_raw:
        if formc.user_id not in formc_counts:
            formc_counts[formc.user_id] = 0
        formc_counts[formc.user_id] += 1
    
    # Second pass: keep latest per user
    for formc, req in supervisor_formC_raw:
        if formc.user_id not in seen_users_c:
            count = formc_counts.get(formc.user_id, 0)
            supervisor_formC.append((formc, req, count))
            seen_users_c.add(formc.user_id)
    
    #supervisor_formA_req=db_session.query(model).filter(FormARequirements.user_id == FormA.user_id).all()

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": supervisor_id,
            },
        ).scalar()
    )

    
    return render_template("supervisor-dashboard.html",     # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,supervisor_role=role,formA=formA,formB=formB,formC=formC,supervisor_formA=supervisor_formA,supervisor_formB=supervisor_formB,supervisor_formC=supervisor_formC)



@app.route('/supervisor_dashboard_previous_forms/<string:user_id>', methods=['GET','POST'])
@role_required('SUPERVISOR', 'REVIEWER', 'ADMIN', 'SUPER_ADMIN')
def supervisor_dashboard_previous_forms(user_id):
    supervisor_id=session.get('id')
    supervisor_role=session.get('supervisor_role')
    if not supervisor_id:
        return jsonify({'error': 'Unauthorized'}), 401

    current_user = get_current_user()
    current_role = str(role_value(current_user) or '').upper()
    student = db_session.query(User).filter_by(user_id=user_id).first()
    if not student:
        abort(404)
    if (
        current_role not in {'ADMIN', 'SUPER_ADMIN'}
        and getattr(student, 'supervisor_id', None) != current_user.user_id
    ):
        abort(403)
    
    formA = (
        db_session.query(FormA)
        .filter(FormA.user_id == user_id)
        .order_by(
            FormA.submitted_at.desc().nullslast(),
            FormA.created_at.desc().nullslast(),
        )
        .all()
    )
    
    # Safe FormB query - only load needed columns to avoid binary column issues
    formB_results = db_session.query(
        FormB.form_id, FormB.user_id, FormB.applicant_name, FormB.student_number,
        FormB.email, FormB.supervisor, FormB.supervisor_email, FormB.submitted_at,
        FormB.recommendation, FormB.supervisor_date, FormB.ethics_status,
        FormB.signature_date, FormB.review_supervisor_signature, FormB.review_date,
        FormB.review_supervisor_signature1, FormB.review_date1, FormB.created_at, FormB.declaration_date,
        FormB.status, FormB.review_form_status, FormB.rejected_or_accepted
    ).filter(FormB.user_id == user_id).order_by(
        FormB.submitted_at.desc().nullslast(),
        FormB.created_at.desc().nullslast(),
    ).all()
    
    # Convert tuples to proxy objects
    formB = []
    for result in formB_results:
        class FormBProxy:
            pass
        proxy = FormBProxy()
        proxy.form_id = result.form_id
        proxy.user_id = result.user_id
        proxy.applicant_name = result.applicant_name
        proxy.student_number = result.student_number
        proxy.email = result.email
        proxy.supervisor = result.supervisor
        proxy.supervisor_email = result.supervisor_email
        proxy.submitted_at = result.submitted_at
        proxy.recommendation = result.recommendation
        proxy.supervisor_date = result.supervisor_date
        proxy.ethics_status = result.ethics_status
        proxy.signature_date = result.signature_date
        proxy.review_supervisor_signature = result.review_supervisor_signature
        proxy.review_date = result.review_date
        proxy.review_supervisor_signature1 = result.review_supervisor_signature1
        proxy.review_date1 = result.review_date1
        proxy.created_at = result.created_at
        proxy.declaration_date = result.declaration_date
        proxy.status = result.status
        proxy.review_form_status = result.review_form_status
        proxy.rejected_or_accepted = result.rejected_or_accepted
        formB.append(proxy)
    
    formC = (
        db_session.query(FormC)
        .filter(FormC.user_id == user_id)
        .order_by(
            FormC.submission_date.desc().nullslast(),
            FormC.created_at.desc().nullslast(),
        )
        .all()
    )

    supervisor_formA = db_session.query(FormA, FormARequirements) \
        .join(User, FormA.user_id == User.user_id) \
        .outerjoin(FormARequirements, FormARequirements.user_id == FormA.user_id) \
        .filter(FormA.user_id==user_id) \
        .order_by(
            FormA.submitted_at.desc().nullslast(),
            FormA.created_at.desc().nullslast(),
            FormA.declaration_date.desc().nullslast(),
        ) \
        .all()
    
    # Safe FormB query with joins - only load needed columns to avoid binary column issues
    supervisor_formB_results = db_session.query(
        FormB.form_id, FormB.user_id, FormB.applicant_name, FormB.student_number,
        FormB.email, FormB.supervisor, FormB.supervisor_email, FormB.submitted_at,
        FormB.recommendation, FormB.supervisor_date, FormB.ethics_status,
        FormB.signature_date, FormB.review_supervisor_signature, FormB.review_date,
        FormB.review_supervisor_signature1, FormB.review_date1, FormB.created_at, FormB.declaration_date,
        FormB.status, FormB.review_form_status, FormB.rejected_or_accepted,
        FormARequirements
    ).join(User, FormB.user_id == User.user_id) \
    .outerjoin(FormARequirements, FormARequirements.user_id == FormB.user_id) \
    .filter(FormB.user_id==user_id) \
    .order_by(
        FormB.submitted_at.desc().nullslast(),
        FormB.created_at.desc().nullslast(),
        FormB.declaration_date.desc().nullslast(),
    ) \
    .all()
    
    # Convert to simplified format: (FormB-like proxy, FormARequirements)
    supervisor_formB = []
    for result in supervisor_formB_results:
        class FormBProxy:
            pass
        proxy = FormBProxy()
        proxy.form_id = result.form_id
        proxy.user_id = result.user_id
        proxy.applicant_name = result.applicant_name
        proxy.student_number = result.student_number
        proxy.email = result.email
        proxy.supervisor = result.supervisor
        proxy.supervisor_email = result.supervisor_email
        proxy.submitted_at = result.submitted_at
        proxy.recommendation = result.recommendation
        proxy.supervisor_date = result.supervisor_date
        proxy.ethics_status = result.ethics_status
        proxy.signature_date = result.signature_date
        proxy.review_supervisor_signature = result.review_supervisor_signature
        proxy.review_date = result.review_date
        proxy.review_supervisor_signature1 = result.review_supervisor_signature1
        proxy.review_date1 = result.review_date1
        proxy.created_at = result.created_at
        proxy.declaration_date = result.declaration_date
        proxy.status = result.status
        proxy.review_form_status = result.review_form_status
        proxy.rejected_or_accepted = result.rejected_or_accepted
        supervisor_formB.append((proxy, result.FormARequirements))
    
    supervisor_formC = db_session.query(FormC, FormARequirements) \
        .join(User, FormC.user_id == User.user_id) \
        .outerjoin(FormARequirements, FormARequirements.user_id == FormC.user_id) \
        .filter(FormC.user_id==user_id) \
        .order_by(
            FormC.submission_date.desc().nullslast(),
            FormC.created_at.desc().nullslast(),
        ) \
        .all()
    
    return render_template("supervisor_form_version_control.html",supervisor_role=supervisor_role,formA=formA,formB=formB,formC=formC,supervisor_formA=supervisor_formA,supervisor_formB=supervisor_formB,supervisor_formC=supervisor_formC)



@app.route('/dean_dashboard', methods=['GET','POST'])
def dean_dashboard():
    role="STUDENT"
    supervisor_formA = db_session.query(FormA).join(User, FormA.user_id == User.user_id).all()
    
    # Safe FormB query - only load needed columns to avoid binary column issues
    supervisor_formB_results = db_session.query(
        FormB.form_id, FormB.user_id, FormB.applicant_name, FormB.student_number,
        FormB.email, FormB.supervisor, FormB.supervisor_email, FormB.submitted_at,
        FormB.recommendation, FormB.supervisor_date, FormB.ethics_status,
        FormB.signature_date, FormB.review_supervisor_signature, FormB.review_date,
        FormB.review_supervisor_signature1, FormB.review_date1, FormB.created_at, FormB.declaration_date
    ).join(User, FormB.user_id == User.user_id).all()
    
    # Convert tuples to proxy objects
    supervisor_formB = []
    for result in supervisor_formB_results:
        class FormBProxy:
            pass
        proxy = FormBProxy()
        proxy.form_id = result.form_id
        proxy.user_id = result.user_id
        proxy.applicant_name = result.applicant_name
        proxy.student_number = result.student_number
        proxy.email = result.email
        proxy.supervisor = result.supervisor
        proxy.supervisor_email = result.supervisor_email
        proxy.submitted_at = result.submitted_at
        proxy.recommendation = result.recommendation
        proxy.supervisor_date = result.supervisor_date
        proxy.ethics_status = result.ethics_status
        proxy.signature_date = result.signature_date
        proxy.review_supervisor_signature = result.review_supervisor_signature
        proxy.review_date = result.review_date
        proxy.review_supervisor_signature1 = result.review_supervisor_signature1
        proxy.review_date1 = result.review_date1
        proxy.created_at = result.created_at
        proxy.declaration_date = result.declaration_date
        supervisor_formB.append(proxy)
    
    supervisor_formC = db_session.query(FormC).join(User, FormC.user_id == User.user_id).all()
    supervisor_formA_req=db_session.query(FormARequirements).filter(FormARequirements.user_id == User.user_id).all()
    students=db_session.query(User).filter(User.role==role).all()
    
    return render_template('dean.html',students=students,supervisor_formA_req=supervisor_formA_req,supervisor_formA=supervisor_formA,supervisor_formB=supervisor_formB,supervisor_formC=supervisor_formC)



@app.route('/supervisor_student', methods=['GET', 'POST'])
def supervisor_student ():
    supervisor_id=session.get('id')
    supervisor_data = (
    db_session.query(User)
    .options(
        joinedload(User.form_a),
        joinedload(User.form_b),
        joinedload(User.form_c),
        joinedload(User.form_a_requirements)
    )
    .filter(User.role == "STUDENT", User.supervisor_id == supervisor_id)
    .all()
        )


    return render_template('students.html',students=supervisor_data)

def validate_reset_token(token):
    user = User.query.filter_by(reset_token=token).first()
    if user and user.reset_token_expiry > datetime.utcnow():
        return user  # Token is valid
    return None  # Token is invalid/expired


# =====================================================================================================
# FILE DOWNLOAD ROUTES - PRODUCTION FIX
# =====================================================================================================

@app.route('/uploads/<path:filename>')
def download_file(filename):
    """
    Serve uploaded files with security checks.
    CRITICAL FIX: This route was missing, causing 404 errors for all uploaded attachments.
    """
    try:
        # Security: Only serve files from the uploads directory
        safe_path = secure_filename(filename)
        if not safe_path or safe_path != filename:
            flash("Invalid file path", "danger")
            return redirect(url_for('dashboard'))
        
        # Check if user is logged in
        if 'id' not in session:
            flash("You must be logged in to access files", "warning")
            return redirect(url_for('login'))
        
        # PRODUCTION FIX: Use the same upload directory logic as file saving
        upload_dir = get_upload_folder()
        
        # Try direct filename first
        file_path = os.path.join(upload_dir, safe_path)
        if not os.path.exists(file_path):
            # Try removing the uploads/form prefix if it exists in the filename
            if safe_path.startswith('uploads/form/'):
                clean_filename = safe_path.replace('uploads/form/', '')
                file_path = os.path.join(upload_dir, clean_filename)
            elif safe_path.startswith('form/'):
                clean_filename = safe_path.replace('form/', '')
                file_path = os.path.join(upload_dir, clean_filename)
        
        if not os.path.exists(file_path):
            print(f"⚠️ File not found: {file_path}")
            flash("File not found. It may have been deleted during deployment.", "danger")
            return redirect(url_for('dashboard'))
            
        # TODO: Add additional security - check if user has permission to access this specific file
        # This would require storing file ownership in database
        
        return send_from_directory(
            upload_dir,
            os.path.basename(file_path),
            as_attachment=True
        )
        
    except Exception as e:
        print(f"⚠️ File download error: {str(e)}")
        flash("Error accessing file", "danger")
        return redirect(url_for('dashboard'))


# =====================================================================================================
# EMAIL REMINDER SYSTEM
# =====================================================================================================

def get_students_with_missing_documents():
    """
    Get students who have submitted forms but have missing documents
    Returns list of (user, missing_files_info) tuples
    """
    students_with_missing_files = []
    
    # Get all users with FormARequirements (students who submitted forms)
    form_requirements = db_session.query(FormARequirements).join(User).all()
    
    for req in form_requirements:
        user = db_session.query(User).filter_by(user_id=req.user_id).first()
        if not user:
            continue
            
        # Check if files exist on filesystem
        upload_folder = get_upload_folder()
        missing_files = {}
        
        file_fields = [
            ('permission_letter', 'Permission Letters'),
            ('research_tools_path', 'Research Tools'),
            ('proposal_path', 'Research Proposal'),
            ('impact_assessment_path', 'Impact Assessment'),
            ('prior_clearance', 'Prior Clearance'),
            ('ethics_evidence_path', 'Ethics Evidence'),
            ('participation_info_sheet', 'Participant Info Sheet'),
            ('pending_note', 'Pending Note')
        ]
        
        for field_name, display_name in file_fields:
            file_data = getattr(req, field_name)
            if file_data:
                # If it's bytes or memoryview, it's a blob and definitely exists
                if isinstance(file_data, (bytes, memoryview)):
                    continue
                    
                # If it's a string, it's a legacy path, check if it exists on disk
                if isinstance(file_data, str):
                    if file_data.startswith('uploads/form/'):
                        clean_path = file_data.replace('uploads/form/', '')
                    elif file_data.startswith('form/'):
                        clean_path = file_data.replace('form/', '')
                    else:
                        clean_path = file_data
                    
                    full_path = os.path.join(upload_folder, clean_path)
                    if not os.path.exists(full_path):
                        missing_files[field_name] = display_name
        
        # Only add students who have missing files
        if missing_files:
            students_with_missing_files.append((user, missing_files))
    
    return students_with_missing_files

@app.route('/send_missing_documents_reminder', methods=['POST'])
def send_missing_documents_reminder():
    """
    Admin route to send email reminders to students with missing documents
    Only sends to students who have submitted forms but are missing files
    """
    try:
        # Check if user is logged in
        user_id = session.get('id')
        if not user_id:
            flash("Please log in to access this feature.", "warning")
            return redirect(url_for('login'))
        
        # Check if user is admin (you can modify this condition based on your admin check)
        user_role = session.get('role')
        if user_role not in ['admin', 'chair', 'dean']:  # Adjust roles as needed
            flash("Unauthorized access. Admin privileges required.", "danger")
            return redirect(url_for('dashboard'))
        
        students_with_missing_files = get_students_with_missing_documents()
        
        if not students_with_missing_files:
            flash("No students with missing documents found.", "info")
            return redirect(request.referrer or url_for('dashboard'))
        
        emails_sent = 0
        failed_emails = 0
        
        for user, missing_files in students_with_missing_files:
            try:
                # Create email message
                missing_files_list = ', '.join(missing_files.values())
                
                message_body = f"""
Dear {user.full_name},

We hope this email finds you well.

Our system has detected that some of your uploaded documents for your ethics application are currently missing from our servers. This may have occurred due to recent system maintenance.

To ensure your application can be processed without delays, please re-upload the missing documents by:

1. Logging into your student dashboard
2. Clicking on the "Reupload Documents" button if you see the missing files alert
3. Selecting your form type and uploading the required documents

If you experience any difficulties or have questions, please don't hesitate to contact our ethics committee.

Thank you for your attention to this matter.

Best regards,
Ethics Committee
"""

                # Send email using existing email function
                send_email(
                    app=current_app,
                    mail=mail,
                    message=message_body,
                    recipient=[user.email]
                )
                
                emails_sent += 1
                print(f"✅ Reminder email sent to {user.email}")
                
            except Exception as e:
                failed_emails += 1
                print(f"⚠️ Failed to send email to {user.email}: {str(e)}")
                continue
        
        # Flash success/failure message
        if emails_sent > 0:
            flash(f"Successfully sent reminder emails to {emails_sent} student(s).", "success")
        if failed_emails > 0:
            flash(f"Failed to send {failed_emails} email(s). Please check email configuration.", "warning")
            
        return redirect(request.referrer or url_for('dashboard'))
        
    except Exception as e:
        print(f"⚠️ Error sending reminder emails: {str(e)}")
        flash("Error sending reminder emails. Please try again later.", "danger")
        return redirect(request.referrer or url_for('dashboard'))

@app.route('/api/missing_documents_count', methods=['GET'])
def get_missing_documents_count():
    """
    API endpoint to get count of students with missing documents
    """
    try:
        # Check if user is logged in
        user_id = session.get('id')
        if not user_id:
            return jsonify({'error': 'Authentication required'}), 401
        
        # Check if user is admin
        user_role = session.get('role')
        if user_role not in ['admin', 'chair', 'dean']:
            return jsonify({'error': 'Admin privileges required'}), 403
        
        students_with_missing_files = get_students_with_missing_documents()
        
        return jsonify({
            'count': len(students_with_missing_files),
            'students': [
                {
                    'name': user.full_name,
                    'email': user.email,
                    'missing_files': list(missing_files.values())
                } 
                for user, missing_files in students_with_missing_files
            ]
        })
        
    except Exception as e:
        print(f"⚠️ Error getting missing documents count: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

# =====================================================================================================
# END OF FORMS
# =====================================================================================================


UPLOAD_FOLDER = "/opt/render/project/src/static/uploads/form"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route("/upload_to", methods=["POST"])
def upload_to():
    file = request.files["file"]
    file.save(os.path.join(UPLOAD_FOLDER, file.filename))
    return "File uploaded!"


@app.route('/export_forms_csv', methods=['GET', 'POST'])
def export_forms_csv():
    """
    Export specific FormA, FormB, and FormC data to an Excel file with formatting
    Only accessible to ADMIN and SUPER_ADMIN roles
    Groups by user_id with alternating colors and styled headers
    Filters by selected year if provided
    """
    print("=== EXPORT FUNCTION CALLED ===")
    
    # Check if user is logged in and is admin/super_admin
    user_id = session.get('id')
    print(f"User ID from session: {user_id}")
    
    if not user_id:
        flash('Please login to access this feature', 'error')
        return redirect(url_for('login_page'))
    
    user = db_session.query(User).filter(User.user_id == user_id).first()
    print(f"User found: {user.full_name if user else 'None'}")
    print(f"User role: {user.role.value if user and user.role else 'None'}")
    
    if not user or not user.role or user.role.value.upper() not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You do not have permission to export forms', 'error')
        return redirect(url_for('chair_landing'))
    
    # Get selected year from form (if provided)
    selected_year = request.form.get('year') if request.method == 'POST' else request.args.get('year')
    print(f"Selected year: {selected_year}")
    
    try:
        print("Starting export process...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        print("openpyxl imported successfully")
        
        # Define the fields to extract
        fields_mapping = {
            'form_type': 'Form Type',
            'applicant_name': 'Applicant Name',
            'student_number': 'Student Number',
            'email': 'Email',
            'supervisor': 'Supervisor',
            'supervisor_email': 'Supervisor Email',
            'student_submission': 'Student Submission Date',
            'recommendation': 'Supervisors Recommendation',
            'supervisor_date': 'Supervisor Recommendation Date',
            'ethics_status': 'Ethics Admin Decision',
            'signature_date': 'Ethics Admin Decision Date',
            'review_supervisor_signature': 'First Reviewer Name',
            'review_recommendation': 'First Reviewer Recommendation',
            'review_date': 'First Reviewer Feedback Date',
            'review_supervisor_signature1': 'Second Reviewer Name',
            'review_recommendation1': 'Second Reviewer Recommendation',
            'review_date1': 'Second Reviewer Feedback Date'
        }
        
        # Query all submitted forms ordered by submission date
        print("Querying forms...")
        
        # FormA query with optional year filter
        form_a_query = db_session.query(FormA).filter(FormA.submitted_at.isnot(None))
        if selected_year and selected_year != 'all':
            form_a_query = form_a_query.filter(extract('year', FormA.submitted_at) == int(selected_year))
        form_a_records = form_a_query.order_by(FormA.submitted_at.desc()).all()
        
        # FormB query with optional year filter - only load needed columns to avoid binary column deserialization issues
        form_b_query = db_session.query(
            FormB.user_id,
            FormB.applicant_name,
            FormB.student_number,
            FormB.email,
            FormB.supervisor,
            FormB.supervisor_email,
            FormB.submitted_at,
            FormB.recommendation,
            FormB.supervisor_date,
            FormB.ethics_status,
            FormB.signature_date,
            FormB.review_supervisor_signature,
            FormB.review_recommendation,
            FormB.review_date,
            FormB.review_supervisor_signature1,
            FormB.review_recommendation1,
            FormB.review_date1
        ).filter(FormB.submitted_at.isnot(None))
        if selected_year and selected_year != 'all':
            form_b_query = form_b_query.filter(extract('year', FormB.submitted_at) == int(selected_year))
        form_b_results = form_b_query.order_by(FormB.submitted_at.desc()).all()
        
        # Convert tuples to objects with attributes for consistent processing
        form_b_records = []
        for result in form_b_results:
            class FormBProxy:
                pass
            proxy = FormBProxy()
            proxy.user_id = result.user_id
            proxy.applicant_name = result.applicant_name
            proxy.student_number = result.student_number
            proxy.email = result.email
            proxy.supervisor = result.supervisor
            proxy.supervisor_email = result.supervisor_email
            proxy.submitted_at = result.submitted_at
            proxy.recommendation = result.recommendation
            proxy.supervisor_date = result.supervisor_date
            proxy.ethics_status = result.ethics_status
            proxy.signature_date = result.signature_date
            proxy.review_supervisor_signature = result.review_supervisor_signature
            proxy.review_recommendation = result.review_recommendation
            proxy.review_date = result.review_date
            proxy.review_supervisor_signature1 = result.review_supervisor_signature1
            proxy.review_recommendation1 = result.review_recommendation1
            proxy.review_date1 = result.review_date1
            form_b_records.append(proxy)
        
        # FormC query with optional year filter
        form_c_query = db_session.query(FormC).filter(FormC.submission_date.isnot(None))
        if selected_year and selected_year != 'all':
            form_c_query = form_c_query.filter(extract('year', FormC.submission_date) == int(selected_year))
        form_c_records = form_c_query.order_by(FormC.submission_date.desc()).all()
        
       
        
        # Helper function to remove timezone from datetime
        def remove_tz(dt):
            """Remove timezone info from datetime object for Excel compatibility"""
            if dt and hasattr(dt, 'replace'):
                return dt.replace(tzinfo=None)
            return dt
        
        # Collect data
        all_data = []
        
        for record in form_a_records:
            all_data.append({
                'form_type': 'Form A',
                'user_id': record.user_id or '',
                'applicant_name': record.applicant_name or '',
                'student_number': record.student_number or '',
                'email': record.email or '',
                'supervisor': record.supervisor or '',
                'supervisor_email': record.supervisor_email or '',
                'student_submission': remove_tz(record.submitted_at) if record.submitted_at else '',
                'recommendation': record.recommendation or '',
                'supervisor_date': remove_tz(record.supervisor_date) if record.supervisor_date else '',
                'ethics_status': record.ethics_status or '',
                'signature_date': remove_tz(record.signature_date) if record.signature_date else '',
                'review_supervisor_signature': record.review_supervisor_signature or '',
                'review_recommendation': record.review_recommendation or '',
                'review_date': record.review_date or '',
                'review_supervisor_signature1': record.review_supervisor_signature1 or '',
                'review_recommendation1': record.review_recommendation1 or '',
                'review_date1': record.review_date1 or ''
            })
        
        for record in form_b_records:
            all_data.append({
                'form_type': 'Form B',
                'user_id': record.user_id or '',
                'applicant_name': record.applicant_name or '',
                'student_number': record.student_number or '',
                'email': record.email or '',
                'supervisor': record.supervisor or '',
                'supervisor_email': record.supervisor_email or '',
                'student_submission': remove_tz(record.submitted_at) if record.submitted_at else '',
                'recommendation': record.recommendation or '',
                'supervisor_date': remove_tz(record.supervisor_date) if record.supervisor_date else '',
                'ethics_status': record.ethics_status or '',
                'signature_date': remove_tz(record.signature_date) if record.signature_date else '',
                'review_supervisor_signature': record.review_supervisor_signature or '',
                'review_recommendation': record.review_recommendation or '',
                'review_date': record.review_date or '',
                'review_supervisor_signature1': record.review_supervisor_signature1 or '',
                'review_recommendation1': record.review_recommendation1 or '',
                'review_date1': record.review_date1 or ''
            })
        
        for record in form_c_records:
            all_data.append({
                'form_type': 'Form C',
                'user_id': record.user_id or '',
                'applicant_name': record.applicant_name or '',
                'student_number': record.student_number or '',
                'email': record.email_address or '',
                'supervisor': record.supervisor_name or '',
                'supervisor_email': record.supervisor_email or '',
                'student_submission': remove_tz(record.submission_date) if record.submission_date else '',
                'recommendation': record.recommendation or '',
                'supervisor_date': remove_tz(record.supervisor_date) if record.supervisor_date else '',
                'ethics_status': record.ethics_status or '',
                'signature_date': remove_tz(record.signature_date) if record.signature_date else '',
                'review_supervisor_signature': record.review_supervisor_signature or '',
                'review_recommendation': record.review_recommendation or '',
                'review_date': record.review_date or '',
                'review_supervisor_signature1': record.review_supervisor_signature1 or '',
                'review_recommendation1': record.review_recommendation1 or '',
                'review_date1': record.review_date1 or ''
            })
        
        print(f"Collected {len(all_data)} total records")
        
        # Create DataFrame
        print("Creating DataFrame...")
        df = pd.DataFrame(all_data)
        if df.empty:
            print("DataFrame is empty!")
            flash('No forms found to export', 'warning')
            return redirect(url_for('chair_landing'))
        
        print(f"DataFrame created with {len(df)} rows")
        
        # Group by user_id and sort by submission date within each group
        print("Grouping by user_id...")
        df = df.sort_values(by=['user_id', 'student_submission'], ascending=[True, False], na_position='last')
        
        print(f"Data grouped by user_id, total groups: {df['user_id'].nunique()}")
        
        # Rename columns (but keep user_id for grouping logic)
        df = df.rename(columns=fields_mapping)
        
        print("Creating Excel workbook...")
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Forms Export"
        
        # Write headers with styling
        headers = list(fields_mapping.values())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Define alternating colors for different user groups
        colors = [
            "E8F4F8",  # Light blue
            "FFF4E6",  # Light orange
            "F0F8E8",  # Light green
            "F8E8F4",  # Light purple
            "FFF0F0",  # Light red
            "F0F0F0",  # Light gray
        ]
        
        # Write data rows with alternating colors per user_id
        current_user_id = None
        color_index = 0
        row_num = 2
        
        for index, row in df.iterrows():
            # Check if we're starting a new user group
            if row['user_id'] != current_user_id:
                current_user_id = row['user_id']
                color_index = (color_index + 1) % len(colors)
            
            # Get the fill color for this user group
            fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
            
            # Write the row data (excluding the user_id column)
            col_num = 1
            for col_name, value in row.items():
                if col_name != 'user_id':  # Skip user_id (don't expose primary key)
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    col_num += 1
            
            row_num += 1
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        print("Saving to BytesIO...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'ethics_forms_export_{timestamp}.xlsx'
        
        print(f"Export complete! Filename: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError as e:
        print(f"Import error - openpyxl not installed: {str(e)}")
        traceback.print_exc()
        flash('Excel export library not installed. Please contact administrator.', 'error')
        return redirect(url_for('chair_landing'))
    except Exception as e:
        print(f"Error exporting forms: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        traceback.print_exc()
        flash(f'Error exporting forms: {str(e)}', 'error')
        return redirect(url_for('chair_landing'))


# Export agenda with reviewer counts per student
@app.route('/export_agenda_csv', methods=['GET', 'POST'])
def export_agenda_csv():
    """
    Export Excel with columns: Student Name, Reviewer 1, Reviewer 2, Count by Reviewer 1, Count by Reviewer 2
    For all FormA, FormB, FormC submissions with reviewer(s).
    """
    import collections
    user_id = session.get('id')
    if not user_id:
        flash('Please login to access this feature', 'error')
        return redirect(url_for('login_page'))
    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user or not user.role or user.role.value.upper() not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You do not have permission to export agenda', 'error')
        return redirect(url_for('chair_landing'))

    # Use only the latest submitted version per student/form type.
    latest_form_a_subq = (
        db_session.query(
            FormA.user_id,
            func.max(FormA.submitted_at).label('latest_submitted_at')
        )
        .filter(FormA.submitted_at.isnot(None))
        .group_by(FormA.user_id)
        .subquery()
    )
    latest_form_b_subq = (
        db_session.query(
            FormB.user_id,
            func.max(FormB.submitted_at).label('latest_submitted_at')
        )
        .filter(FormB.submitted_at.isnot(None))
        .group_by(FormB.user_id)
        .subquery()
    )
    latest_form_c_subq = (
        db_session.query(
            FormC.user_id,
            func.max(FormC.submission_date).label('latest_submission_date')
        )
        .filter(FormC.submission_date.isnot(None))
        .group_by(FormC.user_id)
        .subquery()
    )

    form_a_records = (
        db_session.query(FormA)
        .join(
            latest_form_a_subq,
            and_(
                FormA.user_id == latest_form_a_subq.c.user_id,
                FormA.submitted_at == latest_form_a_subq.c.latest_submitted_at
            )
        )
        .filter(FormA.certificate_issued.is_(None))
        .all()
    )
    form_b_records = (
        db_session.query(FormB)
        .join(
            latest_form_b_subq,
            and_(
                FormB.user_id == latest_form_b_subq.c.user_id,
                FormB.submitted_at == latest_form_b_subq.c.latest_submitted_at
            )
        )
        .filter(FormB.certificate_issued.is_(None))
        .all()
    )
    form_c_records = (
        db_session.query(FormC)
        .join(
            latest_form_c_subq,
            and_(
                FormC.user_id == latest_form_c_subq.c.user_id,
                FormC.submission_date == latest_form_c_subq.c.latest_submission_date
            )
        )
        .filter(FormC.certificate_issued.is_(None))
        .all()
    )


    # Group by student name, aggregate reviewers and recommendations
    student_rows = {}

    def process_form(record, name_field):
        student_name = getattr(record, name_field, None) or ''
        reviewer1_id = getattr(record, 'reviewer_name1', None) or ''
        reviewer2_id = getattr(record, 'reviewer_name2', None) or ''
        reviewer1 = ''
        reviewer2 = ''
        reviewer1_recommendation = getattr(record, 'review_recommendation', None) or ''
        reviewer2_recommendation = getattr(record, 'review_recommendation1', None) or ''
        # Lookup reviewer names from user_id (users table)
        if reviewer1_id:
            reviewer1_user = db_session.query(User).filter(User.user_id == reviewer1_id).first()
            reviewer1 = reviewer1_user.full_name if reviewer1_user else reviewer1_id
        if reviewer2_id:
            reviewer2_user = db_session.query(User).filter(User.user_id == reviewer2_id).first()
            reviewer2 = reviewer2_user.full_name if reviewer2_user else reviewer2_id
        # Group by student name
        if student_name not in student_rows:
            student_rows[student_name] = {
                'Student Name': student_name,
                'Reviewer 1': reviewer1,
                'Reviewer 1 Recommendation': reviewer1_recommendation,
                'Reviewer 2': reviewer2,
                'Reviewer 2 Recommendation': reviewer2_recommendation,
                'Count by Reviewer 1': 1 if reviewer1_recommendation else 0,
                'Count by Reviewer 2': 1 if reviewer2_recommendation else 0,
            }
        else:
            # If already exists, fill in missing reviewer/recommendation if empty
            row = student_rows[student_name]
            if not row['Reviewer 1'] and reviewer1:
                row['Reviewer 1'] = reviewer1
                row['Reviewer 1 Recommendation'] = reviewer1_recommendation
                row['Count by Reviewer 1'] = 1 if reviewer1_recommendation else 0
            if not row['Reviewer 2'] and reviewer2:
                row['Reviewer 2'] = reviewer2
                row['Reviewer 2 Recommendation'] = reviewer2_recommendation
                row['Count by Reviewer 2'] = 1 if reviewer2_recommendation else 0

    # FormA
    for rec in form_a_records:
        process_form(rec, 'applicant_name')
    # FormB
    for rec in form_b_records:
        process_form(rec, 'applicant_name')
    # FormC
    for rec in form_c_records:
        process_form(rec, 'applicant_name')

    rows = list(student_rows.values())

    if not rows:
        flash('No agenda data found to export', 'warning')
        return redirect(url_for('chair_landing'))

    df = pd.DataFrame(rows)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda Export"

    # Write headers
    headers = list(df.columns)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ethics_agenda_export_{timestamp}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/export_reviewer_assignments_csv', methods=['GET', 'POST'])
def export_reviewer_assignments_csv():
    """
    Export reviewer assignments to an Excel workbook.
    Sheet 1 lists all reviewers with total assignment counts.
    Sheet 2 lists every reviewer-to-student assignment.
    Uses alternating colors per reviewer, similar to the forms export.
    """
    user_id = session.get('id')
    if not user_id:
        flash('Please login to access this feature', 'error')
        return redirect(url_for('login_page'))

    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user or not user.role or user.role.value.upper() not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You do not have permission to export reviewer assignments', 'error')
        return redirect(url_for('chair_landing'))

    selected_year = request.form.get('year') if request.method == 'POST' else request.args.get('year')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        def remove_tz(dt):
            if dt and hasattr(dt, 'replace'):
                return dt.replace(tzinfo=None)
            return dt

        # Pull all reviewers so the summary sheet includes reviewers with zero assignments.
        reviewers = (
            db_session.query(User)
            .filter(User.role == UserRole.REVIEWER)
            .order_by(User.full_name.asc())
            .all()
        )
        reviewer_lookup = {str(reviewer.user_id): reviewer for reviewer in reviewers}

        assignment_rows = []
        assignment_counts = {str(reviewer.user_id): 0 for reviewer in reviewers}
        reviewer_students = {str(reviewer.user_id): [] for reviewer in reviewers}

        def process_records(records, form_type, name_attr, student_number_attr, email_attr, submitted_attr):
            for record in records:
                if getattr(record, 'certificate_issued', None):
                    continue

                student_name = getattr(record, name_attr, '') or ''
                student_number = getattr(record, student_number_attr, '') or ''
                student_email = getattr(record, email_attr, '') or ''
                submitted_date = remove_tz(getattr(record, submitted_attr, None))

                for slot_index, reviewer_attr in enumerate(['reviewer_name1', 'reviewer_name2'], start=1):
                    reviewer_id = getattr(record, reviewer_attr, None)
                    if not reviewer_id:
                        continue

                    reviewer_key = str(reviewer_id)
                    reviewer_user = reviewer_lookup.get(reviewer_key)
                    reviewer_name = reviewer_user.full_name if reviewer_user else str(reviewer_id)
                    reviewer_email = reviewer_user.email if reviewer_user else ''

                    assignment_counts[reviewer_key] = assignment_counts.get(reviewer_key, 0) + 1
                    if student_name:
                        reviewer_students.setdefault(reviewer_key, []).append(student_name)
                    assignment_rows.append({
                        'reviewer_id': reviewer_key,
                        'Reviewer Name': reviewer_name,
                        'Reviewer Email': reviewer_email,
                        'Form Type': form_type,
                        'Student Name': student_name,
                        'Student Number': student_number,
                        'Student Email': student_email,
                        'Assignment Slot': f'Reviewer {slot_index}',
                        'Submission Date': submitted_date or '',
                    })

        latest_form_a_subq = (
            db_session.query(
                FormA.user_id,
                func.max(FormA.submitted_at).label('latest_submitted_at')
            )
            .filter(FormA.submitted_at.isnot(None))
            .group_by(FormA.user_id)
            .subquery()
        )
        form_a_query = (
            db_session.query(FormA)
            .join(
                latest_form_a_subq,
                and_(
                    FormA.user_id == latest_form_a_subq.c.user_id,
                    FormA.submitted_at == latest_form_a_subq.c.latest_submitted_at
                )
            )
            .filter(FormA.submitted_at.isnot(None))
        )
        if selected_year and selected_year != 'all':
            form_a_query = form_a_query.filter(extract('year', FormA.submitted_at) == int(selected_year))
        form_a_records = form_a_query.order_by(FormA.submitted_at.desc()).all()

        latest_form_b_subq = (
            db_session.query(
                FormB.user_id,
                func.max(FormB.submitted_at).label('latest_submitted_at')
            )
            .filter(FormB.submitted_at.isnot(None))
            .group_by(FormB.user_id)
            .subquery()
        )
        form_b_query = (
            db_session.query(
                FormB.reviewer_name1,
                FormB.reviewer_name2,
                FormB.applicant_name,
                FormB.student_number,
                FormB.email,
                FormB.submitted_at,
                FormB.certificate_issued
            )
            .join(
                latest_form_b_subq,
                and_(
                    FormB.user_id == latest_form_b_subq.c.user_id,
                    FormB.submitted_at == latest_form_b_subq.c.latest_submitted_at
                )
            )
            .filter(FormB.submitted_at.isnot(None))
        )
        if selected_year and selected_year != 'all':
            form_b_query = form_b_query.filter(extract('year', FormB.submitted_at) == int(selected_year))
        form_b_records = form_b_query.order_by(FormB.submitted_at.desc()).all()

        latest_form_c_subq = (
            db_session.query(
                FormC.user_id,
                func.max(FormC.submission_date).label('latest_submission_date')
            )
            .filter(FormC.submission_date.isnot(None))
            .group_by(FormC.user_id)
            .subquery()
        )
        form_c_query = (
            db_session.query(
                FormC.reviewer_name1,
                FormC.reviewer_name2,
                FormC.applicant_name,
                FormC.student_number,
                FormC.email_address,
                FormC.submission_date,
                FormC.certificate_issued
            )
            .join(
                latest_form_c_subq,
                and_(
                    FormC.user_id == latest_form_c_subq.c.user_id,
                    FormC.submission_date == latest_form_c_subq.c.latest_submission_date
                )
            )
            .filter(FormC.submission_date.isnot(None))
        )
        if selected_year and selected_year != 'all':
            form_c_query = form_c_query.filter(extract('year', FormC.submission_date) == int(selected_year))
        form_c_records = form_c_query.order_by(FormC.submission_date.desc()).all()

        process_records(form_a_records, 'Form A', 'applicant_name', 'student_number', 'email', 'submitted_at')
        process_records(form_b_records, 'Form B', 'applicant_name', 'student_number', 'email', 'submitted_at')
        process_records(form_c_records, 'Form C', 'applicant_name', 'student_number', 'email_address', 'submission_date')

        summary_rows = []
        for reviewer in reviewers:
            reviewer_key = str(reviewer.user_id)
            unique_students = sorted(set(reviewer_students.get(reviewer_key, [])))
            summary_rows.append({
                'reviewer_id': reviewer_key,
                'Reviewer Name': reviewer.full_name or '',
                'Reviewer Email': reviewer.email or '',
                'Total Assignments': len(unique_students),
                'Assigned Students': ', '.join(unique_students),
            })

        if not summary_rows:
            flash('No reviewer data found to export', 'warning')
            return redirect(url_for('chair_landing'))

        summary_df = pd.DataFrame(summary_rows).sort_values(
            by=['Reviewer Name'],
            ascending=[True],
            na_position='last'
        )
        assignments_df = pd.DataFrame(assignment_rows)
        if not assignments_df.empty:
            assignments_df = assignments_df.sort_values(
                by=['Reviewer Name', 'Submission Date', 'Student Name'],
                ascending=[True, False, True],
                na_position='last'
            )

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Reviewer Summary"
        detail_ws = wb.create_sheet(title="Assignments")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        colors = [
            "E8F4F8",
            "FFF4E6",
            "F0F8E8",
            "F8E8F4",
            "FFF0F0",
            "F0F0F0",
        ]

        reviewer_ids_in_order = summary_df['reviewer_id'].tolist()
        reviewer_colors = {
            reviewer_id: colors[index % len(colors)]
            for index, reviewer_id in enumerate(reviewer_ids_in_order)
        }

        def write_sheet(ws, df, visible_columns):
            for col_num, header in enumerate(visible_columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_num, (_, row) in enumerate(df.iterrows(), start=2):
                reviewer_id = row.get('reviewer_id', '')
                fill_color = reviewer_colors.get(reviewer_id, colors[0])
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                for col_num, column_name in enumerate(visible_columns, start=1):
                    value = row.get(column_name, '')
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = fill
                    cell.alignment = body_alignment

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        write_sheet(
            summary_ws,
            summary_df,
            ['Reviewer Name', 'Reviewer Email', 'Total Assignments', 'Assigned Students']
        )

        if assignments_df.empty:
            assignments_df = pd.DataFrame([{
                'reviewer_id': '',
                'Reviewer Name': 'No assignments found',
                'Reviewer Email': '',
                'Total Assignments': '',
                'Form Type': '',
                'Student Name': '',
                'Student Number': '',
                'Student Email': '',
                'Assignment Slot': '',
                'Submission Date': '',
            }])

        write_sheet(
            detail_ws,
            assignments_df,
            [
                'Reviewer Name',
                'Reviewer Email',
                'Form Type',
                'Student Name',
                'Student Number',
                'Student Email',
                'Assignment Slot',
                'Submission Date',
            ]
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'reviewer_assignments_export_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError as e:
        print(f"Import error - openpyxl not installed: {str(e)}")
        traceback.print_exc()
        flash('Excel export library not installed. Please contact administrator.', 'error')
        return redirect(url_for('chair_landing'))
    except Exception as e:
        print(f"Error exporting reviewer assignments: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        traceback.print_exc()
        flash(f'Error exporting reviewer assignments: {str(e)}', 'error')
        return redirect(url_for('chair_landing'))












# Automated Power Bi Configurations


from ROUTES.bi_form_mapping import (
    get_bi_landing_page_context,
    get_bi_configuration_rights_context,
    save_bi_configuration_and_view_rights,
    attach_bi_access_rights_to_records,
    ensure_bi_configuration_rights_table,
)


_bi_rights_schema_ready = False


@app.before_request
def ensure_bi_rights_schema_is_ready():
    """Ensure databases upgraded from older releases can serve BI pages."""
    global _bi_rights_schema_ready
    if not _bi_rights_schema_ready:
        _bi_rights_schema_ready = ensure_bi_configuration_rights_table(db_session)


@app.route(
    '/power/bi/and/reporting',
    methods=[
        'GET',
        'POST',
    ],
)
@role_required(
    'ADMIN',
    'SUPER_ADMIN',
    'REVIEWER',
)
def power_bi_and_reporting():
    user_id = session.get('id')

    user = (
        db_session.query(User)
        .filter(
            User.user_id == user_id
        )
        .first()
    )

    if not user:
        return redirect(
            url_for('login_page')
        )

    role = (
        user.role.value
        if user.role
        else None
    )

    # ==========================================================
    # SAVE CONFIGURATION AND VIEWING RIGHTS
    # ==========================================================

    if request.method == 'POST':
        payload = request.get_json(
            silent=True,
        ) or {}

        users_payload = payload.get(
            'users',
            [],
        )

        updated_by = (
            user.full_name
            or user.email
            or str(user.user_id)
        )

        save_result = (
            save_bi_configuration_and_view_rights(
                db_session=db_session,
                users=users_payload,
                updated_by=updated_by,
            )
        )

        response_status = (
            200
            if save_result.get('success')
            else 400
        )

        return jsonify(
            save_result
        ), response_status

    # ==========================================================
    # MAIN BI TABLE FILTER VALUES
    # ==========================================================

    search_text = (
        request.args.get(
            'search',
            '',
            type=str,
        )
        or ''
    ).strip()

    selected_view = (
        request.args.get(
            'bi_view_name',
            '',
            type=str,
        )
        or ''
    ).strip()

    selected_table = (
        request.args.get(
            'database_table',
            '',
            type=str,
        )
        or ''
    ).strip()

    selected_status = (
        request.args.get(
            'status',
            '',
            type=str,
        )
        or ''
    ).strip()

    # ==========================================================
    # MAIN BI TABLE PAGINATION VALUES
    # ==========================================================

    page = request.args.get(
        'page',
        default=1,
        type=int,
    )

    per_page = request.args.get(
        'per_page',
        default=10,
        type=int,
    )

    if page is None or page < 1:
        page = 1

    if per_page not in {
        10,
        25,
        50,
        100,
    }:
        per_page = 10

    # ==========================================================
    # BUILD MAIN BI TABLE CONTEXT
    # ==========================================================

    bi_context = get_bi_landing_page_context(
        page=page,
        per_page=per_page,
        search_text=search_text,
        selected_view=selected_view,
        selected_table=selected_table,
        selected_status=selected_status,
    )

    bi_context['bi_records'] = (
        attach_bi_access_rights_to_records(
            db_session=db_session,
            bi_records=bi_context.get(
                'bi_records',
                [],
            ),
            user_id=user_id,
        )
    )

    # Admins are the bootstrap owners of BI settings. They must be able to
    # configure reports and assign rights even before a rights row exists.
    if role in {'ADMIN', 'SUPER_ADMIN'}:
        for bi_record in bi_context['bi_records']:
            bi_record['has_bi_config_rights'] = True
            bi_record['has_bi_view_rights'] = True

    # ==========================================================
    # BI RIGHTS FILTER VALUES
    # ==========================================================

    bi_rights_search_text = (
        request.args.get(
            'bi_rights_search',
            '',
            type=str,
        )
        or ''
    ).strip()

    # ==========================================================
    # BI RIGHTS PAGINATION VALUES
    # ==========================================================

    bi_rights_page = request.args.get(
        'bi_rights_page',
        default=1,
        type=int,
    )

    bi_rights_per_page = request.args.get(
        'bi_rights_per_page',
        default=10,
        type=int,
    )

    if (
        bi_rights_page is None
        or bi_rights_page < 1
    ):
        bi_rights_page = 1

    if bi_rights_per_page not in {
        10,
        25,
        50,
        100,
    }:
        bi_rights_per_page = 10

    # ==========================================================
    # BUILD BI RIGHTS CONTEXT
    # ==========================================================

    bi_rights_context = (
        get_bi_configuration_rights_context(
            db_session=db_session,
            page=bi_rights_page,
            per_page=bi_rights_per_page,
            search_text=bi_rights_search_text,
        )
    )

    # ==========================================================
    # RENDER TEMPLATE
    # ==========================================================


    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )




    return render_template(
        'power_bi_landing_home.html',
        role=role,
        **bi_context,
        **bi_rights_context,


        # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )


# Configure BI Template
# Add this import with your other app.py imports:
from ROUTES.bi_form_mapping_configure_bi_template import (
    build_configure_bi_template_context,
    delete_bi_configuration_page,
    save_bi_configuration,
)

import json
@app.route(
    "/power/bi/and/reporting/configure/bi/template",
    methods=["GET", "POST"],
)
@role_required("ADMIN", "SUPER_ADMIN", 'REVIEWER',)
def power_bi_and_reporting_configure_bi_template():
    user_id = session.get("id")

    user = (
        db_session.query(User)
        .filter(User.user_id == user_id)
        .first()
    )

    if not user:
        if request.method == "POST":
            return jsonify({
                "success": False,
                "message": "Your session has expired. Please log in again.",
            }), 401

        return redirect(url_for("login_page"))

    role = user.role.value if user.role else None

    bi_view_name = (
        request.args.get("bi_view_name")
        or (
            request.get_json(silent=True) or {}
        ).get("bi_view_name")
        or ""
    ).strip()

    database_table = (
        request.args.get("database_table")
        or (
            request.get_json(silent=True) or {}
        ).get("database_table")
        or ""
    ).strip()

    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        action = str(payload.get("action") or "save_all").strip().lower()

        configured_by = (
            getattr(user, "email", None)
            or getattr(user, "full_name", None)
            or str(user_id)
        )

        try:
            if action == "delete_page":
                result = delete_bi_configuration_page(
                    db_session,
                    bi_view_name=bi_view_name,
                    database_table=database_table,
                    page_name=payload.get("page_name") or "",
                    configured_by=configured_by,
                )
            else:
                result = save_bi_configuration(
                    db_session,
                    bi_view_name=bi_view_name,
                    database_table=database_table,
                    dashboard_pages=payload.get("dashboard_pages") or {},
                    configured_by=configured_by,
                    config_status=payload.get("config_status") or "Draft",
                )

            return jsonify(result), 200 if result.get("success") else 400

        except ValueError as error:
            db_session.rollback()
            return jsonify({
                "success": False,
                "message": str(error),
            }), 400

        except SQLAlchemyError as error:
            db_session.rollback()
            app.logger.exception(
                "Database error while saving BI configuration"
            )
            return jsonify({
                "success": False,
                "message": "A database error prevented the BI configuration from being saved.",
                "error": str(error),
            }), 500

        except Exception as error:
            db_session.rollback()
            app.logger.exception(
                "Unexpected error while saving BI configuration"
            )
            return jsonify({
                "success": False,
                "message": "The BI configuration could not be saved.",
                "error": str(error),
            }), 500

    if not bi_view_name or not database_table:
        flash(
            "Select a BI view from the BI Dashboard and Reporting page first.",
            "warning",
        )
        return redirect(url_for("power_bi_and_reporting"))

    try:
        configure_context = build_configure_bi_template_context(
            db_session,
            bi_view_name=bi_view_name,
            database_table=database_table,
            preview_limit=250,
        )

    except ValueError as error:
        flash(str(error), "danger")
        return redirect(url_for("power_bi_and_reporting"))

    except SQLAlchemyError as error:
        db_session.rollback()
        app.logger.exception(
            "Database error while opening BI configuration"
        )
        flash(
            "The database table or its columns could not be loaded.",
            "danger",
        )
        return redirect(url_for("power_bi_and_reporting"))

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )

    return render_template(
        "power_bi_landing_home_configure_template.html",
        role=role,
        **configure_context,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )





# View BI Report
from ROUTES.bi_form_mapping_viewbi_report import (
    get_bi_form_mapping_view_report_context,
)


@app.route(
    "/power/bi/and/reporting/view/bi/report",
    methods=["GET"],
)
@role_required("ADMIN", "SUPER_ADMIN", 'REVIEWER',)
def power_bi_and_reporting_view_bi_report():
    user_id = session.get("id")

    user = (
        db_session.query(User)
        .filter(User.user_id == user_id)
        .first()
    )

    if not user:
        return redirect(url_for("login_page"))

    role = user.role.value if user.role else None

    bi_view_name = (
        request.args.get("bi_view_name", "", type=str)
        or ""
    ).strip()

    database_table = (
        request.args.get("database_table", "", type=str)
        or ""
    ).strip()

    if not bi_view_name or not database_table:
        flash(
            "Select a BI report from the BI Dashboard and Reporting page.",
            "warning",
        )
        return redirect(url_for("power_bi_and_reporting"))

    raw_connection = None

    try:
        raw_connection = db_session.get_bind().raw_connection()

        report_context = get_bi_form_mapping_view_report_context(
            raw_connection,
            bi_view_name=bi_view_name,
            database_table=database_table,
            filters=request.args,
            execute_visuals=True,
        )

    except ValueError as error:
        if raw_connection is not None:
            raw_connection.rollback()

        app.logger.exception(
            "Invalid BI report configuration: %s",
            error,
        )

        flash(str(error), "danger")
        return redirect(url_for("power_bi_and_reporting"))

    except Exception as error:
        if raw_connection is not None:
            raw_connection.rollback()

        app.logger.exception(
            "Unable to render BI report: %s",
            error,
        )

        flash(
            "The BI report could not be loaded.",
            "danger",
        )
        return redirect(url_for("power_bi_and_reporting"))

    finally:
        if raw_connection is not None:
            raw_connection.close()

    # Passing the viewing rights
    # ==========================================================
    # CHECK WHETHER USER CAN ACCESS BI REPORTING
    # ==========================================================

    has_any_bi_view_rights = bool(
        db_session.execute(
            text(
                """
                SELECT EXISTS
                (
                    SELECT 1
                    FROM public.users_bi_config_rights
                    WHERE user_id = :user_id
                    AND lower(
                            trim(
                                has_bi_view_rights
                            )
                        ) = 'yes'
                )
                """
            ),
            {
                "user_id": user_id,
            },
        ).scalar()
    )

    return render_template(
        "power_bi_landing_home_view_bi_report.html",
        role=role,
        **report_context,
             # Passing the viewing rights
         has_any_bi_view_rights=has_any_bi_view_rights,
    )






if __name__ == '__main__':
    port = int(os.getenv("PORT", 5010))  
    # PRODUCTION: Set debug=False in production
    debug_mode = os.getenv("DEBUG", "False").lower() == "true"
    app.run(host='0.0.0.0', port=port, debug=debug_mode)
