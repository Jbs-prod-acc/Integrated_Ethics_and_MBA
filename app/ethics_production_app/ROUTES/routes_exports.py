from app_support import *

@app.route('/export_forms_csv', methods=['GET', 'POST'])
def export_forms_csv():
    """
    Export specific FormA, FormB, and FormC data to an Excel file with formatting
    Only accessible to ADMIN and SUPER_ADMIN roles
    Groups by user_id with alternating colors and styled headers
    Filters by selected year if provided
    """
    print("=== EXPORT FUNCTION CALLED ===")
    
    # Check if user is logged in and is admin/super_admin
    user_id = session.get('id')
    print(f"User ID from session: {user_id}")
    
    if not user_id:
        flash('Please login to access this feature', 'error')
        return redirect(url_for('login_page'))
    
    user = db_session.query(User).filter(User.user_id == user_id).first()
    print(f"User found: {user.full_name if user else 'None'}")
    print(f"User role: {user.role.value if user and user.role else 'None'}")
    
    if not user or not user.role or user.role.value.upper() not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You do not have permission to export forms', 'error')
        return redirect(url_for('chair_landing'))
    
    # Get selected year from form (if provided)
    selected_year = request.form.get('year') if request.method == 'POST' else request.args.get('year')
    print(f"Selected year: {selected_year}")
    
    try:
        print("Starting export process...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        print("openpyxl imported successfully")
        
        # Define the fields to extract
        fields_mapping = {
            'form_type': 'Form Type',
            'applicant_name': 'Applicant Name',
            'student_number': 'Student Number',
            'email': 'Email',
            'supervisor': 'Supervisor',
            'supervisor_email': 'Supervisor Email',
            'student_submission': 'Student Submission Date',
            'recommendation': 'Supervisors Recommendation',
            'supervisor_date': 'Supervisor Recommendation Date',
            'ethics_status': 'Ethics Admin Decision',
            'signature_date': 'Ethics Admin Decision Date',
            'review_supervisor_signature': 'First Reviewer Name',
            'review_recommendation': 'First Reviewer Recommendation',
            'review_date': 'First Reviewer Feedback Date',
            'review_supervisor_signature1': 'Second Reviewer Name',
            'review_recommendation1': 'Second Reviewer Recommendation',
            'review_date1': 'Second Reviewer Feedback Date'
        }
        
        # Query all submitted forms ordered by submission date
        print("Querying forms...")
        
        # FormA query with optional year filter
        form_a_query = db_session.query(FormA).filter(FormA.submitted_at.isnot(None))
        if selected_year and selected_year != 'all':
            form_a_query = form_a_query.filter(extract('year', FormA.submitted_at) == int(selected_year))
        form_a_records = form_a_query.order_by(FormA.submitted_at.desc()).all()
        
        # FormB query with optional year filter - only load needed columns to avoid binary column deserialization issues
        form_b_query = db_session.query(
            FormB.user_id,
            FormB.applicant_name,
            FormB.student_number,
            FormB.email,
            FormB.supervisor,
            FormB.supervisor_email,
            FormB.submitted_at,
            FormB.recommendation,
            FormB.supervisor_date,
            FormB.ethics_status,
            FormB.signature_date,
            FormB.review_supervisor_signature,
            FormB.review_recommendation,
            FormB.review_date,
            FormB.review_supervisor_signature1,
            FormB.review_recommendation1,
            FormB.review_date1
        ).filter(FormB.submitted_at.isnot(None))
        if selected_year and selected_year != 'all':
            form_b_query = form_b_query.filter(extract('year', FormB.submitted_at) == int(selected_year))
        form_b_results = form_b_query.order_by(FormB.submitted_at.desc()).all()
        
        # Convert tuples to objects with attributes for consistent processing
        form_b_records = []
        for result in form_b_results:
            class FormBProxy:
                pass
            proxy = FormBProxy()
            proxy.user_id = result.user_id
            proxy.applicant_name = result.applicant_name
            proxy.student_number = result.student_number
            proxy.email = result.email
            proxy.supervisor = result.supervisor
            proxy.supervisor_email = result.supervisor_email
            proxy.submitted_at = result.submitted_at
            proxy.recommendation = result.recommendation
            proxy.supervisor_date = result.supervisor_date
            proxy.ethics_status = result.ethics_status
            proxy.signature_date = result.signature_date
            proxy.review_supervisor_signature = result.review_supervisor_signature
            proxy.review_recommendation = result.review_recommendation
            proxy.review_date = result.review_date
            proxy.review_supervisor_signature1 = result.review_supervisor_signature1
            proxy.review_recommendation1 = result.review_recommendation1
            proxy.review_date1 = result.review_date1
            form_b_records.append(proxy)
        
        # FormC query with optional year filter
        form_c_query = db_session.query(FormC).filter(FormC.submission_date.isnot(None))
        if selected_year and selected_year != 'all':
            form_c_query = form_c_query.filter(extract('year', FormC.submission_date) == int(selected_year))
        form_c_records = form_c_query.order_by(FormC.submission_date.desc()).all()
        
       
        
        # Helper function to remove timezone from datetime
        def remove_tz(dt):
            """Remove timezone info from datetime object for Excel compatibility"""
            if dt and hasattr(dt, 'replace'):
                return dt.replace(tzinfo=None)
            return dt
        
        # Collect data
        all_data = []
        
        for record in form_a_records:
            all_data.append({
                'form_type': 'Form A',
                'user_id': record.user_id or '',
                'applicant_name': record.applicant_name or '',
                'student_number': record.student_number or '',
                'email': record.email or '',
                'supervisor': record.supervisor or '',
                'supervisor_email': record.supervisor_email or '',
                'student_submission': remove_tz(record.submitted_at) if record.submitted_at else '',
                'recommendation': record.recommendation or '',
                'supervisor_date': remove_tz(record.supervisor_date) if record.supervisor_date else '',
                'ethics_status': record.ethics_status or '',
                'signature_date': remove_tz(record.signature_date) if record.signature_date else '',
                'review_supervisor_signature': record.review_supervisor_signature or '',
                'review_recommendation': record.review_recommendation or '',
                'review_date': record.review_date or '',
                'review_supervisor_signature1': record.review_supervisor_signature1 or '',
                'review_recommendation1': record.review_recommendation1 or '',
                'review_date1': record.review_date1 or ''
            })
        
        for record in form_b_records:
            all_data.append({
                'form_type': 'Form B',
                'user_id': record.user_id or '',
                'applicant_name': record.applicant_name or '',
                'student_number': record.student_number or '',
                'email': record.email or '',
                'supervisor': record.supervisor or '',
                'supervisor_email': record.supervisor_email or '',
                'student_submission': remove_tz(record.submitted_at) if record.submitted_at else '',
                'recommendation': record.recommendation or '',
                'supervisor_date': remove_tz(record.supervisor_date) if record.supervisor_date else '',
                'ethics_status': record.ethics_status or '',
                'signature_date': remove_tz(record.signature_date) if record.signature_date else '',
                'review_supervisor_signature': record.review_supervisor_signature or '',
                'review_recommendation': record.review_recommendation or '',
                'review_date': record.review_date or '',
                'review_supervisor_signature1': record.review_supervisor_signature1 or '',
                'review_recommendation1': record.review_recommendation1 or '',
                'review_date1': record.review_date1 or ''
            })
        
        for record in form_c_records:
            all_data.append({
                'form_type': 'Form C',
                'user_id': record.user_id or '',
                'applicant_name': record.applicant_name or '',
                'student_number': record.student_number or '',
                'email': record.email_address or '',
                'supervisor': record.supervisor_name or '',
                'supervisor_email': record.supervisor_email or '',
                'student_submission': remove_tz(record.submission_date) if record.submission_date else '',
                'recommendation': record.recommendation or '',
                'supervisor_date': remove_tz(record.supervisor_date) if record.supervisor_date else '',
                'ethics_status': record.ethics_status or '',
                'signature_date': remove_tz(record.signature_date) if record.signature_date else '',
                'review_supervisor_signature': record.review_supervisor_signature or '',
                'review_recommendation': record.review_recommendation or '',
                'review_date': record.review_date or '',
                'review_supervisor_signature1': record.review_supervisor_signature1 or '',
                'review_recommendation1': record.review_recommendation1 or '',
                'review_date1': record.review_date1 or ''
            })
        
        print(f"Collected {len(all_data)} total records")
        
        # Create DataFrame
        print("Creating DataFrame...")
        df = pd.DataFrame(all_data)
        if df.empty:
            print("DataFrame is empty!")
            flash('No forms found to export', 'warning')
            return redirect(url_for('chair_landing'))
        
        print(f"DataFrame created with {len(df)} rows")
        
        # Group by user_id and sort by submission date within each group
        print("Grouping by user_id...")
        df = df.sort_values(by=['user_id', 'student_submission'], ascending=[True, False], na_position='last')
        
        print(f"Data grouped by user_id, total groups: {df['user_id'].nunique()}")
        
        # Rename columns (but keep user_id for grouping logic)
        df = df.rename(columns=fields_mapping)
        
        print("Creating Excel workbook...")
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Forms Export"
        
        # Write headers with styling
        headers = list(fields_mapping.values())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Define alternating colors for different user groups
        colors = [
            "E8F4F8",  # Light blue
            "FFF4E6",  # Light orange
            "F0F8E8",  # Light green
            "F8E8F4",  # Light purple
            "FFF0F0",  # Light red
            "F0F0F0",  # Light gray
        ]
        
        # Write data rows with alternating colors per user_id
        current_user_id = None
        color_index = 0
        row_num = 2
        
        for index, row in df.iterrows():
            # Check if we're starting a new user group
            if row['user_id'] != current_user_id:
                current_user_id = row['user_id']
                color_index = (color_index + 1) % len(colors)
            
            # Get the fill color for this user group
            fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
            
            # Write the row data (excluding the user_id column)
            col_num = 1
            for col_name, value in row.items():
                if col_name != 'user_id':  # Skip user_id (don't expose primary key)
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    col_num += 1
            
            row_num += 1
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        print("Saving to BytesIO...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'ethics_forms_export_{timestamp}.xlsx'
        
        print(f"Export complete! Filename: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError as e:
        print(f"Import error - openpyxl not installed: {str(e)}")
        traceback.print_exc()
        flash('Excel export library not installed. Please contact administrator.', 'error')
        return redirect(url_for('chair_landing'))
    except Exception as e:
        print(f"Error exporting forms: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        traceback.print_exc()
        flash(f'Error exporting forms: {str(e)}', 'error')
        return redirect(url_for('chair_landing'))


# Export agenda with reviewer counts per student
@app.route('/export_agenda_csv', methods=['GET', 'POST'])
def export_agenda_csv():
    """
    Export Excel with columns: Student Name, Reviewer 1, Reviewer 2, Count by Reviewer 1, Count by Reviewer 2
    For all FormA, FormB, FormC submissions with reviewer(s).
    """
    import collections
    user_id = session.get('id')
    if not user_id:
        flash('Please login to access this feature', 'error')
        return redirect(url_for('login_page'))
    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user or not user.role or user.role.value.upper() not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You do not have permission to export agenda', 'error')
        return redirect(url_for('chair_landing'))

    # Use only the latest submitted version per student/form type.
    latest_form_a_subq = (
        db_session.query(
            FormA.user_id,
            func.max(FormA.submitted_at).label('latest_submitted_at')
        )
        .filter(FormA.submitted_at.isnot(None))
        .group_by(FormA.user_id)
        .subquery()
    )
    latest_form_b_subq = (
        db_session.query(
            FormB.user_id,
            func.max(FormB.submitted_at).label('latest_submitted_at')
        )
        .filter(FormB.submitted_at.isnot(None))
        .group_by(FormB.user_id)
        .subquery()
    )
    latest_form_c_subq = (
        db_session.query(
            FormC.user_id,
            func.max(FormC.submission_date).label('latest_submission_date')
        )
        .filter(FormC.submission_date.isnot(None))
        .group_by(FormC.user_id)
        .subquery()
    )

    form_a_records = (
        db_session.query(FormA)
        .join(
            latest_form_a_subq,
            and_(
                FormA.user_id == latest_form_a_subq.c.user_id,
                FormA.submitted_at == latest_form_a_subq.c.latest_submitted_at
            )
        )
        .filter(FormA.certificate_issued.is_(None))
        .all()
    )
    form_b_records = (
        db_session.query(FormB)
        .join(
            latest_form_b_subq,
            and_(
                FormB.user_id == latest_form_b_subq.c.user_id,
                FormB.submitted_at == latest_form_b_subq.c.latest_submitted_at
            )
        )
        .filter(FormB.certificate_issued.is_(None))
        .all()
    )
    form_c_records = (
        db_session.query(FormC)
        .join(
            latest_form_c_subq,
            and_(
                FormC.user_id == latest_form_c_subq.c.user_id,
                FormC.submission_date == latest_form_c_subq.c.latest_submission_date
            )
        )
        .filter(FormC.certificate_issued.is_(None))
        .all()
    )


    # Group by student name, aggregate reviewers and recommendations
    student_rows = {}

    def process_form(record, name_field):
        student_name = getattr(record, name_field, None) or ''
        reviewer1_id = getattr(record, 'reviewer_name1', None) or ''
        reviewer2_id = getattr(record, 'reviewer_name2', None) or ''
        reviewer1 = ''
        reviewer2 = ''
        reviewer1_recommendation = getattr(record, 'review_recommendation', None) or ''
        reviewer2_recommendation = getattr(record, 'review_recommendation1', None) or ''
        # Lookup reviewer names from user_id (users table)
        if reviewer1_id:
            reviewer1_user = db_session.query(User).filter(User.user_id == reviewer1_id).first()
            reviewer1 = reviewer1_user.full_name if reviewer1_user else reviewer1_id
        if reviewer2_id:
            reviewer2_user = db_session.query(User).filter(User.user_id == reviewer2_id).first()
            reviewer2 = reviewer2_user.full_name if reviewer2_user else reviewer2_id
        # Group by student name
        if student_name not in student_rows:
            student_rows[student_name] = {
                'Student Name': student_name,
                'Reviewer 1': reviewer1,
                'Reviewer 1 Recommendation': reviewer1_recommendation,
                'Reviewer 2': reviewer2,
                'Reviewer 2 Recommendation': reviewer2_recommendation,
                'Count by Reviewer 1': 1 if reviewer1_recommendation else 0,
                'Count by Reviewer 2': 1 if reviewer2_recommendation else 0,
            }
        else:
            # If already exists, fill in missing reviewer/recommendation if empty
            row = student_rows[student_name]
            if not row['Reviewer 1'] and reviewer1:
                row['Reviewer 1'] = reviewer1
                row['Reviewer 1 Recommendation'] = reviewer1_recommendation
                row['Count by Reviewer 1'] = 1 if reviewer1_recommendation else 0
            if not row['Reviewer 2'] and reviewer2:
                row['Reviewer 2'] = reviewer2
                row['Reviewer 2 Recommendation'] = reviewer2_recommendation
                row['Count by Reviewer 2'] = 1 if reviewer2_recommendation else 0

    # FormA
    for rec in form_a_records:
        process_form(rec, 'applicant_name')
    # FormB
    for rec in form_b_records:
        process_form(rec, 'applicant_name')
    # FormC
    for rec in form_c_records:
        process_form(rec, 'applicant_name')

    rows = list(student_rows.values())

    if not rows:
        flash('No agenda data found to export', 'warning')
        return redirect(url_for('chair_landing'))

    df = pd.DataFrame(rows)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda Export"

    # Write headers
    headers = list(df.columns)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ethics_agenda_export_{timestamp}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/export_reviewer_assignments_csv', methods=['GET', 'POST'])
def export_reviewer_assignments_csv():
    """
    Export reviewer assignments to an Excel workbook.
    Sheet 1 lists all reviewers with total assignment counts.
    Sheet 2 lists every reviewer-to-student assignment.
    Uses alternating colors per reviewer, similar to the forms export.
    """
    user_id = session.get('id')
    if not user_id:
        flash('Please login to access this feature', 'error')
        return redirect(url_for('login_page'))

    user = db_session.query(User).filter(User.user_id == user_id).first()
    if not user or not user.role or user.role.value.upper() not in ['ADMIN', 'SUPER_ADMIN']:
        flash('You do not have permission to export reviewer assignments', 'error')
        return redirect(url_for('chair_landing'))

    selected_year = request.form.get('year') if request.method == 'POST' else request.args.get('year')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        def remove_tz(dt):
            if dt and hasattr(dt, 'replace'):
                return dt.replace(tzinfo=None)
            return dt

        # Pull all reviewers so the summary sheet includes reviewers with zero assignments.
        reviewers = (
            db_session.query(User)
            .filter(User.role == UserRole.REVIEWER)
            .order_by(User.full_name.asc())
            .all()
        )
        reviewer_lookup = {str(reviewer.user_id): reviewer for reviewer in reviewers}

        assignment_rows = []
        assignment_counts = {str(reviewer.user_id): 0 for reviewer in reviewers}
        reviewer_students = {str(reviewer.user_id): [] for reviewer in reviewers}

        def process_records(records, form_type, name_attr, student_number_attr, email_attr, submitted_attr):
            for record in records:
                if getattr(record, 'certificate_issued', None):
                    continue

                student_name = getattr(record, name_attr, '') or ''
                student_number = getattr(record, student_number_attr, '') or ''
                student_email = getattr(record, email_attr, '') or ''
                submitted_date = remove_tz(getattr(record, submitted_attr, None))

                for slot_index, reviewer_attr in enumerate(['reviewer_name1', 'reviewer_name2'], start=1):
                    reviewer_id = getattr(record, reviewer_attr, None)
                    if not reviewer_id:
                        continue

                    reviewer_key = str(reviewer_id)
                    reviewer_user = reviewer_lookup.get(reviewer_key)
                    reviewer_name = reviewer_user.full_name if reviewer_user else str(reviewer_id)
                    reviewer_email = reviewer_user.email if reviewer_user else ''

                    assignment_counts[reviewer_key] = assignment_counts.get(reviewer_key, 0) + 1
                    if student_name:
                        reviewer_students.setdefault(reviewer_key, []).append(student_name)
                    assignment_rows.append({
                        'reviewer_id': reviewer_key,
                        'Reviewer Name': reviewer_name,
                        'Reviewer Email': reviewer_email,
                        'Form Type': form_type,
                        'Student Name': student_name,
                        'Student Number': student_number,
                        'Student Email': student_email,
                        'Assignment Slot': f'Reviewer {slot_index}',
                        'Submission Date': submitted_date or '',
                    })

        latest_form_a_subq = (
            db_session.query(
                FormA.user_id,
                func.max(FormA.submitted_at).label('latest_submitted_at')
            )
            .filter(FormA.submitted_at.isnot(None))
            .group_by(FormA.user_id)
            .subquery()
        )
        form_a_query = (
            db_session.query(FormA)
            .join(
                latest_form_a_subq,
                and_(
                    FormA.user_id == latest_form_a_subq.c.user_id,
                    FormA.submitted_at == latest_form_a_subq.c.latest_submitted_at
                )
            )
            .filter(FormA.submitted_at.isnot(None))
        )
        if selected_year and selected_year != 'all':
            form_a_query = form_a_query.filter(extract('year', FormA.submitted_at) == int(selected_year))
        form_a_records = form_a_query.order_by(FormA.submitted_at.desc()).all()

        latest_form_b_subq = (
            db_session.query(
                FormB.user_id,
                func.max(FormB.submitted_at).label('latest_submitted_at')
            )
            .filter(FormB.submitted_at.isnot(None))
            .group_by(FormB.user_id)
            .subquery()
        )
        form_b_query = (
            db_session.query(
                FormB.reviewer_name1,
                FormB.reviewer_name2,
                FormB.applicant_name,
                FormB.student_number,
                FormB.email,
                FormB.submitted_at,
                FormB.certificate_issued
            )
            .join(
                latest_form_b_subq,
                and_(
                    FormB.user_id == latest_form_b_subq.c.user_id,
                    FormB.submitted_at == latest_form_b_subq.c.latest_submitted_at
                )
            )
            .filter(FormB.submitted_at.isnot(None))
        )
        if selected_year and selected_year != 'all':
            form_b_query = form_b_query.filter(extract('year', FormB.submitted_at) == int(selected_year))
        form_b_records = form_b_query.order_by(FormB.submitted_at.desc()).all()

        latest_form_c_subq = (
            db_session.query(
                FormC.user_id,
                func.max(FormC.submission_date).label('latest_submission_date')
            )
            .filter(FormC.submission_date.isnot(None))
            .group_by(FormC.user_id)
            .subquery()
        )
        form_c_query = (
            db_session.query(
                FormC.reviewer_name1,
                FormC.reviewer_name2,
                FormC.applicant_name,
                FormC.student_number,
                FormC.email_address,
                FormC.submission_date,
                FormC.certificate_issued
            )
            .join(
                latest_form_c_subq,
                and_(
                    FormC.user_id == latest_form_c_subq.c.user_id,
                    FormC.submission_date == latest_form_c_subq.c.latest_submission_date
                )
            )
            .filter(FormC.submission_date.isnot(None))
        )
        if selected_year and selected_year != 'all':
            form_c_query = form_c_query.filter(extract('year', FormC.submission_date) == int(selected_year))
        form_c_records = form_c_query.order_by(FormC.submission_date.desc()).all()

        process_records(form_a_records, 'Form A', 'applicant_name', 'student_number', 'email', 'submitted_at')
        process_records(form_b_records, 'Form B', 'applicant_name', 'student_number', 'email', 'submitted_at')
        process_records(form_c_records, 'Form C', 'applicant_name', 'student_number', 'email_address', 'submission_date')

        summary_rows = []
        for reviewer in reviewers:
            reviewer_key = str(reviewer.user_id)
            unique_students = sorted(set(reviewer_students.get(reviewer_key, [])))
            summary_rows.append({
                'reviewer_id': reviewer_key,
                'Reviewer Name': reviewer.full_name or '',
                'Reviewer Email': reviewer.email or '',
                'Total Assignments': len(unique_students),
                'Assigned Students': ', '.join(unique_students),
            })

        if not summary_rows:
            flash('No reviewer data found to export', 'warning')
            return redirect(url_for('chair_landing'))

        summary_df = pd.DataFrame(summary_rows).sort_values(
            by=['Reviewer Name'],
            ascending=[True],
            na_position='last'
        )
        assignments_df = pd.DataFrame(assignment_rows)
        if not assignments_df.empty:
            assignments_df = assignments_df.sort_values(
                by=['Reviewer Name', 'Submission Date', 'Student Name'],
                ascending=[True, False, True],
                na_position='last'
            )

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Reviewer Summary"
        detail_ws = wb.create_sheet(title="Assignments")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        colors = [
            "E8F4F8",
            "FFF4E6",
            "F0F8E8",
            "F8E8F4",
            "FFF0F0",
            "F0F0F0",
        ]

        reviewer_ids_in_order = summary_df['reviewer_id'].tolist()
        reviewer_colors = {
            reviewer_id: colors[index % len(colors)]
            for index, reviewer_id in enumerate(reviewer_ids_in_order)
        }

        def write_sheet(ws, df, visible_columns):
            for col_num, header in enumerate(visible_columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_num, (_, row) in enumerate(df.iterrows(), start=2):
                reviewer_id = row.get('reviewer_id', '')
                fill_color = reviewer_colors.get(reviewer_id, colors[0])
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                for col_num, column_name in enumerate(visible_columns, start=1):
                    value = row.get(column_name, '')
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = fill
                    cell.alignment = body_alignment

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        write_sheet(
            summary_ws,
            summary_df,
            ['Reviewer Name', 'Reviewer Email', 'Total Assignments', 'Assigned Students']
        )

        if assignments_df.empty:
            assignments_df = pd.DataFrame([{
                'reviewer_id': '',
                'Reviewer Name': 'No assignments found',
                'Reviewer Email': '',
                'Total Assignments': '',
                'Form Type': '',
                'Student Name': '',
                'Student Number': '',
                'Student Email': '',
                'Assignment Slot': '',
                'Submission Date': '',
            }])

        write_sheet(
            detail_ws,
            assignments_df,
            [
                'Reviewer Name',
                'Reviewer Email',
                'Form Type',
                'Student Name',
                'Student Number',
                'Student Email',
                'Assignment Slot',
                'Submission Date',
            ]
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'reviewer_assignments_export_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError as e:
        print(f"Import error - openpyxl not installed: {str(e)}")
        traceback.print_exc()
        flash('Excel export library not installed. Please contact administrator.', 'error')
        return redirect(url_for('chair_landing'))
    except Exception as e:
        print(f"Error exporting reviewer assignments: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        traceback.print_exc()
        flash(f'Error exporting reviewer assignments: {str(e)}', 'error')
        return redirect(url_for('chair_landing'))

